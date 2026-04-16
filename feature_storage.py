#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
Feature Storage — v2.5

Writes and reads v2.5 spectral evolution features into hidden Excel sheets.
Follows the same conventions as the v2.4 hidden sheets in mix_analyzer.py:
- openpyxl workbook
- _hidden_sheet_style() aesthetic
- sheet_state = 'hidden'
- Floats rounded to 0.1 dB for compactness.
"""

from __future__ import annotations

from typing import Dict, List, Optional, Tuple, TYPE_CHECKING

import numpy as np

if TYPE_CHECKING:
    from openpyxl import Workbook

from spectral_evolution import (
    TrackFeatures,
    ZoneEnergy,
    SpectralDescriptors,
    PeakTrajectory,
    TransientEvent,
    ZONE_RANGES,
    ZONE_LABELS,
)


# ---------------------------------------------------------------------------
# Style helper (mirrors mix_analyzer._hidden_sheet_style)
# ---------------------------------------------------------------------------

def _v25_sheet_style():
    """Return style objects matching existing hidden sheet conventions."""
    from openpyxl.styles import Font, PatternFill, Border, Side
    return {
        'bg_fill': PatternFill('solid', fgColor='0A0A12'),
        'header_fill': PatternFill('solid', fgColor='1A3A5A'),
        'header_font': Font(name='Calibri', size=9, bold=True, color='E8E8F0'),
        'data_font': Font(name='Calibri', size=9, color='C0C0D0'),
        'dim_font': Font(name='Calibri', size=8, color='888888'),
        'border': Border(
            left=Side(style='thin', color='333344'),
            right=Side(style='thin', color='333344'),
            top=Side(style='thin', color='333344'),
            bottom=Side(style='thin', color='333344'),
        ),
    }


def _styled_cell(ws, row, col, value, font, fill, border):
    """Write a styled cell and return it."""
    c = ws.cell(row=row, column=col, value=value)
    c.font = font
    c.fill = fill
    c.border = border
    return c


# ---------------------------------------------------------------------------
# Downsample helpers
# ---------------------------------------------------------------------------

def _downsample_frames(values: np.ndarray, n_buckets: int = 32) -> List[Optional[float]]:
    """Downsample a 1-D time-series into n_buckets evenly-spaced buckets.

    Args:
        values: 1-D array.
        n_buckets: Number of output buckets.

    Returns:
        List of n_buckets floats (rounded to 0.1) or None.
    """
    values = np.asarray(values, dtype=np.float64)
    if len(values) == 0:
        return [None] * n_buckets
    if len(values) <= n_buckets:
        out = [round(float(v), 1) for v in values]
        out.extend([None] * (n_buckets - len(out)))
        return out
    edges = np.linspace(0, len(values), n_buckets + 1, dtype=int)
    result = []
    for i in range(n_buckets):
        chunk = values[edges[i]:edges[i + 1]]
        if len(chunk) > 0:
            result.append(round(float(np.mean(chunk)), 1))
        else:
            result.append(None)
    return result


def _time_bucket_labels(duration: float, n_buckets: int = 32) -> List[str]:
    """Generate readable time bucket labels."""
    if duration <= 0:
        return [f'T{i + 1}' for i in range(n_buckets)]
    step = duration / n_buckets
    labels = []
    for i in range(n_buckets):
        t_start = i * step
        t_end = (i + 1) * step
        labels.append(f'{t_start:.1f}-{t_end:.1f}s')
    return labels


# ---------------------------------------------------------------------------
# Sheet builders — Phase 1
# ---------------------------------------------------------------------------

def build_v25_zone_energy_sheet(wb: 'Workbook',
                                 features_with_info: List[Tuple[TrackFeatures, dict]],
                                 log_fn=None,
                                 n_buckets: int = 32) -> None:
    """Build hidden _track_zone_energy sheet.

    Layout: stacked blocks — 1 header row + 9 zone rows per track × n_buckets time columns.

    Args:
        wb: openpyxl Workbook.
        features_with_info: List of (TrackFeatures, track_info) tuples.
        log_fn: Logging function.
        n_buckets: Time buckets for downsampling.
    """
    if log_fn is None:
        log_fn = lambda msg: None
    log_fn("    Excel: writing _track_zone_energy sheet (v2.5)...")

    ws = wb.create_sheet('_track_zone_energy')
    ws.sheet_state = 'hidden'
    sty = _v25_sheet_style()

    zone_names = list(ZONE_RANGES.keys())
    row = 1

    for feat, ti in features_with_info:
        track_label = ti.get('name', 'Unknown')
        if ti.get('type') == 'Full Mix':
            track_label = '*** FULL MIX ***'
        duration = float(feat.zone_energy.times[-1]) if len(feat.zone_energy.times) > 0 else 0.0
        time_labels = _time_bucket_labels(duration, n_buckets)

        # Header row: track name + time bucket labels
        _styled_cell(ws, row, 1, track_label, sty['header_font'], sty['header_fill'], sty['border'])
        for ci, tl in enumerate(time_labels, 2):
            _styled_cell(ws, row, ci, tl, sty['dim_font'], sty['header_fill'], sty['border'])
        row += 1

        # One row per zone
        for zone_name in zone_names:
            zone_label = ZONE_LABELS.get(zone_name, zone_name)
            _styled_cell(ws, row, 1, zone_label, sty['dim_font'], sty['bg_fill'], sty['border'])

            zone_data = feat.zone_energy.zones.get(zone_name, np.array([]))
            buckets = _downsample_frames(zone_data, n_buckets)
            for ci, val in enumerate(buckets, 2):
                _styled_cell(ws, row, ci, val, sty['data_font'], sty['bg_fill'], sty['border'])
            row += 1

    log_fn(f"    Excel: _track_zone_energy done — {row - 1} rows x {n_buckets} time buckets")


def build_v25_spectral_descriptors_sheet(wb: 'Workbook',
                                          features_with_info: List[Tuple[TrackFeatures, dict]],
                                          log_fn=None,
                                          n_buckets: int = 32) -> None:
    """Build hidden _track_spectral_descriptors sheet.

    Layout: stacked blocks — 1 header row + 5 descriptor rows per track × n_buckets time columns.
    Descriptors: centroid, spread, flatness, low_rolloff, high_rolloff.

    Args:
        wb: openpyxl Workbook.
        features_with_info: List of (TrackFeatures, track_info) tuples.
        log_fn: Logging function.
        n_buckets: Time buckets for downsampling.
    """
    if log_fn is None:
        log_fn = lambda msg: None
    log_fn("    Excel: writing _track_spectral_descriptors sheet (v2.5)...")

    ws = wb.create_sheet('_track_spectral_descriptors')
    ws.sheet_state = 'hidden'
    sty = _v25_sheet_style()

    descriptor_names = [
        ('centroid', 'Centroid (Hz)'),
        ('spread', 'Spread (Hz)'),
        ('flatness', 'Flatness (0–1)'),
        ('low_rolloff', 'Low Rolloff (Hz)'),
        ('high_rolloff', 'High Rolloff (Hz)'),
    ]

    row = 1

    for feat, ti in features_with_info:
        track_label = ti.get('name', 'Unknown')
        if ti.get('type') == 'Full Mix':
            track_label = '*** FULL MIX ***'
        duration = float(feat.descriptors.times[-1]) if len(feat.descriptors.times) > 0 else 0.0
        time_labels = _time_bucket_labels(duration, n_buckets)

        # Header row
        _styled_cell(ws, row, 1, track_label, sty['header_font'], sty['header_fill'], sty['border'])
        for ci, tl in enumerate(time_labels, 2):
            _styled_cell(ws, row, ci, tl, sty['dim_font'], sty['header_fill'], sty['border'])
        row += 1

        # One row per descriptor
        for attr_name, label in descriptor_names:
            _styled_cell(ws, row, 1, label, sty['dim_font'], sty['bg_fill'], sty['border'])
            data = getattr(feat.descriptors, attr_name, np.array([]))
            buckets = _downsample_frames(data, n_buckets)
            for ci, val in enumerate(buckets, 2):
                _styled_cell(ws, row, ci, val, sty['data_font'], sty['bg_fill'], sty['border'])
            row += 1

    log_fn(f"    Excel: _track_spectral_descriptors done — {row - 1} rows")


# ---------------------------------------------------------------------------
# Sheet builders — Phase 2
# ---------------------------------------------------------------------------

def build_v25_peak_trajectories_sheet(wb: 'Workbook',
                                       features_with_info: List[Tuple[TrackFeatures, dict]],
                                       log_fn=None) -> None:
    """Build hidden _track_peak_trajectories sheet.

    Layout: one block per track — header row then one row per trajectory point.
    Columns: Track, Trajectory#, Frame, Time(s), Freq(Hz), Amplitude(dB).

    Args:
        wb: openpyxl Workbook.
        features_with_info: List of (TrackFeatures, track_info) tuples.
        log_fn: Logging function.
    """
    if log_fn is None:
        log_fn = lambda msg: None
    log_fn("    Excel: writing _track_peak_trajectories sheet (v2.5)...")

    ws = wb.create_sheet('_track_peak_trajectories')
    ws.sheet_state = 'hidden'
    sty = _v25_sheet_style()

    headers = ['Track', 'Traj#', 'Frame', 'Time (s)', 'Freq (Hz)', 'Amp (dB)']
    row = 1
    for ci, h in enumerate(headers, 1):
        _styled_cell(ws, row, ci, h, sty['header_font'], sty['header_fill'], sty['border'])
    row += 1

    for feat, ti in features_with_info:
        track_label = ti.get('name', 'Unknown')
        trajectories = feat.peak_trajectories or []
        for traj_idx, traj in enumerate(trajectories, 1):
            for frame_idx, freq, amp in traj.points:
                _styled_cell(ws, row, 1, track_label, sty['dim_font'], sty['bg_fill'], sty['border'])
                _styled_cell(ws, row, 2, traj_idx, sty['data_font'], sty['bg_fill'], sty['border'])
                _styled_cell(ws, row, 3, frame_idx, sty['data_font'], sty['bg_fill'], sty['border'])
                time_sec = round(frame_idx * (feat.zone_energy.times[1] - feat.zone_energy.times[0]), 3) \
                    if len(feat.zone_energy.times) > 1 else 0.0
                _styled_cell(ws, row, 4, time_sec, sty['data_font'], sty['bg_fill'], sty['border'])
                _styled_cell(ws, row, 5, round(freq, 1), sty['data_font'], sty['bg_fill'], sty['border'])
                _styled_cell(ws, row, 6, round(amp, 1), sty['data_font'], sty['bg_fill'], sty['border'])
                row += 1

    log_fn(f"    Excel: _track_peak_trajectories done — {row - 1} rows")


def build_v25_valley_trajectories_sheet(wb: 'Workbook',
                                         features_with_info: List[Tuple[TrackFeatures, dict]],
                                         log_fn=None) -> None:
    """Build hidden _track_valley_trajectories sheet.

    Same layout as peak trajectories.

    Args:
        wb: openpyxl Workbook.
        features_with_info: List of (TrackFeatures, track_info) tuples.
        log_fn: Logging function.
    """
    if log_fn is None:
        log_fn = lambda msg: None
    log_fn("    Excel: writing _track_valley_trajectories sheet (v2.5)...")

    ws = wb.create_sheet('_track_valley_trajectories')
    ws.sheet_state = 'hidden'
    sty = _v25_sheet_style()

    headers = ['Track', 'Valley#', 'Frame', 'Time (s)', 'Freq (Hz)', 'Amp (dB)']
    row = 1
    for ci, h in enumerate(headers, 1):
        _styled_cell(ws, row, ci, h, sty['header_font'], sty['header_fill'], sty['border'])
    row += 1

    for feat, ti in features_with_info:
        track_label = ti.get('name', 'Unknown')
        trajectories = feat.valley_trajectories or []
        for traj_idx, traj in enumerate(trajectories, 1):
            for frame_idx, freq, amp in traj.points:
                _styled_cell(ws, row, 1, track_label, sty['dim_font'], sty['bg_fill'], sty['border'])
                _styled_cell(ws, row, 2, traj_idx, sty['data_font'], sty['bg_fill'], sty['border'])
                _styled_cell(ws, row, 3, frame_idx, sty['data_font'], sty['bg_fill'], sty['border'])
                time_sec = round(frame_idx * (feat.zone_energy.times[1] - feat.zone_energy.times[0]), 3) \
                    if len(feat.zone_energy.times) > 1 else 0.0
                _styled_cell(ws, row, 4, time_sec, sty['data_font'], sty['bg_fill'], sty['border'])
                _styled_cell(ws, row, 5, round(freq, 1), sty['data_font'], sty['bg_fill'], sty['border'])
                _styled_cell(ws, row, 6, round(amp, 1), sty['data_font'], sty['bg_fill'], sty['border'])
                row += 1

    log_fn(f"    Excel: _track_valley_trajectories done — {row - 1} rows")


def build_v25_crest_by_zone_sheet(wb: 'Workbook',
                                   features_with_info: List[Tuple[TrackFeatures, dict]],
                                   log_fn=None,
                                   n_buckets: int = 32) -> None:
    """Build hidden _track_crest_by_zone sheet.

    Layout: stacked blocks — 1 header row + 9 zone rows per track × n_buckets time columns.

    Args:
        wb: openpyxl Workbook.
        features_with_info: List of (TrackFeatures, track_info) tuples.
        log_fn: Logging function.
        n_buckets: Time buckets for downsampling.
    """
    if log_fn is None:
        log_fn = lambda msg: None
    log_fn("    Excel: writing _track_crest_by_zone sheet (v2.5)...")

    ws = wb.create_sheet('_track_crest_by_zone')
    ws.sheet_state = 'hidden'
    sty = _v25_sheet_style()

    zone_names = list(ZONE_RANGES.keys())
    row = 1

    for feat, ti in features_with_info:
        track_label = ti.get('name', 'Unknown')
        if ti.get('type') == 'Full Mix':
            track_label = '*** FULL MIX ***'
        duration = float(feat.zone_energy.times[-1]) if len(feat.zone_energy.times) > 0 else 0.0
        time_labels = _time_bucket_labels(duration, n_buckets)

        _styled_cell(ws, row, 1, track_label, sty['header_font'], sty['header_fill'], sty['border'])
        for ci, tl in enumerate(time_labels, 2):
            _styled_cell(ws, row, ci, tl, sty['dim_font'], sty['header_fill'], sty['border'])
        row += 1

        crest = feat.crest_by_zone or {}
        for zone_name in zone_names:
            zone_label = ZONE_LABELS.get(zone_name, zone_name)
            _styled_cell(ws, row, 1, zone_label, sty['dim_font'], sty['bg_fill'], sty['border'])
            zone_data = crest.get(zone_name, np.array([]))
            buckets = _downsample_frames(zone_data, n_buckets)
            for ci, val in enumerate(buckets, 2):
                _styled_cell(ws, row, ci, val, sty['data_font'], sty['bg_fill'], sty['border'])
            row += 1

    log_fn(f"    Excel: _track_crest_by_zone done — {row - 1} rows")


def build_v25_transients_sheet(wb: 'Workbook',
                                features_with_info: List[Tuple[TrackFeatures, dict]],
                                log_fn=None) -> None:
    """Build hidden _track_transients sheet.

    Layout: flat table — one row per transient event.
    Columns: Track, Frame, Time(s), Zone, Magnitude(dB).

    Args:
        wb: openpyxl Workbook.
        features_with_info: List of (TrackFeatures, track_info) tuples.
        log_fn: Logging function.
    """
    if log_fn is None:
        log_fn = lambda msg: None
    log_fn("    Excel: writing _track_transients sheet (v2.5)...")

    ws = wb.create_sheet('_track_transients')
    ws.sheet_state = 'hidden'
    sty = _v25_sheet_style()

    headers = ['Track', 'Frame', 'Time (s)', 'Dominant Zone', 'Magnitude (dB)']
    row = 1
    for ci, h in enumerate(headers, 1):
        _styled_cell(ws, row, ci, h, sty['header_font'], sty['header_fill'], sty['border'])
    row += 1

    for feat, ti in features_with_info:
        track_label = ti.get('name', 'Unknown')
        events = feat.transient_events or []
        for ev in events:
            _styled_cell(ws, row, 1, track_label, sty['dim_font'], sty['bg_fill'], sty['border'])
            _styled_cell(ws, row, 2, ev.frame_idx, sty['data_font'], sty['bg_fill'], sty['border'])
            _styled_cell(ws, row, 3, ev.time_sec, sty['data_font'], sty['bg_fill'], sty['border'])
            _styled_cell(ws, row, 4, ev.dominant_zone, sty['data_font'], sty['bg_fill'], sty['border'])
            _styled_cell(ws, row, 5, ev.magnitude_db, sty['data_font'], sty['bg_fill'], sty['border'])
            row += 1

    log_fn(f"    Excel: _track_transients done — {row - 1} rows")


# ---------------------------------------------------------------------------
# Master builder — all v2.5 sheets
# ---------------------------------------------------------------------------

def build_all_v25_sheets(wb: 'Workbook',
                          features_with_info: List[Tuple[TrackFeatures, dict]],
                          log_fn=None,
                          n_buckets: int = 32) -> None:
    """Build all v2.5 hidden sheets at once.

    Args:
        wb: openpyxl Workbook.
        features_with_info: List of (TrackFeatures, track_info) tuples.
        log_fn: Logging function.
        n_buckets: Time buckets for downsampled sheets.
    """
    if log_fn is None:
        log_fn = lambda msg: None
    log_fn("  v2.5: Building spectral evolution hidden sheets...")

    build_v25_zone_energy_sheet(wb, features_with_info, log_fn, n_buckets)
    build_v25_spectral_descriptors_sheet(wb, features_with_info, log_fn, n_buckets)
    build_v25_peak_trajectories_sheet(wb, features_with_info, log_fn)
    build_v25_valley_trajectories_sheet(wb, features_with_info, log_fn)
    build_v25_crest_by_zone_sheet(wb, features_with_info, log_fn, n_buckets)
    build_v25_transients_sheet(wb, features_with_info, log_fn)

    log_fn("  v2.5: All spectral evolution sheets written.")


# ---------------------------------------------------------------------------
# Automation Map sheet (v2.5.1)
# ---------------------------------------------------------------------------

def build_automation_map_sheet(
    wb: 'Workbook',
    automation_maps: dict,
    n_buckets: int = 64,
    duration: float = 0.0,
    log_fn=None,
) -> None:
    """Build hidden _track_automation_map sheet.

    Layout per track:
      - One row per automation curve (Volume, Utility.Gain, Speaker, ...)
      - One row for effective_gain (combined)
      - One row for is_audible (TRUE/FALSE)
    Columns: track_name | param_name | device_name | t0 | t1 | ... | t(n-1)

    Args:
        wb: openpyxl Workbook.
        automation_maps: Dict[str, TrackAutomationMap].
        n_buckets: Number of time buckets.
        duration: Total song duration in seconds (for bucket labels).
        log_fn: Logging function.
    """
    from automation_map import resample_effective_gain, resample_audibility, AUDIBILITY_THRESHOLD

    if log_fn is None:
        log_fn = lambda msg: None
    log_fn("    Excel: writing _track_automation_map sheet (v2.5.1)...")

    ws = wb.create_sheet('_track_automation_map')
    ws.sheet_state = 'hidden'
    sty = _v25_sheet_style()

    # Time bucket labels
    if duration > 0:
        step = duration / n_buckets
        time_labels = [f'{i * step:.1f}-{(i + 1) * step:.1f}s' for i in range(n_buckets)]
    else:
        time_labels = [f'T{i + 1}' for i in range(n_buckets)]

    # Bucket center times for resampling
    if duration > 0:
        step = duration / n_buckets
        bucket_centers = np.array([(i + 0.5) * step for i in range(n_buckets)])
    else:
        bucket_centers = np.linspace(0, 1, n_buckets)

    # Header row
    row = 1
    headers = ['Track', 'Parameter', 'Device'] + time_labels
    for col_idx, header in enumerate(headers, 1):
        _styled_cell(ws, row, col_idx, header,
                     sty['header_font'], sty['header_fill'], sty['border'])
    row += 1

    for track_name, auto_map in automation_maps.items():
        # Individual curves — resample each to bucket resolution
        for curve in auto_map.curves:
            _styled_cell(ws, row, 1, track_name,
                         sty['data_font'], sty['bg_fill'], sty['border'])
            _styled_cell(ws, row, 2, curve.param_name,
                         sty['data_font'], sty['bg_fill'], sty['border'])
            _styled_cell(ws, row, 3, curve.device_name,
                         sty['data_font'], sty['bg_fill'], sty['border'])

            # Resample this curve's values onto bucket centers
            if len(curve.times_beats) > 0 and len(auto_map.effective_gain_times) > 0:
                # Use the automation map's time grid to interpolate
                curve_resampled = _downsample_frames(
                    _resample_curve_to_buckets(curve, auto_map, bucket_centers),
                    n_buckets
                )
            else:
                curve_resampled = [None] * n_buckets

            for col_idx, val in enumerate(curve_resampled, 4):
                _styled_cell(ws, row, col_idx, val,
                             sty['data_font'], sty['bg_fill'], sty['border'])
            row += 1

        # effective_gain row
        _styled_cell(ws, row, 1, track_name,
                     sty['header_font'], sty['header_fill'], sty['border'])
        _styled_cell(ws, row, 2, 'effective_gain',
                     sty['header_font'], sty['header_fill'], sty['border'])
        _styled_cell(ws, row, 3, '—',
                     sty['header_font'], sty['header_fill'], sty['border'])

        gain_resampled = resample_effective_gain(auto_map, bucket_centers)
        for col_idx, val in enumerate(gain_resampled, 4):
            _styled_cell(ws, row, col_idx, round(float(val), 4),
                         sty['data_font'], sty['bg_fill'], sty['border'])
        row += 1

        # is_audible row
        _styled_cell(ws, row, 1, track_name,
                     sty['header_font'], sty['header_fill'], sty['border'])
        _styled_cell(ws, row, 2, 'is_audible',
                     sty['header_font'], sty['header_fill'], sty['border'])
        _styled_cell(ws, row, 3, '—',
                     sty['header_font'], sty['header_fill'], sty['border'])

        audible_resampled = resample_audibility(auto_map, bucket_centers)
        for col_idx, val in enumerate(audible_resampled, 4):
            _styled_cell(ws, row, col_idx, bool(val),
                         sty['data_font'], sty['bg_fill'], sty['border'])
        row += 1

    log_fn(f"    Excel: _track_automation_map done — {row - 1} rows x {n_buckets} time buckets")


def _resample_curve_to_buckets(
    curve,
    auto_map,
    bucket_centers: np.ndarray,
) -> np.ndarray:
    """Resample a single AutomationCurve onto bucket center times."""
    from automation_map import _interpolate_at

    src_times = auto_map.effective_gain_times
    if len(src_times) == 0:
        return np.full(len(bucket_centers), curve.values[0] if len(curve.values) > 0 else 0.0)

    # The curve stores times in beats; the auto_map has a seconds grid.
    # We need to interpolate the curve values at the bucket times.
    # Use the curve's beat-time values interpolated at beat equivalents of bucket_centers.
    tempo_ratio = src_times[-1] / (auto_map.effective_gain.shape[0] - 1) if len(src_times) > 1 else 1.0

    # Convert bucket_centers (seconds) back to beats for interpolation
    if len(src_times) > 1:
        total_seconds = src_times[-1]
        total_beats_est = auto_map.effective_gain.shape[0]  # grid length ~ beats
        if total_seconds > 0:
            bucket_beats = bucket_centers * (total_beats_est / total_seconds)
        else:
            bucket_beats = bucket_centers
    else:
        bucket_beats = bucket_centers

    return _interpolate_at(curve.times_beats, curve.values, bucket_beats)
