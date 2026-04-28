"""banque_loader — read-only access to ableton/banque_midi_qrust.xlsx.

Why this exists:

The banque (10 sheets : drum_mapping, rhythm_patterns, scales,
chord_progressions, qrust_profiles, velocity_dynamics, tempo_reference,
bassline_patterns, notes_usage) is hand-curated by the user. It is the
**concrete material library** that mix + composition agents draw from
instead of inventing their own notes / patterns / velocities (today's
70/30 problem).

This loader slices the workbook so each agent receives only the slice
relevant to its scope. The .xlsx itself stays untouched — single source
of truth, edited by the user only.

Usage examples (intended consumer agents):

    motif-decider    -> get_qrust_profile("Acid Drops") + get_drum_pattern("Dark techno", "Kick")
    rhythm-decider   -> list_qrust_profiles() + get_tempo_for_genre("Industrial groovy")
    harmony-decider  -> list_qrust_starred_scales() + list_chord_progressions("Dark Phrygien")
    performance-decider (Phase 2.8) -> get_velocity_range("NIN heavy", "Kick")
    bassline (any layer with role='bass')       -> list_basslines("Acid Drops style")

Mirrors the song_loader / catalog_loader patterns.
"""
from __future__ import annotations

from functools import lru_cache
from pathlib import Path
from typing import Any, Optional

try:
    import openpyxl  # type: ignore[import-untyped]
except ImportError as exc:  # pragma: no cover
    raise ImportError(
        "openpyxl is required to load the banque MIDI. "
        "Install with: pip install openpyxl"
    ) from exc


_BANQUE_XLSX = (
    Path(__file__).resolve().parents[2]
    / "ableton"
    / "banque_midi_qrust.xlsx"
)


# Each sheet's data starts at row 4 (1-indexed) — rows 0-1 are titles,
# row 2 is blank, row 3 is the header. This is the user's convention
# verified across all 10 sheets.
_HEADER_ROW_IDX = 3  # 0-indexed
_DATA_START_IDX = 4  # 0-indexed


@lru_cache(maxsize=1)
def load_banque() -> dict:
    """Load and cache the banque workbook into a dict-of-list-of-dict structure.

    Each sheet becomes `{sheet_name: [row_dict, …]}`. Row dicts use the
    header row's column names as keys. Empty rows are skipped.

    Loaded once per process; subsequent calls are free. Tests may call
    `load_banque.cache_clear()` between modifications.
    """
    if not _BANQUE_XLSX.exists():
        raise FileNotFoundError(
            f"banque MIDI not found at {_BANQUE_XLSX}. "
            f"Has the user added it via the ableton/ folder?"
        )

    wb = openpyxl.load_workbook(_BANQUE_XLSX, data_only=True, read_only=True)
    out: dict[str, list[dict]] = {}

    try:
        for sheet_name in wb.sheetnames:
            ws = wb[sheet_name]
            rows = list(ws.iter_rows(values_only=True))

            # Pull the header row. If the sheet is too short (e.g. 00_Index
            # has its title at row 0 + content rows from row 4), we treat
            # row 3 as headers regardless.
            if len(rows) <= _HEADER_ROW_IDX:
                out[sheet_name] = []
                continue

            headers_raw = rows[_HEADER_ROW_IDX]
            headers = [
                str(h).strip() if h is not None else f"col_{i}"
                for i, h in enumerate(headers_raw)
            ]

            data: list[dict] = []
            for row in rows[_DATA_START_IDX:]:
                # Skip rows that are entirely empty
                if all(c is None or (isinstance(c, str) and not c.strip())
                       for c in row):
                    continue
                row_dict = {
                    headers[i]: row[i]
                    for i in range(min(len(headers), len(row)))
                }
                data.append(row_dict)
            out[sheet_name] = data
    finally:
        wb.close()

    return out


# ============================================================================
# 00 Index + meta
# ============================================================================


def get_meta() -> dict:
    """Return the index sheet (sheet → description map)."""
    banque = load_banque()
    index_rows = banque.get("00_Index", [])
    return {
        "sheet_count": len(banque),
        "sheets": list(banque.keys()),
        "index_entries": [
            {"num": r.get("#"), "name": r.get("Onglet"), "desc": r.get("Contenu")}
            for r in index_rows
            if r.get("#") and r.get("Onglet")
        ],
    }


# ============================================================================
# 01 Drum mapping
# ============================================================================


def get_drum_mapping() -> list[dict]:
    """Return the full drum mapping table (MIDI # → role labels).

    Each entry: {midi: int, note: str, gm: str, ableton: str, qrust: str, notes: str}
    """
    banque = load_banque()
    out: list[dict] = []
    for r in banque.get("01_Drum_Mapping", []):
        midi = r.get("MIDI #")
        if midi is None:
            continue
        out.append({
            "midi": int(midi),
            "note": r.get("Note (Ableton)") or "",
            "gm": r.get("GM Standard") or "",
            "ableton": r.get("Drum Rack Ableton (808)") or "",
            "qrust": r.get("Industrial / Qrust") or "",
            "notes": r.get("Notes") or "",
        })
    return out


def get_drum_note_for_role(role: str) -> Optional[int]:
    """Convenience : map a role name like 'kick', 'snare', 'hat' to a MIDI #.

    Returns the canonical MIDI note for that role, prioritizing the
    qrust column over ableton over gm. Aliases are listed most-specific
    first so "kick" returns 36 (Kick principal) rather than 35 (Sub kick).
    """
    needle = role.lower().strip()
    role_aliases = {
        # role: list of (column, canonical-substring) — searched in order,
        # first hit wins. Most specific first.
        "kick":     [("ableton", "kick principal")],
        "sub_kick": [("qrust", "sub kick"), ("ableton", "kick alt / sub")],
        "snare":    [("ableton", "snare"), ("gm", "acoustic snare")],
        "hat":      [("ableton", "closed hat"), ("gm", "closed hi-hat")],
        "open_hat": [("ableton", "open hat"), ("gm", "open hi-hat")],
        "clap":     [("ableton", "clap"), ("gm", "hand clap")],
        "rim":      [("ableton", "rim shot"), ("gm", "side stick")],
        "tom":      [("ableton", "low tom")],
        "ride":     [("ableton", "ride")],
        "crash":    [("ableton", "crash")],
        "perc":     [("qrust", "metal hit"), ("qrust", "industrial")],
        "shaker":   [("qrust", "shaker")],
        "cowbell":  [("qrust", "cowbell"), ("ableton", "cowbell")],
    }
    aliases = role_aliases.get(needle, [(None, needle)])
    for column, substring in aliases:
        for entry in get_drum_mapping():
            if column is None:
                # generic fallback: search across all label columns
                haystack = " ".join([
                    entry["qrust"], entry["ableton"], entry["gm"]
                ]).lower()
            else:
                haystack = (entry.get(column) or "").lower()
            if substring in haystack:
                return entry["midi"]
    return None


# ============================================================================
# 02 Rhythm patterns
# ============================================================================


def list_rhythm_genres() -> list[str]:
    """Distinct genre names appearing in 02_Rhythm_Patterns."""
    banque = load_banque()
    seen = []
    for r in banque.get("02_Rhythm_Patterns", []):
        g = r.get("Genre")
        if g and g not in seen:
            seen.append(g)
    return seen


def get_rhythm_pattern(
    genre: str, element: Optional[str] = None
) -> list[dict]:
    """Return rhythm pattern rows matching the given genre.

    The 16-step pattern is in column "1" as a 16-char string of the form
    "X...X...X...X..." (X = hit, g = ghost, a = accent, . = silence).
    """
    banque = load_banque()
    out: list[dict] = []
    for r in banque.get("02_Rhythm_Patterns", []):
        if r.get("Genre") != genre:
            continue
        if element is not None and r.get("Élément") != element:
            continue
        pattern = r.get("1") or ""
        out.append({
            "genre": r.get("Genre"),
            "element": r.get("Élément") or "",
            "pattern_16": pattern,
            "tempo_range": r.get("2") or "",
            "notes": r.get("3") or "",
        })
    return out


def parse_pattern_16(pattern: str) -> list[dict]:
    """Decode a 16-step pattern string into a list of step events.

    Each event: {step: int (0..15), kind: "hit"|"ghost"|"accent", char: str}
    Silences ('.') are excluded. Useful for the motif-decider to convert
    a banque pattern to actual MIDI notes at the right beat positions.
    """
    out: list[dict] = []
    char_map = {"X": "hit", "g": "ghost", "a": "accent"}
    for i, c in enumerate(pattern[:16]):
        kind = char_map.get(c)
        if kind is None:
            continue
        out.append({"step": i, "kind": kind, "char": c})
    return out


# ============================================================================
# 03 Scales / modes
# ============================================================================


def list_scales() -> list[dict]:
    """All scales/modes documented in 03_Scales_Modes."""
    banque = load_banque()
    out: list[dict] = []
    for r in banque.get("03_Scales_Modes", []):
        if not r.get("Nom"):
            continue
        out.append({
            "family": r.get("Famille") or "",
            "name": r.get("Nom") or "",
            "intervals": r.get("Intervalles") or "",
            "notes_in_c": r.get("Notes en C") or "",
            "mood": r.get("Mood / Caractère") or "",
            "usage": r.get("Usage typique") or "",
            "qrust_rating": r.get("Qrust ⚡") or "",  # "★", "★★", "★★★", or ""
        })
    return out


def get_scale(name: str) -> Optional[dict]:
    """Lookup a scale by exact name. Case-insensitive."""
    needle = name.strip().lower()
    for s in list_scales():
        if s["name"].strip().lower() == needle:
            return s
    return None


def list_qrust_starred_scales(min_stars: int = 1) -> list[dict]:
    """Scales rated ≥ min_stars in the Qrust ⚡ column.

    Star count is derived from the "★" character count in the rating.
    """
    out: list[dict] = []
    for s in list_scales():
        rating = s["qrust_rating"]
        if not rating:
            continue
        star_count = rating.count("★")
        if star_count >= min_stars:
            out.append({**s, "star_count": star_count})
    return sorted(out, key=lambda x: -x["star_count"])


def parse_intervals(intervals_str: str) -> tuple[int, ...]:
    """Decode '0,2,3,5,7,8,10' into a tuple of semitone offsets from root.

    Robust to spaces and to the special '0..11' chromatic shorthand.
    Returns empty tuple if unparseable.
    """
    s = intervals_str.strip()
    if s == "0..11":
        return tuple(range(12))
    try:
        return tuple(
            int(x.strip()) for x in s.split(",") if x.strip()
        )
    except ValueError:
        return ()


# ============================================================================
# 04 Chord progressions
# ============================================================================


def list_chord_progressions(mood: Optional[str] = None) -> list[dict]:
    """All chord progressions, optionally filtered by mood."""
    banque = load_banque()
    out: list[dict] = []
    for r in banque.get("04_Chord_Progressions", []):
        if not r.get("Progression"):
            continue
        if mood is not None and r.get("Mood") != mood:
            continue
        out.append({
            "mood": r.get("Mood") or "",
            "progression": r.get("Progression") or "",
            "roman": r.get("Chiffrage romain") or "",
            "chords_in_cm": r.get("Accords (Cm)") or "",
            "style": r.get("Style typique") or "",
            "notes": r.get("Notes") or "",
        })
    return out


# ============================================================================
# 05 Qrust profiles (full presets)
# ============================================================================


def list_qrust_profiles() -> list[dict]:
    """All Qrust profiles — each is a complete preset (tempo + scale +
    kick/snare/hat patterns + velocity range)."""
    banque = load_banque()
    out: list[dict] = []
    for r in banque.get("05_Qrust_Profiles", []):
        if not r.get("Profil"):
            continue
        out.append({
            "name": r.get("Profil") or "",
            "tempo_sweet": r.get("Tempo sweet"),
            "tempo_range": r.get("Tempo range") or "",
            "scale": r.get("Scale préf.") or "",
            "kick_beats": r.get("Kick (beats)") or "",
            "snare_beats": r.get("Snare/Clap") or "",
            "hat_pattern": r.get("Hat pattern") or "",
            "velocity_range": r.get("Velocity range") or "",
            "description": r.get("Description") or r.get("col_8") or "",
        })
    return out


def get_qrust_profile(name: str) -> Optional[dict]:
    """Lookup a profile by exact name. Case-insensitive."""
    needle = name.strip().lower()
    for p in list_qrust_profiles():
        if p["name"].strip().lower() == needle:
            return p
    return None


# ============================================================================
# 06 Velocity dynamics
# ============================================================================


def get_velocity_range(
    style: str, element: Optional[str] = None
) -> list[dict]:
    """Return velocity ranges for a given style (and optional element).

    Each entry: {style, element, vel_min, vel_max, variance, notes}.
    """
    banque = load_banque()
    out: list[dict] = []
    for r in banque.get("06_Velocity_Dynamics", []):
        if r.get("Style") != style:
            continue
        if element is not None and r.get("Élément") != element:
            continue
        vmin = r.get("Vélocité min")
        vmax = r.get("Vélocité max")
        if vmin is None or vmax is None:
            continue
        out.append({
            "style": r.get("Style"),
            "element": r.get("Élément") or "",
            "vel_min": int(vmin),
            "vel_max": int(vmax),
            "variance": r.get("Variance") or "",
            "notes": r.get("Notes") or "",
        })
    return out


# ============================================================================
# 07 Tempo reference
# ============================================================================


def get_tempo_for_genre(genre: str) -> Optional[dict]:
    """Lookup tempo info for a given genre. Case-insensitive exact match."""
    banque = load_banque()
    needle = genre.strip().lower()
    for r in banque.get("07_Tempo_Reference", []):
        g = (r.get("Genre") or "").strip().lower()
        if g == needle:
            return {
                "genre": r.get("Genre"),
                "tempo_min": int(r["Tempo min"]) if r.get("Tempo min") else None,
                "tempo_max": int(r["Tempo max"]) if r.get("Tempo max") else None,
                "sweet_spot": int(r["Sweet spot"]) if r.get("Sweet spot") else None,
                "time_signature": r.get("Time sig") or "",
            }
    return None


def list_tempo_genres() -> list[str]:
    """Distinct genre names in 07_Tempo_Reference."""
    banque = load_banque()
    return [r["Genre"] for r in banque.get("07_Tempo_Reference", []) if r.get("Genre")]


# ============================================================================
# 08 Bassline patterns
# ============================================================================


def list_basslines(style: Optional[str] = None) -> list[dict]:
    """All bassline patterns, optionally filtered by style."""
    banque = load_banque()
    out: list[dict] = []
    for r in banque.get("08_Bassline_Patterns", []):
        if not r.get("Style"):
            continue
        if style is not None and r.get("Style") != style:
            continue
        out.append({
            "style": r.get("Style") or "",
            "rhythm": r.get("Pattern rythmique") or "",
            "degrees": r.get("Notes (degrés)") or "",
            "description": r.get("Description") or "",
            "example_in_cm": r.get("Ex. en C mineur") or "",
        })
    return out


# ============================================================================
# Public API
# ============================================================================


__all__ = [
    "load_banque",
    "get_meta",
    "get_drum_mapping",
    "get_drum_note_for_role",
    "list_rhythm_genres",
    "get_rhythm_pattern",
    "parse_pattern_16",
    "list_scales",
    "get_scale",
    "list_qrust_starred_scales",
    "parse_intervals",
    "list_chord_progressions",
    "list_qrust_profiles",
    "get_qrust_profile",
    "get_velocity_range",
    "get_tempo_for_genre",
    "list_tempo_genres",
    "list_basslines",
]


if __name__ == "__main__":
    meta = get_meta()
    print(f"Banque MIDI Qrust — {meta['sheet_count']} sheets")
    for entry in meta["index_entries"]:
        print(f"  {entry['num']} {entry['name']}: {entry['desc'][:60]}")
    print()
    drums = get_drum_mapping()
    print(f"Drum mapping: {len(drums)} entries (MIDI {drums[0]['midi']}-{drums[-1]['midi']})")
    print(f"  kick → MIDI {get_drum_note_for_role('kick')}")
    print(f"  snare → MIDI {get_drum_note_for_role('snare')}")
    print(f"  hat → MIDI {get_drum_note_for_role('hat')}")
    print()
    profile = get_qrust_profile("Acid Drops (cible)")
    print(f"Profile 'Acid Drops': tempo {profile['tempo_sweet']} BPM, scale {profile['scale']}")
    print()
    starred = list_qrust_starred_scales(min_stars=3)
    print(f"3-star Qrust scales: {[s['name'] for s in starred]}")
