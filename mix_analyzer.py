#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
Mix Analyzer v2.2 - Visual audio mix analysis tool
Generates detailed Excel reports for audio tracks to aid mixing and mastering decisions.

Usage:
    python mix_analyzer.py

Dependencies:
    pip install librosa pyloudnorm soundfile numpy scipy matplotlib openpyxl
"""

import os
import sys
import json
import threading
import traceback
import webbrowser
import subprocess
from pathlib import Path
from datetime import datetime

import numpy as np
import soundfile as sf
import librosa
import librosa.display
import pyloudnorm as pyln
import matplotlib
matplotlib.use('Agg')
import matplotlib.pyplot as plt
from scipy import signal

import tkinter as tk
from tkinter import ttk, filedialog, messagebox, scrolledtext


# ============================================================================
# CATEGORIES - 25 options organized hierarchically
# ============================================================================

CATEGORIES = {
    'Drums': [
        'Kick',
        'Snare / Clap',
        'Hi-Hat / Cymbal',
        'Tom',
        'Percussion',
        'Drum Loop / Bus',
    ],
    'Bass': [
        'Sub Bass',
        'Bass (standard)',
        'Acid Bass',
        '808 / Pitched Bass',
    ],
    'Synth': [
        'Lead Synth',
        'Pluck / Stab',
        'Pad / Drone',
        'Arpeggio / Sequence',
        'Texture / Atmosphere',
    ],
    'Guitar': [
        'Guitar Clean',
        'Guitar Distorted',
        'Guitar Acoustic',
    ],
    'Vocal': [
        'Lead Vocal',
        'Backing / Harmony Vocal',
        'Vocal FX / Chop',
    ],
    'FX & Other': [
        'FX / Riser / Impact',
        'Noise / Ambience',
        'Sample / Loop',
        'Other',
    ],
}

# Flat list for dropdowns
ALL_CATEGORIES = []
for family, items in CATEGORIES.items():
    for item in items:
        ALL_CATEGORIES.append(item)
ALL_CATEGORIES.append('(not set)')

# Category -> family mapping
CATEGORY_FAMILY = {}
for family, items in CATEGORIES.items():
    for item in items:
        CATEGORY_FAMILY[item] = family
CATEGORY_FAMILY['(not set)'] = 'Unknown'


# ============================================================================
# TRACK TYPES
# ============================================================================

TRACK_TYPES = ['Individual', 'BUS', 'Full Mix']


# ============================================================================
# MUSICAL STYLES - 30 styles mapped to 7 analytical families
# ============================================================================

STYLES = [
    'Generic',
    'Acoustic / Folk',
    'Ambient / Drone',
    'Bass Music (Dubstep, DnB)',
    'Blues',
    'Classical / Orchestral',
    'Country',
    'Dance / EDM',
    'Electronic (General)',
    'Experimental',
    'Film Score / Cinematic',
    'Funk / Soul',
    'Hip-Hop / Rap',
    'House / Deep House',
    'Indie / Alternative',
    'Industrial',
    'Jazz',
    'Latin',
    'Lo-Fi',
    'Metal',
    'Pop',
    'Punk',
    'R&B',
    'Reggae / Dub',
    'Rock',
    'Singer-Songwriter',
    'Synthwave / Retrowave',
    'Techno',
    'Trance',
    'Trap',
    'World Music',
]

STYLE_FAMILY = {
    'Generic':                     'generic',
    'Acoustic / Folk':             'acoustic',
    'Blues':                       'acoustic',
    'Country':                     'acoustic',
    'Singer-Songwriter':           'acoustic',
    'Jazz':                        'acoustic',
    'Classical / Orchestral':      'acoustic',
    'Rock':                        'rock',
    'Indie / Alternative':         'rock',
    'Punk':                        'rock',
    'Metal':                       'rock',
    'Ambient / Drone':             'electronic_soft',
    'Lo-Fi':                       'electronic_soft',
    'Experimental':                'electronic_soft',
    'Film Score / Cinematic':      'electronic_soft',
    'Dance / EDM':                 'electronic_dance',
    'House / Deep House':          'electronic_dance',
    'Techno':                      'electronic_dance',
    'Trance':                      'electronic_dance',
    'Electronic (General)':        'electronic_dance',
    'Industrial':                  'electronic_aggressive',
    'Bass Music (Dubstep, DnB)':   'electronic_aggressive',
    'Synthwave / Retrowave':       'electronic_aggressive',
    'Hip-Hop / Rap':               'urban',
    'Trap':                        'urban',
    'R&B':                         'urban',
    'Funk / Soul':                 'urban',
    'Reggae / Dub':                'urban',
    'Pop':                         'pop',
    'Latin':                       'pop',
    'World Music':                 'pop',
}

# Analytical profiles per family
FAMILY_PROFILES = {
    'generic':                {'target_lufs_mix': -14, 'typical_crest_mix': 10, 'density_tolerance': 'normal'},
    'acoustic':               {'target_lufs_mix': -16, 'typical_crest_mix': 14, 'density_tolerance': 'low'},
    'rock':                   {'target_lufs_mix': -10, 'typical_crest_mix': 10, 'density_tolerance': 'high'},
    'electronic_soft':        {'target_lufs_mix': -16, 'typical_crest_mix': 14, 'density_tolerance': 'low'},
    'electronic_dance':       {'target_lufs_mix': -9,  'typical_crest_mix': 8,  'density_tolerance': 'high'},
    'electronic_aggressive':  {'target_lufs_mix': -8,  'typical_crest_mix': 8,  'density_tolerance': 'very_high'},
    'urban':                  {'target_lufs_mix': -10, 'typical_crest_mix': 9,  'density_tolerance': 'high'},
    'pop':                    {'target_lufs_mix': -10, 'typical_crest_mix': 9,  'density_tolerance': 'normal'},
}


# ============================================================================
# MIX COMPLETION STATES
# ============================================================================

MIX_STATES = [
    'Rough mix',
    'Mix in progress',
    'Pre-final mix',
    'Final mix',
    'Pre-master',
    'Final master',
]

MASTER_PLUGINS = [
    'EQ',
    'Compression',
    'Limiting',
    'Saturation',
    'Stereo Imager',
    'Other',
]

LOUDNESS_TARGETS = [
    '-14 LUFS (Streaming standard)',
    '-12 LUFS (Intermediate)',
    '-10 LUFS (Hot)',
    '-8 LUFS (Very hot)',
    'Custom / Not defined',
]


# ============================================================================
# AUTO-DETECTION PATTERNS (conservative, short patterns at start of filename)
# ============================================================================

AUTO_DETECT_PATTERNS = [
    # Kick variations
    (['kick', 'bd', 'bassdrum', 'bass drum', 'kik'], 'Kick'),
    # Snare
    (['snare', 'snr', 'sn '], 'Snare / Clap'),
    (['clap', 'clp'], 'Snare / Clap'),
    # Hats
    (['hat', 'hh', 'hihat', 'hi-hat', 'cymbal', 'ride', 'crash'], 'Hi-Hat / Cymbal'),
    # Toms
    (['tom', 'floor tom'], 'Tom'),
    # Perc
    (['perc', 'tambourine', 'shaker', 'conga', 'bongo', 'clave', 'cowbell'], 'Percussion'),
    # Sub
    (['sub bass', 'subbass', 'sub_bass', 'sub-bass', 'sub '], 'Sub Bass'),
    # Acid
    (['acid', '303', 'tb303', 'tb-303'], 'Acid Bass'),
    # 808
    (['808'], '808 / Pitched Bass'),
    # Bass
    (['bass', 'bs '], 'Bass (standard)'),
    # Lead
    (['lead', 'ld '], 'Lead Synth'),
    # Pluck
    (['pluck', 'stab'], 'Pluck / Stab'),
    # Pad
    (['pad', 'drone'], 'Pad / Drone'),
    # Arp
    (['arp', 'seq', 'sequence'], 'Arpeggio / Sequence'),
    # Texture
    (['texture', 'atmo', 'ambient'], 'Texture / Atmosphere'),
    # Guitar
    (['gtr clean', 'clean gtr', 'guitar clean'], 'Guitar Clean'),
    (['gtr dist', 'dist gtr', 'guitar dist', 'metal gtr'], 'Guitar Distorted'),
    (['acoustic gtr', 'acoustic guitar', 'gtr acoustic'], 'Guitar Acoustic'),
    (['gtr', 'guitar'], 'Guitar Clean'),  # fallback
    # Vocal
    (['lead vox', 'lead vocal', 'vox lead', 'main vocal'], 'Lead Vocal'),
    (['backing', 'harmony vox', 'vox harm', 'bv '], 'Backing / Harmony Vocal'),
    (['vox chop', 'vox fx', 'vocal fx', 'vocal chop'], 'Vocal FX / Chop'),
    (['vox', 'vocal'], 'Lead Vocal'),  # fallback
    # FX
    (['riser', 'impact', 'fx '], 'FX / Riser / Impact'),
    (['noise', 'ambience'], 'Noise / Ambience'),
    (['loop', 'sample'], 'Sample / Loop'),
]


def auto_detect_category(filename, project_name=None):
    """Attempt to detect category from filename. Returns '(not set)' if no match.
    If project_name is provided, it is stripped from the filename before matching
    to avoid false positives (e.g. song 'Acid Drops' matching 'Acid Bass')."""
    name_lower = filename.lower().replace('_', ' ').replace('-', ' ')
    for ext in ['.wav', '.aiff', '.aif', '.flac', '.ogg']:
        if name_lower.endswith(ext):
            name_lower = name_lower[:-len(ext)]
            break
    if project_name:
        pn = project_name.lower().replace('_', ' ').replace('-', ' ').strip()
        if pn and pn in name_lower:
            name_lower = name_lower.replace(pn, ' ').strip()
    for patterns, category in AUTO_DETECT_PATTERNS:
        for pattern in patterns:
            if pattern in name_lower:
                return category
    return '(not set)'
# ============================================================================
# AUDIO ANALYSIS ENGINE
# ============================================================================

FREQ_BANDS = [
    ('sub',      20,    60),
    ('bass',     60,    250),
    ('low_mid',  250,   500),
    ('mid',      500,   2000),
    ('high_mid', 2000,  4000),
    ('presence', 4000,  8000),
    ('air',      8000,  20000),
]

BAND_LABELS = {
    'sub':      'Sub (20-60 Hz)',
    'bass':     'Bass (60-250 Hz)',
    'low_mid':  'Low-Mid (250-500 Hz)',
    'mid':      'Mid (500-2000 Hz)',
    'high_mid': 'High-Mid (2-4 kHz)',
    'presence': 'Presence (4-8 kHz)',
    'air':      'Air (8-20 kHz)',
}

# High-resolution bands for masking matrix (~third-octave, 22 bands)
FREQ_BANDS_HIRES = [
    ('20-32 Hz',     20,     32),
    ('32-50 Hz',     32,     50),
    ('50-80 Hz',     50,     80),
    ('80-125 Hz',    80,     125),
    ('125-160 Hz',   125,    160),
    ('160-200 Hz',   160,    200),
    ('200-250 Hz',   200,    250),
    ('250-315 Hz',   250,    315),
    ('315-400 Hz',   315,    400),
    ('400-500 Hz',   400,    500),
    ('500-630 Hz',   500,    630),
    ('630-800 Hz',   630,    800),
    ('800-1k Hz',    800,    1000),
    ('1-1.25 kHz',   1000,   1250),
    ('1.25-1.6 kHz', 1250,   1600),
    ('1.6-2 kHz',    1600,   2000),
    ('2-2.5 kHz',    2000,   2500),
    ('2.5-3.15 kHz', 2500,   3150),
    ('3.15-4 kHz',   3150,   4000),
    ('4-5 kHz',      4000,   5000),
    ('5-8 kHz',      5000,   8000),
    ('8-20 kHz',     8000,   20000),
]


def load_audio(filepath):
    """Load audio file. Returns (data, sr, is_stereo)."""
    data, sr = sf.read(filepath, always_2d=True)
    is_stereo = data.shape[1] >= 2
    if data.shape[1] > 2:
        data = data[:, :2]
    return data.astype(np.float32), sr, is_stereo


def to_mono(data):
    if data.ndim == 2 and data.shape[1] > 1:
        return np.mean(data, axis=1)
    return data.flatten()


def db(x, floor=-120):
    x = np.asarray(x, dtype=np.float64)
    with np.errstate(divide='ignore'):
        result = 20 * np.log10(np.maximum(np.abs(x), 1e-12))
    return np.maximum(result, floor)


def analyze_loudness(data, sr):
    mono = to_mono(data)
    audio_for_lufs = data if data.ndim == 2 else mono.reshape(-1, 1)

    meter = pyln.Meter(sr)
    try:
        lufs_integrated = float(meter.integrated_loudness(audio_for_lufs))
    except Exception:
        lufs_integrated = -float('inf')

    block_size = int(3 * sr)
    hop = int(0.1 * sr)
    st_values = []
    if len(mono) > block_size:
        for i in range(0, len(mono) - block_size, hop):
            block = audio_for_lufs[i:i+block_size]
            try:
                lufs_st = meter.integrated_loudness(block)
                if np.isfinite(lufs_st):
                    st_values.append(lufs_st)
            except Exception:
                pass

    if st_values:
        lufs_st_max = float(max(st_values))
        lra = float(np.percentile(st_values, 95) - np.percentile(st_values, 10))
    else:
        lufs_st_max = lufs_integrated
        lra = 0.0

    peak = float(np.max(np.abs(data)))
    peak_db = float(db(peak))

    try:
        oversampled = signal.resample_poly(mono, 4, 1)
        true_peak = float(np.max(np.abs(oversampled)))
        true_peak_db = float(db(true_peak))
    except Exception:
        true_peak_db = peak_db

    rms = float(np.sqrt(np.mean(mono ** 2)))
    rms_db = float(db(rms))

    crest_factor = peak_db - rms_db if rms > 0 else 0.0
    plr = peak_db - lufs_integrated if np.isfinite(lufs_integrated) else 0.0
    psr = peak_db - lufs_st_max if np.isfinite(lufs_st_max) else 0.0

    return {
        'peak_db': peak_db,
        'true_peak_db': true_peak_db,
        'rms_db': rms_db,
        'lufs_integrated': lufs_integrated,
        'lufs_short_term_max': lufs_st_max,
        'lra': lra,
        'crest_factor': crest_factor,
        'plr': plr,
        'psr': psr,
    }


def analyze_spectrum(mono, sr):
    n_fft = 8192
    S = np.abs(librosa.stft(mono, n_fft=n_fft, hop_length=n_fft//4))
    spectrum_mean = np.mean(S, axis=1)
    freqs = librosa.fft_frequencies(sr=sr, n_fft=n_fft)

    band_energies = {}
    total_energy = np.sum(spectrum_mean ** 2) + 1e-12
    for name, flow, fhigh in FREQ_BANDS:
        mask = (freqs >= flow) & (freqs < fhigh)
        if np.any(mask):
            energy = float(np.sum(spectrum_mean[mask] ** 2))
            band_energies[name] = 100 * energy / total_energy
        else:
            band_energies[name] = 0.0

    dominant_band = max(band_energies, key=band_energies.get)

    try:
        centroid = float(np.mean(librosa.feature.spectral_centroid(y=mono, sr=sr)))
        rolloff = float(np.mean(librosa.feature.spectral_rolloff(y=mono, sr=sr, roll_percent=0.85)))
        flatness = float(np.mean(librosa.feature.spectral_flatness(y=mono)))
    except Exception:
        centroid = rolloff = flatness = 0.0

    spectrum_db = db(spectrum_mean / (np.max(spectrum_mean) + 1e-12))
    peaks, properties = signal.find_peaks(spectrum_db, height=-20, distance=20, prominence=6)
    peak_freqs = freqs[peaks][:8]
    peak_heights = spectrum_db[peaks][:8]
    peak_list = sorted(
        [(float(f), float(h)) for f, h in zip(peak_freqs, peak_heights)],
        key=lambda x: x[1], reverse=True
    )[:6]

    return {
        'freqs': freqs,
        'spectrum_mean': spectrum_mean,
        'spectrum_db_normalized': spectrum_db,
        'band_energies': band_energies,
        'dominant_band': dominant_band,
        'centroid': centroid,
        'rolloff': rolloff,
        'flatness': flatness,
        'peaks': peak_list,
    }


def compute_hires_band_energies(mono, sr):
    """Compute band energies for FREQ_BANDS_HIRES (used in masking matrix)."""
    n_fft = 8192
    S = np.abs(librosa.stft(mono, n_fft=n_fft, hop_length=n_fft // 4))
    spectrum_mean = np.mean(S, axis=1)
    freqs = librosa.fft_frequencies(sr=sr, n_fft=n_fft)
    total_energy = np.sum(spectrum_mean ** 2) + 1e-12
    energies = {}
    for idx_b, (label, flow, fhigh) in enumerate(FREQ_BANDS_HIRES):
        # Use <= for last band to include Nyquist
        if idx_b == len(FREQ_BANDS_HIRES) - 1:
            mask = (freqs >= flow) & (freqs <= fhigh)
        else:
            mask = (freqs >= flow) & (freqs < fhigh)
        if np.any(mask):
            energies[label] = 100 * float(np.sum(spectrum_mean[mask] ** 2)) / total_energy
        else:
            energies[label] = 0.0
    return energies


def analyze_temporal(mono, sr):
    frame_length = 2048
    hop_length = 512
    rms = librosa.feature.rms(y=mono, frame_length=frame_length, hop_length=hop_length)[0]
    rms_times = librosa.frames_to_time(np.arange(len(rms)), sr=sr, hop_length=hop_length)

    try:
        onset_frames = librosa.onset.onset_detect(y=mono, sr=sr, hop_length=hop_length, units='frames')
        onset_times = librosa.frames_to_time(onset_frames, sr=sr, hop_length=hop_length)
    except Exception:
        onset_times = np.array([])

    hist_values, hist_bins = np.histogram(mono, bins=100, range=(-1, 1))

    return {
        'rms_envelope': rms,
        'rms_times': rms_times,
        'onset_times': onset_times,
        'num_onsets': len(onset_times),
        'hist_values': hist_values,
        'hist_bins': hist_bins,
    }


def analyze_tempo_dynamic(mono, sr):
    """
    Enhanced tempo analysis with confidence score and tempogram.
    Returns median tempo, range, confidence, and tempogram data.
    """
    result = {
        'tempo_median': 0.0,
        'tempo_min': 0.0,
        'tempo_max': 0.0,
        'tempo_std': 0.0,
        'confidence': 0.0,
        'confidence_label': 'unreliable',
        'tempogram': None,
        'tempogram_times': None,
        'tempo_over_time': None,
        'reliable': False,
    }

    try:
        # Onset strength envelope - foundation for tempo detection
        hop_length = 512
        onset_env = librosa.onset.onset_strength(y=mono, sr=sr, hop_length=hop_length)

        # Onset strength must be meaningful - if too low, content is not percussive enough
        onset_strength_mean = float(np.mean(onset_env))
        onset_strength_max = float(np.max(onset_env)) if len(onset_env) > 0 else 0.0

        if onset_strength_max < 0.01 or onset_strength_mean < 0.001:
            # Essentially no transients - tempo is meaningless
            result['confidence_label'] = 'not applicable (non-percussive)'
            return result

        # Global tempo estimate
        tempo_global, beats = librosa.beat.beat_track(onset_envelope=onset_env, sr=sr, hop_length=hop_length)
        tempo_global = float(tempo_global) if np.isscalar(tempo_global) else (float(tempo_global[0]) if len(tempo_global) else 0.0)

        # Dynamic tempo via tempogram
        try:
            tempogram = librosa.feature.tempogram(onset_envelope=onset_env, sr=sr, hop_length=hop_length)
            tempogram_times = librosa.times_like(tempogram, sr=sr, hop_length=hop_length)

            # Extract dominant tempo per frame
            # Tempogram bins correspond to tempo values via fourier_tempo_frequencies
            try:
                tempo_frequencies = librosa.tempo_frequencies(tempogram.shape[0], hop_length=hop_length, sr=sr)
            except Exception:
                tempo_frequencies = None

            if tempo_frequencies is not None and len(tempo_frequencies) == tempogram.shape[0]:
                # For each time frame, find the tempo with max correlation
                # Restrict to reasonable tempo range (30-300 BPM)
                valid_bins = (tempo_frequencies >= 30) & (tempo_frequencies <= 300)
                if np.any(valid_bins):
                    restricted = tempogram[valid_bins]
                    restricted_freqs = tempo_frequencies[valid_bins]
                    dominant_idx = np.argmax(restricted, axis=0)
                    tempo_over_time = restricted_freqs[dominant_idx]

                    # Also compute confidence per frame (peak prominence vs mean)
                    peak_values = np.max(restricted, axis=0)
                    mean_values = np.mean(restricted, axis=0)
                    per_frame_confidence = peak_values / (mean_values + 1e-9)

                    # Overall confidence: median of per-frame confidence, normalized
                    median_confidence = float(np.median(per_frame_confidence))

                    # Stats on tempo over time
                    # Filter out outliers using median-based approach
                    tempo_median = float(np.median(tempo_over_time))
                    tempo_min = float(np.percentile(tempo_over_time, 5))
                    tempo_max = float(np.percentile(tempo_over_time, 95))
                    tempo_std = float(np.std(tempo_over_time))

                    result['tempogram'] = tempogram
                    result['tempogram_times'] = tempogram_times
                    result['tempo_over_time'] = tempo_over_time
                    result['tempo_median'] = tempo_median
                    result['tempo_min'] = tempo_min
                    result['tempo_max'] = tempo_max
                    result['tempo_std'] = tempo_std

                    # Confidence scoring (heuristic)
                    # A track needs: strong AND frequent onsets, consistent tempo, clear peak prominence
                    # Count significant onsets per second to detect truly percussive content
                    try:
                        onset_frames = librosa.onset.onset_detect(onset_envelope=onset_env, sr=sr, hop_length=hop_length)
                        duration_s = len(mono) / sr
                        onsets_per_sec = len(onset_frames) / duration_s if duration_s > 0 else 0
                    except Exception:
                        onsets_per_sec = 0

                    # Require at least 1 onset per second for tempo to be meaningful (slowest = 60 BPM = 1/s)
                    if onsets_per_sec < 0.8:
                        # Sustained content - tempo is meaningless regardless of what librosa says
                        result['tempo_median'] = tempo_global
                        result['confidence'] = 0.1
                        result['confidence_label'] = 'not applicable (sustained/non-rhythmic content)'
                        result['reliable'] = False
                        return result

                    # Normal confidence computation for rhythmic content
                    confidence = min(1.0, (
                        min(1.0, onset_strength_mean * 50) * 0.25 +
                        min(1.0, median_confidence / 4.0) * 0.35 +
                        max(0.0, 1.0 - tempo_std / 30.0) * 0.25 +
                        min(1.0, onsets_per_sec / 2.0) * 0.15
                    ))

                    # Hard penalty if tempo is outside typical musical range
                    # Most music is 60-180 BPM. Outside this, detection is almost certainly wrong.
                    if tempo_median < 55 or tempo_median > 200:
                        confidence *= 0.25  # drastically reduce
                    elif tempo_median < 70 or tempo_median > 180:
                        confidence *= 0.6

                    result['confidence'] = float(confidence)

                    if confidence >= 0.7:
                        result['confidence_label'] = 'high'
                        result['reliable'] = True
                    elif confidence >= 0.4:
                        result['confidence_label'] = 'medium'
                        result['reliable'] = True
                    elif confidence >= 0.2:
                        result['confidence_label'] = 'low'
                        result['reliable'] = False
                    else:
                        result['confidence_label'] = 'very low (likely unreliable)'
                        result['reliable'] = False
        except Exception:
            # Fallback to global tempo
            result['tempo_median'] = tempo_global
            result['tempo_min'] = tempo_global
            result['tempo_max'] = tempo_global
            result['confidence'] = 0.3
            result['confidence_label'] = 'low (global estimate only)'

    except Exception:
        pass

    return result


def analyze_musical(mono, sr):
    try:
        chroma = librosa.feature.chroma_cqt(y=mono, sr=sr, hop_length=1024)
    except Exception:
        try:
            chroma = librosa.feature.chroma_stft(y=mono, sr=sr, hop_length=1024)
        except Exception:
            chroma = np.zeros((12, 10))

    note_names = ['C', 'C#', 'D', 'D#', 'E', 'F', 'F#', 'G', 'G#', 'A', 'A#', 'B']
    chroma_mean = np.mean(chroma, axis=1)
    dominant_note = note_names[int(np.argmax(chroma_mean))]

    tonal_strength = float(np.max(chroma_mean) / (np.mean(chroma_mean) + 1e-9))

    return {
        'chroma': chroma,
        'dominant_note': dominant_note,
        'tonal_strength': tonal_strength,
        'is_tonal': tonal_strength > 1.8,
    }


def analyze_stereo(data, sr):
    """Stereo analysis including spectral panorama (vectorscope per frequency)."""
    if data.shape[1] < 2:
        return {
            'is_stereo': False,
            'correlation': 1.0,
            'width_overall': 0.0,
            'width_per_band': {name: 0.0 for name, _, _ in FREQ_BANDS},
            'mid': to_mono(data),
            'side': np.zeros(len(to_mono(data))),
            'pan_per_freq': None,
            'pan_freqs': None,
            'pan_energy': None,
        }

    L = data[:, 0]
    R = data[:, 1]

    if np.std(L) > 0 and np.std(R) > 0:
        correlation = float(np.corrcoef(L, R)[0, 1])
    else:
        correlation = 1.0

    M = 0.5 * (L + R)
    S = 0.5 * (L - R)

    energy_m = np.sum(M ** 2) + 1e-12
    energy_s = np.sum(S ** 2)
    width_overall = float(energy_s / (energy_m + energy_s))

    width_per_band = {}
    n_fft = 4096
    SL = np.abs(librosa.stft(L, n_fft=n_fft))
    SR = np.abs(librosa.stft(R, n_fft=n_fft))
    SM = np.abs(librosa.stft(M, n_fft=n_fft))
    SS = np.abs(librosa.stft(S, n_fft=n_fft))
    freqs = librosa.fft_frequencies(sr=sr, n_fft=n_fft)

    for name, flow, fhigh in FREQ_BANDS:
        mask = (freqs >= flow) & (freqs < fhigh)
        if np.any(mask):
            em = np.sum(SM[mask] ** 2) + 1e-12
            es = np.sum(SS[mask] ** 2)
            width_per_band[name] = float(es / (em + es))
        else:
            width_per_band[name] = 0.0

    # Spectral panorama (vectorscope per frequency)
    # For each frequency bin, compute the pan position -1 (L) to +1 (R)
    # pan = (R - L) / (R + L) in magnitude
    SL_mean = np.mean(SL, axis=1)
    SR_mean = np.mean(SR, axis=1)
    total = SL_mean + SR_mean + 1e-12
    pan_per_freq = (SR_mean - SL_mean) / total  # -1 (full left) to +1 (full right)
    pan_energy = SL_mean + SR_mean  # total energy at each freq (for color intensity)

    return {
        'is_stereo': True,
        'correlation': correlation,
        'width_overall': width_overall,
        'width_per_band': width_per_band,
        'mid': M,
        'side': S,
        'pan_per_freq': pan_per_freq,
        'pan_freqs': freqs,
        'pan_energy': pan_energy,
    }


def detect_anomalies(analysis):
    """Detect objective anomalies. Returns a list of (severity, description) tuples."""
    anomalies = []
    L = analysis['loudness']
    S = analysis['spectrum']
    stereo = analysis['stereo']

    # Clipping risk
    if L['peak_db'] > -0.3:
        anomalies.append(('critical', f"Peak level at {L['peak_db']:+.2f} dBFS - clipping risk"))
    elif L['peak_db'] > -1.0:
        anomalies.append(('warning', f"Peak level at {L['peak_db']:+.2f} dBFS - very little headroom"))

    # True peak
    if L['true_peak_db'] > 0.0:
        anomalies.append(('critical', f"True Peak at {L['true_peak_db']:+.2f} dBFS - inter-sample clipping"))

    # Phase
    if stereo['is_stereo']:
        if stereo['correlation'] < -0.3:
            anomalies.append(('critical', f"Phase correlation {stereo['correlation']:+.2f} - serious mono compatibility issue"))
        elif stereo['correlation'] < 0.0:
            anomalies.append(('warning', f"Phase correlation {stereo['correlation']:+.2f} - mono compatibility concern"))

    # Very low energy
    if L['rms_db'] < -60:
        anomalies.append(('warning', f"RMS level very low ({L['rms_db']:.1f} dBFS) - track may be nearly silent"))

    # Strong resonance peaks
    strong_peaks = [p for p in S['peaks'] if p[1] > -3 and p[0] > 100]
    if len(strong_peaks) >= 2:
        freqs_str = ', '.join(f"{p[0]:.0f}Hz" for p in strong_peaks[:3])
        anomalies.append(('warning', f"Strong resonance peaks detected at: {freqs_str}"))

    # Extreme compression
    if L['crest_factor'] < 5 and L['rms_db'] > -30:
        anomalies.append(('warning', f"Very low crest factor ({L['crest_factor']:.1f} dB) - heavy compression"))

    # Extreme stereo imbalance
    if stereo['is_stereo']:
        if stereo['width_overall'] > 0.6:
            anomalies.append(('info', f"Very wide stereo image ({stereo['width_overall']:.2f}) - verify mono compatibility"))

    return anomalies


def describe_characteristics(analysis):
    """Objective descriptive characteristics of the track. No prescriptions."""
    chars = []
    L = analysis['loudness']
    S = analysis['spectrum']
    T = analysis['temporal']
    M = analysis['musical']
    stereo = analysis['stereo']
    tempo = analysis.get('tempo', {})
    duration = analysis['duration']

    # Tonal character
    if M['is_tonal']:
        chars.append(('tonal', f"Tonal content (strongest note class: {M['dominant_note']}, tonal strength {M['tonal_strength']:.1f})"))
    else:
        chars.append(('atonal', f"Non-tonal / percussive or noisy content (tonal strength {M['tonal_strength']:.1f})"))

    # Temporal profile
    onsets_per_sec = T['num_onsets'] / duration if duration > 0 else 0
    if onsets_per_sec < 0.3:
        chars.append(('sustained', f"Sustained profile ({T['num_onsets']} transients over {duration:.1f}s = {onsets_per_sec:.2f}/s)"))
    elif onsets_per_sec > 3.0:
        chars.append(('rhythmic', f"Dense rhythmic profile ({T['num_onsets']} transients = {onsets_per_sec:.2f}/s)"))
    else:
        chars.append(('transient', f"Transient profile ({T['num_onsets']} transients = {onsets_per_sec:.2f}/s)"))

    # Spectral character
    chars.append(('spectrum', f"Dominant band: {BAND_LABELS[S['dominant_band']]} ({S['band_energies'][S['dominant_band']]:.1f}% of energy)"))
    chars.append(('spectrum', f"Spectral centroid: {S['centroid']:.0f} Hz, rolloff 85%: {S['rolloff']:.0f} Hz"))

    # Dynamics
    if L['crest_factor'] < 6:
        chars.append(('dynamics', f"Very compressed dynamics (crest factor {L['crest_factor']:.1f} dB)"))
    elif L['crest_factor'] < 10:
        chars.append(('dynamics', f"Moderately compressed dynamics (crest factor {L['crest_factor']:.1f} dB)"))
    elif L['crest_factor'] < 15:
        chars.append(('dynamics', f"Preserved dynamics (crest factor {L['crest_factor']:.1f} dB)"))
    else:
        chars.append(('dynamics', f"High dynamic range (crest factor {L['crest_factor']:.1f} dB)"))

    # Stereo
    if stereo['is_stereo']:
        w = stereo['width_overall']
        if w < 0.05:
            chars.append(('stereo', f"Quasi-mono stereo image (width {w:.2f})"))
        elif w < 0.15:
            chars.append(('stereo', f"Narrow stereo image (width {w:.2f})"))
        elif w < 0.30:
            chars.append(('stereo', f"Moderate stereo image (width {w:.2f})"))
        else:
            chars.append(('stereo', f"Wide stereo image (width {w:.2f})"))
        chars.append(('stereo', f"Phase correlation {stereo['correlation']:+.2f}"))
    else:
        chars.append(('stereo', "Mono track"))

    # Tempo (only if computed and reliable)
    if tempo.get('reliable'):
        tm = tempo['tempo_median']
        tmin = tempo['tempo_min']
        tmax = tempo['tempo_max']
        conf = tempo['confidence_label']
        if tmax - tmin > 5:
            chars.append(('tempo', f"Dynamic tempo: median {tm:.1f} BPM, range {tmin:.1f}-{tmax:.1f} BPM (confidence: {conf})"))
        else:
            chars.append(('tempo', f"Stable tempo: {tm:.1f} BPM (confidence: {conf})"))
    elif tempo.get('confidence_label') != 'not computed (individual track)':
        label = tempo.get('confidence_label', 'unreliable')
        chars.append(('tempo', f"Tempo detection: {label}"))
    # If tempo is 'not computed', just skip it (no line added)

    return chars


def analyze_multiband_timeline(mono, sr, n_segments=200):
    """
    Compute energy per frequency band over time.
    Returns a dict with time axis and energy array per band.
    """
    n_fft = 2048
    hop_length = max(1, len(mono) // n_segments)
    if hop_length < 256:
        hop_length = 256

    # Compute magnitude spectrogram
    S = np.abs(librosa.stft(mono, n_fft=n_fft, hop_length=hop_length))
    freqs = librosa.fft_frequencies(sr=sr, n_fft=n_fft)
    times = librosa.times_like(S, sr=sr, hop_length=hop_length)

    # For each band, compute energy per frame
    band_energy_timeline = {}
    for name, flow, fhigh in FREQ_BANDS:
        mask = (freqs >= flow) & (freqs < fhigh)
        if np.any(mask):
            # Energy per frame in dB for readability
            band_energy = np.sum(S[mask] ** 2, axis=0)
            # Convert to dB, floor at -80
            with np.errstate(divide='ignore'):
                band_db = 10 * np.log10(np.maximum(band_energy, 1e-10))
            band_energy_timeline[name] = band_db
        else:
            band_energy_timeline[name] = np.full(len(times), -80.0)

    return {
        'times': times,
        'bands': band_energy_timeline,
    }


def analyze_dynamic_range_timeline(mono, sr):
    """
    Sliding window peak vs RMS analysis.
    Returns time series of peak, RMS, and instantaneous crest factor.
    """
    window_size = int(0.050 * sr)  # 50 ms windows
    hop_size = int(0.020 * sr)     # 20 ms hop

    if len(mono) < window_size:
        return {
            'times': np.array([0.0]),
            'peak_db': np.array([db(np.max(np.abs(mono)))]),
            'rms_db': np.array([db(np.sqrt(np.mean(mono ** 2)))]),
            'crest_instant': np.array([0.0]),
        }

    n_windows = (len(mono) - window_size) // hop_size + 1
    peaks = np.zeros(n_windows)
    rmss = np.zeros(n_windows)
    times = np.zeros(n_windows)

    for i in range(n_windows):
        start = i * hop_size
        end = start + window_size
        block = mono[start:end]
        peaks[i] = np.max(np.abs(block))
        rmss[i] = np.sqrt(np.mean(block ** 2))
        times[i] = (start + window_size / 2) / sr

    peak_db = db(peaks)
    rms_db = db(rmss)
    crest_instant = peak_db - rms_db

    return {
        'times': times,
        'peak_db': peak_db,
        'rms_db': rms_db,
        'crest_instant': crest_instant,
    }


def analyze_structure_sections(mono, sr, n_sections_target=14):
    """
    Detect musical section boundaries using multi-feature novelty detection.
    Only used for Full Mix tracks.
    Returns boundaries (in seconds) and energy envelope.
    Uses two passes: chroma agglomerative + spectral flux novelty curve,
    then merges boundaries. Target: 12-16 sections for industrial music.
    """
    result = {
        'boundaries': [],
        'energy_envelope': None,
        'envelope_times': None,
        'success': False,
    }

    try:
        hop_length = 2048

        # --- Pass 1: Chroma-based agglomerative segmentation ---
        chroma = librosa.feature.chroma_cqt(y=mono, sr=sr, hop_length=hop_length)
        try:
            bounds_chroma = librosa.segment.agglomerative(chroma, k=n_sections_target)
            bt_chroma = librosa.frames_to_time(bounds_chroma, sr=sr, hop_length=hop_length)
        except Exception:
            bt_chroma = np.array([])

        # --- Pass 2: Spectral flux novelty curve ---
        novelty_sf = np.array([])  # M4.2: initialize before try for use in merge step
        try:
            S = np.abs(librosa.stft(mono, n_fft=2048, hop_length=hop_length))
            # Compute novelty from spectral flux
            novelty_sf = np.sqrt(np.sum(np.maximum(0, np.diff(S, axis=1)) ** 2, axis=0))
            # Smooth with a small kernel for finer detection
            kernel_size = max(1, int(sr / hop_length * 0.5))  # ~0.5s smoothing
            if kernel_size > 1:
                kernel = np.ones(kernel_size) / kernel_size
                novelty_sf = np.convolve(novelty_sf, kernel, mode='same')
            # Normalize
            novelty_sf = novelty_sf / (np.max(novelty_sf) + 1e-12)
            # Adaptive threshold: mean + 0.4 * std (lower than default for more sensitivity)
            threshold = np.mean(novelty_sf) + 0.4 * np.std(novelty_sf)
            min_distance = max(1, int(sr / hop_length * 2.0))  # at least 2s between boundaries
            peaks, _ = signal.find_peaks(novelty_sf, height=threshold,
                                          distance=min_distance, prominence=0.05)
            bt_novelty = librosa.frames_to_time(peaks, sr=sr, hop_length=hop_length)
        except Exception:
            bt_novelty = np.array([])

        # --- Pass 3: MFCC-based novelty for timbral changes ---
        try:
            mfcc = librosa.feature.mfcc(y=mono, sr=sr, n_mfcc=13, hop_length=hop_length)
            mfcc_delta = np.sqrt(np.sum(np.diff(mfcc, axis=1) ** 2, axis=0))
            kernel_size_m = max(1, int(sr / hop_length * 0.8))
            if kernel_size_m > 1:
                kernel_m = np.ones(kernel_size_m) / kernel_size_m
                mfcc_delta = np.convolve(mfcc_delta, kernel_m, mode='same')
            mfcc_delta = mfcc_delta / (np.max(mfcc_delta) + 1e-12)
            threshold_m = np.mean(mfcc_delta) + 0.5 * np.std(mfcc_delta)
            min_dist_m = max(1, int(sr / hop_length * 3.0))
            peaks_m, _ = signal.find_peaks(mfcc_delta, height=threshold_m,
                                             distance=min_dist_m, prominence=0.05)
            bt_mfcc = librosa.frames_to_time(peaks_m, sr=sr, hop_length=hop_length)
        except Exception:
            bt_mfcc = np.array([])

        # --- M4.2: Merge all boundaries with minimum spacing and section cap ---
        MIN_DISTANCE_SECONDS = 10.0  # M4.2: minimum 10s between sections
        MAX_SECTIONS_ABSOLUTE = 16   # M4.2: absolute cap on section count

        all_bounds = np.concatenate([bt_chroma, bt_novelty, bt_mfcc])

        # Compute salience for each boundary using spectral flux novelty curve
        all_salience = np.ones(len(all_bounds))
        if len(novelty_sf) > 0 and len(all_bounds) > 0:
            for i, b in enumerate(all_bounds):
                frame_idx = min(int(b * sr / hop_length), len(novelty_sf) - 1)
                all_salience[i] = novelty_sf[max(0, frame_idx)]

        # Sort by time
        sort_idx = np.argsort(all_bounds)
        all_bounds = all_bounds[sort_idx]
        all_salience = all_salience[sort_idx]

        # M4.2: Filter sections too close together (keep most salient)
        if len(all_bounds) > 1:
            keep = [0]
            for i in range(1, len(all_bounds)):
                if all_bounds[i] - all_bounds[keep[-1]] >= MIN_DISTANCE_SECONDS:
                    keep.append(i)
                else:
                    # Two sections too close: keep the more salient one
                    if all_salience[i] > all_salience[keep[-1]]:
                        keep[-1] = i
            all_bounds = all_bounds[keep]
            all_salience = all_salience[keep]

        # M4.2: Cap at MAX_SECTIONS_ABSOLUTE (keep most salient)
        if len(all_bounds) > MAX_SECTIONS_ABSOLUTE:
            top_idx = np.argsort(all_salience)[-MAX_SECTIONS_ABSOLUTE:]
            all_bounds = np.sort(all_bounds[top_idx])

        result['boundaries'] = all_bounds.tolist() if len(all_bounds) > 0 else []

        # Overall energy envelope (RMS)
        frame_length = 8192
        envelope_hop = 2048
        rms = librosa.feature.rms(y=mono, frame_length=frame_length, hop_length=envelope_hop)[0]
        envelope_times = librosa.frames_to_time(np.arange(len(rms)), sr=sr, hop_length=envelope_hop)
        with np.errstate(divide='ignore'):
            rms_db = 20 * np.log10(np.maximum(rms, 1e-10))
        result['energy_envelope'] = rms_db
        result['envelope_times'] = envelope_times
        result['success'] = True
    except Exception:
        pass

    return result


def compute_difference_spectrogram(full_mix_mono, individuals_monos, sr):
    """
    Compute the spectrogram difference between Full Mix and the sum of individual tracks.
    Returns the difference in dB, the sum spectrogram, and the mix spectrogram.
    Used to reveal the effect of master bus processing.
    """
    result = {
        'success': False,
        'diff_db': None,
        'times': None,
        'freqs': None,
    }

    if not individuals_monos:
        return result

    try:
        # Align lengths: use the minimum length
        min_len = min(len(full_mix_mono), *[len(m) for m in individuals_monos])
        if min_len < sr:  # Need at least 1 second
            return result

        mix = full_mix_mono[:min_len]
        # Sum all individuals, aligned to the same length
        summed = np.zeros(min_len, dtype=np.float32)
        for ind in individuals_monos:
            summed[:min(len(ind), min_len)] += ind[:min(len(ind), min_len)]

        # Compute magnitude spectrograms
        n_fft = 2048
        hop_length = 512
        S_mix = np.abs(librosa.stft(mix, n_fft=n_fft, hop_length=hop_length))
        S_sum = np.abs(librosa.stft(summed, n_fft=n_fft, hop_length=hop_length))

        # Compute dB difference (mix - sum in dB domain)
        S_mix_db = 20 * np.log10(np.maximum(S_mix, 1e-10))
        S_sum_db = 20 * np.log10(np.maximum(S_sum, 1e-10))
        diff_db = S_mix_db - S_sum_db

        # Clip to reasonable range for visualization
        diff_db = np.clip(diff_db, -24, 24)

        times = librosa.times_like(S_mix, sr=sr, hop_length=hop_length)
        freqs = librosa.fft_frequencies(sr=sr, n_fft=n_fft)

        result['success'] = True
        result['diff_db'] = diff_db
        result['times'] = times
        result['freqs'] = freqs
    except Exception:
        pass

    return result


def analyze_track(filepath, compute_tempo=False):
    """Complete analysis of a single track."""
    data, sr, is_stereo = load_audio(filepath)
    mono = to_mono(data)
    duration = len(mono) / sr

    result = {
        'filepath': filepath,
        'filename': os.path.basename(filepath),
        'duration': duration,
        'sample_rate': sr,
        'is_stereo': is_stereo,
        'num_channels': data.shape[1],
    }

    result['loudness'] = analyze_loudness(data, sr)
    result['spectrum'] = analyze_spectrum(mono, sr)
    result['temporal'] = analyze_temporal(mono, sr)
    # Tempo is only computed for Full Mix tracks (too unreliable on isolated tracks)
    if compute_tempo:
        result['tempo'] = analyze_tempo_dynamic(mono, sr)
    else:
        result['tempo'] = {
            'tempo_median': 0.0, 'tempo_min': 0.0, 'tempo_max': 0.0, 'tempo_std': 0.0,
            'confidence': 0.0, 'confidence_label': 'not computed (individual track)',
            'tempogram': None, 'tempogram_times': None, 'tempo_over_time': None,
            'reliable': False,
        }
    result['musical'] = analyze_musical(mono, sr)
    result['stereo'] = analyze_stereo(data, sr)
    # New analyses for v1.6
    result['multiband_timeline'] = analyze_multiband_timeline(mono, sr)
    result['dynamic_range_timeline'] = analyze_dynamic_range_timeline(mono, sr)
    result['anomalies'] = detect_anomalies(result)
    result['characteristics'] = describe_characteristics(result)

    result['_mono'] = mono
    result['_data'] = data

    return result
# ============================================================================
# CHART GENERATION - Enhanced with tempogram and spectral panorama
# ============================================================================

# Cyberpunk industrial theme colors
THEME = {
    'bg':           '#0a0a12',
    'panel':        '#1a1a24',
    'fg':           '#e8e8f0',
    'fg_dim':       '#8888a0',
    'accent1':      '#00d9ff',  # cyan
    'accent2':      '#b967ff',  # violet
    'accent3':      '#ff3d8b',  # magenta
    'accent4':      '#00ff9f',  # green
    'warning':      '#ffaa00',
    'critical':     '#ff3333',
    'grid':         '#333344',
}

plt.rcParams.update({
    'figure.facecolor': THEME['bg'],
    'axes.facecolor':   THEME['panel'],
    'savefig.facecolor': THEME['bg'],
    'axes.edgecolor':   THEME['fg_dim'],
    'axes.labelcolor':  THEME['fg'],
    'text.color':       THEME['fg'],
    'xtick.color':      THEME['fg_dim'],
    'ytick.color':      THEME['fg_dim'],
    'grid.color':       THEME['grid'],
    'grid.alpha':       0.5,
    'axes.grid':        True,
    'figure.dpi':       100,
    'font.family':      'sans-serif',
})

# Image quality presets for matplotlib renders
IMAGE_PRESETS = {
    'standard': {'dpi': 200, 'width': 1600, 'height': 900},
    'high':     {'dpi': 400, 'width': 3200, 'height': 1800},
}

# Excel display sizing — consistent regardless of DPI
EXCEL_IMAGE_MAX_WIDTH = 1400   # px, ~80% of Full HD usable width
EXCEL_IMAGE_MAX_HEIGHT = 800   # px, avoids excessive vertical space
EXCEL_ROW_HEIGHT_PX = 20       # default Excel row height in pixels


def make_page_header(fig, title, track_name, track_info=None):
    """Uniform page header. track_info: dict with type, category, etc."""
    fig.suptitle(title, fontsize=15, color=THEME['accent1'], fontweight='bold', y=0.975)
    display_name = track_name if len(track_name) <= 60 else track_name[:57] + '...'
    fig.text(0.5, 0.938, f"Track: {display_name}", ha='center', fontsize=9, color=THEME['fg_dim'])
    if track_info:
        info_parts = []
        if track_info.get('type'):
            info_parts.append(f"Type: {track_info['type']}")
        if track_info.get('category') and track_info['category'] != '(not set)':
            info_parts.append(f"Category: {track_info['category']}")
        if track_info.get('parent_bus') and track_info['parent_bus'] != 'None':
            info_parts.append(f"Parent BUS: {track_info['parent_bus']}")
        if info_parts:
            fig.text(0.5, 0.915, ' | '.join(info_parts), ha='center', fontsize=8,
                     color=THEME['accent2'], fontweight='bold')


def page_identity(analysis, track_info, style_name):
    """Page 1 - Identity and key metrics."""
    fig = plt.figure(figsize=(11, 8.5))
    make_page_header(fig, 'TRACK IDENTITY', analysis['filename'], track_info)

    ax = fig.add_subplot(111)
    ax.axis('off')

    L = analysis['loudness']
    S = analysis['spectrum']
    stereo = analysis['stereo']
    tempo = analysis['tempo']

    channels_str = 'Stereo' if analysis['is_stereo'] else 'Mono'

    col1 = [
        ('Duration', f"{analysis['duration']:.2f} s"),
        ('Sample rate', f"{analysis['sample_rate']} Hz"),
        ('Channels', channels_str),
        ('', ''),
        ('Peak level', f"{L['peak_db']:+.2f} dBFS"),
        ('True Peak', f"{L['true_peak_db']:+.2f} dBFS"),
        ('RMS level', f"{L['rms_db']:+.2f} dBFS"),
        ('', ''),
        ('Integrated LUFS', f"{L['lufs_integrated']:+.2f} LUFS" if np.isfinite(L['lufs_integrated']) else '—'),
        ('Short-term LUFS max', f"{L['lufs_short_term_max']:+.2f} LUFS" if np.isfinite(L['lufs_short_term_max']) else '—'),
        ('Loudness Range (LRA)', f"{L['lra']:.2f} LU"),
    ]

    col2 = [
        ('Crest factor', f"{L['crest_factor']:.2f} dB"),
        ('PLR (Peak to Loudness)', f"{L['plr']:.2f} dB"),
        ('PSR (Peak to Short-term)', f"{L['psr']:.2f} dB"),
        ('', ''),
        ('Dominant band', BAND_LABELS.get(S['dominant_band'], S['dominant_band'])),
        ('Spectral centroid', f"{S['centroid']:.0f} Hz"),
        ('Spectral rolloff (85%)', f"{S['rolloff']:.0f} Hz"),
        ('', ''),
        ('Tempo (median)', f"{tempo['tempo_median']:.1f} BPM" if tempo['reliable'] else 'Not reliable'),
        ('Phase correlation', f"{stereo['correlation']:.3f}" if stereo['is_stereo'] else '—'),
        ('Stereo width', f"{stereo['width_overall']:.3f}" if stereo['is_stereo'] else '—'),
    ]

    y0 = 0.87
    dy = 0.055
    for i, (label, value) in enumerate(col1):
        y = y0 - i * dy
        if label:
            ax.text(0.04, y, label, fontsize=10, color=THEME['fg_dim'], transform=ax.transAxes)
            ax.text(0.33, y, value, fontsize=11, color=THEME['fg'], fontweight='bold', transform=ax.transAxes)

    for i, (label, value) in enumerate(col2):
        y = y0 - i * dy
        if label:
            ax.text(0.52, y, label, fontsize=10, color=THEME['fg_dim'], transform=ax.transAxes)
            ax.text(0.82, y, value, fontsize=11, color=THEME['fg'], fontweight='bold', transform=ax.transAxes)

    ax.text(0.04, 0.04, f"Style context: {style_name}",
            fontsize=9, color=THEME['fg_dim'], transform=ax.transAxes)
    ax.text(0.04, 0.01, f"Generated: {datetime.now().strftime('%Y-%m-%d %H:%M')}",
            fontsize=9, color=THEME['fg_dim'], transform=ax.transAxes)

    plt.tight_layout(rect=[0, 0.03, 1, 0.90])
    return fig


def page_temporal(analysis, track_info):
    fig = plt.figure(figsize=(11, 8.5))
    make_page_header(fig, 'TEMPORAL VIEW', analysis['filename'], track_info)

    mono = analysis['_mono']
    sr = analysis['sample_rate']
    T = analysis['temporal']
    times = np.arange(len(mono)) / sr

    if len(mono) > 200000:
        step = len(mono) // 100000
        times_vis = times[::step]
        mono_vis = mono[::step]
    else:
        times_vis = times
        mono_vis = mono

    ax1 = fig.add_subplot(2, 1, 1)
    ax1.fill_between(times_vis, mono_vis, -mono_vis, color=THEME['accent1'], alpha=0.25, linewidth=0)
    ax1.plot(times_vis, mono_vis, color=THEME['accent1'], linewidth=0.3, alpha=0.6)
    ax1.plot(T['rms_times'], T['rms_envelope'], color=THEME['accent3'], linewidth=1.5, label='RMS')
    ax1.plot(T['rms_times'], -T['rms_envelope'], color=THEME['accent3'], linewidth=1.5)
    onset_display = T['onset_times']
    if len(onset_display) > 150:
        step = len(onset_display) // 150
        onset_display = onset_display[::step]
    for onset_t in onset_display:
        ax1.axvline(onset_t, color=THEME['accent4'], alpha=0.35, linewidth=0.4)
    ax1.set_xlabel('Time (s)')
    ax1.set_ylabel('Amplitude')
    ax1.set_title(f"Waveform with RMS envelope  |  Transients detected: {T['num_onsets']}",
                  color=THEME['accent1'], fontsize=10)
    ax1.set_ylim(-1.05, 1.05)
    ax1.legend(loc='upper right', fontsize=8)

    ax2 = fig.add_subplot(2, 1, 2)
    bin_centers = 0.5 * (T['hist_bins'][:-1] + T['hist_bins'][1:])
    ax2.bar(bin_centers, T['hist_values'], width=(T['hist_bins'][1] - T['hist_bins'][0]),
            color=THEME['accent2'], alpha=0.7, edgecolor=THEME['accent2'])
    ax2.set_yscale('log')
    ax2.set_xlabel('Amplitude')
    ax2.set_ylabel('Count')
    ax2.set_title('Amplitude distribution histogram', color=THEME['accent1'], fontsize=10)
    ax2.set_xlim(-1.05, 1.05)

    plt.tight_layout(rect=[0, 0.03, 1, 0.90])
    return fig


def page_spectral(analysis, track_info):
    fig = plt.figure(figsize=(11, 8.5))
    make_page_header(fig, 'SPECTRAL VIEW', analysis['filename'], track_info)

    S = analysis['spectrum']

    ax1 = fig.add_subplot(1, 1, 1)
    freqs = S['freqs']
    spec_db = S['spectrum_db_normalized']
    mask = freqs > 0
    ax1.semilogx(freqs[mask], spec_db[mask], color=THEME['accent1'], linewidth=1.2)
    ax1.fill_between(freqs[mask], spec_db[mask], -100, color=THEME['accent1'], alpha=0.2)

    for pf, ph in S['peaks']:
        if pf > 20:
            ax1.plot(pf, ph, 'o', color=THEME['accent3'], markersize=6)
            ax1.annotate(f"{pf:.0f}Hz", (pf, ph), textcoords="offset points",
                         xytext=(5, 5), fontsize=7, color=THEME['accent3'])

    ax1.set_xlim(20, 20000)
    ax1.set_ylim(-80, 5)
    ax1.set_xlabel('Frequency (Hz)')
    ax1.set_ylabel('Magnitude (dB)')
    ax1.set_title('Average spectrum (FFT)', color=THEME['accent1'], fontsize=10)

    band_colors = ['#2a1a4a', '#1a3a2a', '#4a3a1a', '#4a1a2a', '#4a1a3a', '#2a1a4a', '#1a3a4a']
    for i, (name, flow, fhigh) in enumerate(FREQ_BANDS):
        ax1.axvspan(flow, fhigh, alpha=0.15, color=band_colors[i % len(band_colors)])

    plt.tight_layout(rect=[0, 0.03, 1, 0.90])
    return fig


def page_spectrogram(analysis, track_info):
    fig = plt.figure(figsize=(11, 8.5))
    make_page_header(fig, 'SPECTRO-TEMPORAL VIEW', analysis['filename'], track_info)

    mono = analysis['_mono']
    sr = analysis['sample_rate']

    ax1 = fig.add_subplot(2, 1, 1)
    D = librosa.amplitude_to_db(np.abs(librosa.stft(mono, n_fft=2048, hop_length=512)), ref=np.max)
    img1 = librosa.display.specshow(D, sr=sr, x_axis='time', y_axis='log', ax=ax1,
                                      cmap='magma', vmin=-80, vmax=0)
    ax1.set_title('Spectrogram (log scale)', color=THEME['accent1'], fontsize=10)
    fig.colorbar(img1, ax=ax1, format='%+2.0f dB', pad=0.01)

    ax2 = fig.add_subplot(2, 1, 2)
    mel = librosa.feature.melspectrogram(y=mono, sr=sr, n_mels=128, fmax=sr/2)
    mel_db = librosa.power_to_db(mel, ref=np.max)
    img2 = librosa.display.specshow(mel_db, sr=sr, x_axis='time', y_axis='mel', ax=ax2,
                                      cmap='magma', vmin=-80, vmax=0)
    ax2.set_title('Mel spectrogram (human perception)', color=THEME['accent1'], fontsize=10)
    fig.colorbar(img2, ax=ax2, format='%+2.0f dB', pad=0.01)

    plt.tight_layout(rect=[0, 0.03, 1, 0.90])
    return fig


def page_musical(analysis, track_info):
    """Page 5 - Musical analysis.
    For individual tracks: chromagram + tonal info only.
    For Full Mix tracks: chromagram + tempogram + tempo stats.
    """
    fig = plt.figure(figsize=(11, 8.5))
    make_page_header(fig, 'MUSICAL ANALYSIS', analysis['filename'], track_info)

    M = analysis['musical']
    tempo = analysis['tempo']
    sr = analysis['sample_rate']
    is_full_mix = track_info.get('type') == 'Full Mix'

    if is_full_mix and tempo.get('tempogram') is not None:
        # Full Mix: chromagram + tempogram + stats
        # Chromagram (top)
        ax1 = fig.add_subplot(3, 1, 1)
        img = librosa.display.specshow(M['chroma'], sr=sr, x_axis='time', y_axis='chroma',
                                         ax=ax1, cmap='magma')
        ax1.set_title('Chromagram (pitch classes over time)', color=THEME['accent1'], fontsize=10)
        fig.colorbar(img, ax=ax1, pad=0.01)

        # Tempogram (middle)
        ax2 = fig.add_subplot(3, 1, 2)
        tg = tempo['tempogram']
        tg_times = tempo['tempogram_times']
        try:
            tempo_freqs = librosa.tempo_frequencies(tg.shape[0], hop_length=512, sr=sr)
            valid = (tempo_freqs >= 30) & (tempo_freqs <= 250)
            if np.any(valid):
                tg_display = tg[valid]
                freqs_display = tempo_freqs[valid]
                img = ax2.imshow(tg_display, aspect='auto', origin='lower', cmap='magma',
                                  extent=[tg_times[0], tg_times[-1], freqs_display[-1], freqs_display[0]])
                if tempo.get('tempo_over_time') is not None and len(tempo['tempo_over_time']) == len(tg_times):
                    ax2.plot(tg_times, tempo['tempo_over_time'], color=THEME['accent4'],
                             linewidth=1.5, alpha=0.9, label='Detected tempo')
                    ax2.legend(loc='upper right', fontsize=8)
                ax2.set_ylabel('Tempo (BPM)')
                ax2.set_xlabel('Time (s)')
                ax2.set_ylim(30, 250)
                fig.colorbar(img, ax=ax2, pad=0.01)
                conf_text = f"Confidence: {tempo['confidence_label']}"
                ax2.set_title(f"Tempogram - tempo evolution over time  |  {conf_text}",
                              color=THEME['accent1'], fontsize=10)
        except Exception:
            ax2.text(0.5, 0.5, 'Tempogram unavailable',
                     ha='center', va='center', transform=ax2.transAxes, color=THEME['fg_dim'])

        # Info panel (bottom)
        ax3 = fig.add_subplot(3, 1, 3)
        ax3.axis('off')

        info_lines = []
        if M['is_tonal']:
            info_lines.append(('Dominant pitch class', M['dominant_note']))
            info_lines.append(('Tonal strength', f"{M['tonal_strength']:.2f} (tonal content)"))
        else:
            info_lines.append(('Tonal character', 'Non-tonal / percussive mix'))
            info_lines.append(('Tonal strength', f"{M['tonal_strength']:.2f}"))

        if tempo['reliable']:
            info_lines.append(('Tempo (median)', f"{tempo['tempo_median']:.1f} BPM"))
            if tempo['tempo_max'] - tempo['tempo_min'] > 5:
                info_lines.append(('Tempo range', f"{tempo['tempo_min']:.1f} - {tempo['tempo_max']:.1f} BPM"))
                info_lines.append(('Tempo variability', f"std = {tempo['tempo_std']:.1f} BPM"))
            info_lines.append(('Confidence', tempo['confidence_label']))

        info_lines.append(('Transients count', f"{analysis['temporal']['num_onsets']}"))

        y0 = 0.90
        for i, (label, value) in enumerate(info_lines):
            y = y0 - i * 0.11
            ax3.text(0.08, y, label, fontsize=11, color=THEME['fg_dim'], transform=ax3.transAxes)
            ax3.text(0.45, y, value, fontsize=12, color=THEME['accent1'], fontweight='bold',
                     transform=ax3.transAxes)
    else:
        # Individual track: chromagram only (larger) + simple info
        ax1 = fig.add_subplot(2, 1, 1)
        img = librosa.display.specshow(M['chroma'], sr=sr, x_axis='time', y_axis='chroma',
                                         ax=ax1, cmap='magma')
        ax1.set_title('Chromagram (pitch classes over time)', color=THEME['accent1'], fontsize=10)
        fig.colorbar(img, ax=ax1, pad=0.01)

        # Info panel
        ax2 = fig.add_subplot(2, 1, 2)
        ax2.axis('off')

        info_lines = []
        if M['is_tonal']:
            info_lines.append(('Dominant pitch class', M['dominant_note']))
            info_lines.append(('Tonal strength', f"{M['tonal_strength']:.2f} (tonal content detected)"))
        else:
            info_lines.append(('Tonal character', 'Non-tonal / percussive'))
            info_lines.append(('Tonal strength', f"{M['tonal_strength']:.2f}"))

        info_lines.append(('Transients count', f"{analysis['temporal']['num_onsets']}"))

        y0 = 0.85
        for i, (label, value) in enumerate(info_lines):
            y = y0 - i * 0.12
            ax2.text(0.10, y, label, fontsize=12, color=THEME['fg_dim'], transform=ax2.transAxes)
            ax2.text(0.50, y, value, fontsize=14, color=THEME['accent1'], fontweight='bold',
                     transform=ax2.transAxes)

        # Note about tempo
        ax2.text(0.10, 0.15,
                 'Tempo detection is only performed on the Full Mix track, '
                 'where it reflects the complete rhythmic context.',
                 fontsize=9, color=THEME['fg_dim'], transform=ax2.transAxes,
                 style='italic', wrap=True)

    plt.tight_layout(rect=[0, 0.03, 1, 0.90])
    return fig


def page_stereo(analysis, track_info):
    """Page 6 - Stereo analysis with spectral panorama (new)."""
    fig = plt.figure(figsize=(11, 8.5))
    make_page_header(fig, 'STEREO ANALYSIS', analysis['filename'], track_info)

    stereo = analysis['stereo']
    sr = analysis['sample_rate']

    if not stereo['is_stereo']:
        ax = fig.add_subplot(111)
        ax.axis('off')
        ax.text(0.5, 0.5, 'MONO TRACK\n\nStereo analysis not applicable',
                ha='center', va='center',
                fontsize=24, color=THEME['fg_dim'], transform=ax.transAxes)
        plt.tight_layout(rect=[0, 0.03, 1, 0.90])
        return fig

    # Top-left: Width per band
    ax1 = fig.add_subplot(2, 3, 1)
    band_names_short = ['Sub', 'Bass', 'L-Mid', 'Mid', 'H-Mid', 'Pres', 'Air']
    width_values = [stereo['width_per_band'][name] for name, _, _ in FREQ_BANDS]
    ax1.barh(range(len(band_names_short)), width_values, color=THEME['accent1'], alpha=0.7,
             edgecolor=THEME['accent2'])
    ax1.set_yticks(range(len(band_names_short)))
    ax1.set_yticklabels(band_names_short, fontsize=8)
    ax1.set_xlabel('Width', fontsize=8)
    ax1.set_title('Width per band', color=THEME['accent1'], fontsize=9)
    ax1.set_xlim(0, 1)

    # Top-center: Correlation + width big display
    ax2 = fig.add_subplot(2, 3, 2)
    ax2.axis('off')
    corr = stereo['correlation']
    width = stereo['width_overall']
    corr_color = THEME['accent4'] if corr > 0.3 else (THEME['warning'] if corr > -0.2 else THEME['critical'])

    ax2.text(0.5, 0.85, 'Phase correlation', ha='center', fontsize=10, color=THEME['fg_dim'], transform=ax2.transAxes)
    ax2.text(0.5, 0.62, f"{corr:+.3f}", ha='center', fontsize=26, color=corr_color, fontweight='bold', transform=ax2.transAxes)
    ax2.text(0.5, 0.35, 'Overall width', ha='center', fontsize=10, color=THEME['fg_dim'], transform=ax2.transAxes)
    ax2.text(0.5, 0.12, f"{width:.3f}", ha='center', fontsize=26, color=THEME['accent1'], fontweight='bold', transform=ax2.transAxes)

    # Top-right: Lissajous
    ax3 = fig.add_subplot(2, 3, 3)
    L = analysis['_data'][:, 0]
    R = analysis['_data'][:, 1]
    step = max(1, len(L) // 4000)
    ax3.scatter(L[::step], R[::step], s=1, color=THEME['accent1'], alpha=0.3)
    ax3.set_xlim(-1, 1)
    ax3.set_ylim(-1, 1)
    ax3.set_xlabel('L', fontsize=8)
    ax3.set_ylabel('R', fontsize=8)
    ax3.set_title('Lissajous (L vs R)', color=THEME['accent1'], fontsize=9)
    ax3.axhline(0, color=THEME['fg_dim'], linewidth=0.5)
    ax3.axvline(0, color=THEME['fg_dim'], linewidth=0.5)
    ax3.set_aspect('equal')

    # Bottom-left: Mid/Side spectrum
    ax4 = fig.add_subplot(2, 3, 4)
    n_fft = 4096
    SM_spec = np.mean(np.abs(librosa.stft(stereo['mid'], n_fft=n_fft)), axis=1)
    SS_spec = np.mean(np.abs(librosa.stft(stereo['side'], n_fft=n_fft)), axis=1)
    freqs_ms = librosa.fft_frequencies(sr=sr, n_fft=n_fft)
    SM_db = db(SM_spec / (np.max(SM_spec) + 1e-12))
    SS_db = db(SS_spec / (np.max(SM_spec) + 1e-12))
    mask = freqs_ms > 20
    ax4.semilogx(freqs_ms[mask], SM_db[mask], color=THEME['accent1'], label='Mid', linewidth=1.2)
    ax4.semilogx(freqs_ms[mask], SS_db[mask], color=THEME['accent3'], label='Side', linewidth=1.2)
    ax4.set_xlim(20, 20000)
    ax4.set_ylim(-80, 5)
    ax4.set_xlabel('Freq (Hz)', fontsize=8)
    ax4.set_ylabel('Mag (dB)', fontsize=8)
    ax4.set_title('Mid / Side spectrum', color=THEME['accent1'], fontsize=9)
    ax4.legend(loc='upper right', fontsize=7)

    # Bottom-center + bottom-right: Spectral panorama (NEW)
    ax5 = fig.add_subplot(2, 3, (5, 6))
    if stereo.get('pan_per_freq') is not None:
        pan = stereo['pan_per_freq']
        energy = stereo['pan_energy']
        pan_freqs = stereo['pan_freqs']
        freq_mask = (pan_freqs >= 20) & (pan_freqs <= 20000)

        # Normalize energy for color intensity
        energy_norm = energy[freq_mask] / (np.max(energy[freq_mask]) + 1e-12)
        energy_db = 20 * np.log10(np.maximum(energy_norm, 1e-4))
        # Clamp to -60 to 0
        energy_db = np.clip(energy_db, -60, 0)

        scatter = ax5.scatter(pan_freqs[freq_mask], pan[freq_mask],
                               c=energy_db, cmap='magma',
                               s=8, alpha=0.9, vmin=-60, vmax=0)
        ax5.set_xscale('log')
        ax5.set_xlim(20, 20000)
        ax5.set_ylim(-1.1, 1.1)
        ax5.axhline(0, color=THEME['fg_dim'], linewidth=0.8, linestyle='--', alpha=0.7)
        ax5.axhline(1, color=THEME['fg_dim'], linewidth=0.4, alpha=0.3)
        ax5.axhline(-1, color=THEME['fg_dim'], linewidth=0.4, alpha=0.3)
        ax5.set_xlabel('Frequency (Hz)', fontsize=9)
        ax5.set_ylabel('Pan (-1 = Left, +1 = Right)', fontsize=9)
        ax5.set_title('Spectral Panorama - where each frequency sits in stereo field',
                      color=THEME['accent1'], fontsize=10)
        ax5.set_yticks([-1, -0.5, 0, 0.5, 1])
        ax5.set_yticklabels(['L', '-0.5', 'C', '+0.5', 'R'])
        cbar = fig.colorbar(scatter, ax=ax5, pad=0.01)
        cbar.set_label('Energy (dB)', fontsize=8)

    plt.tight_layout(rect=[0, 0.03, 1, 0.90])
    return fig


def page_characteristics(analysis, track_info, style_name):
    """Page 7 - Detected characteristics (descriptive, not prescriptive)."""
    fig = plt.figure(figsize=(11, 8.5))
    make_page_header(fig, 'DETECTED CHARACTERISTICS', analysis['filename'], track_info)

    ax = fig.add_subplot(111)
    ax.axis('off')

    chars = analysis['characteristics']
    anomalies = analysis['anomalies']

    y = 0.88

    # Anomalies first (if any)
    if anomalies:
        critical_anoms = [a for a in anomalies if a[0] == 'critical']
        warning_anoms = [a for a in anomalies if a[0] == 'warning']
        info_anoms = [a for a in anomalies if a[0] == 'info']

        if critical_anoms:
            ax.text(0.04, y, '[CRITICAL]', fontsize=13, color=THEME['critical'],
                    fontweight='bold', transform=ax.transAxes)
            y -= 0.05
            for sev, desc in critical_anoms:
                ax.text(0.07, y, '- ' + desc, fontsize=10, color=THEME['fg'],
                        transform=ax.transAxes)
                y -= 0.04
            y -= 0.02

        if warning_anoms:
            ax.text(0.04, y, '[WARNING]', fontsize=13, color=THEME['warning'],
                    fontweight='bold', transform=ax.transAxes)
            y -= 0.05
            for sev, desc in warning_anoms:
                ax.text(0.07, y, '- ' + desc, fontsize=10, color=THEME['fg'],
                        transform=ax.transAxes)
                y -= 0.04
            y -= 0.02

        if info_anoms:
            ax.text(0.04, y, '[NOTICE]', fontsize=13, color=THEME['accent1'],
                    fontweight='bold', transform=ax.transAxes)
            y -= 0.05
            for sev, desc in info_anoms:
                ax.text(0.07, y, '- ' + desc, fontsize=10, color=THEME['fg'],
                        transform=ax.transAxes)
                y -= 0.04
            y -= 0.03

    # Descriptive characteristics
    ax.text(0.04, y, 'OBJECTIVE OBSERVATIONS',
            fontsize=13, color=THEME['accent2'], fontweight='bold', transform=ax.transAxes)
    y -= 0.05

    for cat, desc in chars:
        if y < 0.05:
            break
        ax.text(0.07, y, '- ' + desc, fontsize=10, color=THEME['fg'], transform=ax.transAxes)
        y -= 0.045

    # Footer
    ax.text(0.04, 0.02,
            f"Note: This page presents objective measurements only. "
            f"Style-specific interpretation and recommendations should be made via AI analysis of the full report set.",
            fontsize=8, color=THEME['fg_dim'], style='italic', transform=ax.transAxes, wrap=True)

    plt.tight_layout(rect=[0, 0.03, 1, 0.90])
    return fig


def page_multiband_timeline(analysis, track_info):
    """Page 8 - Multiband energy timeline: energy per frequency band over time."""
    fig = plt.figure(figsize=(11, 8.5))
    make_page_header(fig, 'MULTIBAND ENERGY TIMELINE', analysis['filename'], track_info)

    mbt = analysis['multiband_timeline']
    times = mbt['times']
    bands = mbt['bands']

    # Single large plot with all 7 bands as separate lines
    ax = fig.add_subplot(1, 1, 1)

    band_colors_list = [
        '#b967ff',  # sub - violet
        '#8a5bff',  # bass - blue-violet
        '#00d9ff',  # low-mid - cyan
        '#00ff9f',  # mid - green
        '#ffcc00',  # high-mid - yellow
        '#ff8800',  # presence - orange
        '#ff3d8b',  # air - magenta
    ]

    for i, (name, flow, fhigh) in enumerate(FREQ_BANDS):
        if name in bands:
            ax.plot(times, bands[name],
                     label=BAND_LABELS[name],
                     color=band_colors_list[i],
                     linewidth=1.2, alpha=0.9)

    ax.set_xlabel('Time (s)', fontsize=11)
    ax.set_ylabel('Energy (dB)', fontsize=11)
    ax.set_title('Energy per frequency band across the track duration',
                 color=THEME['accent1'], fontsize=12, pad=15)
    ax.legend(loc='lower right', fontsize=9, ncol=2,
              framealpha=0.85, facecolor=THEME['panel'],
              edgecolor=THEME['fg_dim'], labelcolor=THEME['fg'])
    ax.grid(True, alpha=0.3)

    # Add contextual note
    fig.text(0.5, 0.04,
             'This view reveals arrangement dynamics: watch when each frequency range enters, '
             'dominates, or disappears. Useful for identifying sections, detecting sub drops, '
             'and spotting masking in time.',
             ha='center', fontsize=8, color=THEME['fg_dim'], style='italic', wrap=True)

    plt.tight_layout(rect=[0, 0.07, 1, 0.90])
    return fig


def page_dynamic_range_map(analysis, track_info):
    """Page 9 - Dynamic range map: peak vs RMS sliding + crest factor over time."""
    fig = plt.figure(figsize=(11, 8.5))
    make_page_header(fig, 'DYNAMIC RANGE MAP', analysis['filename'], track_info)

    drt = analysis['dynamic_range_timeline']
    times = drt['times']
    peak_db = drt['peak_db']
    rms_db = drt['rms_db']
    crest = drt['crest_instant']

    # Two subplots: peak+RMS on top, crest factor on bottom
    ax1 = fig.add_subplot(2, 1, 1)
    ax1.fill_between(times, peak_db, rms_db,
                      color=THEME['accent1'], alpha=0.2, label='Dynamic headroom')
    ax1.plot(times, peak_db, color=THEME['accent3'], linewidth=1.3, label='Peak (50ms window)')
    ax1.plot(times, rms_db, color=THEME['accent4'], linewidth=1.3, label='RMS (50ms window)')
    ax1.set_xlabel('Time (s)', fontsize=10)
    ax1.set_ylabel('Level (dB)', fontsize=10)
    ax1.set_title('Sliding Peak vs RMS - headroom between the two lines shows dynamic range',
                   color=THEME['accent1'], fontsize=11, pad=10)
    ax1.legend(loc='lower right', fontsize=9,
               framealpha=0.85, facecolor=THEME['panel'],
               edgecolor=THEME['fg_dim'], labelcolor=THEME['fg'])
    ax1.grid(True, alpha=0.3)
    ax1.set_ylim(-80, 5)

    # Crest factor over time
    ax2 = fig.add_subplot(2, 1, 2)
    # Color the crest line by severity
    ax2.fill_between(times, crest, 0,
                      where=(crest >= 12), color=THEME['accent4'], alpha=0.3,
                      label='High dynamic (>12 dB)', interpolate=True)
    ax2.fill_between(times, crest, 0,
                      where=((crest >= 6) & (crest < 12)), color=THEME['warning'], alpha=0.3,
                      label='Moderate (6-12 dB)', interpolate=True)
    ax2.fill_between(times, crest, 0,
                      where=(crest < 6), color=THEME['critical'], alpha=0.3,
                      label='Compressed (<6 dB)', interpolate=True)
    ax2.plot(times, crest, color=THEME['fg'], linewidth=1.2)
    ax2.axhline(12, color=THEME['accent4'], linewidth=0.6, alpha=0.5, linestyle='--')
    ax2.axhline(6, color=THEME['warning'], linewidth=0.6, alpha=0.5, linestyle='--')

    # Global crest factor reference line
    global_crest = analysis['loudness']['crest_factor']
    ax2.axhline(global_crest, color=THEME['accent1'], linewidth=1.5, linestyle=':',
                label=f'Global crest factor: {global_crest:.1f} dB')

    ax2.set_xlabel('Time (s)', fontsize=10)
    ax2.set_ylabel('Crest factor (dB)', fontsize=10)
    ax2.set_title('Instantaneous crest factor - reveals where compression acts most',
                   color=THEME['accent1'], fontsize=11, pad=10)
    ax2.legend(loc='upper right', fontsize=8,
               framealpha=0.85, facecolor=THEME['panel'],
               edgecolor=THEME['fg_dim'], labelcolor=THEME['fg'])
    ax2.grid(True, alpha=0.3)
    ax2.set_ylim(-2, max(30, np.max(crest) + 3))

    plt.tight_layout(rect=[0, 0.03, 1, 0.90])
    return fig



# ============================================================================
# EXCEL REPORT
# ============================================================================

METRIC_GLOSSARY = {
    'LUFS': (
        "Integrated loudness in LUFS (Loudness Units Full Scale).\n"
        "Target: -14 LUFS for Spotify, YouTube, Apple Music.\n\n"
        "Interpretation:\n"
        "\u2022 -14 to -12: Ideal streaming range\n"
        "\u2022 -16 to -14: Slightly quiet, acceptable\n"
        "\u2022 Below -16: May sound quiet vs other tracks\n"
        "\u2022 Above -10: Will be turned down by streaming platforms"
    ),
    'Peak (dB)': (
        "Maximum inter-sample peak level in dBTP (True Peak).\n"
        "Target: Below -1.0 dBTP for safe streaming.\n\n"
        "Interpretation:\n"
        "\u2022 Below -1.5 dBTP: Safe for all codecs\n"
        "\u2022 -1.5 to -1.0 dBTP: Acceptable\n"
        "\u2022 -1.0 to -0.5 dBTP: Risk of clipping on lossy codecs\n"
        "\u2022 Above -0.5 dBTP: Likely to clip on MP3/AAC/Ogg"
    ),
    'Crest (dB)': (
        "Dynamic range indicator (Peak - RMS in dB).\n"
        "Higher = more dynamic, lower = more compressed.\n\n"
        "Interpretation:\n"
        "\u2022 Above 12 dB: High dynamics (acoustic, classical)\n"
        "\u2022 8-12 dB: Moderate dynamics (typical mix)\n"
        "\u2022 6-8 dB: Compressed (modern pop/rock)\n"
        "\u2022 Below 6 dB: Heavily compressed (loudness war)"
    ),
    'Stereo Width': (
        "Stereo image width (Mid/Side energy ratio).\n"
        "0 = Mono, 0.5 = balanced, >0.7 = very wide.\n\n"
        "Interpretation:\n"
        "\u2022 0-0.2: Essentially mono (bass, kick)\n"
        "\u2022 0.2-0.5: Narrow stereo (vocals, snare)\n"
        "\u2022 0.5-0.8: Normal stereo (guitars, keys)\n"
        "\u2022 0.8-1.0: Wide stereo (pads, ambience)\n"
        "\u2022 Above 1.0: Out-of-phase content detected"
    ),
    'Dom. Band': (
        "Frequency band with highest energy concentration.\n"
        "Reveals the spectral center of gravity for this track.\n"
        "Useful for identifying masking conflicts between tracks."
    ),
    'Centroid (Hz)': (
        "Spectral centroid \u2014 brightness indicator in Hz.\n"
        "Low values = dark/warm, high values = bright/harsh.\n\n"
        "Typical ranges:\n"
        "\u2022 Below 1500 Hz: Dark/warm mix\n"
        "\u2022 1500-3000 Hz: Balanced\n"
        "\u2022 Above 3000 Hz: Bright/aggressive"
    ),
    'Duration (s)': "Track length in seconds.",
    'PLR': (
        "Peak-to-Loudness Ratio (Peak dBFS minus LUFS).\n"
        "Measures available headroom.\n\n"
        "Interpretation:\n"
        "\u2022 Above 12 dB: Very dynamic\n"
        "\u2022 8-12 dB: Healthy headroom\n"
        "\u2022 6-8 dB: Moderate\n"
        "\u2022 Below 6 dB: May clip on normalization"
    ),
    'PSR': (
        "Peak-to-Short-term Loudness Ratio.\n"
        "Measures instantaneous headroom.\n\n"
        "Interpretation:\n"
        "\u2022 Above 10 dB: Very dynamic transients\n"
        "\u2022 6-10 dB: Healthy transient headroom\n"
        "\u2022 Below 6 dB: Heavily limited/compressed"
    ),
    'LRA': (
        "Loudness Range in LU (Loudness Units).\n"
        "Measures variation in loudness over time.\n\n"
        "Interpretation:\n"
        "\u2022 Above 10 LU: High variation (classical, film)\n"
        "\u2022 6-10 LU: Moderate variation (typical music)\n"
        "\u2022 3-6 LU: Low variation (pop, EDM)\n"
        "\u2022 Below 3 LU: Very consistent (heavily processed)"
    ),
    'Phase Correlation': (
        "Phase correlation between L and R channels.\n"
        "+1 = Mono (identical), 0 = Unrelated, -1 = Out of phase.\n\n"
        "Interpretation:\n"
        "\u2022 0.8 to 1.0: Mono-compatible, safe\n"
        "\u2022 0.5 to 0.8: Good stereo, mono-compatible\n"
        "\u2022 0.0 to 0.5: Wide stereo, check mono\n"
        "\u2022 Below 0.0: Phase issues, will cancel in mono"
    ),
    'Flatness': (
        "Spectral flatness (0 to 1).\n"
        "0 = purely tonal, 1 = noise-like.\n\n"
        "Interpretation:\n"
        "\u2022 Below 0.05: Tonal/pitched content\n"
        "\u2022 0.05-0.2: Mixed content\n"
        "\u2022 Above 0.2: Noisy/percussive content"
    ),
    'RMS': (
        "Root Mean Square level in dBFS.\n"
        "Represents average perceived loudness.\n"
        "Lower than LUFS due to different weighting."
    ),
    'True Peak': (
        "Maximum inter-sample peak level in dBTP.\n"
        "Can exceed 0 dBFS due to reconstruction.\n"
        "Critical for broadcast/streaming compliance.\n\n"
        "Target: Below -1.0 dBTP for safe delivery."
    ),
    'Track': "Filename of the analyzed audio bounce.",
    'Type': "Track role: Individual (single instrument), BUS (submix), Full Mix (master bounce).",
    'Category': "Instrument category for grouping and masking analysis.",
    # Dashboard alternate header names
    'True Peak (dBFS)': (
        "Maximum inter-sample peak level in dBTP.\n"
        "Target: Below -1.0 dBTP for safe streaming.\n\n"
        "\u2022 Below -1.5: Safe for all codecs\n"
        "\u2022 -1.0 to -0.5: Risk of clipping on lossy codecs\n"
        "\u2022 Above -0.5: Likely to clip on MP3/AAC/Ogg"
    ),
    'RMS (dB)': (
        "Root Mean Square level in dBFS.\n"
        "Represents average perceived loudness.\n"
        "Compare with LUFS for loudness context."
    ),
    'PLR (dB)': (
        "Peak-to-Loudness Ratio (Peak dBFS minus LUFS).\n"
        "Higher = more headroom.\n"
        "\u2022 Above 8 dB: Healthy  \u2022 Below 6 dB: Risk of clipping"
    ),
    'PSR (dB)': (
        "Peak-to-Short-term Loudness Ratio.\n"
        "Measures instantaneous headroom.\n"
        "\u2022 Above 6 dB: Healthy  \u2022 Below 6 dB: Heavily limited"
    ),
    'LRA (LU)': (
        "Loudness Range in LU (Loudness Units).\n"
        "Measures loudness variation over time.\n"
        "\u2022 Above 10: Dynamic  \u2022 3-6: Consistent  \u2022 Below 3: Flat"
    ),
    'Width': (
        "Stereo image width (Mid/Side energy ratio).\n"
        "0 = Mono, 0.5 = balanced, >0.7 = very wide.\n"
        "Bass/kick should be narrow, pads/FX can be wide."
    ),
    'Correlation': (
        "Phase correlation between L and R channels.\n"
        "+1 = Mono, 0 = Unrelated, -1 = Out of phase.\n"
        "\u2022 Above 0.5: Mono-safe  \u2022 Below 0.0: Phase problems"
    ),
    'Rolloff (Hz)': (
        "Spectral rolloff at 85% energy threshold.\n"
        "Frequency below which 85% of spectral energy lies.\n"
        "Lower = darker mix, higher = brighter mix."
    ),
    'Family': "Instrument family grouping (e.g. Rhythmic, Tonal, FX).\nUsed for category-level analysis.",
    # Health Score
    'Health Score': (
        "Overall mix health indicator (0-100).\n"
        "Combines multiple metrics weighted by importance.\n\n"
        "Interpretation:\n"
        "\u2022 90-100: Excellent, ready for release\n"
        "\u2022 75-89: Good, minor issues possible\n"
        "\u2022 50-74: Fair, review flagged areas\n"
        "\u2022 Below 50: Needs attention, check anomalies"
    ),
    'Mix Health Score': (
        "Overall mix health indicator (0-100).\n"
        "Combines Loudness, Dynamics, Spectral, Stereo,\n"
        "and Anomaly sub-scores. Tracks improvement over versions."
    ),
    # Spectral bands
    'Sub Energy %': (
        "Sub band energy (20-60 Hz).\n"
        "Contains sub-bass fundamentals.\n"
        "Excessive sub energy causes muddiness on small speakers."
    ),
    'Bass Energy %': (
        "Bass band energy (60-250 Hz).\n"
        "Contains fundamentals of bass and kick.\n"
        "\u2022 High: Strong bass presence\n"
        "\u2022 Low: Thin or bright mix"
    ),
    'Low-Mid Energy %': (
        "Low-Mid band energy (250-500 Hz).\n"
        "Contains warmth and body of instruments.\n"
        "Excess causes muddiness and boominess."
    ),
    'Mid Energy %': (
        "Mid band energy (500-2000 Hz).\n"
        "Contains presence of vocals and melodic instruments.\n"
        "Critical for clarity and definition."
    ),
    'High-Mid Energy %': (
        "High-Mid band energy (2-6 kHz).\n"
        "Contains attack and presence.\n"
        "Excess causes harshness and listening fatigue."
    ),
    'Presence Energy %': (
        "Presence band energy (6-12 kHz).\n"
        "Contains sibilance and air.\n"
        "Important for clarity and detail."
    ),
    'Air Energy %': (
        "Air band energy (12-20 kHz).\n"
        "Contains highest harmonics and air.\n"
        "Adds sparkle and openness to the mix."
    ),
    # Version Tracking specific metrics
    'Full Mix LUFS': (
        "Integrated loudness of the full mix bounce.\n"
        "Track this across versions to monitor loudness changes.\n"
        "Target: -14 LUFS for streaming platforms."
    ),
    'Full Mix True Peak (dBFS)': (
        "True peak level of the full mix bounce.\n"
        "Should decrease or stay below -1.0 dBTP across versions.\n"
        "Rising peaks may indicate over-processing."
    ),
    'Full Mix Crest (dB)': (
        "Crest factor of the full mix bounce.\n"
        "Track dynamic range across mix iterations.\n"
        "Decreasing values suggest increasing compression."
    ),
    'Full Mix PLR': (
        "Peak-to-Loudness Ratio of the full mix.\n"
        "Monitors headroom across versions.\n"
        "Should remain above 6 dB for safe normalization."
    ),
    'Full Mix Width': (
        "Stereo width of the full mix bounce.\n"
        "Track stereo image consistency across versions.\n"
        "Large changes may indicate stereo processing shifts."
    ),
    'Avg Individual Crest (dB)': (
        "Average crest factor across individual tracks.\n"
        "Monitors overall dynamic range of the session.\n"
        "Decreasing trend may indicate over-compression."
    ),
    'Anomaly count': (
        "Total number of detected anomalies.\n"
        "Should decrease across mix iterations.\n"
        "Rising count suggests new issues introduced."
    ),
    'Track count': (
        "Number of tracks analyzed.\n"
        "Track changes in session composition across versions."
    ),
    # Track Comparison specific
    'LUFS': (
        "Integrated loudness in LUFS (Loudness Units Full Scale).\n"
        "Target: -14 LUFS for Spotify, YouTube, Apple Music.\n\n"
        "Interpretation:\n"
        "\u2022 -14 to -12: Ideal streaming range\n"
        "\u2022 -16 to -14: Slightly quiet, acceptable\n"
        "\u2022 Below -16: May sound quiet vs other tracks\n"
        "\u2022 Above -10: Will be turned down by streaming platforms"
    ),
    'Peak (dBFS)': (
        "Peak level in dBFS (decibels Full Scale).\n"
        "Should stay below 0 dBFS to avoid clipping.\n"
        "\u2022 Below -1.0: Safe headroom\n"
        "\u2022 Above -0.5: Clipping risk"
    ),
    'Crest Factor (dB)': (
        "Dynamic range indicator (Peak - RMS in dB).\n"
        "Higher = more dynamic, lower = more compressed.\n"
        "\u2022 Above 12: Dynamic  \u2022 6-8: Compressed  \u2022 Below 6: Over-compressed"
    ),
    'Dominant Band': (
        "Frequency band with highest energy concentration.\n"
        "Reveals the spectral center of gravity.\n"
        "Compare across tracks to identify masking conflicts."
    ),
}

# M7.4: Anomaly explanations for contextual comments
ANOMALY_COMMENTS = {
    'clipping risk': (
        "Peak level is at or near 0 dBFS.\n"
        "Impact: Audible distortion, inter-sample clipping.\n"
        "Suggestion: Reduce gain or add a limiter with -1 dB ceiling."
    ),
    'very little headroom': (
        "Peak level is close to 0 dBFS with minimal margin.\n"
        "Impact: Risk of clipping on lossy codec conversion.\n"
        "Suggestion: Pull back the master fader or reduce gain by 1-2 dB."
    ),
    'inter-sample clipping': (
        "True Peak exceeds 0 dBFS between samples.\n"
        "Impact: Distortion on D/A conversion and lossy encoding.\n"
        "Suggestion: Use a True Peak limiter set to -1.0 dBTP ceiling."
    ),
    'serious mono compatibility': (
        "Phase correlation is strongly negative.\n"
        "Impact: Severe signal cancellation when summed to mono.\n"
        "Suggestion: Check for inverted polarity or excessive stereo widening."
    ),
    'mono compatibility concern': (
        "Phase correlation is low (below 0.3).\n"
        "Impact: May sound thin or hollow in mono playback.\n"
        "Suggestion: Check stereo widening plugins and verify phase alignment."
    ),
    'nearly silent': (
        "RMS level is extremely low.\n"
        "Impact: Track may be inaudible in the mix.\n"
        "Suggestion: Verify this is intentional, or increase gain."
    ),
    'resonance peaks': (
        "Strong narrow-band energy buildup detected.\n"
        "Impact: Fatiguing frequencies, uneven tonal balance.\n"
        "Suggestion: Apply a narrow EQ cut at the flagged frequencies."
    ),
    'heavy compression': (
        "Crest factor indicates very low dynamic range.\n"
        "Impact: Flat, lifeless sound lacking punch and transients.\n"
        "Suggestion: Ease compression ratio/threshold or use parallel compression."
    ),
    'wide stereo image': (
        "Stereo width is very high.\n"
        "Impact: May collapse or phase-cancel in mono playback.\n"
        "Suggestion: Check mono compatibility and reduce widening if needed."
    ),
}


def _safe_sheet_name(name, max_len=31):
    """Sanitize name for Excel sheet name rules."""
    for ch in '[]:*?/\\':
        name = name.replace(ch, '_')
    return name[:max_len] if len(name) > max_len else name


def _fig_to_image(fig, quality='standard'):
    """Render matplotlib figure to openpyxl Image via temp PNG.
    Returns (Image, tmp_path, row_span) where row_span is the number of
    Excel rows the image occupies at its display size.
    quality: 'standard' (200 DPI, 1600x900) or 'high' (400 DPI, 3200x1800)."""
    import tempfile
    from openpyxl.drawing.image import Image as XlImage

    preset = IMAGE_PRESETS.get(quality, IMAGE_PRESETS['standard'])
    dpi = preset['dpi']
    target_width = preset['width']
    target_height = preset['height']

    fig.set_size_inches(target_width / dpi, target_height / dpi)
    tmp = tempfile.NamedTemporaryFile(suffix='.png', delete=False)
    tmp_path = tmp.name
    tmp.close()
    fig.savefig(tmp_path, dpi=dpi, facecolor=fig.get_facecolor(), bbox_inches='tight')
    plt.close(fig)
    img = XlImage(tmp_path)

    # Scale display size to fit within Excel max bounds, preserving aspect ratio
    orig_w, orig_h = img.width, img.height
    scale = min(EXCEL_IMAGE_MAX_WIDTH / orig_w, EXCEL_IMAGE_MAX_HEIGHT / orig_h, 1.0)
    img.width = int(orig_w * scale)
    img.height = int(orig_h * scale)

    row_span = int(img.height / EXCEL_ROW_HEIGHT_PX) + 2
    return img, tmp_path, row_span


def _xl_write_header(ws, title, subtitle=''):
    """Write sheet header in row 1-2 with formatting."""
    from openpyxl.styles import Font, PatternFill, Alignment
    _init_ma_fonts()
    header_fill = PatternFill('solid', fgColor='0A0A12')
    accent_font = MA_FONT_TITLE
    sub_font = Font(name='Calibri', size=11, color='8888A0')

    ws.merge_cells('A1:J1')
    ws['A1'] = title
    ws['A1'].font = accent_font
    ws['A1'].fill = header_fill
    ws['A1'].alignment = Alignment(horizontal='left')
    if subtitle:
        ws.merge_cells('A2:J2')
        ws['A2'] = subtitle
        ws['A2'].font = sub_font
        ws['A2'].fill = header_fill
    return 4  # next row to write data


def _xl_add_nav_row(ws, row, sheet_names_ordered, current_idx):
    """Add prev/next/Index/Anomalies navigation links to a track sheet."""
    from openpyxl.styles import Font, Alignment
    nav_font = MA_FONT_LINK
    sep_font = MA_FONT_SMALL

    col = 3  # Start in column C to leave A-B for metrics
    # Back to Index
    c = ws.cell(row=row, column=col, value='< Index')
    c.font = nav_font
    c.hyperlink = '#Index!A1'
    col += 1

    ws.cell(row=row, column=col, value=' | ').font = sep_font
    col += 1

    # Anomalies
    c = ws.cell(row=row, column=col, value='Anomalies')
    c.font = nav_font
    c.hyperlink = '#Anomalies!A1'
    col += 1

    ws.cell(row=row, column=col, value=' | ').font = sep_font
    col += 1

    # Summary
    c = ws.cell(row=row, column=col, value='Summary')
    c.font = nav_font
    c.hyperlink = '#Summary!A1'
    col += 1

    ws.cell(row=row, column=col, value=' | ').font = sep_font
    col += 1

    # Previous track
    if current_idx > 0:
        prev_name = sheet_names_ordered[current_idx - 1]
        c = ws.cell(row=row, column=col, value=f'< Prev')
        c.font = nav_font
        c.hyperlink = f'#{prev_name}!A1'
    else:
        ws.cell(row=row, column=col, value='< Prev').font = Font(name='Calibri', size=9, color='333344')
    col += 1

    ws.cell(row=row, column=col, value=' | ').font = sep_font
    col += 1

    # Next track
    if current_idx < len(sheet_names_ordered) - 1:
        next_name = sheet_names_ordered[current_idx + 1]
        c = ws.cell(row=row, column=col, value='Next >')
        c.font = nav_font
        c.hyperlink = f'#{next_name}!A1'
    else:
        ws.cell(row=row, column=col, value='Next >').font = Font(name='Calibri', size=9, color='333344')

    return row + 1


def _xl_add_sheet_nav(ws, row, current_sheet=None, nav_targets=None):
    """M7.5: Add a navigation bar to non-track sheets.
    Provides links to Index, Summary, Dashboard, Anomalies, etc.
    nav_targets: optional list of (label, sheet_name) to override defaults."""
    from openpyxl.styles import Font
    _init_ma_fonts()
    nav_font = Font(name='Calibri', size=9, color='00D9FF', underline='single')
    sep_font = MA_FONT_SMALL
    dim_font = Font(name='Calibri', size=9, color='333344')

    if nav_targets is None:
        nav_targets = [
            ('Index', 'Index'),
            ('Summary', 'Summary'),
            ('Dashboard', 'Dashboard'),
            ('Anomalies', 'Anomalies'),
            ('Health Score', 'Mix Health Score'),
            ('AI Context', 'AI Context'),
        ]
    col = 1
    for i, (label, sheet_name) in enumerate(nav_targets):
        if i > 0:
            ws.cell(row=row, column=col, value=' | ').font = sep_font
            col += 1
        c = ws.cell(row=row, column=col, value=label)
        if current_sheet and sheet_name == current_sheet:
            c.font = dim_font  # current sheet not clickable
        else:
            c.font = nav_font
            c.hyperlink = f'#{sheet_name}!A1'
        col += 1
    return row


def _xl_add_comment(cell, text, width=300, height=150):
    """Add a comment (tooltip) to a cell with optional dimensions."""
    from openpyxl.comments import Comment
    comment = Comment(text, 'Mix Analyzer')
    comment.width = width
    comment.height = height
    cell.comment = comment


def _apply_clean_layout(ws):
    """
    M5.1: Apply clean layout to a worksheet.
    Hides Excel gridlines and row/column headers for a clean,
    dashboard-like appearance consistent with the cyberpunk theme.

    Called on every worksheet created in Mix Analyzer reports.
    """
    ws.sheet_view.showGridLines = False
    ws.sheet_view.showRowColHeaders = False


# ============================================================================
# AI CONTEXT SHEET (v1.8) — Dense consolidated metrics for AI ingestion
# ============================================================================

def encode_anomalies(anomaly_list):
    """Encode a list of (severity, description) anomaly tuples into compact codes.
    Returns 'OK' if no anomalies, otherwise codes joined with ' | '.
    """
    import re
    if not anomaly_list:
        return 'OK'

    codes = []
    for sev, desc in anomaly_list:
        m = re.search(r'Peak level at ([+\-\d.]+) dBFS.*clipping risk', desc)
        if m:
            codes.append(f'PEAK_CLIP:{m.group(1)}')
            continue
        m = re.search(r'Peak level at ([+\-\d.]+) dBFS.*little headroom', desc)
        if m:
            codes.append(f'PEAK_HOT:{m.group(1)}')
            continue
        m = re.search(r'True Peak at ([+\-\d.]+) dBFS', desc)
        if m:
            codes.append(f'TP_OVER:{m.group(1)}')
            continue
        m = re.search(r'Phase correlation ([+\-\d.]+).*serious', desc)
        if m:
            codes.append(f'PHASE_CRIT:{m.group(1)}')
            continue
        m = re.search(r'Phase correlation ([+\-\d.]+).*mono compatibility', desc)
        if m:
            codes.append(f'PHASE:{m.group(1)}')
            continue
        m = re.search(r'RMS level very low \(([+\-\d.]+) dBFS\)', desc)
        if m:
            codes.append(f'RMS_LOW:{m.group(1)}')
            continue
        m = re.search(r'Strong resonance peaks detected at: (.+)', desc)
        if m:
            freq_nums = re.findall(r'(\d+)', m.group(1))
            codes.append(f'RES:{",".join(freq_nums)}')
            continue
        m = re.search(r'Very low crest factor \(([+\-\d.]+) dB\)', desc)
        if m:
            codes.append(f'CREST_LOW:{m.group(1)}')
            continue
        m = re.search(r'Very wide stereo image \(([+\-\d.]+)\)', desc)
        if m:
            codes.append(f'WIDTH_HIGH:{m.group(1)}')
            continue
        short = desc[:40].replace('|', '/').strip()
        codes.append(f'WARN:{short}')

    return ' | '.join(codes) if codes else 'OK'


def build_ai_context_sheet(workbook, analyses_with_info, style_name, log_fn=None,
                           nav_targets=None):
    """Build the AI Context sheet — dense consolidated metrics for AI ingestion."""
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter
    _init_ma_fonts()

    if log_fn is None:
        log_fn = lambda msg: None

    log_fn("    Excel: writing AI Context sheet...")

    ws = workbook.create_sheet('AI Context')
    _apply_clean_layout(ws)
    ws.sheet_properties.tabColor = '00D9FF'

    bg_fill = PatternFill('solid', fgColor='0A0A12')
    panel_fill = PatternFill('solid', fgColor='1A1A24')
    header_fill = PatternFill('solid', fgColor='1A3A5A')
    fullmix_fill = PatternFill('solid', fgColor='2A1A3A')
    section_fill = PatternFill('solid', fgColor='1A2A1A')
    thin_border = Border(
        left=Side(style='thin', color='333344'),
        right=Side(style='thin', color='333344'),
        top=Side(style='thin', color='333344'),
        bottom=Side(style='thin', color='333344'),
    )

    accent_font = MA_FONT_SUBHEADING
    header_font = MA_FONT_TABLE_HEADER
    data_font = MA_FONT_BODY
    dim_font = MA_FONT_DIM

    individuals = [(a, ti) for a, ti in analyses_with_info if ti.get('type') == 'Individual']
    buses = [(a, ti) for a, ti in analyses_with_info if ti.get('type') == 'BUS']
    full_mixes = [(a, ti) for a, ti in analyses_with_info if ti.get('type') == 'Full Mix']

    n_ind = len(individuals)
    n_bus = len(buses)
    n_fm = len(full_mixes)
    track_count_str = f'{n_ind} Individual'
    if n_bus:
        track_count_str += f' + {n_bus} BUS'
    if n_fm:
        track_count_str += f' + {n_fm} Full Mix'

    # ---- Header block ----
    ws.merge_cells('A1:Z1')
    ws['A1'] = 'AI CONTEXT \u2014 CONSOLIDATED TRACK METRICS'
    ws['A1'].font = MA_FONT_TITLE
    ws['A1'].fill = bg_fill

    row = 2
    for label, val in [('Generated:', datetime.now().strftime('%Y-%m-%d %H:%M')),
                        ('Tracks:', track_count_str),
                        ('Style:', style_name)]:
        ws.cell(row=row, column=1, value=label).font = dim_font
        ws.cell(row=row, column=2, value=val).font = dim_font
        ws.cell(row=row, column=1).fill = bg_fill
        ws.cell(row=row, column=2).fill = bg_fill
        row += 1

    # M7.5: Navigation bar
    row += 1
    _xl_add_sheet_nav(ws, row, current_sheet='AI Context', nav_targets=nav_targets)
    row += 1

    # ---- Anomaly codes legend ----
    row += 1
    ws.cell(row=row, column=1, value='ANOMALY CODES LEGEND').font = accent_font
    ws.cell(row=row, column=1).fill = bg_fill
    row += 1

    for code, desc in [
        ('OK', 'No anomalies detected'),
        ('RES:f1,f2,f3', 'Resonance peaks at listed frequencies (Hz)'),
        ('PHASE:val', 'Phase correlation concern (mono compat warning)'),
        ('PHASE_CRIT:val', 'Phase correlation critical (serious mono compat issue)'),
        ('RMS_LOW:val', 'RMS level very low \u2014 track nearly silent (dBFS)'),
        ('PEAK_HOT:val', 'Peak level close to clipping (dBFS)'),
        ('PEAK_CLIP:val', 'Peak level at/above clipping threshold (dBFS)'),
        ('TP_OVER:val', 'True peak exceeds 0 dBFS \u2014 inter-sample clipping'),
        ('CREST_LOW:val', 'Very low crest factor \u2014 heavy compression (dB)'),
        ('WIDTH_HIGH:val', 'Very wide stereo image \u2014 verify mono compat (0-1)'),
        ('WARN:text', 'Generic warning (fallback)'),
    ]:
        ws.cell(row=row, column=1, value=code).font = Font(
            name='Calibri', size=10, bold=True, color='FFAA00')
        ws.cell(row=row, column=2, value=desc).font = dim_font
        ws.cell(row=row, column=1).fill = bg_fill
        ws.cell(row=row, column=2).fill = bg_fill
        row += 1

    # ---- Column legend ----
    row += 1
    ws.cell(row=row, column=1, value='COLUMN LEGEND').font = accent_font
    ws.cell(row=row, column=1).fill = bg_fill
    row += 1

    column_schema = [
        ('track_name',     'Audio file name'),
        ('type',           'Individual / BUS / Full Mix'),
        ('category',       'Instrument category (e.g. Kick, Pad / Drone)'),
        ('family',         'Category family (e.g. Drums, Synth)'),
        ('lufs_int',       'Integrated LUFS (loudness units full scale)'),
        ('lufs_st_max',    'Maximum short-term LUFS (3s window)'),
        ('peak_db',        'Sample peak level (dBFS)'),
        ('true_peak_db',   'Inter-sample true peak level (dBFS, 4x oversampled)'),
        ('rms_db',         'RMS level (dBFS, average energy)'),
        ('crest_db',       'Crest factor: peak minus RMS (dB)'),
        ('plr_db',         'Peak-to-Loudness Ratio: peak minus LUFS (dB)'),
        ('psr_db',         'Peak-to-Short-term Ratio (dB)'),
        ('lra_lu',         'Loudness Range \u2014 macro dynamics (LU)'),
        ('dom_band',       'Frequency band with highest energy'),
        ('centroid_hz',    'Spectral centroid \u2014 brightness indicator (Hz)'),
        ('rolloff_hz',     '85% spectral rolloff frequency (Hz)'),
        ('flatness',       'Spectral flatness (0=tonal, 1=noise)'),
        ('pct_sub',        'Energy % in Sub band (20-60 Hz)'),
        ('pct_bass',       'Energy % in Bass band (60-250 Hz)'),
        ('pct_low_mid',    'Energy % in Low-Mid band (250-500 Hz)'),
        ('pct_mid',        'Energy % in Mid band (500-2000 Hz)'),
        ('pct_high_mid',   'Energy % in High-Mid band (2-4 kHz)'),
        ('pct_presence',   'Energy % in Presence band (4-8 kHz)'),
        ('pct_air',        'Energy % in Air band (8-20 kHz)'),
        ('stereo_width',   'Mid/Side energy ratio (0=mono, 1=full side)'),
        ('phase_corr',     'L/R phase correlation (+1=perfect mono compat)'),
        ('is_stereo',      'TRUE or FALSE'),
        ('dom_note',       'Dominant pitch class (e.g. C#, A)'),
        ('tonal_strength', 'Tonal peak-to-mean ratio (>1.8 = tonal)'),
        ('is_tonal',       'TRUE or FALSE'),
        ('tempo_bpm',      'Median detected tempo (BPM)'),
        ('tempo_conf',     'Tempo confidence label'),
        ('tempo_reliable', 'TRUE if tempo is reliable'),
        ('num_onsets',     'Number of detected transients'),
        ('duration_s',     'Track duration (seconds)'),
        ('sample_rate_hz', 'Audio sample rate (Hz)'),
        ('num_channels',   'Number of audio channels'),
        ('anomaly_codes',  'Compact anomaly codes joined with |'),
    ]

    for col_name, col_desc in column_schema:
        ws.cell(row=row, column=1, value=col_name).font = Font(
            name='Calibri', size=10, bold=True, color='00D9FF')
        ws.cell(row=row, column=2, value=col_desc).font = dim_font
        ws.cell(row=row, column=1).fill = bg_fill
        ws.cell(row=row, column=2).fill = bg_fill
        row += 1

    # ---- Track table ----
    row += 1
    table_header_row = row
    col_names = [c[0] for c in column_schema]

    for col_idx, col_name in enumerate(col_names, 1):
        c = ws.cell(row=row, column=col_idx, value=col_name)
        c.font = header_font
        c.fill = header_fill
        c.border = thin_border
        c.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
    row += 1

    def _extract_row_values(a, ti):
        L = a['loudness']
        S = a['spectrum']
        st = a['stereo']
        M = a['musical']
        tempo = a.get('tempo', {})
        cat = ti.get('category', '(not set)')
        family = CATEGORY_FAMILY.get(cat, 'Unknown')
        lufs_int = round(L['lufs_integrated'], 2) if np.isfinite(L['lufs_integrated']) else None
        lufs_st = round(L['lufs_short_term_max'], 2) if np.isfinite(L['lufs_short_term_max']) else None
        return [
            a['filename'], ti['type'], cat, family,
            lufs_int, lufs_st,
            round(L['peak_db'], 2), round(L['true_peak_db'], 2),
            round(L['rms_db'], 2), round(L['crest_factor'], 2),
            round(L['plr'], 2), round(L['psr'], 2), round(L['lra'], 2),
            BAND_LABELS.get(S['dominant_band'], S['dominant_band']),
            round(S['centroid'], 1), round(S['rolloff'], 1),
            round(S['flatness'], 6),
            round(S['band_energies'].get('sub', 0.0), 2),
            round(S['band_energies'].get('bass', 0.0), 2),
            round(S['band_energies'].get('low_mid', 0.0), 2),
            round(S['band_energies'].get('mid', 0.0), 2),
            round(S['band_energies'].get('high_mid', 0.0), 2),
            round(S['band_energies'].get('presence', 0.0), 2),
            round(S['band_energies'].get('air', 0.0), 2),
            round(st['width_overall'], 4) if st['is_stereo'] else 0.0,
            round(st['correlation'], 4) if st['is_stereo'] else 1.0,
            'TRUE' if st['is_stereo'] else 'FALSE',
            M['dominant_note'], round(M['tonal_strength'], 2),
            'TRUE' if M['is_tonal'] else 'FALSE',
            round(tempo.get('tempo_median', 0.0), 1),
            tempo.get('confidence_label', ''),
            'TRUE' if tempo.get('reliable', False) else 'FALSE',
            a['temporal']['num_onsets'],
            round(a['duration'], 2), a['sample_rate'], a['num_channels'],
            encode_anomalies(a.get('anomalies', [])),
        ]

    def _write_track_row(ws, row, vals, fill, font=None):
        if font is None:
            font = data_font
        for col_idx, v in enumerate(vals, 1):
            c = ws.cell(row=row, column=col_idx, value=v)
            c.font = font
            c.border = thin_border
            c.fill = fill
            if col_idx >= 5 and col_idx <= 37:
                c.alignment = Alignment(horizontal='center')

    for a, ti in individuals:
        _write_track_row(ws, row, _extract_row_values(a, ti), panel_fill)
        row += 1
    for a, ti in buses:
        _write_track_row(ws, row, _extract_row_values(a, ti), panel_fill)
        row += 1
    for a, ti in full_mixes:
        vals = _extract_row_values(a, ti)
        vals[0] = '*** FULL MIX ***'
        _write_track_row(ws, row, vals, fullmix_fill,
                         Font(name='Calibri', size=10, bold=True, color='E8E8F0'))
        row += 1

    last_col_letter = get_column_letter(len(col_names))
    ws.auto_filter.ref = f'A{table_header_row}:{last_col_letter}{max(row - 1, table_header_row + 1)}'

    # ---- Health Score breakdown ----
    row += 2
    ws.cell(row=row, column=1, value='MIX HEALTH SCORE BREAKDOWN').font = accent_font
    ws.cell(row=row, column=1).fill = bg_fill
    row += 1

    full_mix = full_mixes[0][0] if full_mixes else None
    try:
        loud_score, _, _ = _calc_loudness_score(individuals, full_mix)
        dyn_score, _, _ = _calc_dynamics_score(individuals, full_mix)
        spec_score, _, _ = _calc_spectral_balance_score(individuals, full_mix)
        stereo_score, _, _ = _calc_stereo_image_score(individuals, full_mix)
        anom_score, _, _ = _calc_anomalies_score(analyses_with_info)

        categories_hs = [
            ('SCORE_LOUDNESS', loud_score, 0.20),
            ('SCORE_DYNAMICS', dyn_score, 0.20),
            ('SCORE_SPECTRAL', spec_score, 0.25),
            ('SCORE_STEREO', stereo_score, 0.15),
            ('SCORE_ANOMALIES', anom_score, 0.20),
        ]
        global_score = round(sum(s * w for _, s, w in categories_hs), 1)

        for col_idx, h in enumerate(['Key', 'Value', 'Notes'], 1):
            c = ws.cell(row=row, column=col_idx, value=h)
            c.font = header_font
            c.fill = header_fill
            c.border = thin_border
        row += 1

        ws.cell(row=row, column=1, value='SCORE_TOTAL').font = accent_font
        ws.cell(row=row, column=2, value=global_score).font = Font(
            name='Calibri', size=12, bold=True, color='00FF9F')
        ws.cell(row=row, column=3, value='').font = dim_font
        for ci in range(1, 4):
            ws.cell(row=row, column=ci).fill = bg_fill
            ws.cell(row=row, column=ci).border = thin_border
        row += 1

        for name, score, weight in categories_hs:
            contrib = round(score * weight, 1)
            ws.cell(row=row, column=1, value=name).font = data_font
            ws.cell(row=row, column=2, value=round(score, 1)).font = data_font
            ws.cell(row=row, column=2).alignment = Alignment(horizontal='center')
            ws.cell(row=row, column=3, value=f'weight={weight:.2f}, contrib={contrib}').font = dim_font
            for ci in range(1, 4):
                ws.cell(row=row, column=ci).fill = bg_fill
                ws.cell(row=row, column=ci).border = thin_border
            row += 1
    except Exception:
        ws.cell(row=row, column=1, value='Health score computation unavailable.').font = dim_font
        ws.cell(row=row, column=1).fill = bg_fill
        row += 1

    # ---- Per-family aggregates ----
    row += 2
    ws.cell(row=row, column=1, value='PER-FAMILY AGGREGATES').font = accent_font
    ws.cell(row=row, column=1).fill = bg_fill
    row += 1

    family_groups = {}
    for a, ti in individuals:
        cat = ti.get('category', '(not set)')
        family = CATEGORY_FAMILY.get(cat, 'Unknown')
        if family not in family_groups:
            family_groups[family] = []
        family_groups[family].append((a, ti))

    for family_name, tracks in sorted(family_groups.items()):
        n = len(tracks)
        ws.cell(row=row, column=1, value=f'FAMILY: {family_name} (n={n})').font = Font(
            name='Calibri', size=11, bold=True, color='00D9FF')
        ws.cell(row=row, column=1).fill = section_fill
        ws.cell(row=row, column=1).border = thin_border
        row += 1

        lufs_vals = [a['loudness']['lufs_integrated'] for a, _ in tracks
                     if np.isfinite(a['loudness']['lufs_integrated'])]
        crest_vals = [a['loudness']['crest_factor'] for a, _ in tracks]
        width_vals = [a['stereo']['width_overall'] for a, _ in tracks if a['stereo']['is_stereo']]

        lufs_mean = round(float(np.mean(lufs_vals)), 2) if lufs_vals else None
        lufs_std = round(float(np.std(lufs_vals)), 2) if len(lufs_vals) >= 2 else 0.0
        crest_mean = round(float(np.mean(crest_vals)), 2) if crest_vals else None
        width_mean = round(float(np.mean(width_vals)), 3) if width_vals else None

        dom_bands = [a['spectrum']['dominant_band'] for a, _ in tracks]
        if dom_bands:
            from collections import Counter
            consensus_band = Counter(dom_bands).most_common(1)[0][0]
            consensus_label = BAND_LABELS.get(consensus_band, consensus_band)
        else:
            consensus_label = 'N/A'

        for col_idx, h in enumerate(['lufs_mean', 'lufs_std', 'crest_mean', 'width_mean', 'dom_band_consensus'], 1):
            c = ws.cell(row=row, column=col_idx, value=h)
            c.font = Font(name='Calibri', size=9, bold=True, color='8888A0')
            c.fill = bg_fill
            c.border = thin_border
        row += 1
        for col_idx, v in enumerate([lufs_mean, lufs_std, crest_mean, width_mean, consensus_label], 1):
            c = ws.cell(row=row, column=col_idx, value=v)
            c.font = data_font
            c.fill = bg_fill
            c.border = thin_border
            c.alignment = Alignment(horizontal='center')
        row += 1

        track_names = ', '.join(a['filename'] for a, _ in tracks)
        ws.cell(row=row, column=1, value='Tracks:').font = dim_font
        ws.cell(row=row, column=1).fill = bg_fill
        c = ws.cell(row=row, column=2, value=track_names)
        c.font = dim_font
        c.fill = bg_fill
        c.alignment = Alignment(wrap_text=True)
        row += 2

    # Column widths
    ws.column_dimensions['A'].width = 45
    ws.column_dimensions['B'].width = 18
    for col_idx in range(3, len(col_names) + 1):
        ws.column_dimensions[get_column_letter(col_idx)].width = 14

    ws.freeze_panes = f'A{table_header_row + 1}'
    _apply_dark_background(ws)
    log_fn("    Excel: AI Context sheet done.")


def generate_freq_conflicts_sheet(wb, analyses_with_info, default_threshold=15.0,
                                   default_min_tracks=2, log_fn=None,
                                   nav_targets=None):
    """
    Génère le sheet 'Freq Conflicts' (P3.1) dans le workbook donné.
    Utilise FREQ_BANDS_HIRES pour les bandes de fréquence.
    Filtre les pistes : Individual uniquement, exclut BUS et Full Mix.
    """
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.formatting.rule import ColorScaleRule, DataBarRule, FormulaRule
    _init_ma_fonts()
    from openpyxl.utils import get_column_letter

    if log_fn is None:
        log_fn = lambda msg: None

    log_fn("    Excel: writing Freq Conflicts sheet (P3.1)...")

    # Filter: Individual tracks only
    individuals = [(a, ti) for a, ti in analyses_with_info if ti.get('type') == 'Individual']
    if not individuals:
        ws = wb.create_sheet('Freq Conflicts')
        _apply_clean_layout(ws)
        ws.sheet_properties.tabColor = 'FF3333'
        _xl_write_header(ws, 'FREQUENCY CONFLICT DETECTOR', 'No individual tracks found.')
        _apply_dark_background(ws)
        return

    # Compute hires band energies for each individual track
    track_names = []
    all_energies = []  # list of dicts {band_label: energy_value}
    for a, ti in individuals:
        track_names.append(a['filename'])
        hires = compute_hires_band_energies(a['_mono'], a['sample_rate'])
        all_energies.append(hires)

    band_labels = [label for label, _, _ in FREQ_BANDS_HIRES]
    n_bands = len(band_labels)
    n_tracks = len(track_names)

    # Build energy matrix and normalize per band (% of max in that band)
    import numpy as np
    matrix = np.zeros((n_bands, n_tracks))
    for t_idx, energies in enumerate(all_energies):
        for b_idx, label in enumerate(band_labels):
            matrix[b_idx, t_idx] = energies.get(label, 0.0)

    # Normalize each band row: percentage of max energy in that band
    norm_matrix = np.zeros_like(matrix)
    for b_idx in range(n_bands):
        row_max = np.max(matrix[b_idx]) if np.max(matrix[b_idx]) > 0 else 1.0
        norm_matrix[b_idx] = 100.0 * matrix[b_idx] / row_max

    # Theme styles
    bg_fill = PatternFill('solid', fgColor='0A0A12')
    panel_fill = PatternFill('solid', fgColor='1A1A24')
    header_fill = PatternFill('solid', fgColor='1A3A5A')
    accent_font = MA_FONT_SUBHEADING
    header_font = MA_FONT_TABLE_HEADER
    data_font = MA_FONT_BODY
    dim_font = MA_FONT_DIM
    thin_border = Border(
        left=Side(style='thin', color='333344'),
        right=Side(style='thin', color='333344'),
        top=Side(style='thin', color='333344'),
        bottom=Side(style='thin', color='333344'),
    )
    conflict_fill = PatternFill('solid', fgColor='FF3333')
    ok_fill = PatternFill('solid', fgColor='00FF9F')

    # Create sheet
    ws = wb.create_sheet('Freq Conflicts')
    _apply_clean_layout(ws)
    ws.sheet_properties.tabColor = 'FF3D8B'

    # Row 1: Title
    ws.merge_cells('A1:J1')
    ws['A1'] = 'FREQUENCY CONFLICT DETECTOR'
    ws['A1'].font = MA_FONT_TITLE
    ws['A1'].fill = bg_fill

    # Row 2: Threshold parameter
    ws['A2'] = 'Conflict threshold (% of max band energy)'
    ws['A2'].font = dim_font
    ws['A2'].fill = bg_fill
    ws['B2'] = default_threshold
    ws['B2'].font = accent_font
    ws['B2'].fill = panel_fill
    ws['B2'].border = thin_border

    # Row 3: Min tracks parameter
    ws['A3'] = 'Min tracks for conflict'
    ws['A3'].font = dim_font
    ws['A3'].fill = bg_fill
    ws['B3'] = default_min_tracks
    ws['B3'].font = accent_font
    ws['B3'].fill = panel_fill
    ws['B3'].border = thin_border

    # Row 4: M7.5 Navigation bar
    _xl_add_sheet_nav(ws, 4, nav_targets=nav_targets)
    # Row 5: Headers
    header_row = 5
    ws.cell(row=header_row, column=1, value='Frequency Band').font = header_font
    ws.cell(row=header_row, column=1).fill = header_fill
    ws.cell(row=header_row, column=1).border = thin_border
    ws.cell(row=header_row, column=1).alignment = Alignment(horizontal='center', vertical='center')

    for t_idx, name in enumerate(track_names):
        col = t_idx + 2
        c = ws.cell(row=header_row, column=col, value=name[:20])
        c.font = header_font
        c.fill = header_fill
        c.border = thin_border
        c.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)

    conflict_count_col = n_tracks + 2
    status_col = n_tracks + 3

    ws.cell(row=header_row, column=conflict_count_col, value='Conflict count').font = header_font
    ws.cell(row=header_row, column=conflict_count_col).fill = header_fill
    ws.cell(row=header_row, column=conflict_count_col).border = thin_border
    ws.cell(row=header_row, column=conflict_count_col).alignment = Alignment(horizontal='center')

    ws.cell(row=header_row, column=status_col, value='Status').font = header_font
    ws.cell(row=header_row, column=status_col).fill = header_fill
    ws.cell(row=header_row, column=status_col).border = thin_border
    ws.cell(row=header_row, column=status_col).alignment = Alignment(horizontal='center')

    # Data rows (one per band)
    data_start_row = 6
    for b_idx, label in enumerate(band_labels):
        row = data_start_row + b_idx
        # Band label
        c = ws.cell(row=row, column=1, value=label)
        c.font = data_font
        c.fill = bg_fill
        c.border = thin_border

        # Track energy values (normalized %)
        for t_idx in range(n_tracks):
            col = t_idx + 2
            val = round(norm_matrix[b_idx, t_idx], 1)
            c = ws.cell(row=row, column=col, value=val)
            c.font = data_font
            c.fill = bg_fill
            c.border = thin_border
            c.number_format = '0.0'

        # Conflict count formula: COUNTIF over track columns where value >= threshold in B2
        first_track_col = get_column_letter(2)
        last_track_col = get_column_letter(n_tracks + 1)
        countif_range = f'{first_track_col}{row}:{last_track_col}{row}'
        formula_count = f'=COUNTIF({countif_range},">="&$B$2)'
        c = ws.cell(row=row, column=conflict_count_col, value=formula_count)
        c.font = data_font
        c.fill = bg_fill
        c.border = thin_border
        c.alignment = Alignment(horizontal='center')

        # Status formula: IF conflict_count >= B3 then "CONFLICT" else "OK"
        cc_letter = get_column_letter(conflict_count_col)
        formula_status = f'=IF({cc_letter}{row}>=$B$3,"CONFLICT","OK")'
        c = ws.cell(row=row, column=status_col, value=formula_status)
        c.font = data_font
        c.fill = bg_fill
        c.border = thin_border
        c.alignment = Alignment(horizontal='center')

    data_end_row = data_start_row + n_bands - 1

    # --- Conditional Formatting ---
    # 1. Color scale on data cells (green -> yellow -> red)
    data_range = (f'{get_column_letter(2)}{data_start_row}:'
                  f'{get_column_letter(n_tracks + 1)}{data_end_row}')
    ws.conditional_formatting.add(
        data_range,
        ColorScaleRule(
            start_type='num', start_value=0, start_color='00FF9F',
            mid_type='num', mid_value=50, mid_color='FFFF00',
            end_type='num', end_value=100, end_color='FF3333'))

    # 2. Status column: red if CONFLICT, green if OK
    status_range = f'{get_column_letter(status_col)}{data_start_row}:{get_column_letter(status_col)}{data_end_row}'
    ws.conditional_formatting.add(
        status_range,
        FormulaRule(
            formula=[f'{get_column_letter(status_col)}{data_start_row}="CONFLICT"'],
            fill=conflict_fill))
    ws.conditional_formatting.add(
        status_range,
        FormulaRule(
            formula=[f'{get_column_letter(status_col)}{data_start_row}="OK"'],
            fill=ok_fill,
            font=Font(name='Calibri', size=10, bold=True, color='000000')))

    # 3. Data bars on conflict count column
    cc_range = f'{get_column_letter(conflict_count_col)}{data_start_row}:{get_column_letter(conflict_count_col)}{data_end_row}'
    ws.conditional_formatting.add(
        cc_range,
        DataBarRule(start_type='num', start_value=0,
                    end_type='num', end_value=n_tracks,
                    color='FF3D8B'))

    # --- Autofilter ---
    last_col_letter = get_column_letter(status_col)
    ws.auto_filter.ref = f'A{header_row}:{last_col_letter}{data_end_row}'

    # --- Freeze panes (row 6, col B = headers + band labels visible) ---
    ws.freeze_panes = 'B6'

    # --- Column widths ---
    ws.column_dimensions['A'].width = 20
    for t_idx in range(n_tracks):
        ws.column_dimensions[get_column_letter(t_idx + 2)].width = 12
    ws.column_dimensions[get_column_letter(conflict_count_col)].width = 15
    ws.column_dimensions[get_column_letter(status_col)].width = 12

    _apply_dark_background(ws)
    log_fn("    Excel: Freq Conflicts sheet done.")


def generate_track_comparison_sheet(workbook, analyses_with_info, log_fn=None):
    """
    Génère le sheet 'Track Comparison' (P3.2) dans le workbook donné.
    Crée des sélecteurs de pistes (data validations) et un tableau
    comparatif avec deltas calculés via formules INDEX/MATCH.
    Filtre les pistes : Individual uniquement, exclut BUS et Full Mix.
    """
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.formatting.rule import ColorScaleRule
    from openpyxl.utils import get_column_letter
    from openpyxl.worksheet.datavalidation import DataValidation
    _init_ma_fonts()

    if log_fn is None:
        log_fn = lambda msg: None

    log_fn("    Excel: writing Track Comparison sheet (P3.2)...")

    # Filter: Individual tracks only
    individuals = [(a, ti) for a, ti in analyses_with_info if ti.get('type') == 'Individual']
    if not individuals:
        ws = workbook.create_sheet('Track Comparison')
        _apply_clean_layout(ws)
        ws.sheet_properties.tabColor = 'FF8B3D'
        _xl_write_header(ws, 'TRACK COMPARISON TOOL', 'No individual tracks found.')
        _apply_dark_background(ws)
        return

    track_names = [a['filename'] for a, ti in individuals]

    # --- Metrics definition ---
    # (display_name, key_path, is_numeric)
    # key_path is used to build the data table; the column index in the hidden
    # data table maps to these metrics in order.
    band_keys = [name for name, _, _ in FREQ_BANDS]
    band_display = {
        'sub': 'Sub Energy %', 'bass': 'Bass Energy %',
        'low_mid': 'Low-Mid Energy %', 'mid': 'Mid Energy %',
        'high_mid': 'High-Mid Energy %', 'presence': 'Presence Energy %',
        'air': 'Air Energy %',
    }
    metrics = [
        ('LUFS', True),
        ('Peak (dBFS)', True),
        ('Crest Factor (dB)', True),
        ('Stereo Width', True),
        ('PLR (dB)', True),
        ('PSR (dB)', True),
        ('Dominant Band', False),
    ]
    for bk in band_keys:
        metrics.append((band_display[bk], True))

    n_metrics = len(metrics)

    # --- Build hidden data sheet ---
    ws_data = workbook.create_sheet('_track_data')
    _apply_clean_layout(ws_data)
    ws_data.sheet_state = 'hidden'

    # Header row
    data_headers = ['Track Name'] + [m[0] for m in metrics]
    for col, h in enumerate(data_headers, 1):
        ws_data.cell(row=1, column=col, value=h)

    # Data rows
    for r_idx, (a, ti) in enumerate(individuals):
        row = r_idx + 2
        L = a['loudness']
        S = a['spectrum']
        st = a['stereo']
        ws_data.cell(row=row, column=1, value=a['filename'])
        col = 2
        # LUFS
        ws_data.cell(row=row, column=col,
                     value=round(L['lufs_integrated'], 2) if np.isfinite(L['lufs_integrated']) else None)
        col += 1
        # Peak
        ws_data.cell(row=row, column=col, value=round(L['peak_db'], 2))
        col += 1
        # Crest
        ws_data.cell(row=row, column=col, value=round(L['crest_factor'], 2))
        col += 1
        # Width
        ws_data.cell(row=row, column=col,
                     value=round(st['width_overall'], 3) if st['is_stereo'] else None)
        col += 1
        # PLR
        ws_data.cell(row=row, column=col, value=round(L['plr'], 2))
        col += 1
        # PSR
        ws_data.cell(row=row, column=col, value=round(L['psr'], 2))
        col += 1
        # Dominant Band
        ws_data.cell(row=row, column=col,
                     value=BAND_LABELS.get(S['dominant_band'], S['dominant_band']))
        col += 1
        # Band energies (7 bands)
        for bk in band_keys:
            ws_data.cell(row=row, column=col,
                         value=round(S['band_energies'].get(bk, 0.0), 2))
            col += 1

    n_data_rows = len(individuals)
    data_range_end = n_data_rows + 1  # last data row in _track_data

    # --- Theme styles ---
    bg_fill = PatternFill('solid', fgColor='0A0A12')
    panel_fill = PatternFill('solid', fgColor='1A1A24')
    header_fill = PatternFill('solid', fgColor='1A3A5A')
    accent_font = MA_FONT_SUBHEADING
    header_font = MA_FONT_TABLE_HEADER
    data_font = MA_FONT_BODY
    dim_font = MA_FONT_DIM
    thin_border = Border(
        left=Side(style='thin', color='333344'),
        right=Side(style='thin', color='333344'),
        top=Side(style='thin', color='333344'),
        bottom=Side(style='thin', color='333344'),
    )

    # --- Main sheet ---
    ws = workbook.create_sheet('Track Comparison')
    _apply_clean_layout(ws)
    ws.sheet_properties.tabColor = 'FF8B3D'

    # Row 1: Title
    ws.merge_cells('A1:K1')
    ws['A1'] = 'TRACK COMPARISON TOOL'
    ws['A1'].font = MA_FONT_TITLE
    ws['A1'].fill = bg_fill

    # --- Section 1: Track selectors (rows 2-6) ---
    track_list_str = ','.join(track_names)
    track_list_with_empty = ',' + track_list_str  # leading comma = empty option

    dv_required = DataValidation(type='list', formula1=f'"{track_list_str}"', allow_blank=False)
    dv_required.error = 'Please select a track from the list.'
    dv_required.errorTitle = 'Invalid track'
    ws.add_data_validation(dv_required)

    dv_optional = DataValidation(type='list', formula1=f'"{track_list_with_empty}"', allow_blank=True)
    dv_optional.error = 'Please select a track from the list or leave empty.'
    dv_optional.errorTitle = 'Invalid track'
    ws.add_data_validation(dv_optional)

    selectors = [
        ('Track A (reference)', track_names[0] if len(track_names) > 0 else '', dv_required),
        ('Track B', track_names[1] if len(track_names) > 1 else (track_names[0] if track_names else ''), dv_required),
        ('Track C (optional)', '', dv_optional),
        ('Track D (optional)', '', dv_optional),
    ]

    for i, (label, default, dv) in enumerate(selectors):
        row = i + 2
        ws.cell(row=row, column=1, value=label).font = dim_font
        ws.cell(row=row, column=1).fill = bg_fill
        c = ws.cell(row=row, column=2, value=default)
        c.font = accent_font
        c.fill = panel_fill
        c.border = thin_border
        dv.add(f'B{row}')

    # Note row
    ws.cell(row=6, column=1,
            value='Track A is the reference. Deltas show how B/C/D differ from A.').font = dim_font
    ws.cell(row=6, column=1).fill = bg_fill

    # Row 7: empty separator

    # Row 8: empty separator

    # Row 8: M7.5 Navigation bar
    _xl_add_sheet_nav(ws, 8)

    # --- Section 3: Comparison table (row 9 = headers, row 10+ = data) ---
    header_row = 9
    # Columns: A=Metric, B=Track A, C=Track B, D=Δ B-A, E=% B-A,
    #           F=Track C, G=Δ C-A, H=% C-A, I=Track D, J=Δ D-A, K=% D-A
    comp_headers = ['Metric', 'Track A', 'Track B', 'Δ B-A', '% B-A',
                    'Track C', 'Δ C-A', '% C-A', 'Track D', 'Δ D-A', '% D-A']
    for col, h in enumerate(comp_headers, 1):
        c = ws.cell(row=header_row, column=col, value=h)
        c.font = header_font
        c.fill = header_fill
        c.border = thin_border
        c.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)

    # --- Build INDEX/MATCH formulas ---
    # _track_data sheet: column A = names, columns B onwards = metrics
    # MATCH($B$2, _track_data!$A:$A, 0) gives the row of Track A
    # INDEX(_track_data!B:B, match_result) gives the metric value

    data_sheet = '_track_data'
    name_col_ref = f"'{data_sheet}'!$A$2:$A${data_range_end}"

    data_start_row = 10

    for m_idx, (metric_name, is_numeric) in enumerate(metrics):
        row = data_start_row + m_idx
        metric_col_idx = m_idx + 2  # column in _track_data (1-based, +1 for Track Name col)
        metric_col_letter = get_column_letter(metric_col_idx)
        metric_range = f"'{data_sheet}'!${metric_col_letter}$2:${metric_col_letter}${data_range_end}"

        # Metric name
        c = ws.cell(row=row, column=1, value=metric_name)
        c.font = data_font
        c.fill = bg_fill
        c.border = thin_border
        if metric_name in METRIC_GLOSSARY:
            _xl_add_comment(c, METRIC_GLOSSARY[metric_name])

        # Track A value (col B) - always shown
        match_a = f'MATCH($B$2,{name_col_ref},0)'
        formula_a = f'=IFERROR(INDEX({metric_range},{match_a}),"")'
        c = ws.cell(row=row, column=2, value=formula_a)
        c.font = data_font
        c.fill = bg_fill
        c.border = thin_border

        # Track B (col C), Delta B-A (col D), % B-A (col E)
        match_b = f'MATCH($B$3,{name_col_ref},0)'
        formula_b = f'=IFERROR(INDEX({metric_range},{match_b}),"")'
        c = ws.cell(row=row, column=3, value=formula_b)
        c.font = data_font
        c.fill = bg_fill
        c.border = thin_border

        if is_numeric:
            # Delta = B - A
            formula_delta_b = f'=IFERROR(C{row}-B{row},"")'
            c = ws.cell(row=row, column=4, value=formula_delta_b)
            c.font = data_font
            c.fill = bg_fill
            c.border = thin_border
            c.number_format = '+0.00;-0.00;0.00'
            # % delta = (B-A)/ABS(A)*100
            formula_pct_b = f'=IFERROR(IF(B{row}=0,"",D{row}/ABS(B{row})*100),"")'
            c = ws.cell(row=row, column=5, value=formula_pct_b)
            c.font = data_font
            c.fill = bg_fill
            c.border = thin_border
            c.number_format = '+0.0"%";-0.0"%";0.0"%"'
        else:
            # Dominant Band: text comparison
            formula_delta_b = f'=IF(C{row}=B{row},"match","differ")'
            c = ws.cell(row=row, column=4, value=formula_delta_b)
            c.font = data_font
            c.fill = bg_fill
            c.border = thin_border
            # No % for text
            ws.cell(row=row, column=5, value='').fill = bg_fill

        # Track C (col F), Delta C-A (col G), % C-A (col H) — conditional on B4 not empty
        match_c = f'MATCH($B$4,{name_col_ref},0)'
        formula_c = f'=IF($B$4="","",IFERROR(INDEX({metric_range},{match_c}),""))'
        c = ws.cell(row=row, column=6, value=formula_c)
        c.font = data_font
        c.fill = bg_fill
        c.border = thin_border

        if is_numeric:
            formula_delta_c = f'=IF($B$4="","",IFERROR(F{row}-B{row},""))'
            c = ws.cell(row=row, column=7, value=formula_delta_c)
            c.font = data_font
            c.fill = bg_fill
            c.border = thin_border
            c.number_format = '+0.00;-0.00;0.00'
            formula_pct_c = f'=IF($B$4="","",IFERROR(IF(B{row}=0,"",G{row}/ABS(B{row})*100),""))'
            c = ws.cell(row=row, column=8, value=formula_pct_c)
            c.font = data_font
            c.fill = bg_fill
            c.border = thin_border
            c.number_format = '+0.0"%";-0.0"%";0.0"%"'
        else:
            formula_delta_c = f'=IF($B$4="","",IF(F{row}=B{row},"match","differ"))'
            c = ws.cell(row=row, column=7, value=formula_delta_c)
            c.font = data_font
            c.fill = bg_fill
            c.border = thin_border
            ws.cell(row=row, column=8, value='').fill = bg_fill

        # Track D (col I), Delta D-A (col J), % D-A (col K) — conditional on B5 not empty
        match_d = f'MATCH($B$5,{name_col_ref},0)'
        formula_d = f'=IF($B$5="","",IFERROR(INDEX({metric_range},{match_d}),""))'
        c = ws.cell(row=row, column=9, value=formula_d)
        c.font = data_font
        c.fill = bg_fill
        c.border = thin_border

        if is_numeric:
            formula_delta_d = f'=IF($B$5="","",IFERROR(I{row}-B{row},""))'
            c = ws.cell(row=row, column=10, value=formula_delta_d)
            c.font = data_font
            c.fill = bg_fill
            c.border = thin_border
            c.number_format = '+0.00;-0.00;0.00'
            formula_pct_d = f'=IF($B$5="","",IFERROR(IF(B{row}=0,"",J{row}/ABS(B{row})*100),""))'
            c = ws.cell(row=row, column=11, value=formula_pct_d)
            c.font = data_font
            c.fill = bg_fill
            c.border = thin_border
            c.number_format = '+0.0"%";-0.0"%";0.0"%"'
        else:
            formula_delta_d = f'=IF($B$5="","",IF(I{row}=B{row},"match","differ"))'
            c = ws.cell(row=row, column=10, value=formula_delta_d)
            c.font = data_font
            c.fill = bg_fill
            c.border = thin_border
            ws.cell(row=row, column=11, value='').fill = bg_fill

    data_end_row = data_start_row + n_metrics - 1

    # --- Conditional Formatting ---
    # Color scale on delta columns (D, G, J): red-white-green centered on 0
    for delta_col_letter in ['D', 'G', 'J']:
        delta_range = f'{delta_col_letter}{data_start_row}:{delta_col_letter}{data_end_row}'
        ws.conditional_formatting.add(
            delta_range,
            ColorScaleRule(
                start_type='min', start_color='FF3333',
                mid_type='num', mid_value=0, mid_color='FFFFFF',
                end_type='max', end_color='00FF9F'))

    # Color scale on band energy values (rows for band metrics, columns B, C, F, I)
    # Band metrics start at index 7 (after Dominant Band) in the metrics list
    band_start_row = data_start_row + 7  # first band metric row
    band_end_row = data_end_row
    for val_col_letter in ['B', 'C', 'F', 'I']:
        band_range = f'{val_col_letter}{band_start_row}:{val_col_letter}{band_end_row}'
        ws.conditional_formatting.add(
            band_range,
            ColorScaleRule(
                start_type='num', start_value=0, start_color='00FF9F',
                mid_type='num', mid_value=50, mid_color='FFFF00',
                end_type='num', end_value=100, end_color='FF3333'))

    # --- Freeze panes (row 10, col A = headers visible) ---
    ws.freeze_panes = 'A10'

    # --- Column widths ---
    ws.column_dimensions['A'].width = 20
    for col_idx in range(2, 12):
        ws.column_dimensions[get_column_letter(col_idx)].width = 12

    _apply_dark_background(ws_data)
    _apply_dark_background(ws)
    log_fn("    Excel: Track Comparison sheet done.")


def _calc_loudness_score(individuals, full_mix):
    """Loudness sub-score (0-100). Returns (score, details_list)."""
    details = []
    scores = []

    # Coherence: std of individual LUFS
    ind_lufs = [a['loudness']['lufs_integrated'] for a, _ in individuals
                if np.isfinite(a['loudness']['lufs_integrated'])]
    if len(ind_lufs) >= 2:
        lufs_std = float(np.std(ind_lufs))
        coh_score = max(0, 100 - lufs_std * 10)
        details.append(('LUFS std (individuals)', round(lufs_std, 2), '< 4 LU', round(coh_score, 1)))
    else:
        coh_score = 70
        details.append(('LUFS std (individuals)', 'N/A', '< 4 LU', coh_score))
    scores.append((coh_score, 0.4))

    # Full Mix target distance from -14 LUFS
    if full_mix:
        fm_lufs = full_mix['loudness']['lufs_integrated']
        if np.isfinite(fm_lufs):
            dist = abs(fm_lufs - (-14))
            target_score = max(0, 100 - dist * 5)
            details.append(('Full Mix LUFS', round(fm_lufs, 2), '-14 LUFS', round(target_score, 1)))
        else:
            target_score = 50
            details.append(('Full Mix LUFS', 'N/A', '-14 LUFS', target_score))
    else:
        target_score = 70
        details.append(('Full Mix LUFS', 'No Full Mix', '-14 LUFS', target_score))
    scores.append((target_score, 0.4))

    # True peak safety
    if full_mix:
        tp = full_mix['loudness'].get('true_peak_db', full_mix['loudness']['peak_db'])
        if tp < -1.0:
            peak_score = 100
        else:
            peak_score = max(0, 100 - (tp + 1) * 50)
        details.append(('True Peak (Full Mix)', round(tp, 2), '< -1.0 dBFS', round(peak_score, 1)))
    else:
        peak_score = 70
        details.append(('True Peak (Full Mix)', 'No Full Mix', '< -1.0 dBFS', peak_score))
    scores.append((peak_score, 0.2))

    total = sum(s * w for s, w in scores)
    note = 'Tight LUFS' if total >= 75 else ('Moderate spread' if total >= 50 else 'Wide LUFS spread')
    return round(total, 1), details, note


def _calc_dynamics_score(individuals, full_mix):
    """Dynamics sub-score (0-100). Returns (score, details_list, note)."""
    details = []
    scores = []

    # Crest factor mean
    ind_crests = [a['loudness']['crest_factor'] for a, _ in individuals]
    if ind_crests:
        crest_mean = float(np.mean(ind_crests))
        # Linear 0 at crest=4, 100 at crest=14
        crest_score = max(0, min(100, (crest_mean - 4) * 10))
        details.append(('Crest factor mean', round(crest_mean, 2), '8-14 dB', round(crest_score, 1)))
    else:
        crest_score = 50
        details.append(('Crest factor mean', 'N/A', '8-14 dB', crest_score))
    scores.append((crest_score, 0.4))

    # PLR Full Mix
    if full_mix:
        plr = full_mix['loudness']['plr']
        # Peak at 100 when PLR=12, degrade symmetrically
        plr_score = max(0, 100 - abs(plr - 12) * 10)
        details.append(('PLR (Full Mix)', round(plr, 2), '11-14 dB', round(plr_score, 1)))
    else:
        plr_score = 60
        details.append(('PLR (Full Mix)', 'No Full Mix', '11-14 dB', plr_score))
    scores.append((plr_score, 0.4))

    # Variance of crest factors
    if len(ind_crests) >= 2:
        crest_std = float(np.std(ind_crests))
        var_score = max(0, 100 - crest_std * 5)
        details.append(('Crest std (individuals)', round(crest_std, 2), '< 4 dB', round(var_score, 1)))
    else:
        var_score = 70
        details.append(('Crest std (individuals)', 'N/A', '< 4 dB', var_score))
    scores.append((var_score, 0.2))

    total = sum(s * w for s, w in scores)
    note = 'Good dynamics' if total >= 75 else ('Low crest' if crest_score < 50 else 'Moderate')
    return round(total, 1), details, note


def _calc_spectral_balance_score(individuals, full_mix):
    """Spectral Balance sub-score (0-100). Returns (score, details_list, note)."""
    details = []

    # Use Full Mix if available, otherwise average individuals
    if full_mix:
        energies = full_mix['spectrum']['band_energies']
        source = 'Full Mix'
    elif individuals:
        # Average band energies across individuals
        avg = {name: 0.0 for name, _, _ in FREQ_BANDS}
        for a, _ in individuals:
            for name, _, _ in FREQ_BANDS:
                avg[name] += a['spectrum']['band_energies'].get(name, 0.0)
        for name, _, _ in FREQ_BANDS:
            avg[name] /= len(individuals)
        energies = avg
        source = 'Individual avg'
    else:
        return 50.0, [('Source', 'No tracks', '-', 50)], 'No data'

    band_values = [energies.get(name, 0.0) for name, _, _ in FREQ_BANDS]
    total_e = sum(band_values)
    if total_e <= 0:
        return 50.0, [('Total energy', 0, '> 0', 50)], 'No energy'

    # Spectral entropy (normalized)
    probs = [v / total_e for v in band_values if v > 0]
    if probs:
        entropy = -sum(p * np.log2(p) for p in probs)
        max_entropy = np.log2(len(FREQ_BANDS))
        entropy_norm = (entropy / max_entropy) * 100 if max_entropy > 0 else 50
    else:
        entropy_norm = 0
    details.append(('Spectral entropy', f'{entropy_norm:.1f}%', '> 70%', round(entropy_norm, 1)))

    score = entropy_norm

    # Penalty if Sub Energy = 0
    sub_e = energies.get('sub', 0.0)
    if sub_e <= 0.5:
        score = max(0, score - 10)
        details.append(('Sub Energy', f'{sub_e:.1f}%', '> 0.5%', '-10 penalty'))
    else:
        details.append(('Sub Energy', f'{sub_e:.1f}%', '> 0.5%', 'OK'))

    # Penalty if Air Energy = 0
    air_e = energies.get('air', 0.0)
    if air_e <= 0.5:
        score = max(0, score - 10)
        details.append(('Air Energy', f'{air_e:.1f}%', '> 0.5%', '-10 penalty'))
    else:
        details.append(('Air Energy', f'{air_e:.1f}%', '> 0.5%', 'OK'))

    # Penalty if any band > 35%
    max_band = max(band_values)
    max_band_name = [name for name, _, _ in FREQ_BANDS
                     if energies.get(name, 0) == max_band][0]
    if max_band > 35:
        penalty = min(20, (max_band - 35) * 2)
        score = max(0, score - penalty)
        details.append((f'Dominant band ({BAND_LABELS[max_band_name]})',
                        f'{max_band:.1f}%', '< 35%', f'-{penalty:.0f} penalty'))
    else:
        details.append(('Max band energy', f'{max_band:.1f}%', '< 35%', 'OK'))

    score = max(0, min(100, score))
    note = 'Balanced' if score >= 75 else ('Uneven' if score < 50 else 'Moderate')
    return round(score, 1), details, note


def _calc_stereo_image_score(individuals, full_mix):
    """Stereo Image sub-score (0-100). Returns (score, details_list, note)."""
    details = []
    scores = []

    # Width score (60 pts weight)
    if full_mix and full_mix['stereo']['is_stereo']:
        width = full_mix['stereo']['width_overall']
        if 0.4 <= width <= 0.7:
            width_score = 100
        elif width < 0.4:
            width_score = max(0, 100 - (0.4 - width) * 200)
        else:
            width_score = max(0, 100 - (width - 0.7) * 200)
        details.append(('Stereo Width (Full Mix)', round(width, 3), '0.40-0.70', round(width_score, 1)))
    elif individuals:
        widths = [a['stereo']['width_overall'] for a, _ in individuals if a['stereo']['is_stereo']]
        if widths:
            avg_w = float(np.mean(widths))
            if 0.3 <= avg_w <= 0.7:
                width_score = 100
            elif avg_w < 0.3:
                width_score = max(0, 100 - (0.3 - avg_w) * 200)
            else:
                width_score = max(0, 100 - (avg_w - 0.7) * 200)
            details.append(('Avg Width (individuals)', round(avg_w, 3), '0.30-0.70', round(width_score, 1)))
        else:
            width_score = 50
            details.append(('Stereo Width', 'Mono tracks', '0.40-0.70', width_score))
    else:
        width_score = 50
        details.append(('Stereo Width', 'No data', '0.40-0.70', width_score))
    scores.append((width_score, 0.6))

    # Sub mono check (40 pts): sub should be more mono than high
    src = full_mix if full_mix else (individuals[0][0] if individuals else None)
    if src and src['stereo']['is_stereo']:
        sub_w = src['stereo']['width_per_band'].get('sub', 0.0)
        high_w = src['stereo']['width_per_band'].get('presence', 0.5)
        if sub_w <= high_w:
            mono_score = 100
        else:
            mono_score = max(0, 100 - (sub_w - high_w) * 200)
        details.append(('Sub width vs High width', f'{sub_w:.2f} vs {high_w:.2f}',
                        'Sub < High', round(mono_score, 1)))
    else:
        mono_score = 70
        details.append(('Sub mono check', 'N/A', 'Sub < High', mono_score))
    scores.append((mono_score, 0.4))

    total = sum(s * w for s, w in scores)
    note = 'Good image' if total >= 75 else ('Too narrow' if width_score < 50 else 'Wide')
    return round(total, 1), details, note


def _calc_anomalies_score(analyses_with_info):
    """Anomalies sub-score (0-100). Returns (score, details_list, note)."""
    details = []
    total_anomalies = 0
    critical_count = 0
    warning_count = 0

    for a, ti in analyses_with_info:
        if ti.get('type') == 'BUS':
            continue
        try:
            anoms = detect_anomalies(a)
            for sev, desc in anoms:
                total_anomalies += 1
                if sev == 'critical':
                    critical_count += 1
                else:
                    warning_count += 1
        except Exception:
            pass

    # Score: 100 - penalties
    penalty = critical_count * 15 + warning_count * 5
    score = max(0, 100 - penalty)

    details.append(('Total anomalies', total_anomalies, '0', round(score, 1)))
    details.append(('Critical', critical_count, '0', f'-{critical_count * 15} pts'))
    details.append(('Warnings', warning_count, '0', f'-{warning_count * 5} pts'))

    note = f'{total_anomalies} anomalies' if total_anomalies > 0 else 'Clean'
    return round(score, 1), details, note


# M5.2: Cyberpunk dark background
_DARK_BG_FILL = None  # Lazy-initialized to avoid top-level openpyxl import

def _apply_dark_background(ws, max_row=None, max_col=None):
    """
    M5.2: Apply cyberpunk dark background to unstyled cells.

    Preserves existing fills (color scales, data bars, custom highlights).
    Only applies dark background to cells that don't already have a
    custom fill defined.

    Args:
        ws: openpyxl worksheet
        max_row: max row to apply (default: ws.max_row)
        max_col: max col to apply (default: ws.max_column)
    """
    global _DARK_BG_FILL
    if _DARK_BG_FILL is None:
        from openpyxl.styles import PatternFill
        _DARK_BG_FILL = PatternFill(start_color='0A0A0F',
                                     end_color='0A0A0F',
                                     fill_type='solid')

    if max_row is None:
        max_row = ws.max_row
    if max_col is None:
        max_col = ws.max_column

    for row in ws.iter_rows(min_row=1, max_row=max_row,
                             min_col=1, max_col=max_col):
        for cell in row:
            existing_fill = cell.fill
            has_custom_fill = (
                existing_fill is not None
                and existing_fill.fill_type == 'solid'
                and existing_fill.fgColor is not None
                and existing_fill.fgColor.value not in (None, '00000000', 'FFFFFFFF')
            )
            if not has_custom_fill:
                cell.fill = _DARK_BG_FILL


# =============================================================================
# M5.3: Centralized theme palette and typography
# =============================================================================

MA_THEME = {
    # Backgrounds
    'bg_primary': '0A0A0F',       # Dark principal (from M5.2)
    'bg_secondary': '15151F',     # Dark secondaire
    'bg_highlight': '252535',     # Dark highlight

    # Text colors
    'text_primary': 'E8E8F0',    # Blanc cassé — corps principal
    'text_heading': '00D9FF',    # Cyan brillant — titres
    'text_subheading': '66B8D9', # Cyan atténué — sous-titres
    'text_secondary': '8A8AA0',  # Gris violet — notes
    'text_muted': '505068',      # Gris foncé — désactivé

    # Accent colors (sheets Phase 3)
    'accent_pink': 'FF3D8B',     # P3.1 Freq Conflicts
    'accent_orange': 'FF8B3D',   # P3.2 Track Comparison
    'accent_green': '3DFFAA',    # P3.3 Health Score
    'accent_cyan': '3DAAFF',     # P3.4 Version Tracking

    # Status colors
    'status_success': '3DFFAA',
    'status_warning': 'FFD93D',
    'status_error': 'FF3D6E',
    'status_info': '3DAAFF',

    # Typography
    'font_family': 'Calibri',
    'font_size_title': 16,
    'font_size_heading': 13,
    'font_size_subheading': 11,
    'font_size_body': 10,
    'font_size_small': 9,
    'font_size_tiny': 8,
}

# Lazy-initialized Font constants (avoid top-level openpyxl import)
_MA_FONTS_INITIALIZED = False
MA_FONT_TITLE = None
MA_FONT_HEADING = None
MA_FONT_SUBHEADING = None
MA_FONT_BODY = None
MA_FONT_BODY_BOLD = None
MA_FONT_SMALL = None
MA_FONT_TABLE_HEADER = None
MA_FONT_DIM = None
MA_FONT_LINK = None


def _init_ma_fonts():
    """Initialize MA_FONT_* constants on first use."""
    global _MA_FONTS_INITIALIZED
    global MA_FONT_TITLE, MA_FONT_HEADING, MA_FONT_SUBHEADING
    global MA_FONT_BODY, MA_FONT_BODY_BOLD, MA_FONT_SMALL
    global MA_FONT_TABLE_HEADER, MA_FONT_DIM, MA_FONT_LINK
    if _MA_FONTS_INITIALIZED:
        return
    from openpyxl.styles import Font
    MA_FONT_TITLE = Font(
        name=MA_THEME['font_family'],
        size=MA_THEME['font_size_title'],
        bold=True,
        color=MA_THEME['text_heading']
    )
    MA_FONT_HEADING = Font(
        name=MA_THEME['font_family'],
        size=MA_THEME['font_size_heading'],
        bold=True,
        color=MA_THEME['text_heading']
    )
    MA_FONT_SUBHEADING = Font(
        name=MA_THEME['font_family'],
        size=MA_THEME['font_size_subheading'],
        bold=True,
        color=MA_THEME['text_subheading']
    )
    MA_FONT_BODY = Font(
        name=MA_THEME['font_family'],
        size=MA_THEME['font_size_body'],
        color=MA_THEME['text_primary']
    )
    MA_FONT_BODY_BOLD = Font(
        name=MA_THEME['font_family'],
        size=MA_THEME['font_size_body'],
        bold=True,
        color=MA_THEME['text_primary']
    )
    MA_FONT_SMALL = Font(
        name=MA_THEME['font_family'],
        size=MA_THEME['font_size_small'],
        color=MA_THEME['text_secondary']
    )
    MA_FONT_TABLE_HEADER = Font(
        name=MA_THEME['font_family'],
        size=MA_THEME['font_size_subheading'],
        bold=True,
        color=MA_THEME['text_primary']
    )
    MA_FONT_DIM = Font(
        name=MA_THEME['font_family'],
        size=MA_THEME['font_size_body'],
        color=MA_THEME['text_secondary']
    )
    MA_FONT_LINK = Font(
        name=MA_THEME['font_family'],
        size=MA_THEME['font_size_small'],
        color=MA_THEME['text_heading'],
        underline='single'
    )
    _MA_FONTS_INITIALIZED = True


def generate_health_score_sheet(workbook, analyses_with_info, log_fn=None,
                                nav_targets=None):
    """
    Génère le sheet 'Mix Health Score' (P3.3) dans le workbook donné.
    Calcule un score global de santé du mix sur 100, décomposé en
    5 sous-scores : Loudness, Dynamics, Spectral Balance, Stereo Image,
    Anomalies. Filtre les pistes : Individual + Full Mix, exclut BUS.
    """
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.formatting.rule import ColorScaleRule
    _init_ma_fonts()
    from openpyxl.utils import get_column_letter
    import datetime

    if log_fn is None:
        log_fn = lambda msg: None

    log_fn("    Excel: writing Mix Health Score sheet (P3.3)...")

    individuals = [(a, ti) for a, ti in analyses_with_info if ti.get('type') == 'Individual']
    full_mixes = [(a, ti) for a, ti in analyses_with_info if ti.get('type') == 'Full Mix']
    full_mix = full_mixes[0][0] if full_mixes else None

    if not individuals and not full_mix:
        ws = workbook.create_sheet('Mix Health Score')
        _apply_clean_layout(ws)
        ws.sheet_properties.tabColor = '3DFFAA'
        _xl_write_header(ws, 'MIX HEALTH SCORE', 'No tracks available for scoring.')
        _apply_dark_background(ws)
        return

    # Compute sub-scores
    loud_score, loud_details, loud_note = _calc_loudness_score(individuals, full_mix)
    dyn_score, dyn_details, dyn_note = _calc_dynamics_score(individuals, full_mix)
    spec_score, spec_details, spec_note = _calc_spectral_balance_score(individuals, full_mix)
    stereo_score, stereo_details, stereo_note = _calc_stereo_image_score(individuals, full_mix)
    anom_score, anom_details, anom_note = _calc_anomalies_score(analyses_with_info)

    categories = [
        ('Loudness', loud_score, 0.20, loud_note, loud_details),
        ('Dynamics', dyn_score, 0.20, dyn_note, dyn_details),
        ('Spectral Balance', spec_score, 0.25, spec_note, spec_details),
        ('Stereo Image', stereo_score, 0.15, stereo_note, stereo_details),
        ('Anomalies', anom_score, 0.20, anom_note, anom_details),
    ]

    global_score = round(sum(s * w for _, s, w, _, _ in categories), 1)

    # Theme styles
    bg_fill = PatternFill('solid', fgColor='0A0A12')
    panel_fill = PatternFill('solid', fgColor='1A1A24')
    header_fill = PatternFill('solid', fgColor='1A3A5A')
    accent_font = MA_FONT_SUBHEADING
    header_font = MA_FONT_TABLE_HEADER
    data_font = MA_FONT_BODY
    dim_font = MA_FONT_DIM
    thin_border = Border(
        left=Side(style='thin', color='333344'),
        right=Side(style='thin', color='333344'),
        top=Side(style='thin', color='333344'),
        bottom=Side(style='thin', color='333344'),
    )

    # Score color
    if global_score >= 80:
        score_color = '00FF9F'
    elif global_score >= 60:
        score_color = 'AAFF00'
    elif global_score >= 40:
        score_color = 'FFAA00'
    elif global_score >= 20:
        score_color = 'FF3333'
    else:
        score_color = '990000'

    # Create sheet
    ws = workbook.create_sheet('Mix Health Score')
    _apply_clean_layout(ws)
    ws.sheet_properties.tabColor = '3DFFAA'

    # --- Section 1: Global score (rows 1-8) ---
    ws['A1'] = 'MIX HEALTH SCORE'
    ws['A1'].font = Font(name='Calibri', size=18, bold=True, color='00D9FF')
    ws['A1'].fill = bg_fill
    # M7.5: Navigation bar
    _xl_add_sheet_nav(ws, 2, current_sheet='Mix Health Score', nav_targets=nav_targets)

    # Score display
    ws.merge_cells('A3:D3')
    ws['A3'] = global_score
    ws['A3'].font = Font(name='Calibri', size=36, bold=True, color=score_color)
    ws['A3'].fill = bg_fill
    ws['A3'].alignment = Alignment(horizontal='center', vertical='center')
    ws['A3'].number_format = '0.0"/100"'
    _xl_add_comment(ws['A3'], METRIC_GLOSSARY['Health Score'])

    ws['A5'] = f'Calculated on: {datetime.date.today().isoformat()}'
    ws['A5'].font = dim_font
    ws['A5'].fill = bg_fill
    ws['A6'] = f'Tracks analyzed: {len(individuals)} Individual' + \
               (f' + 1 Full Mix' if full_mix else '')
    ws['A6'].font = dim_font
    ws['A6'].fill = bg_fill
    ws['A7'] = 'Indicative score based on technical heuristics. ' \
               'Not an artistic judgment. Use as a tracking tool between mix iterations.'
    ws['A7'].font = Font(name='Calibri', size=10, italic=True, color='8888A0')
    ws['A7'].fill = bg_fill

    # --- Section 2: Sub-scores table (rows 10-17) ---
    cat_header_row = 10
    cat_headers = ['Category', 'Score', 'Weight', 'Notes']
    for col, h in enumerate(cat_headers, 1):
        c = ws.cell(row=cat_header_row, column=col, value=h)
        c.font = header_font
        c.fill = header_fill
        c.border = thin_border
        c.alignment = Alignment(horizontal='center', vertical='center')

    _health_cat_comments = {
        'Loudness': "Evaluates LUFS targets, peak levels, and headroom.\nWeight: 20% of total score.",
        'Dynamics': "Evaluates crest factor, PLR, and LRA.\nWeight: 20% of total score.",
        'Spectral Balance': "Evaluates frequency distribution and centroid.\nWeight: 25% of total score.",
        'Stereo Image': "Evaluates stereo width and phase correlation.\nWeight: 15% of total score.",
        'Anomalies': "Penalty based on detected issues.\nWeight: 20% of total score.",
    }

    for i, (name, score, weight, note, _) in enumerate(categories):
        row = cat_header_row + 1 + i
        c = ws.cell(row=row, column=1, value=name)
        c.font = data_font
        c.fill = bg_fill
        c.border = thin_border
        if name in _health_cat_comments:
            _xl_add_comment(c, _health_cat_comments[name])

        c = ws.cell(row=row, column=2, value=score)
        c.font = accent_font
        c.fill = bg_fill
        c.border = thin_border
        c.alignment = Alignment(horizontal='center')
        c.number_format = '0.0'

        c = ws.cell(row=row, column=3, value=f'{int(weight * 100)}%')
        c.font = data_font
        c.fill = bg_fill
        c.border = thin_border
        c.alignment = Alignment(horizontal='center')

        c = ws.cell(row=row, column=4, value=note)
        c.font = dim_font
        c.fill = bg_fill
        c.border = thin_border

    # Separator row 16
    sep_row = cat_header_row + 1 + len(categories)

    # Weighted total row
    total_row = sep_row + 1
    c = ws.cell(row=total_row, column=1, value='WEIGHTED TOTAL')
    c.font = MA_FONT_SUBHEADING
    c.fill = bg_fill
    c.border = thin_border

    c = ws.cell(row=total_row, column=2, value=global_score)
    c.font = Font(name='Calibri', size=14, bold=True, color=score_color)
    c.fill = bg_fill
    c.border = thin_border
    c.alignment = Alignment(horizontal='center')
    c.number_format = '0.0'

    c = ws.cell(row=total_row, column=3, value='100%')
    c.font = data_font
    c.fill = bg_fill
    c.border = thin_border
    c.alignment = Alignment(horizontal='center')

    # Color scale on sub-score column (B11:B15)
    score_start = cat_header_row + 1
    score_end = score_start + len(categories) - 1
    ws.conditional_formatting.add(
        f'B{score_start}:B{score_end}',
        ColorScaleRule(
            start_type='num', start_value=0, start_color='FF3333',
            mid_type='num', mid_value=50, mid_color='FFAA00',
            end_type='num', end_value=100, end_color='00FF9F'))
    # M7.3: Data bars for sub-scores (0-100 scale)
    from openpyxl.formatting.rule import DataBarRule
    ws.conditional_formatting.add(
        f'B{score_start}:B{score_end}',
        DataBarRule(start_type='num', start_value=0,
                    end_type='num', end_value=100, color='00D9FF'))

    # --- Section 3: Details per category (from row 20) ---
    detail_row = 20

    for cat_name, _, _, _, cat_details in categories:
        # Sub-header
        c = ws.cell(row=detail_row, column=1, value=f'{cat_name} Details')
        c.font = MA_FONT_SUBHEADING
        c.fill = panel_fill
        c.border = thin_border
        for col in range(2, 5):
            c2 = ws.cell(row=detail_row, column=col)
            c2.fill = panel_fill
            c2.border = thin_border
        # Detail column headers
        detail_row += 1
        for col, h in enumerate(['Metric', 'Value', 'Ideal Range', 'Contribution'], 1):
            c = ws.cell(row=detail_row, column=col, value=h)
            c.font = MA_FONT_BODY_BOLD
            c.fill = header_fill
            c.border = thin_border
            c.alignment = Alignment(horizontal='center')

        detail_row += 1
        for metric, value, ideal, contrib in cat_details:
            c = ws.cell(row=detail_row, column=1, value=metric)
            c.font = data_font
            c.fill = bg_fill
            c.border = thin_border

            c = ws.cell(row=detail_row, column=2, value=value)
            c.font = data_font
            c.fill = bg_fill
            c.border = thin_border
            c.alignment = Alignment(horizontal='center')

            c = ws.cell(row=detail_row, column=3, value=ideal)
            c.font = dim_font
            c.fill = bg_fill
            c.border = thin_border
            c.alignment = Alignment(horizontal='center')

            c = ws.cell(row=detail_row, column=4, value=contrib)
            c.font = data_font
            c.fill = bg_fill
            c.border = thin_border
            c.alignment = Alignment(horizontal='center')

            detail_row += 1

        detail_row += 1  # blank row between sections

    # --- Freeze panes (row 11) ---
    ws.freeze_panes = 'A11'

    # --- Column widths ---
    ws.column_dimensions['A'].width = 25
    ws.column_dimensions['B'].width = 10
    ws.column_dimensions['C'].width = 12
    ws.column_dimensions['D'].width = 30

    _apply_dark_background(ws)
    log_fn("    Excel: Mix Health Score sheet done.")


def _find_previous_reports(output_folder, song_name):
    """Find previous Mix Analyzer .xlsx reports in the output folder matching the song name."""
    import glob
    import re
    pattern = os.path.join(output_folder, f'{song_name}_MixAnalyzer_*.xlsx')
    files = glob.glob(pattern)
    results = []
    date_re = re.compile(r'_MixAnalyzer_(\d{4}-\d{2}-\d{2})(?:_\d{2}-\d{2})?')
    for f in files:
        m = date_re.search(os.path.basename(f))
        if m:
            results.append((m.group(1), f))
    results.sort(key=lambda x: x[0])
    return results


def _extract_metrics_from_report(xlsx_path, log_fn=None):
    """Extract key metrics from a previous Mix Analyzer .xlsx report.
    Returns a dict of metric_name -> value, or None if extraction fails."""
    try:
        from openpyxl import load_workbook
        wb = load_workbook(xlsx_path, read_only=True, data_only=True)
        metrics = {}

        # Read from Summary sheet
        if 'Summary' in wb.sheetnames:
            ws = wb['Summary']
            # Headers at row 4: Track, Type, Category, LUFS, Peak, Crest, Width, ...
            # Find Full Mix row and Individual rows
            fm_lufs = None
            fm_peak = None
            fm_crest = None
            fm_width = None
            ind_crests = []
            track_count = 0
            for row in ws.iter_rows(min_row=5, max_col=10, values_only=True):
                if row[0] is None:
                    break
                track_type = row[1]
                if track_type == 'Full Mix':
                    fm_lufs = row[3]
                    fm_peak = row[4]
                    fm_crest = row[5]
                    fm_width = row[6]
                elif track_type == 'Individual':
                    track_count += 1
                    if row[5] is not None:
                        try:
                            ind_crests.append(float(row[5]))
                        except (ValueError, TypeError):
                            pass

            if fm_lufs is not None:
                metrics['Full Mix LUFS'] = fm_lufs
            if fm_peak is not None:
                metrics['Full Mix Peak (dBFS)'] = fm_peak
            if fm_crest is not None:
                metrics['Full Mix Crest (dB)'] = fm_crest
            if fm_width is not None and fm_width != 'mono':
                try:
                    metrics['Full Mix Width'] = float(fm_width)
                except (ValueError, TypeError):
                    pass
            if ind_crests:
                metrics['Avg Individual Crest (dB)'] = round(sum(ind_crests) / len(ind_crests), 2)
            metrics['Track count'] = track_count

        # Read from Dashboard for PLR, True Peak
        if 'Dashboard' in wb.sheetnames:
            ws_dash = wb['Dashboard']
            # Headers at row 4: Track, Type, ..., True Peak at col 7, PLR at col 10
            # Find header row first
            header_row_vals = None
            for row in ws_dash.iter_rows(min_row=1, max_row=10, max_col=20, values_only=True):
                if row and row[0] == 'Track':
                    header_row_vals = list(row)
                    break
            if header_row_vals:
                tp_idx = None
                plr_idx = None
                for i, h in enumerate(header_row_vals):
                    if h and 'True Peak' in str(h):
                        tp_idx = i
                    if h and h == 'PLR (dB)':
                        plr_idx = i
                type_idx = header_row_vals.index('Type') if 'Type' in header_row_vals else 1

                for row in ws_dash.iter_rows(min_row=5, max_col=20, values_only=True):
                    if row[0] is None:
                        break
                    if type_idx < len(row) and row[type_idx] == 'Full Mix':
                        if tp_idx is not None and tp_idx < len(row) and row[tp_idx] is not None:
                            try:
                                metrics['Full Mix True Peak (dBFS)'] = float(row[tp_idx])
                            except (ValueError, TypeError):
                                pass
                        if plr_idx is not None and plr_idx < len(row) and row[plr_idx] is not None:
                            try:
                                metrics['Full Mix PLR'] = float(row[plr_idx])
                            except (ValueError, TypeError):
                                pass

        # Read Mix Health Score if present
        if 'Mix Health Score' in wb.sheetnames:
            ws_hs = wb['Mix Health Score']
            # Global score is in A3
            score_val = ws_hs['A3'].value
            if score_val is not None:
                try:
                    metrics['Mix Health Score'] = float(score_val)
                except (ValueError, TypeError):
                    pass

        # Count anomalies from Anomalies sheet
        if 'Anomalies' in wb.sheetnames:
            ws_anom = wb['Anomalies']
            anom_count = 0
            for row in ws_anom.iter_rows(min_row=5, max_col=3, values_only=True):
                if row[0] is None:
                    break
                anom_count += 1
            if anom_count > 0:
                metrics['Anomaly count'] = anom_count
            else:
                metrics['Anomaly count'] = 0

        wb.close()
        return metrics
    except Exception as e:
        if log_fn:
            log_fn(f"    Warning: could not read {os.path.basename(xlsx_path)}: {e}")
        return None


def _compute_current_metrics(analyses_with_info):
    """Compute version tracking metrics from the current in-memory analyses."""
    metrics = {}
    individuals = [(a, ti) for a, ti in analyses_with_info if ti.get('type') == 'Individual']
    full_mixes = [(a, ti) for a, ti in analyses_with_info if ti.get('type') == 'Full Mix']
    full_mix = full_mixes[0][0] if full_mixes else None

    if full_mix:
        L = full_mix['loudness']
        st = full_mix['stereo']
        metrics['Full Mix LUFS'] = round(L['lufs_integrated'], 2) if np.isfinite(L['lufs_integrated']) else None
        metrics['Full Mix True Peak (dBFS)'] = round(L.get('true_peak_db', L['peak_db']), 2)
        metrics['Full Mix Crest (dB)'] = round(L['crest_factor'], 2)
        metrics['Full Mix PLR'] = round(L['plr'], 2)
        metrics['Full Mix Peak (dBFS)'] = round(L['peak_db'], 2)
        metrics['Full Mix Width'] = round(st['width_overall'], 3) if st['is_stereo'] else None

    if individuals:
        ind_crests = [a['loudness']['crest_factor'] for a, _ in individuals]
        metrics['Avg Individual Crest (dB)'] = round(float(np.mean(ind_crests)), 2)

    # Anomaly count
    anom_count = 0
    for a, ti in analyses_with_info:
        if ti.get('type') == 'BUS':
            continue
        try:
            anoms = detect_anomalies(a)
            anom_count += len(anoms)
        except Exception:
            pass
    metrics['Anomaly count'] = anom_count

    # Health score
    try:
        loud_s, _, _ = _calc_loudness_score(individuals, full_mix)
        dyn_s, _, _ = _calc_dynamics_score(individuals, full_mix)
        spec_s, _, _ = _calc_spectral_balance_score(individuals, full_mix)
        stereo_s, _, _ = _calc_stereo_image_score(individuals, full_mix)
        anom_s, _, _ = _calc_anomalies_score(analyses_with_info)
        health = round(loud_s * 0.20 + dyn_s * 0.20 + spec_s * 0.25 + stereo_s * 0.15 + anom_s * 0.20, 1)
        metrics['Mix Health Score'] = health
    except Exception:
        pass

    metrics['Track count'] = len(individuals)

    return metrics


def _compute_trend(first_val, last_val, metric_name):
    """Compute trend symbol based on metric direction preference.
    Returns one of: '↗' (improving), '↘' (worsening), '→' (stable), '—' (insufficient data)."""
    if first_val is None or last_val is None:
        return '—'
    try:
        first_val = float(first_val)
        last_val = float(last_val)
    except (ValueError, TypeError):
        return '—'

    if first_val == 0:
        return '→'

    pct_change = abs((last_val - first_val) / abs(first_val)) * 100
    if pct_change < 5:
        return '→'

    delta = last_val - first_val

    # Metrics where "higher is better"
    higher_is_better = {'Mix Health Score', 'Full Mix Crest (dB)', 'Avg Individual Crest (dB)',
                        'Track count'}
    # Metrics where "lower is better"
    lower_is_better = {'Anomaly count', 'Full Mix True Peak (dBFS)', 'Full Mix Peak (dBFS)'}
    # Metrics where "closer to target" is better
    target_metrics = {'Full Mix LUFS': -14.0, 'Full Mix PLR': 12.0, 'Full Mix Width': 0.55}

    if metric_name in higher_is_better:
        return '↗' if delta > 0 else '↘'
    elif metric_name in lower_is_better:
        return '↗' if delta < 0 else '↘'
    elif metric_name in target_metrics:
        target = target_metrics[metric_name]
        old_dist = abs(first_val - target)
        new_dist = abs(last_val - target)
        return '↗' if new_dist < old_dist else ('↘' if new_dist > old_dist else '→')
    else:
        return '→'


def _text_sparkline(values):
    """Generate a Unicode text sparkline from a list of numeric values."""
    blocks = " ▁▂▃▄▅▆▇█"
    nums = []
    for v in values:
        try:
            f = float(v)
            import math
            if math.isfinite(f):
                nums.append(f)
        except (ValueError, TypeError):
            pass
    if len(nums) < 2:
        return "—"
    min_v, max_v = min(nums), max(nums)
    if max_v == min_v:
        return "▅" * len(nums)
    return "".join(blocks[min(8, int((n - min_v) / (max_v - min_v) * 8))] for n in nums)


def _add_version_sparklines(ws, n_versions, tracked_metrics,
                             header_row, data_start_row, sparkline_col):
    """Add sparklines to the Version Tracking sheet (M6.5).
    Tries native Excel SparklineGroup first, falls back to Unicode text."""
    from openpyxl.utils import get_column_letter
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

    bg_fill = PatternFill('solid', fgColor='0A0A12')
    thin_border = Border(
        left=Side(style='thin', color='333344'),
        right=Side(style='thin', color='333344'),
        top=Side(style='thin', color='333344'),
        bottom=Side(style='thin', color='333344'),
    )

    # Color per metric group
    metric_colors = {
        'Full Mix LUFS': '00FF9F',
        'Full Mix True Peak (dBFS)': 'FF3D8B',
        'Full Mix Crest (dB)': 'FFD93D',
        'Full Mix PLR': '00D4AA',
        'Full Mix Width': '7B68EE',
        'Avg Individual Crest (dB)': 'FF6B35',
        'Anomaly count': 'FF5252',
        'Mix Health Score': '00D9FF',
        'Track count': '8888A0',
    }

    # Data range: columns 2 through n_versions+1, rows data_start_row onward
    first_data_col = get_column_letter(2)
    last_data_col = get_column_letter(n_versions + 1)
    spark_col_letter = get_column_letter(sparkline_col)

    # Try native Excel sparklines
    native_ok = False
    try:
        from openpyxl.worksheet.sparkline import SparklineGroup, Sparkline
        from openpyxl.styles.colors import Color

        for m_idx, metric_name in enumerate(tracked_metrics):
            row = data_start_row + m_idx
            data_range = f'{first_data_col}{row}:{last_data_col}{row}'
            sqref = f'{spark_col_letter}{row}'

            color_hex = metric_colors.get(metric_name, '00FF9F')
            group = SparklineGroup(
                type='line',
                sparklines=[Sparkline(sqref=sqref, dataRange=data_range)],
                colorSeries=Color(rgb=f'FF{color_hex}'),
                displayEmptyCellsAs='gap',
                high=True,
                low=True,
                last=True,
            )
            ws.sparkline_groups.append(group)

        native_ok = True
    except Exception:
        pass

    # Fallback: Unicode text sparklines
    if not native_ok:
        for m_idx, metric_name in enumerate(tracked_metrics):
            row = data_start_row + m_idx
            values = []
            for v_idx in range(n_versions):
                cell_val = ws.cell(row=row, column=v_idx + 2).value
                values.append(cell_val)
            text = _text_sparkline(values)
            color_hex = metric_colors.get(metric_name, '00FF9F')
            c = ws.cell(row=row, column=sparkline_col, value=text)
            c.font = Font(name='Consolas', size=11, color=color_hex)
            c.fill = bg_fill
            c.border = thin_border
            c.alignment = Alignment(horizontal='center')


def generate_version_tracking_sheet(workbook, analyses_with_info,
                                     output_folder=None, song_name=None,
                                     log_fn=None):
    """
    Génère le sheet 'Version Tracking' (P3.4) dans le workbook donné.
    Détecte automatiquement les rapports Mix Analyzer précédents dans
    output_folder qui matchent le pattern {song_name}_MixAnalyzer_*_GLOBAL.xlsx,
    extrait leurs métriques clés, et affiche l'évolution chronologique
    des indicateurs critiques.
    """
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    _init_ma_fonts()
    from openpyxl.formatting.rule import ColorScaleRule
    from openpyxl.utils import get_column_letter
    import datetime

    if log_fn is None:
        log_fn = lambda msg: None

    log_fn("    Excel: writing Version Tracking sheet (P3.4)...")

    # Theme styles
    bg_fill = PatternFill('solid', fgColor='0A0A12')
    panel_fill = PatternFill('solid', fgColor='1A1A24')
    header_fill = PatternFill('solid', fgColor='1A3A5A')
    accent_font = MA_FONT_SUBHEADING
    header_font = MA_FONT_TABLE_HEADER
    data_font = MA_FONT_BODY
    dim_font = MA_FONT_DIM
    thin_border = Border(
        left=Side(style='thin', color='333344'),
        right=Side(style='thin', color='333344'),
        top=Side(style='thin', color='333344'),
        bottom=Side(style='thin', color='333344'),
    )

    ws = workbook.create_sheet('Version Tracking')
    _apply_clean_layout(ws)
    ws.sheet_properties.tabColor = '3DAAFF'

    # --- Gather version data ---
    current_date = datetime.date.today().isoformat()
    current_metrics = _compute_current_metrics(analyses_with_info)

    # Find previous reports
    previous_reports = []
    if output_folder and song_name:
        found = _find_previous_reports(output_folder, song_name)
        for date_str, path in found:
            # Skip the current report if it already exists on disk (same date)
            if date_str == current_date:
                continue
            extracted = _extract_metrics_from_report(path, log_fn=log_fn)
            if extracted is not None:
                previous_reports.append((date_str, path, extracted))

    # Build ordered version list: previous + current
    versions = []
    for date_str, path, metrics in previous_reports:
        versions.append((date_str, os.path.basename(path), metrics))
    versions.append((current_date, '[current report]', current_metrics))

    n_versions = len(versions)

    # Metrics to track (in order)
    tracked_metrics = [
        'Full Mix LUFS',
        'Full Mix True Peak (dBFS)',
        'Full Mix Crest (dB)',
        'Full Mix PLR',
        'Full Mix Width',
        'Avg Individual Crest (dB)',
        'Anomaly count',
        'Mix Health Score',
        'Track count',
    ]

    # --- Section 1: Header (rows 1-6) ---
    ws['A1'] = 'VERSION TRACKING'
    ws['A1'].font = Font(name='Calibri', size=18, bold=True, color='00D9FF')
    ws['A1'].fill = bg_fill
    # M7.5: Navigation bar
    _xl_add_sheet_nav(ws, 2)

    subtitle = f'Mix evolution over time for: {song_name}' if song_name else 'Mix evolution over time'
    ws['A3'] = subtitle
    ws['A3'].font = dim_font
    ws['A3'].fill = bg_fill

    ws['A5'] = f'Versions detected: {n_versions}'
    ws['A5'].font = data_font
    ws['A5'].fill = bg_fill

    if n_versions >= 2:
        ws['A6'] = f'Oldest: {versions[0][0]}  |  Most recent: {versions[-1][0]}'
    else:
        ws['A6'] = f'Current version: {current_date}'
    ws['A6'].font = dim_font
    ws['A6'].fill = bg_fill

    if n_versions == 1:
        ws['A7'] = ('No previous versions detected. This is the first Mix Analyzer report '
                    'for this song in this folder. Subsequent reports will appear here for comparison.')
        ws['A7'].font = Font(name='Calibri', size=10, italic=True, color='8888A0')
        ws['A7'].fill = bg_fill

    # --- Section 2: Evolution table (row 8 = headers) ---
    header_row = 8

    # Column A: Metric
    c = ws.cell(row=header_row, column=1, value='Metric')
    c.font = header_font
    c.fill = header_fill
    c.border = thin_border

    # Columns B onwards: versions
    for v_idx, (date_str, fname, _) in enumerate(versions):
        col = v_idx + 2
        label = f'{date_str}'
        if fname == '[current report]':
            label += ' (current)'
        c = ws.cell(row=header_row, column=col, value=label)
        c.font = header_font
        c.fill = header_fill
        c.border = thin_border
        c.alignment = Alignment(horizontal='center', wrap_text=True)

    # Delta columns after versions
    delta_abs_col = n_versions + 2
    delta_pct_col = n_versions + 3
    trend_col = n_versions + 4
    sparkline_col = n_versions + 5

    _vt_header_comments = {
        'Δ first→last': "Absolute change from first to most recent version.\nPositive = increased, negative = decreased.",
        'Δ %': "Percentage change from first to most recent version.\nUseful for comparing relative magnitude of changes.",
        'Trend': "Direction indicator: ↑ increasing, ↓ decreasing, → stable.\nBased on overall trajectory across versions.",
        'Sparkline': "Visual trend line across all versions.\nShows the evolution pattern at a glance.",
    }
    for col, label in [(delta_abs_col, 'Δ first→last'),
                       (delta_pct_col, 'Δ %'),
                       (trend_col, 'Trend'),
                       (sparkline_col, 'Sparkline')]:
        c = ws.cell(row=header_row, column=col, value=label)
        c.font = header_font
        c.fill = header_fill
        c.border = thin_border
        c.alignment = Alignment(horizontal='center')
        if label in _vt_header_comments:
            _xl_add_comment(c, _vt_header_comments[label])

    # Data rows
    data_start_row = 9
    for m_idx, metric_name in enumerate(tracked_metrics):
        row = data_start_row + m_idx

        # Metric name
        c = ws.cell(row=row, column=1, value=metric_name)
        c.font = data_font
        c.fill = bg_fill
        c.border = thin_border
        if metric_name in METRIC_GLOSSARY:
            _xl_add_comment(c, METRIC_GLOSSARY[metric_name])

        # Version values
        first_val = None
        last_val = None
        for v_idx, (_, _, metrics) in enumerate(versions):
            col = v_idx + 2
            val = metrics.get(metric_name)
            c = ws.cell(row=row, column=col, value=val)
            c.font = data_font
            c.fill = bg_fill
            c.border = thin_border
            c.alignment = Alignment(horizontal='center')
            if val is not None:
                if isinstance(val, float):
                    c.number_format = '0.00'
                if first_val is None:
                    first_val = val
                last_val = val

        # Delta absolute
        if n_versions >= 2 and first_val is not None and last_val is not None:
            try:
                delta = float(last_val) - float(first_val)
                c = ws.cell(row=row, column=delta_abs_col, value=round(delta, 2))
                c.number_format = '+0.00;-0.00;0.00'
            except (ValueError, TypeError):
                c = ws.cell(row=row, column=delta_abs_col, value='—')
        else:
            c = ws.cell(row=row, column=delta_abs_col, value='—')
        c.font = data_font
        c.fill = bg_fill
        c.border = thin_border
        c.alignment = Alignment(horizontal='center')

        # Delta percentage
        if n_versions >= 2 and first_val is not None and last_val is not None:
            try:
                fv = float(first_val)
                lv = float(last_val)
                if fv != 0:
                    pct = round((lv - fv) / abs(fv) * 100, 1)
                    c = ws.cell(row=row, column=delta_pct_col, value=f'{pct:+.1f}%')
                else:
                    c = ws.cell(row=row, column=delta_pct_col, value='—')
            except (ValueError, TypeError):
                c = ws.cell(row=row, column=delta_pct_col, value='—')
        else:
            c = ws.cell(row=row, column=delta_pct_col, value='—')
        c.font = data_font
        c.fill = bg_fill
        c.border = thin_border
        c.alignment = Alignment(horizontal='center')

        # Trend
        if n_versions >= 2:
            trend = _compute_trend(first_val, last_val, metric_name)
        else:
            trend = '—'
        c = ws.cell(row=row, column=trend_col, value=trend)
        c.font = Font(name='Calibri', size=12, bold=True,
                      color='00FF9F' if trend == '↗' else
                      ('FF3333' if trend == '↘' else '8888A0'))
        c.fill = bg_fill
        c.border = thin_border
        c.alignment = Alignment(horizontal='center')

    data_end_row = data_start_row + len(tracked_metrics) - 1

    # --- Conditional formatting on delta column ---
    if n_versions >= 2:
        delta_range = f'{get_column_letter(delta_abs_col)}{data_start_row}:{get_column_letter(delta_abs_col)}{data_end_row}'
        ws.conditional_formatting.add(
            delta_range,
            ColorScaleRule(
                start_type='min', start_color='FF3333',
                mid_type='num', mid_value=0, mid_color='FFFFFF',
                end_type='max', end_color='00FF9F'))

    # --- Section 4: Source files (after data rows) ---
    source_row = data_end_row + 3

    c = ws.cell(row=source_row, column=1, value='Source files used for this comparison:')
    c.font = accent_font
    c.fill = panel_fill
    c.border = thin_border
    for col in range(2, 4):
        ws.cell(row=source_row, column=col).fill = panel_fill

    source_row += 1
    for col, h in enumerate(['Date', 'File', 'Status'], 1):
        c = ws.cell(row=source_row, column=col, value=h)
        c.font = header_font
        c.fill = header_fill
        c.border = thin_border

    source_row += 1
    for date_str, fname, _ in versions:
        c = ws.cell(row=source_row, column=1, value=date_str)
        c.font = data_font
        c.fill = bg_fill
        c.border = thin_border

        c = ws.cell(row=source_row, column=2, value=fname)
        c.font = data_font
        c.fill = bg_fill
        c.border = thin_border

        status = '[in-memory, current report]' if fname == '[current report]' else 'read from disk'
        c = ws.cell(row=source_row, column=3, value=status)
        c.font = dim_font
        c.fill = bg_fill
        c.border = thin_border
        source_row += 1

    # --- Sparklines for version evolution (M6.5) ---
    if n_versions >= 2:
        _add_version_sparklines(ws, n_versions, tracked_metrics,
                                header_row, data_start_row, sparkline_col)

    # --- Freeze panes (row 9) ---
    ws.freeze_panes = 'A9'

    # --- Column widths ---
    ws.column_dimensions['A'].width = 30
    for v_idx in range(n_versions):
        ws.column_dimensions[get_column_letter(v_idx + 2)].width = 14
    ws.column_dimensions[get_column_letter(delta_abs_col)].width = 14
    ws.column_dimensions[get_column_letter(delta_pct_col)].width = 10
    ws.column_dimensions[get_column_letter(trend_col)].width = 8
    ws.column_dimensions[get_column_letter(sparkline_col)].width = 14

    _apply_dark_background(ws)
    log_fn("    Excel: Version Tracking sheet done.")


def _downsample_for_chart(times, values, target_points=600):
    """Downsample arrays to target_points for Excel chart performance."""
    if len(times) <= target_points:
        return times, values
    step = max(1, len(times) // target_points)
    return times[::step], values[::step]


def _write_peak_rms_chart_data(ws, analysis, start_row, track_label):
    """Write downsampled Peak/RMS/Crest data to a hidden sheet for charting.
    Returns (start_row, end_row) of written data."""
    from openpyxl.styles import Font
    drt = analysis['dynamic_range_timeline']
    times_ds, peak_ds = _downsample_for_chart(drt['times'], drt['peak_db'])
    _, rms_ds = _downsample_for_chart(drt['times'], drt['rms_db'])
    _, crest_ds = _downsample_for_chart(drt['times'], drt['crest_instant'])
    global_crest = analysis['loudness']['crest_factor']

    # Header row
    dim_font = Font(name='Calibri', size=9, color='888888')
    headers = [
        f'Time (s) [{track_label}]', 'Peak (dB)', 'RMS (dB)',
        'High (>12dB)', 'Moderate (6-12dB)', 'Compressed (<6dB)',
        '12 dB threshold', '6 dB threshold', f'Global: {global_crest:.1f} dB',
    ]
    for col, h in enumerate(headers, 1):
        ws.cell(row=start_row, column=col, value=h).font = dim_font
    data_start = start_row + 1

    for i in range(len(times_ds)):
        c = float(crest_ds[i])
        ws.cell(row=data_start + i, column=1, value=round(float(times_ds[i]), 3))
        ws.cell(row=data_start + i, column=2, value=round(float(peak_ds[i]), 2))
        ws.cell(row=data_start + i, column=3, value=round(float(rms_ds[i]), 2))
        # Crest zone columns (Option A: three series)
        ws.cell(row=data_start + i, column=4, value=round(c, 2) if c >= 12 else None)
        ws.cell(row=data_start + i, column=5, value=round(c, 2) if 6 <= c < 12 else None)
        ws.cell(row=data_start + i, column=6, value=round(c, 2) if c < 6 else None)
        # Threshold reference lines (constant)
        ws.cell(row=data_start + i, column=7, value=12)
        ws.cell(row=data_start + i, column=8, value=6)
        ws.cell(row=data_start + i, column=9, value=round(global_crest, 1))

    data_end = data_start + len(times_ds) - 1
    return start_row, data_end


def _create_peak_rms_linechart(ws_data, header_row, data_end, track_name):
    """Create a native Excel LineChart for Peak vs RMS over time."""
    from openpyxl.chart import LineChart, Reference

    chart = LineChart()
    chart.title = f"Peak vs RMS — {track_name}"
    chart.style = 13
    chart.y_axis.title = "Level (dB)"
    chart.x_axis.title = "Time (s)"
    chart.y_axis.scaling.min = -80
    chart.y_axis.scaling.max = 0
    chart.y_axis.delete = False
    chart.x_axis.delete = False
    chart.legend.position = 'b'

    data_start = header_row + 1

    # Category axis (time)
    cats = Reference(ws_data, min_col=1, min_row=data_start, max_row=data_end)

    # Peak series
    peak_ref = Reference(ws_data, min_col=2, min_row=header_row, max_row=data_end)
    chart.add_data(peak_ref, titles_from_data=True)

    # RMS series
    rms_ref = Reference(ws_data, min_col=3, min_row=header_row, max_row=data_end)
    chart.add_data(rms_ref, titles_from_data=True)

    chart.set_categories(cats)

    # Styling
    chart.series[0].graphicalProperties.line.solidFill = "FF3D8B"
    chart.series[0].graphicalProperties.line.width = 12000  # EMU (~1pt)
    chart.series[1].graphicalProperties.line.solidFill = "00FF9F"
    chart.series[1].graphicalProperties.line.width = 12000

    # No markers
    for s in chart.series:
        s.graphicalProperties.line.dashStyle = None
        s.smooth = False

    chart.width = 28
    chart.height = 12

    return chart


def _create_crest_areachart(ws_data, header_row, data_end, track_name):
    """Create a native Excel AreaChart for Crest Factor with color-coded zones."""
    from openpyxl.chart import AreaChart, LineChart, Reference
    from openpyxl.chart.series import DataPoint
    from openpyxl.drawing.line import LineProperties, LineEndProperties

    chart = AreaChart()
    chart.title = f"Crest Factor — {track_name}"
    chart.style = 13
    chart.y_axis.title = "Crest (dB)"
    chart.x_axis.title = "Time (s)"
    chart.y_axis.scaling.min = 0
    chart.y_axis.scaling.max = 30
    chart.y_axis.delete = False
    chart.x_axis.delete = False
    chart.legend.position = 'b'
    chart.grouping = "standard"

    data_start = header_row + 1

    # Category axis (time) — column 1
    cats = Reference(ws_data, min_col=1, min_row=data_start, max_row=data_end)

    # Three zone series (columns 4, 5, 6)
    # High (>12 dB) — green
    high_ref = Reference(ws_data, min_col=4, min_row=header_row, max_row=data_end)
    chart.add_data(high_ref, titles_from_data=True)
    # Moderate (6-12 dB) — yellow/orange
    mid_ref = Reference(ws_data, min_col=5, min_row=header_row, max_row=data_end)
    chart.add_data(mid_ref, titles_from_data=True)
    # Compressed (<6 dB) — red
    low_ref = Reference(ws_data, min_col=6, min_row=header_row, max_row=data_end)
    chart.add_data(low_ref, titles_from_data=True)

    chart.set_categories(cats)

    # Style zone series
    # High dynamic — green
    chart.series[0].graphicalProperties.solidFill = "00FF9F"
    chart.series[0].graphicalProperties.line.solidFill = "00FF9F"
    chart.series[0].graphicalProperties.line.width = 8000
    # Moderate — amber/yellow
    chart.series[1].graphicalProperties.solidFill = "FFAA00"
    chart.series[1].graphicalProperties.line.solidFill = "FFAA00"
    chart.series[1].graphicalProperties.line.width = 8000
    # Compressed — red
    chart.series[2].graphicalProperties.solidFill = "FF3333"
    chart.series[2].graphicalProperties.line.solidFill = "FF3333"
    chart.series[2].graphicalProperties.line.width = 8000

    # Overlay threshold and global crest lines via a secondary LineChart
    line_overlay = LineChart()

    # 12 dB threshold line (column 7)
    thresh12_ref = Reference(ws_data, min_col=7, min_row=header_row, max_row=data_end)
    line_overlay.add_data(thresh12_ref, titles_from_data=True)
    line_overlay.series[0].graphicalProperties.line.solidFill = "00FF9F"
    line_overlay.series[0].graphicalProperties.line.width = 10000
    line_overlay.series[0].graphicalProperties.line.dashStyle = "dash"

    # 6 dB threshold line (column 8)
    thresh6_ref = Reference(ws_data, min_col=8, min_row=header_row, max_row=data_end)
    line_overlay.add_data(thresh6_ref, titles_from_data=True)
    line_overlay.series[1].graphicalProperties.line.solidFill = "FFAA00"
    line_overlay.series[1].graphicalProperties.line.width = 10000
    line_overlay.series[1].graphicalProperties.line.dashStyle = "dash"

    # Global crest factor line (column 9)
    global_ref = Reference(ws_data, min_col=9, min_row=header_row, max_row=data_end)
    line_overlay.add_data(global_ref, titles_from_data=True)
    line_overlay.series[2].graphicalProperties.line.solidFill = "00D9FF"
    line_overlay.series[2].graphicalProperties.line.width = 15000
    line_overlay.series[2].graphicalProperties.line.dashStyle = "sysDot"

    # No markers on line overlay
    for s in line_overlay.series:
        s.smooth = False

    line_overlay.y_axis.scaling.min = 0
    line_overlay.y_axis.scaling.max = 30
    line_overlay.y_axis.delete = True  # hide secondary axis

    # Combine area + line charts
    chart += line_overlay

    chart.width = 28
    chart.height = 12

    return chart


def _write_spectral_chart_data(ws, analysis, start_row, track_label):
    """Write spectral band energy data to a hidden sheet for charting.
    Returns (header_row, end_row) of written data."""
    from openpyxl.styles import Font
    dim_font = Font(name='Calibri', size=9, color='888888')

    S = analysis['spectrum']
    headers = [f'Band [{track_label}]', 'Energy (%)']
    for col, h in enumerate(headers, 1):
        ws.cell(row=start_row, column=col, value=h).font = dim_font

    data_start = start_row + 1
    for i, (name, _, _) in enumerate(FREQ_BANDS):
        ws.cell(row=data_start + i, column=1, value=BAND_LABELS[name])
        ws.cell(row=data_start + i, column=2,
                value=round(S['band_energies'].get(name, 0.0), 2))

    data_end = data_start + len(FREQ_BANDS) - 1
    return start_row, data_end


def _create_spectral_barchart(ws_data, header_row, data_end, track_name):
    """Create a native Excel BarChart for spectral distribution by frequency band."""
    from openpyxl.chart import BarChart, Reference
    from openpyxl.chart.series import DataPoint

    chart = BarChart()
    chart.type = "col"
    chart.style = 10
    chart.title = f"Spectral Balance — {track_name}"
    chart.y_axis.title = "% of total energy"
    chart.x_axis.title = "Frequency Band"
    chart.y_axis.scaling.min = 0
    chart.y_axis.delete = False
    chart.x_axis.delete = False
    chart.legend = None

    data_start = header_row + 1

    # Data reference (Energy column = col 2)
    data_ref = Reference(ws_data, min_col=2, min_row=header_row, max_row=data_end)
    chart.add_data(data_ref, titles_from_data=True)

    # Category labels (Band names = col 1)
    cats = Reference(ws_data, min_col=1, min_row=data_start, max_row=data_end)
    chart.set_categories(cats)

    # Per-band colors: warm (low freq) -> cool (high freq)
    band_colors = [
        "FF3D8B",  # Sub — magenta/pink
        "FF6B35",  # Bass — orange
        "FFD93D",  # Low-Mid — yellow
        "00FF9F",  # Mid — green (theme accent)
        "00D4AA",  # High-Mid — teal
        "00B4D8",  # Presence — cyan
        "7B68EE",  # Air — medium slate blue
    ]
    for i, color in enumerate(band_colors):
        pt = DataPoint(idx=i)
        pt.graphicalProperties.solidFill = color
        chart.series[0].data_points.append(pt)

    chart.width = 20
    chart.height = 12

    return chart


# -- M6.4: Multi-Track Comparison Chart helpers --

_COMPARISON_METRICS = [
    ('Loudness',      lambda a: a['loudness']['lufs_integrated'], -60, 0, False),
    ('Dynamic Range', lambda a: a['loudness']['crest_factor'],      0, 20, False),
    ('Bass',          lambda a: a['spectrum']['band_energies'].get('sub', 0) +
                                a['spectrum']['band_energies'].get('bass', 0), 0, 100, False),
    ('Mid',           lambda a: a['spectrum']['band_energies'].get('low_mid', 0) +
                                a['spectrum']['band_energies'].get('mid', 0) +
                                a['spectrum']['band_energies'].get('high_mid', 0), 0, 100, False),
    ('High',          lambda a: a['spectrum']['band_energies'].get('presence', 0) +
                                a['spectrum']['band_energies'].get('air', 0), 0, 100, False),
    ('Stereo Width',  lambda a: a['stereo']['width_overall'] * 100 if a['stereo']['is_stereo'] else 0,
                      0, 100, False),
]


def _normalize_metric(value, min_val, max_val):
    """Clamp and normalize a metric to 0-100 scale."""
    import math
    if not math.isfinite(value):
        return 0.0
    normalized = (value - min_val) / (max_val - min_val) if max_val != min_val else 0.0
    return round(max(0.0, min(100.0, normalized * 100)), 1)


def _write_comparison_chart_data(ws, analyses_with_info, start_row):
    """Write normalized multi-track comparison data to _chart_data.
    Layout: metrics as rows, tracks as columns (works for both Radar and GroupedBar).
    Returns (header_row, end_row, n_tracks)."""
    from openpyxl.styles import Font
    dim_font = Font(name='Calibri', size=9, color='888888')

    # Select tracks: Full Mix first, then BUS, then individuals, max 12
    full_mixes = [(a, ti) for a, ti in analyses_with_info if ti.get('type') == 'Full Mix']
    buses = [(a, ti) for a, ti in analyses_with_info if ti.get('type') == 'BUS']
    indivs = [(a, ti) for a, ti in analyses_with_info if ti.get('type') == 'Individual']
    selected = (full_mixes + buses + indivs)[:12]

    if len(selected) < 2:
        return None, None, 0

    # Header row: "Metric" + track names
    ws.cell(row=start_row, column=1, value='Metric [Comparison]').font = dim_font
    for col_idx, (a, ti) in enumerate(selected):
        ws.cell(row=start_row, column=col_idx + 2,
                value=a['filename'][:25]).font = dim_font

    # Data rows: one per metric
    data_start = start_row + 1
    for m_idx, (metric_name, extractor, m_min, m_max, _) in enumerate(_COMPARISON_METRICS):
        row = data_start + m_idx
        ws.cell(row=row, column=1, value=metric_name)
        for col_idx, (a, ti) in enumerate(selected):
            try:
                raw = extractor(a)
                val = _normalize_metric(raw, m_min, m_max)
            except Exception:
                val = 0.0
            ws.cell(row=row, column=col_idx + 2, value=val)

    data_end = data_start + len(_COMPARISON_METRICS) - 1
    return start_row, data_end, len(selected)


def _create_comparison_radarchart(ws_data, header_row, data_end, n_tracks):
    """Create a RadarChart for multi-track comparison (best for ≤6 tracks)."""
    from openpyxl.chart import RadarChart, Reference

    chart = RadarChart()
    chart.type = "filled"
    chart.style = 10
    chart.title = "Track Profiles — Multi-Dimensional Comparison"
    chart.y_axis.scaling.min = 0
    chart.y_axis.scaling.max = 100
    chart.y_axis.delete = False

    data_start = header_row + 1

    # Categories = metric names (column 1)
    cats = Reference(ws_data, min_col=1, min_row=data_start, max_row=data_end)
    chart.set_categories(cats)

    # One series per track (columns 2, 3, ...)
    track_colors = [
        "FF3D8B", "00FF9F", "00D4AA", "FFD93D", "7B68EE", "FF6B35",
        "00B4D8", "E040FB", "76FF03", "FF5252", "18FFFF", "FFAB40",
    ]
    for i in range(n_tracks):
        ref = Reference(ws_data, min_col=i + 2, min_row=header_row, max_row=data_end)
        chart.add_data(ref, titles_from_data=True)
        chart.series[i].graphicalProperties.solidFill = track_colors[i % len(track_colors)]
        chart.series[i].graphicalProperties.line.solidFill = track_colors[i % len(track_colors)]
        chart.series[i].graphicalProperties.line.width = 15000

    chart.legend.position = 'b'
    chart.width = 22
    chart.height = 16

    return chart


def _create_comparison_grouped_barchart(ws_data, header_row, data_end, n_tracks):
    """Create a GroupedBarChart for multi-track comparison (>6 tracks)."""
    from openpyxl.chart import BarChart, Reference

    chart = BarChart()
    chart.type = "col"
    chart.grouping = "clustered"
    chart.style = 10
    chart.title = "Track Profiles — Multi-Dimensional Comparison"
    chart.y_axis.title = "Normalized Score (0-100)"
    chart.x_axis.title = "Metric"
    chart.y_axis.scaling.min = 0
    chart.y_axis.scaling.max = 100
    chart.y_axis.delete = False
    chart.x_axis.delete = False

    data_start = header_row + 1

    # Categories = metric names (column 1)
    cats = Reference(ws_data, min_col=1, min_row=data_start, max_row=data_end)
    chart.set_categories(cats)

    # One series per track
    track_colors = [
        "FF3D8B", "00FF9F", "00D4AA", "FFD93D", "7B68EE", "FF6B35",
        "00B4D8", "E040FB", "76FF03", "FF5252", "18FFFF", "FFAB40",
    ]
    for i in range(n_tracks):
        ref = Reference(ws_data, min_col=i + 2, min_row=header_row, max_row=data_end)
        chart.add_data(ref, titles_from_data=True)
        chart.series[i].graphicalProperties.solidFill = track_colors[i % len(track_colors)]

    chart.legend.position = 'b'
    chart.width = 28
    chart.height = 14

    return chart


def _create_comparison_chart(ws_data, header_row, data_end, n_tracks):
    """Auto-select RadarChart (≤6 tracks) or GroupedBarChart (>6 tracks)."""
    if n_tracks <= 6:
        return _create_comparison_radarchart(ws_data, header_row, data_end, n_tracks)
    return _create_comparison_grouped_barchart(ws_data, header_row, data_end, n_tracks)


def generate_excel_report(analyses_with_info, output_path, style_name,
                           full_mix_info=None, ai_prompt='', log_fn=None,
                           export_mode='full',
                           image_quality='standard'):
    """
    Generate complete Excel report.
    analyses_with_info: list of (analysis, track_info) tuples
    export_mode: 'full' (all sheets + individual tracks),
                 'globals' (all global sheets, no individual tracks),
                 'ai_optimized' (AI Context + complementary globals only)
    image_quality: 'standard' (200 DPI) or 'high' (400 DPI, sharper images)
    """
    import tempfile
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.drawing.image import Image as XlImage
    _init_ma_fonts()
    from openpyxl.utils import get_column_letter

    if log_fn is None:
        log_fn = lambda msg: None

    # --- Export mode: determine which sheets to generate ---
    # Sheets always generated regardless of mode
    ALWAYS_SHEETS = {'Index', 'AI Context'}

    # Additional global sheets per mode
    ALL_GLOBAL_SHEETS = [
        'Dashboard', 'Summary', 'Anomalies', 'Full Mix Context',
        'Global Comparison', 'Full Mix Analysis', 'AI Prompt',
        'Freq Conflicts', 'Track Comparison', 'Mix Health Score',
        'Version History',
    ]

    # AI-optimized: only sheets with data NOT covered by AI Context
    AI_OPT_SHEETS = [
        'Anomalies',           # Full descriptions vs compact codes in AI Context
        'Full Mix Context',    # User context (state, plugins, target, note) not in AI Context
        'Mix Health Score',    # Detailed breakdown beyond the 6 scores in AI Context
        'Freq Conflicts',      # Structured conflict data not in AI Context
        'AI Prompt',           # Pure text, negligible weight, essential for AI workflow
        # EXCLUDED from AI-optimized mode:
        # Dashboard       — 19-col metrics, strict subset of AI Context 38 columns (redundant)
        # Summary         — 10-col metrics, strict subset of AI Context (redundant)
        # Global Comparison — 100% visual (heavy images), data in AI Context
        # Full Mix Analysis — mostly visual (5 images + 3 charts), metrics in AI Context row
        # Track Comparison  — interactive Excel tool, raw data already in AI Context
        # Version History   — historical evolution, not current analysis, heavy visual
    ]

    if export_mode == 'full':
        sheets_to_generate = ALWAYS_SHEETS | set(ALL_GLOBAL_SHEETS)
        generate_individual = True
    elif export_mode == 'globals':
        sheets_to_generate = ALWAYS_SHEETS | set(ALL_GLOBAL_SHEETS)
        generate_individual = False
    elif export_mode == 'ai_optimized':
        sheets_to_generate = ALWAYS_SHEETS | set(AI_OPT_SHEETS)
        generate_individual = False
    else:
        raise ValueError(f"Unknown export mode: {export_mode}")

    skipped_sheets = (set(ALL_GLOBAL_SHEETS) | ALWAYS_SHEETS) - sheets_to_generate
    if not generate_individual:
        skipped_sheets.add('Individual track sheets')

    log_fn(f"    Excel: export mode = {export_mode}")
    log_fn(f"    Excel: sheets to generate = {sorted(sheets_to_generate)}")
    if skipped_sheets:
        log_fn(f"    Excel: sheets skipped = {sorted(skipped_sheets)}")
    log_fn(f"    Excel: individual track sheets = {generate_individual}")

    # Navigation targets adapted per mode
    if export_mode == 'ai_optimized':
        nav_targets = [
            ('Index', 'Index'),
            ('Anomalies', 'Anomalies'),
            ('Health Score', 'Mix Health Score'),
            ('AI Context', 'AI Context'),
        ]
    else:
        nav_targets = None  # use default (Index|Summary|Dashboard|Anomalies|Health|AI Context)

    wb = Workbook()
    tmp_files = []

    # Hidden sheet for native chart data (Peak/RMS timelines)
    ws_chart_data = wb.create_sheet('_chart_data')
    _apply_clean_layout(ws_chart_data)
    ws_chart_data.sheet_state = 'hidden'
    chart_data_row = 1  # tracks where to write next block of chart data

    # Theme colors for Excel
    bg_fill = PatternFill('solid', fgColor='0A0A12')
    panel_fill = PatternFill('solid', fgColor='1A1A24')
    header_fill = PatternFill('solid', fgColor='1A3A5A')
    accent_font = MA_FONT_SUBHEADING
    header_font = MA_FONT_TABLE_HEADER
    data_font = MA_FONT_BODY
    dim_font = MA_FONT_DIM
    warn_font = Font(name='Calibri', size=10, bold=True, color=MA_THEME['status_warning'])
    crit_font = Font(name='Calibri', size=10, bold=True, color=MA_THEME['status_error'])
    thin_border = Border(
        left=Side(style='thin', color='333344'),
        right=Side(style='thin', color='333344'),
        top=Side(style='thin', color='333344'),
        bottom=Side(style='thin', color='333344'),
    )

    # Separate tracks
    individuals = [(a, ti) for a, ti in analyses_with_info if ti.get('type') == 'Individual']
    buses = [(a, ti) for a, ti in analyses_with_info if ti.get('type') == 'BUS']
    full_mixes = [(a, ti) for a, ti in analyses_with_info if ti.get('type') == 'Full Mix']

    # ---- SHEET 1: Index ----
    log_fn("    Excel: writing Index sheet...")
    ws_index = wb.active
    _apply_clean_layout(ws_index)
    ws_index.title = 'Index'
    ws_index.sheet_properties.tabColor = '00D9FF'
    mode_suffix = {
        'full': '',
        'globals': ' | Globals only',
        'ai_optimized': ' | AI-optimized export',
    }.get(export_mode, '')
    row = _xl_write_header(ws_index, 'MIX ANALYZER \u2014 REPORT INDEX',
                            f'Style: {style_name}{mode_suffix} | Generated: {datetime.now().strftime("%Y-%m-%d %H:%M")}')
    # M7.5: Navigation bar
    _xl_add_sheet_nav(ws_index, row - 1, current_sheet='Index', nav_targets=nav_targets)

    # Track list with hyperlinks
    headers = ['#', 'Track Name', 'Type', 'Category', 'Sheet Link']
    for col, h in enumerate(headers, 1):
        c = ws_index.cell(row=row, column=col, value=h)
        c.font = header_font
        c.fill = header_fill
        c.border = thin_border
    row += 1

    sheet_names = {}  # track_name -> sheet_name
    for idx, (a, ti) in enumerate(analyses_with_info, 1):
        sname = _safe_sheet_name(os.path.splitext(ti['name'])[0])
        # Avoid duplicate sheet names
        base = sname
        counter = 1
        while sname in sheet_names.values():
            sname = f"{base[:28]}_{counter}"
            counter += 1
        sheet_names[ti['name']] = sname

    if generate_individual:
        for idx, (a, ti) in enumerate(analyses_with_info, 1):
            sname = sheet_names[ti['name']]
            ws_index.cell(row=row, column=1, value=idx).font = data_font
            ws_index.cell(row=row, column=2, value=ti['name']).font = data_font
            ws_index.cell(row=row, column=3, value=ti['type']).font = accent_font if ti['type'] == 'Full Mix' else data_font
            ws_index.cell(row=row, column=4, value=ti.get('category', '')).font = data_font
            link_cell = ws_index.cell(row=row, column=5, value=sname)
            link_cell.font = Font(name='Calibri', size=10, color='00D9FF', underline='single')
            link_cell.hyperlink = f"#{sname}!A1"
            for col in range(1, 6):
                ws_index.cell(row=row, column=col).border = thin_border
                ws_index.cell(row=row, column=col).fill = panel_fill
            row += 1
    else:
        mode_desc = 'AI-optimized mode' if export_mode == 'ai_optimized' else 'compact mode'
        c = ws_index.cell(row=row, column=2, value=f'Individual track sheets: disabled ({mode_desc})')
        c.font = dim_font
        c.fill = panel_fill
        row += 1

    # Link to special sheets — only list sheets actually generated
    row += 1
    special_sheet_order = ['Dashboard', 'AI Context', 'Summary', 'Anomalies', 'Full Mix Context',
                           'Global Comparison', 'Full Mix Analysis', 'AI Prompt',
                           'Freq Conflicts', 'Track Comparison', 'Mix Health Score',
                           'Version History']
    for special_name in special_sheet_order:
        if special_name not in sheets_to_generate:
            continue
        ws_index.cell(row=row, column=2, value=special_name).font = data_font
        link_cell = ws_index.cell(row=row, column=5, value=special_name)
        link_cell.font = Font(name='Calibri', size=10, color='00D9FF', underline='single')
        link_cell.hyperlink = f"#{special_name}!A1"
        row += 1

    ws_index.column_dimensions['B'].width = 45
    ws_index.column_dimensions['C'].width = 12
    ws_index.column_dimensions['D'].width = 20
    ws_index.column_dimensions['E'].width = 25

    # M7.5: link font for track names (used by multiple sheets)
    _link_font = Font(name='Calibri', size=10, color='00D9FF', underline='single')
    from openpyxl.formatting.rule import ColorScaleRule, DataBarRule, IconSetRule

    # ---- SHEET 2: Summary ----
    # Mode ai_optimized: EXCLUDED (10-col metrics, strict subset of AI Context)
    if 'Summary' in sheets_to_generate:
        log_fn("    Excel: writing Summary sheet...")
        ws_sum = wb.create_sheet('Summary')
        _apply_clean_layout(ws_sum)
        ws_sum.sheet_properties.tabColor = 'B967FF'
        row = _xl_write_header(ws_sum, 'SUMMARY — GLOBAL METRICS', f'{len(analyses_with_info)} tracks analyzed')
        # M7.5: Navigation bar
        _xl_add_sheet_nav(ws_sum, row - 1, current_sheet='Summary', nav_targets=nav_targets)

        sum_headers = ['Track', 'Type', 'Category', 'LUFS', 'Peak (dB)', 'Crest (dB)',
                       'Stereo Width', 'Dom. Band', 'Centroid (Hz)', 'Duration (s)']
        for col, h in enumerate(sum_headers, 1):
            c = ws_sum.cell(row=row, column=col, value=h)
            c.font = header_font
            c.fill = header_fill
            c.border = thin_border
            c.alignment = Alignment(horizontal='center', vertical='center')
            # Add glossary comment
            if h in METRIC_GLOSSARY:
                _xl_add_comment(c, METRIC_GLOSSARY[h])
        sum_header_row = row
        row += 1

        for a, ti in analyses_with_info:
            L = a['loudness']
            S = a['spectrum']
            st = a['stereo']
            vals = [
                a['filename'],
                ti['type'],
                ti.get('category', ''),
                round(L['lufs_integrated'], 2) if np.isfinite(L['lufs_integrated']) else None,
                round(L['peak_db'], 2),
                round(L['crest_factor'], 2),
                round(st['width_overall'], 3) if st['is_stereo'] else 'mono',
                BAND_LABELS.get(S['dominant_band'], S['dominant_band']),
                round(S['centroid'], 0),
                round(a['duration'], 1),
            ]
            for col, v in enumerate(vals, 1):
                c = ws_sum.cell(row=row, column=col, value=v)
                c.font = data_font
                c.border = thin_border
                c.fill = panel_fill
                if col >= 4:
                    c.alignment = Alignment(horizontal='center')
            # M7.5: Make track name clickable → individual sheet
            if generate_individual and ti['name'] in sheet_names:
                sname = sheet_names[ti['name']]
                track_cell = ws_sum.cell(row=row, column=1)
                track_cell.hyperlink = f"#{sname}!A1"
                track_cell.font = _link_font
            row += 1

        # Enriched conditional formatting (Phase 2)
        if len(analyses_with_info) > 0:
            data_start = sum_header_row + 1
            data_end = data_start + len(analyses_with_info) - 1
            # LUFS: color scale (red=quiet -> yellow -> green=loud)
            ws_sum.conditional_formatting.add(
                f'D{data_start}:D{data_end}',
                ColorScaleRule(start_type='min', start_color='FF3333',
                               mid_type='percentile', mid_value=50, mid_color='FFAA00',
                               end_type='max', end_color='00FF9F'))
            # LUFS: data bars
            ws_sum.conditional_formatting.add(
                f'D{data_start}:D{data_end}',
                DataBarRule(start_type='min', end_type='max', color='00D9FF'))
            # Peak: color scale (red=hot -> green=safe)
            ws_sum.conditional_formatting.add(
                f'E{data_start}:E{data_end}',
                ColorScaleRule(start_type='max', start_color='FF3333',
                               end_type='min', end_color='00FF9F'))
            # Crest factor: color scale + data bars
            ws_sum.conditional_formatting.add(
                f'F{data_start}:F{data_end}',
                ColorScaleRule(start_type='min', start_color='FF3333',
                               mid_type='percentile', mid_value=50, mid_color='FFAA00',
                               end_type='max', end_color='00D9FF'))
            ws_sum.conditional_formatting.add(
                f'F{data_start}:F{data_end}',
                DataBarRule(start_type='min', end_type='max', color='B967FF'))
            # Stereo Width: data bars
            ws_sum.conditional_formatting.add(
                f'G{data_start}:G{data_end}',
                DataBarRule(start_type='min', end_type='max', color='00FF9F'))
            # Centroid: color scale (low=warm -> high=bright)
            ws_sum.conditional_formatting.add(
                f'I{data_start}:I{data_end}',
                ColorScaleRule(start_type='min', start_color='B967FF',
                               end_type='max', end_color='FF3D8B'))

            # M7.3: Icon sets for instant visual status
            # Crest Factor (col F): traffic lights — red <6, yellow 6-12, green >12
            ws_sum.conditional_formatting.add(
                f'F{data_start}:F{data_end}',
                IconSetRule(icon_style='3TrafficLights1', type='num',
                            values=[0, 6, 12], showValue=True, reverse=False))
            # Peak (col E): 3Symbols — ✓ safe <-1.5, ! caution -1.5 to -0.5, ✗ risk >-0.5
            ws_sum.conditional_formatting.add(
                f'E{data_start}:E{data_end}',
                IconSetRule(icon_style='3Symbols2', type='num',
                            values=[-100, -1.5, -0.5], showValue=True, reverse=True))

            # M7.3: Multi-criteria formula alerts on Summary rows
            from openpyxl.formatting.rule import FormulaRule
            # Streaming risk: LUFS > -10 AND Peak > -1 → red highlight
            ws_sum.conditional_formatting.add(
                f'A{data_start}:J{data_end}',
                FormulaRule(
                    formula=[f'AND($D{data_start}>-10,$E{data_start}>-1)'],
                    fill=PatternFill(start_color='FF5252', end_color='FF5252', fill_type='solid'),
                    font=Font(color='FFFFFF', bold=True)))
            # Over-compressed + quiet: Crest < 6 AND LUFS < -16 → yellow highlight
            ws_sum.conditional_formatting.add(
                f'A{data_start}:J{data_end}',
                FormulaRule(
                    formula=[f'AND($F{data_start}<6,$D{data_start}<-16)'],
                    fill=PatternFill(start_color='FFD93D', end_color='FFD93D', fill_type='solid')))

        # Auto-filter on Summary
        ws_sum.auto_filter.ref = f'A{sum_header_row}:J{max(row - 1, sum_header_row + 1)}'

        for col_idx in range(1, 11):
            ws_sum.column_dimensions[get_column_letter(col_idx)].width = 16
        ws_sum.column_dimensions['A'].width = 40
        _apply_dark_background(ws_sum)

    # ---- SHEET 3: Anomalies ----
    log_fn("    Excel: writing Anomalies sheet...")
    ws_anom = wb.create_sheet('Anomalies')
    _apply_clean_layout(ws_anom)
    ws_anom.sheet_properties.tabColor = 'FF3333'
    row = _xl_write_header(ws_anom, 'ANOMALIES')
    # M7.5: Navigation bar
    _xl_add_sheet_nav(ws_anom, row - 1, current_sheet='Anomalies', nav_targets=nav_targets)

    anom_headers = ['Track', 'Type', 'Severity', 'Description']
    _anom_header_comments = {
        'Severity': "CRITICAL: Likely audible issue requiring fix.\nWARNING: Potential issue, verify by listening.\nINFO: Informational, may be intentional.",
        'Description': "Hover over individual descriptions for\nexplanation, impact, and suggestions.",
    }
    for col, h in enumerate(anom_headers, 1):
        c = ws_anom.cell(row=row, column=col, value=h)
        c.font = header_font
        c.fill = header_fill
        c.border = thin_border
        if h in _anom_header_comments:
            _xl_add_comment(c, _anom_header_comments[h])
    row += 1

    crit_fill = PatternFill('solid', fgColor='2A0A0A')
    warn_fill = PatternFill('solid', fgColor='2A1A0A')
    for a, ti in analyses_with_info:
        if a.get('anomalies'):
            for sev, desc in a['anomalies']:
                row_fill = crit_fill if sev == 'critical' else (warn_fill if sev == 'warning' else panel_fill)
                track_cell = ws_anom.cell(row=row, column=1, value=a['filename'])
                track_cell.font = data_font
                # M7.5: Link anomaly track name → individual sheet
                if generate_individual and ti['name'] in sheet_names:
                    sname = sheet_names[ti['name']]
                    track_cell.hyperlink = f"#{sname}!A1"
                    track_cell.font = _link_font
                ws_anom.cell(row=row, column=2, value=ti['type']).font = data_font
                sev_cell = ws_anom.cell(row=row, column=3, value=sev.upper())
                sev_cell.font = crit_font if sev == 'critical' else warn_font
                desc_cell = ws_anom.cell(row=row, column=4, value=desc)
                desc_cell.font = data_font
                # M7.4: Add contextual comment for known anomaly types
                for anom_key, anom_text in ANOMALY_COMMENTS.items():
                    if anom_key in desc.lower():
                        _xl_add_comment(desc_cell, anom_text)
                        break
                for col in range(1, 5):
                    ws_anom.cell(row=row, column=col).border = thin_border
                    ws_anom.cell(row=row, column=col).fill = row_fill
                row += 1

    ws_anom.auto_filter.ref = f'A4:D{max(row - 1, 5)}'
    ws_anom.column_dimensions['A'].width = 40
    ws_anom.column_dimensions['B'].width = 12
    ws_anom.column_dimensions['C'].width = 12
    ws_anom.column_dimensions['D'].width = 70
    _apply_dark_background(ws_anom)

    # ---- SHEET 4: Full Mix Context ----
    log_fn("    Excel: writing Full Mix Context sheet...")
    ws_ctx = wb.create_sheet('Full Mix Context')
    _apply_clean_layout(ws_ctx)
    ws_ctx.sheet_properties.tabColor = 'B967FF'
    row = _xl_write_header(ws_ctx, 'FULL MIX CONTEXT')
    # M7.5: Navigation bar
    _xl_add_sheet_nav(ws_ctx, row - 1, nav_targets=nav_targets)

    if full_mix_info:
        ctx_items = [
            ('Mix State', full_mix_info.get('state', 'Not specified')),
            ('Active Plugins', ', '.join(full_mix_info.get('plugins', [])) or 'None'),
            ('Loudness Target', full_mix_info.get('loudness_target', 'Not specified')),
            ('Note', full_mix_info.get('note', '') or 'None'),
        ]
        for label, val in ctx_items:
            ws_ctx.cell(row=row, column=1, value=label).font = accent_font
            ws_ctx.cell(row=row, column=2, value=val).font = data_font
            ws_ctx.cell(row=row, column=1).fill = panel_fill
            ws_ctx.cell(row=row, column=2).fill = panel_fill
            ws_ctx.cell(row=row, column=1).border = thin_border
            ws_ctx.cell(row=row, column=2).border = thin_border
            row += 1
    else:
        ws_ctx.cell(row=row, column=1, value='No Full Mix context configured.').font = dim_font

    ws_ctx.column_dimensions['A'].width = 20
    ws_ctx.column_dimensions['B'].width = 60
    _apply_dark_background(ws_ctx)

    # ---- SHEET 5+: One sheet per track (Individual + BUS) ----
    if generate_individual:
        track_sheets = [(a, ti) for a, ti in analyses_with_info
                        if ti['type'] in ('Individual', 'BUS')]
        # Build ordered list of sheet names for navigation
        track_sheet_names_ordered = []
        for a_t, ti_t in track_sheets:
            track_sheet_names_ordered.append(
                sheet_names.get(ti_t['name'], _safe_sheet_name(os.path.splitext(ti_t['name'])[0])))

        # Metric glossary keys for per-track sheets
        _trk_glossary = {
            'Duration': 'Duration (s)', 'LUFS Integrated': 'LUFS', 'LUFS Short-term Max': 'LUFS',
            'True Peak': 'True Peak', 'Peak': 'Peak (dB)', 'RMS': 'RMS',
            'Crest Factor': 'Crest (dB)', 'PLR': 'PLR', 'PSR': 'PSR', 'LRA': 'LRA',
            'Centroid': 'Centroid (Hz)', 'Flatness': 'Flatness',
            'Phase Correlation': 'Phase Correlation', 'Stereo Width': 'Stereo Width',
        }

        for sheet_idx, (a, ti) in enumerate(track_sheets):
            sname = sheet_names.get(ti['name'], _safe_sheet_name(os.path.splitext(ti['name'])[0]))
            log_fn(f"    Excel: writing sheet {sheet_idx + 1}/{len(track_sheets)}: {sname}")
            ws_trk = wb.create_sheet(sname)
            _apply_clean_layout(ws_trk)
            ws_trk.sheet_properties.tabColor = 'FF3D8B' if ti['type'] == 'BUS' else '00D9FF'
            row = _xl_write_header(ws_trk, ti['name'],
                                    f"Type: {ti['type']} | Category: {ti.get('category', '')}")

            # Navigation links (Phase 2)
            row = _xl_add_nav_row(ws_trk, row, track_sheet_names_ordered, sheet_idx)
            row += 1

            # Metrics table
            L = a['loudness']
            S = a['spectrum']
            st = a['stereo']
            metrics = [
                ('Duration', f"{a['duration']:.1f} s"),
                ('LUFS Integrated', f"{L['lufs_integrated']:+.2f}" if np.isfinite(L['lufs_integrated']) else '-'),
                ('LUFS Short-term Max', f"{L['lufs_short_term_max']:+.2f}" if np.isfinite(L['lufs_short_term_max']) else '-'),
                ('True Peak', f"{L['true_peak_db']:+.2f} dBFS"),
                ('Peak', f"{L['peak_db']:+.2f} dB"),
                ('RMS', f"{L['rms_db']:+.2f} dB"),
                ('Crest Factor', f"{L['crest_factor']:.2f} dB"),
                ('PLR', f"{L['plr']:.2f} dB"),
                ('PSR', f"{L['psr']:.2f} dB"),
                ('LRA', f"{L['lra']:.2f} LU"),
                ('Dominant Band', BAND_LABELS.get(S['dominant_band'], S['dominant_band'])),
                ('Centroid', f"{S['centroid']:.0f} Hz"),
                ('Rolloff 85%', f"{S['rolloff']:.0f} Hz"),
                ('Flatness', f"{S['flatness']:.4f}"),
                ('Phase Correlation', f"{st['correlation']:+.3f}" if st['is_stereo'] else 'Mono'),
                ('Stereo Width', f"{st['width_overall']:.3f}" if st['is_stereo'] else 'Mono'),
            ]
            for label, val in metrics:
                c_label = ws_trk.cell(row=row, column=1, value=label)
                c_label.font = accent_font
                c_label.fill = panel_fill
                c_label.border = thin_border
                c_label.alignment = Alignment(horizontal='left', vertical='center')
                # Add glossary comment (Phase 2)
                glossary_key = _trk_glossary.get(label)
                if glossary_key and glossary_key in METRIC_GLOSSARY:
                    _xl_add_comment(c_label, METRIC_GLOSSARY[glossary_key])

                c_val = ws_trk.cell(row=row, column=2, value=val)
                c_val.font = data_font
                c_val.fill = panel_fill
                c_val.border = thin_border
                c_val.alignment = Alignment(horizontal='right', vertical='center')
                row += 1

            # Anomalies for this track
            if a.get('anomalies'):
                row += 1
                ws_trk.cell(row=row, column=1, value='ANOMALIES').font = Font(
                    name='Calibri', size=12, bold=True, color='FFAA00')
                row += 1
                for sev, desc in a['anomalies']:
                    c_sev = ws_trk.cell(row=row, column=1, value=sev.upper())
                    c_sev.font = crit_font if sev == 'critical' else warn_font
                    c_sev.fill = panel_fill
                    c_sev.border = thin_border
                    c_desc = ws_trk.cell(row=row, column=2, value=desc)
                    c_desc.font = data_font
                    c_desc.fill = panel_fill
                    c_desc.border = thin_border
                    # M7.4: Contextual comment for anomaly
                    for anom_key, anom_text in ANOMALY_COMMENTS.items():
                        if anom_key in desc.lower():
                            _xl_add_comment(c_desc, anom_text)
                            break
                    row += 1

            # Embed matplotlib visualizations as images
            row += 2
            img_row = row
            page_fns = [
                ('Identity', lambda: page_identity(a, ti, style_name)),
                ('Temporal', lambda: page_temporal(a, ti)),
                ('Spectral', lambda: page_spectral(a, ti)),
                ('Spectrogram', lambda: page_spectrogram(a, ti)),
                ('Musical', lambda: page_musical(a, ti)),
                ('Stereo', lambda: page_stereo(a, ti)),
                ('Multiband Timeline', lambda: page_multiband_timeline(a, ti)),
                ('Characteristics', lambda: page_characteristics(a, ti, style_name)),
            ]

            for page_name, page_fn in page_fns:
                try:
                    fig = page_fn()
                    img, tmp_path, rspan = _fig_to_image(fig, quality=image_quality)
                    tmp_files.append(tmp_path)
                    ws_trk.add_image(img, f'A{img_row}')
                    img_row += rspan
                except Exception:
                    pass

            # Native Excel LineChart for Peak vs RMS (M6.1)
            try:
                hdr_row, end_row = _write_peak_rms_chart_data(
                    ws_chart_data, a, chart_data_row, sname)
                peak_rms_chart = _create_peak_rms_linechart(
                    ws_chart_data, hdr_row, end_row, a['filename'])
                ws_trk.add_chart(peak_rms_chart, f'A{img_row}')
                chart_data_row = end_row + 2
                img_row += 24  # chart occupies ~24 rows
            except Exception:
                pass

            # Native Excel AreaChart for Crest Factor (M6.2)
            try:
                crest_chart = _create_crest_areachart(
                    ws_chart_data, hdr_row, end_row, a['filename'])
                ws_trk.add_chart(crest_chart, f'A{img_row}')
                img_row += 24
            except Exception:
                pass

            # Native Excel BarChart for Spectral Distribution (M6.3)
            try:
                spec_hdr, spec_end = _write_spectral_chart_data(
                    ws_chart_data, a, chart_data_row, sname)
                spectral_chart = _create_spectral_barchart(
                    ws_chart_data, spec_hdr, spec_end, a['filename'])
                ws_trk.add_chart(spectral_chart, f'A{img_row}')
                chart_data_row = spec_end + 2
                img_row += 24
            except Exception:
                pass

            ws_trk.column_dimensions['A'].width = 25
            ws_trk.column_dimensions['B'].width = 40
            _apply_dark_background(ws_trk)

        log_fn(f"    Excel: {len(track_sheets)} individual track sheets generated.")
    else:
        log_fn(f"    Excel: Individual track sheets skipped ({export_mode} mode).")

    # ---- SHEET: Global Comparison ----
    # Mode ai_optimized: EXCLUDED (100% visual, heavy images, data in AI Context)
    if 'Global Comparison' not in sheets_to_generate:
        log_fn("    Excel: Global Comparison skipped (not in export mode).")
    else:
        log_fn("    Excel: writing Global Comparison sheet...")
        ws_global = wb.create_sheet('Global Comparison')
        _apply_clean_layout(ws_global)
        ws_global.sheet_properties.tabColor = '00FF9F'
        row = _xl_write_header(ws_global, 'GLOBAL COMPARISON',
                                'Masking matrix, spectral balance, LUFS/Crest comparisons (excludes BUS)')
        # M7.5: Navigation bar
        _xl_add_sheet_nav(ws_global, row - 1, nav_targets=nav_targets)

        if individuals:
            # Masking matrix as image
            try:
                fig = plt.figure(figsize=(16, 9))
                fig.suptitle('FREQUENCY MASKING MATRIX (High Resolution)', fontsize=14,
                             color=THEME['accent1'], fontweight='bold', y=0.97)
                ax = fig.add_subplot(111)
                hires_labels = [label for label, _, _ in FREQ_BANDS_HIRES]
                n_bands = len(FREQ_BANDS_HIRES)
                matrix = np.zeros((len(individuals), n_bands))
                for i, (a_i, ti_i) in enumerate(individuals):
                    hires = compute_hires_band_energies(a_i['_mono'], a_i['sample_rate'])
                    for j, (label, _, _) in enumerate(FREQ_BANDS_HIRES):
                        matrix[i, j] = hires[label]
                im = ax.imshow(matrix, aspect='auto', cmap='magma', interpolation='nearest')
                track_labels = [a_i['filename'][:35] for a_i, _ in individuals]
                ax.set_yticks(range(len(individuals)))
                ax.set_yticklabels(track_labels, fontsize=6)
                ax.set_xticks(range(n_bands))
                ax.set_xticklabels(hires_labels, rotation=45, ha='right', fontsize=7)
                fig.colorbar(im, ax=ax, label='% of track energy', pad=0.01)
                plt.tight_layout(rect=[0, 0.02, 1, 0.90])

                img, tmp_path, rspan = _fig_to_image(fig, quality=image_quality)
                tmp_files.append(tmp_path)
                ws_global.add_image(img, f'A{row}')
                row += rspan
            except Exception:
                pass

            # LUFS comparison bar chart
            try:
                fig = plt.figure(figsize=(16, 9))
                fig.suptitle('LUFS COMPARISON', fontsize=14, color=THEME['accent1'],
                             fontweight='bold', y=0.97)
                ax = fig.add_subplot(111)
                names = [a_i['filename'][:30] for a_i, _ in individuals]
                lufs_vals = [a_i['loudness']['lufs_integrated'] for a_i, _ in individuals]
                lufs_vals = [v if np.isfinite(v) else -70 for v in lufs_vals]
                colors = [THEME['accent1'] if v > -20 else THEME['accent4'] for v in lufs_vals]
                ax.barh(range(len(names)), lufs_vals, color=colors, edgecolor=THEME['fg_dim'], linewidth=0.3)
                ax.set_yticks(range(len(names)))
                ax.set_yticklabels(names, fontsize=7)
                ax.set_xlabel('Integrated LUFS', fontsize=11)
                ax.invert_yaxis()
                ax.grid(True, alpha=0.3, axis='x')
                plt.tight_layout(rect=[0, 0.02, 1, 0.90])

                img, tmp_path, rspan = _fig_to_image(fig, quality=image_quality)
                tmp_files.append(tmp_path)
                ws_global.add_image(img, f'A{row}')
                row += rspan
            except Exception:
                pass

            # Crest factor comparison
            try:
                fig = plt.figure(figsize=(16, 9))
                fig.suptitle('CREST FACTOR COMPARISON', fontsize=14, color=THEME['accent1'],
                             fontweight='bold', y=0.97)
                ax = fig.add_subplot(111)
                crest_vals = [a_i['loudness']['crest_factor'] for a_i, _ in individuals]
                bar_colors = []
                for v in crest_vals:
                    if v < 6:
                        bar_colors.append(THEME['critical'])
                    elif v < 12:
                        bar_colors.append(THEME['warning'])
                    else:
                        bar_colors.append(THEME['accent4'])
                ax.barh(range(len(names)), crest_vals, color=bar_colors,
                         edgecolor=THEME['fg_dim'], linewidth=0.3)
                ax.set_yticks(range(len(names)))
                ax.set_yticklabels(names, fontsize=7)
                ax.set_xlabel('Crest Factor (dB)', fontsize=11)
                ax.axvline(6, color=THEME['critical'], linewidth=1, linestyle='--', alpha=0.7)
                ax.axvline(12, color=THEME['accent4'], linewidth=1, linestyle='--', alpha=0.7)
                ax.invert_yaxis()
                ax.grid(True, alpha=0.3, axis='x')
                plt.tight_layout(rect=[0, 0.02, 1, 0.90])

                img, tmp_path, rspan = _fig_to_image(fig, quality=image_quality)
                tmp_files.append(tmp_path)
                ws_global.add_image(img, f'A{row}')
                row += rspan
            except Exception:
                pass

            # Spectral balance comparison
            try:
                fig = plt.figure(figsize=(16, 9))
                fig.suptitle('SPECTRAL BALANCE COMPARISON', fontsize=14, color=THEME['accent1'],
                             fontweight='bold', y=0.97)
                ax = fig.add_subplot(111)
                band_names_7 = [BAND_LABELS[name] for name, _, _ in FREQ_BANDS]
                band_x = np.arange(len(FREQ_BANDS))
                bar_width = 0.8 / max(len(individuals), 1)
                for idx_t, (a_i, ti_i) in enumerate(individuals[:12]):
                    offsets = band_x + idx_t * bar_width - 0.4
                    vals = [a_i['spectrum']['band_energies'][name] for name, _, _ in FREQ_BANDS]
                    ax.bar(offsets, vals, width=bar_width, label=a_i['filename'][:20], alpha=0.8)
                ax.set_xticks(band_x)
                ax.set_xticklabels(band_names_7, rotation=25, ha='right', fontsize=9)
                ax.set_ylabel('% energy')
                ax.legend(fontsize=6, ncol=3, loc='upper right')
                ax.grid(True, alpha=0.3, axis='y')
                plt.tight_layout(rect=[0, 0.02, 1, 0.90])

                img, tmp_path, rspan = _fig_to_image(fig, quality=image_quality)
                tmp_files.append(tmp_path)
                ws_global.add_image(img, f'A{row}')
                row += rspan
            except Exception:
                pass

            # Native Excel chart for Multi-Track Comparison (M6.4)
            try:
                comp_hdr, comp_end, comp_n = _write_comparison_chart_data(
                    ws_chart_data, analyses_with_info, chart_data_row)
                if comp_hdr is not None and comp_n >= 2:
                    comparison_chart = _create_comparison_chart(
                        ws_chart_data, comp_hdr, comp_end, comp_n)
                    ws_global.add_chart(comparison_chart, f'A{row}')
                    chart_data_row = comp_end + 2
                    row += 30
            except Exception:
                pass

        _apply_dark_background(ws_global)

    # ---- SHEET: Full Mix Analysis ----
    # Mode ai_optimized: EXCLUDED (mostly visual, metrics in AI Context row)
    if 'Full Mix Analysis' not in sheets_to_generate:
        log_fn("    Excel: Full Mix Analysis skipped (not in export mode).")
    else:
        log_fn("    Excel: writing Full Mix Analysis sheet...")
        ws_fm = wb.create_sheet('Full Mix Analysis')
        _apply_clean_layout(ws_fm)
        ws_fm.sheet_properties.tabColor = 'B967FF'

        # M7.5: Navigation bar
        _xl_add_sheet_nav(ws_fm, 3, nav_targets=nav_targets)

        if full_mixes:
            a_fm, ti_fm = full_mixes[0]
            row = _xl_write_header(ws_fm, 'FULL MIX ANALYSIS', a_fm['filename'])

            L = a_fm['loudness']
            S = a_fm['spectrum']
            st = a_fm['stereo']
            tempo = a_fm.get('tempo', {})

            fm_metrics = [
                ('LUFS Integrated', f"{L['lufs_integrated']:+.2f}" if np.isfinite(L['lufs_integrated']) else '-'),
                ('LUFS Short-term Max', f"{L['lufs_short_term_max']:+.2f}" if np.isfinite(L['lufs_short_term_max']) else '-'),
                ('True Peak', f"{L['true_peak_db']:+.2f} dBFS"),
                ('Crest Factor', f"{L['crest_factor']:.2f} dB"),
                ('PLR', f"{L['plr']:.2f} dB"),
                ('LRA', f"{L['lra']:.2f} LU"),
                ('Dominant Band', BAND_LABELS.get(S['dominant_band'], S['dominant_band'])),
                ('Centroid', f"{S['centroid']:.0f} Hz"),
                ('Phase Correlation', f"{st['correlation']:+.3f}" if st['is_stereo'] else 'Mono'),
                ('Stereo Width', f"{st['width_overall']:.3f}" if st['is_stereo'] else 'Mono'),
                ('Tempo (median)', f"{tempo.get('tempo_median', 0):.1f} BPM"),
                ('Tempo range', f"{tempo.get('tempo_min', 0):.0f}-{tempo.get('tempo_max', 0):.0f} BPM"),
            ]
            for label, val in fm_metrics:
                c1 = ws_fm.cell(row=row, column=1, value=label)
                c1.font = accent_font
                c1.fill = panel_fill
                c1.border = thin_border
                c1.alignment = Alignment(horizontal='left', vertical='center')
                glossary_key_fm = {'LUFS Integrated': 'LUFS', 'Crest Factor': 'Crest (dB)',
                                   'PLR': 'PLR', 'LRA': 'LRA', 'Centroid': 'Centroid (Hz)',
                                   'Phase Correlation': 'Phase Correlation',
                                   'Stereo Width': 'Stereo Width', 'True Peak': 'True Peak'}.get(label)
                if glossary_key_fm and glossary_key_fm in METRIC_GLOSSARY:
                    _xl_add_comment(c1, METRIC_GLOSSARY[glossary_key_fm])
                c2 = ws_fm.cell(row=row, column=2, value=val)
                c2.font = data_font
                c2.fill = panel_fill
                c2.border = thin_border
                c2.alignment = Alignment(horizontal='right', vertical='center')
                row += 1

            # Structure detection
            row += 1
            try:
                full_mix_structure = analyze_structure_sections(a_fm['_mono'], a_fm['sample_rate'])
                if full_mix_structure and full_mix_structure.get('success'):
                    ws_fm.cell(row=row, column=1, value='SECTION BOUNDARIES').font = Font(
                        name='Calibri', size=12, bold=True, color='B967FF')
                    row += 1
                    bounds = full_mix_structure.get('boundaries', [])
                    valid_bounds = [b for b in bounds if b > 0.5]
                    for i, b in enumerate(valid_bounds):
                        ws_fm.cell(row=row, column=1, value=f'Section {i + 1}').font = data_font
                        ws_fm.cell(row=row, column=2, value=f'{b:.1f} s').font = data_font
                        row += 1
            except Exception:
                pass

            # Full Mix visualizations
            row += 2
            fm_pages = [
                ('Identity', lambda: page_identity(a_fm, ti_fm, style_name)),
                ('Temporal', lambda: page_temporal(a_fm, ti_fm)),
                ('Spectral', lambda: page_spectral(a_fm, ti_fm)),
                ('Spectrogram', lambda: page_spectrogram(a_fm, ti_fm)),
                ('Multiband', lambda: page_multiband_timeline(a_fm, ti_fm)),
            ]
            for page_name, page_fn in fm_pages:
                try:
                    fig = page_fn()
                    img, tmp_path, rspan = _fig_to_image(fig, quality=image_quality)
                    tmp_files.append(tmp_path)
                    ws_fm.add_image(img, f'A{row}')
                    row += rspan
                except Exception:
                    pass

            # Native Excel LineChart for Peak vs RMS (M6.1)
            try:
                hdr_row, end_row = _write_peak_rms_chart_data(
                    ws_chart_data, a_fm, chart_data_row, 'FullMix')
                peak_rms_chart = _create_peak_rms_linechart(
                    ws_chart_data, hdr_row, end_row, a_fm['filename'])
                ws_fm.add_chart(peak_rms_chart, f'A{row}')
                chart_data_row = end_row + 2
                row += 24
            except Exception:
                pass

            # Native Excel AreaChart for Crest Factor (M6.2)
            try:
                crest_chart = _create_crest_areachart(
                    ws_chart_data, hdr_row, end_row, a_fm['filename'])
                ws_fm.add_chart(crest_chart, f'A{row}')
                row += 24
            except Exception:
                pass

            # Native Excel BarChart for Spectral Distribution (M6.3)
            try:
                spec_hdr, spec_end = _write_spectral_chart_data(
                    ws_chart_data, a_fm, chart_data_row, 'FullMix')
                spectral_chart = _create_spectral_barchart(
                    ws_chart_data, spec_hdr, spec_end, a_fm['filename'])
                ws_fm.add_chart(spectral_chart, f'A{row}')
                chart_data_row = spec_end + 2
                row += 24
            except Exception:
                pass

            ws_fm.column_dimensions['A'].width = 25
            ws_fm.column_dimensions['B'].width = 40
        else:
            _xl_write_header(ws_fm, 'FULL MIX ANALYSIS', 'No Full Mix track detected.')
        _apply_dark_background(ws_fm)

    # ---- SHEET 8: AI Prompt ----
    # Included in all modes (text only, negligible weight)
    if 'AI Prompt' not in sheets_to_generate:
        log_fn("    Excel: AI Prompt skipped (not in export mode).")
    else:
        log_fn("    Excel: writing AI Prompt sheet...")
        ws_ai = wb.create_sheet('AI Prompt')
        _apply_clean_layout(ws_ai)
        ws_ai.sheet_properties.tabColor = '00FF9F'
        row = _xl_write_header(ws_ai, 'AI ANALYSIS PROMPT',
                                'Copy this text and paste it into Claude along with this report file.')
        # M7.5: Navigation bar
        _xl_add_sheet_nav(ws_ai, row - 1, nav_targets=nav_targets)
        row += 1

        if ai_prompt:
            ws_ai.cell(row=row, column=1, value=ai_prompt).font = data_font
            ws_ai.cell(row=row, column=1).alignment = Alignment(wrap_text=True, vertical='top')
            ws_ai.column_dimensions['A'].width = 120
            ws_ai.row_dimensions[row].height = 600
        else:
            ws_ai.cell(row=row, column=1, value='No AI prompt available.').font = dim_font
        _apply_dark_background(ws_ai)

    # ---- SHEET: Dashboard (Phase 2) ----
    # Mode ai_optimized: EXCLUDED (19-col metrics, strict subset of AI Context 38 columns)
    if 'Dashboard' not in sheets_to_generate:
        log_fn("    Excel: Dashboard skipped (not in export mode).")
    else:
        # Flat data table with all numeric metrics for filtering
        log_fn("    Excel: writing Dashboard sheet...")
        ws_dash = wb.create_sheet('Dashboard')
        _apply_clean_layout(ws_dash)
        ws_dash.sheet_properties.tabColor = 'FFAA00'
        row = _xl_write_header(ws_dash, 'DASHBOARD \u2014 ALL METRICS',
                                'Use filters to slice by Category, Type, or value ranges')
        # M7.5: Navigation bar
        _xl_add_sheet_nav(ws_dash, row - 1, current_sheet='Dashboard', nav_targets=nav_targets)

        dash_headers = [
            'Track', 'Type', 'Category', 'Family',
            'LUFS', 'Peak (dB)', 'True Peak (dBFS)', 'RMS (dB)',
            'Crest (dB)', 'PLR (dB)', 'PSR (dB)', 'LRA (LU)',
            'Width', 'Correlation', 'Centroid (Hz)', 'Rolloff (Hz)',
            'Flatness', 'Dom. Band', 'Duration (s)',
        ]
        for col, h in enumerate(dash_headers, 1):
            c = ws_dash.cell(row=row, column=col, value=h)
            c.font = header_font
            c.fill = header_fill
            c.border = thin_border
            c.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
            if h in METRIC_GLOSSARY:
                _xl_add_comment(c, METRIC_GLOSSARY[h])
        dash_header_row = row
        row += 1

        for a, ti in analyses_with_info:
            L = a['loudness']
            S = a['spectrum']
            st = a['stereo']
            cat = ti.get('category', '(not set)')
            family = CATEGORY_FAMILY.get(cat, 'Unknown')
            vals = [
                a['filename'],
                ti['type'],
                cat,
                family,
                round(L['lufs_integrated'], 2) if np.isfinite(L['lufs_integrated']) else None,
                round(L['peak_db'], 2),
                round(L['true_peak_db'], 2),
                round(L['rms_db'], 2),
                round(L['crest_factor'], 2),
                round(L['plr'], 2),
                round(L['psr'], 2),
                round(L['lra'], 2),
                round(st['width_overall'], 3) if st['is_stereo'] else None,
                round(st['correlation'], 3) if st['is_stereo'] else None,
                round(S['centroid'], 0),
                round(S['rolloff'], 0),
                round(S['flatness'], 4),
                BAND_LABELS.get(S['dominant_band'], S['dominant_band']),
                round(a['duration'], 1),
            ]
            for col, v in enumerate(vals, 1):
                c = ws_dash.cell(row=row, column=col, value=v)
                c.font = data_font
                c.border = thin_border
                c.fill = panel_fill
                if col >= 5:
                    c.alignment = Alignment(horizontal='center')
            # M7.5: Make track name clickable → individual sheet
            if generate_individual and ti['name'] in sheet_names:
                sname = sheet_names[ti['name']]
                dash_track_cell = ws_dash.cell(row=row, column=1)
                dash_track_cell.hyperlink = f"#{sname}!A1"
                dash_track_cell.font = _link_font
            row += 1

        # Apply auto-filter on the Dashboard table
        dash_data_end = max(row - 1, dash_header_row + 1)
        last_col = get_column_letter(len(dash_headers))
        ws_dash.auto_filter.ref = f'A{dash_header_row}:{last_col}{dash_data_end}'

        # Enriched conditional formatting on Dashboard
        if len(analyses_with_info) > 0:
            ds = dash_header_row + 1
            de = dash_data_end
            # LUFS (E): data bars + color scale
            ws_dash.conditional_formatting.add(
                f'E{ds}:E{de}',
                DataBarRule(start_type='min', end_type='max', color='00D9FF'))
            ws_dash.conditional_formatting.add(
                f'E{ds}:E{de}',
                ColorScaleRule(start_type='min', start_color='FF3333',
                               mid_type='percentile', mid_value=50, mid_color='FFAA00',
                               end_type='max', end_color='00FF9F'))
            # Peak (F): color scale
            ws_dash.conditional_formatting.add(
                f'F{ds}:F{de}',
                ColorScaleRule(start_type='max', start_color='FF3333',
                               end_type='min', end_color='00FF9F'))
            # Crest (I): data bars + color scale
            ws_dash.conditional_formatting.add(
                f'I{ds}:I{de}',
                DataBarRule(start_type='min', end_type='max', color='B967FF'))
            ws_dash.conditional_formatting.add(
                f'I{ds}:I{de}',
                ColorScaleRule(start_type='min', start_color='FF3333',
                               mid_type='percentile', mid_value=50, mid_color='FFAA00',
                               end_type='max', end_color='00D9FF'))
            # PLR (J): data bars
            ws_dash.conditional_formatting.add(
                f'J{ds}:J{de}',
                DataBarRule(start_type='min', end_type='max', color='00FF9F'))
            # PSR (K): data bars
            ws_dash.conditional_formatting.add(
                f'K{ds}:K{de}',
                DataBarRule(start_type='min', end_type='max', color='00D9FF'))
            # LRA (L): data bars
            ws_dash.conditional_formatting.add(
                f'L{ds}:L{de}',
                DataBarRule(start_type='min', end_type='max', color='FFAA00'))
            # Width (M): data bars
            ws_dash.conditional_formatting.add(
                f'M{ds}:M{de}',
                DataBarRule(start_type='min', end_type='max', color='00FF9F'))
            # Correlation (N): color scale (low=red -> high=green)
            ws_dash.conditional_formatting.add(
                f'N{ds}:N{de}',
                ColorScaleRule(start_type='min', start_color='FF3333',
                               mid_type='percentile', mid_value=50, mid_color='FFAA00',
                               end_type='max', end_color='00FF9F'))
            # Centroid (O): color scale
            ws_dash.conditional_formatting.add(
                f'O{ds}:O{de}',
                ColorScaleRule(start_type='min', start_color='B967FF',
                               end_type='max', end_color='FF3D8B'))

            # M7.3: Icon sets on Dashboard
            # Crest Factor (col I): traffic lights — red <6, yellow 6-12, green >12
            ws_dash.conditional_formatting.add(
                f'I{ds}:I{de}',
                IconSetRule(icon_style='3TrafficLights1', type='num',
                            values=[0, 6, 12], showValue=True, reverse=False))
            # True Peak (col G): 3Symbols — safe <-1.5, caution, risk >-0.5
            ws_dash.conditional_formatting.add(
                f'G{ds}:G{de}',
                IconSetRule(icon_style='3Symbols2', type='num',
                            values=[-100, -1.5, -0.5], showValue=True, reverse=True))
            # Correlation (col N): traffic lights — red <0, yellow 0-0.5, green >0.5
            ws_dash.conditional_formatting.add(
                f'N{ds}:N{de}',
                IconSetRule(icon_style='3TrafficLights1', type='num',
                            values=[-1, 0, 0.5], showValue=True, reverse=False))

            # M7.3: Multi-criteria formula alerts on Dashboard rows
            from openpyxl.formatting.rule import FormulaRule
            # Streaming risk: LUFS > -10 AND True Peak > -1 → red row
            ws_dash.conditional_formatting.add(
                f'A{ds}:S{de}',
                FormulaRule(
                    formula=[f'AND($E{ds}>-10,$G{ds}>-1)'],
                    fill=PatternFill(start_color='FF5252', end_color='FF5252', fill_type='solid'),
                    font=Font(color='FFFFFF', bold=True)))
            # Over-compressed + quiet: Crest < 6 AND LUFS < -16 → yellow row
            ws_dash.conditional_formatting.add(
                f'A{ds}:S{de}',
                FormulaRule(
                    formula=[f'AND($I{ds}<6,$E{ds}<-16)'],
                    fill=PatternFill(start_color='FFD93D', end_color='FFD93D', fill_type='solid')))
            # Near-clipping: True Peak > -0.3 → orange row
            ws_dash.conditional_formatting.add(
                f'A{ds}:S{de}',
                FormulaRule(
                    formula=[f'$G{ds}>-0.3'],
                    fill=PatternFill(start_color='FF8B3D', end_color='FF8B3D', fill_type='solid')))

        # Column widths for Dashboard
        ws_dash.column_dimensions['A'].width = 40
        for col_idx in range(2, len(dash_headers) + 1):
            ws_dash.column_dimensions[get_column_letter(col_idx)].width = 14
        ws_dash.row_dimensions[dash_header_row].height = 30
        _apply_dark_background(ws_dash)

        # Move Dashboard to be the 2nd sheet (after Index)
        wb.move_sheet(ws_dash, offset=-(len(wb.sheetnames) - 2))

    # ---- AI Context Sheet (v1.8) — always included ----
    build_ai_context_sheet(wb, analyses_with_info, style_name, log_fn=log_fn,
                           nav_targets=nav_targets)
    ws_ai_ctx = wb['AI Context']
    # Position AI Context: 3rd if Dashboard present, 2nd otherwise
    if 'Dashboard' in sheets_to_generate:
        wb.move_sheet(ws_ai_ctx, offset=-(len(wb.sheetnames) - 3))
    else:
        wb.move_sheet(ws_ai_ctx, offset=-(len(wb.sheetnames) - 2))

    # ---- P3.1: Frequency Conflict Detector ----
    # Mode ai_optimized: INCLUDED (structured conflict data not in AI Context)
    if 'Freq Conflicts' in sheets_to_generate:
        generate_freq_conflicts_sheet(wb, analyses_with_info, log_fn=log_fn,
                                       nav_targets=nav_targets)
    else:
        log_fn("    Excel: Freq Conflicts skipped (not in export mode).")

    # ---- P3.2: Track Comparison Tool ----
    # Mode ai_optimized: EXCLUDED (interactive Excel, raw data in AI Context)
    if 'Track Comparison' in sheets_to_generate:
        generate_track_comparison_sheet(wb, analyses_with_info, log_fn=log_fn)
    else:
        log_fn("    Excel: Track Comparison skipped (not in export mode).")

    # ---- P3.3: Mix Health Score ----
    # Mode ai_optimized: INCLUDED (detailed breakdown beyond scores in AI Context)
    if 'Mix Health Score' in sheets_to_generate:
        generate_health_score_sheet(wb, analyses_with_info, log_fn=log_fn,
                                     nav_targets=nav_targets)
    else:
        log_fn("    Excel: Mix Health Score skipped (not in export mode).")

    # ---- P3.4: Version Tracking ----
    # Mode ai_optimized: EXCLUDED (historical, not current analysis, heavy visual)
    if 'Version History' in sheets_to_generate:
        import re as _re
        _output_dir = os.path.dirname(output_path) if output_path else None
        _song_name = None
        _base = os.path.basename(output_path) if output_path else ''
        _m = _re.match(r'^(.+?)_MixAnalyzer_', _base)
        if _m:
            _song_name = _m.group(1)
        generate_version_tracking_sheet(wb, analyses_with_info,
                                         output_folder=_output_dir,
                                         song_name=_song_name,
                                         log_fn=log_fn)
    else:
        log_fn("    Excel: Version Tracking skipped (not in export mode).")

    # ---- P2.5: Polish cyberpunk theme on Index and special sheets ----
    # Apply background fill to empty rows in Index
    for r in range(1, ws_index.max_row + 1):
        for col in range(1, 6):
            c = ws_index.cell(row=r, column=col)
            if c.fill == PatternFill():
                c.fill = bg_fill
    _apply_dark_background(ws_index)

    # Save workbook
    log_fn("    Excel: saving workbook...")
    wb.save(output_path)

    # Cleanup temp image files
    for tmp_path in tmp_files:
        try:
            os.unlink(tmp_path)
        except Exception:
            pass

    # Log final report stats
    try:
        file_size = os.path.getsize(output_path)
        sheet_count = len([s for s in wb.sheetnames if not s.startswith('_')])
        size_str = f"{file_size / 1024:.0f} KB" if file_size < 1024 * 1024 else f"{file_size / (1024*1024):.1f} MB"
        log_fn(f"    Excel: saved to {os.path.basename(output_path)} ({size_str}, {sheet_count} visible sheets)")
    except Exception:
        log_fn(f"    Excel: saved to {os.path.basename(output_path)}")


def _format_row(idx, type_tag, analysis, track_info):
    """Helper to build a row for the summary table."""
    L = analysis['loudness']
    S = analysis['spectrum']
    st = analysis['stereo']
    lufs = f"{L['lufs_integrated']:+.1f}" if np.isfinite(L['lufs_integrated']) else '-'
    width = f"{st['width_overall']:.2f}" if st['is_stereo'] else 'mono'
    name = analysis['filename']
    if len(name) > 32:
        name = name[:29] + '...'
    cat = track_info.get('category', '(not set)')
    if len(cat) > 16:
        cat = cat[:13] + '...'
    dom_band = BAND_LABELS.get(S['dominant_band'], S['dominant_band'])
    if len(dom_band) > 16:
        dom_band = dom_band[:13] + '...'
    return [
        str(idx),
        type_tag,
        name,
        cat,
        lufs,
        f"{L['peak_db']:+.1f}",
        f"{L['crest_factor']:.1f}",
        dom_band,
        width,
    ]
# ============================================================================
# TKINTER USER INTERFACE - Multi-tab with list+details pattern
# ============================================================================

# M8.2: Centralized cyberpunk color palette
THEME_COLORS = {
    # Backgrounds
    'bg_primary':       '#0D0D0D',      # Main window (near-black)
    'bg_secondary':     '#1A1A2E',      # Cards / panels
    'bg_tertiary':      '#252542',      # Hover / input highlight
    'bg_input':         '#1E1E2E',      # Input field backgrounds

    # Accent colors
    'accent_primary':   '#00FF9F',      # Turquoise neon (main accent)
    'accent_secondary': '#FF3D8B',      # Pink neon (alerts, emphasis)
    'accent_tertiary':  '#00D4AA',      # Cyan (links, secondary)
    'accent_cyan':      '#00D9FF',      # Bright cyan (log text, data)
    'accent_violet':    '#B967FF',      # Violet (decorative)
    'accent_warning':   '#FFD93D',      # Yellow (warnings)
    'accent_error':     '#FF5252',      # Red (errors / critical)

    # Text
    'text_primary':     '#FFFFFF',      # Primary text (white)
    'text_secondary':   '#B8B8B8',      # Secondary text (light grey)
    'text_muted':       '#666666',      # Muted text (grey)
    'text_disabled':    '#444444',      # Disabled text

    # Borders & lines
    'border_default':   '#333333',      # Standard borders
    'border_focus':     '#00FF9F',      # Focused element border
    'border_subtle':    '#222222',      # Subtle dividers

    # Interactive states
    'hover':            '#2A2A4A',      # Hover background
    'active':           '#353560',      # Active / pressed
    'selected':         '#1A3A3A',      # Selected item background
}

# Tkinter theme colors — derived from THEME_COLORS for UI widgets
UI_THEME = {
    'bg':           THEME_COLORS['bg_primary'],
    'panel':        THEME_COLORS['bg_secondary'],
    'panel_light':  THEME_COLORS['bg_tertiary'],
    'bg_input':     THEME_COLORS['bg_input'],
    'fg':           THEME_COLORS['text_primary'],
    'fg_dim':       THEME_COLORS['text_secondary'],
    'fg_muted':     THEME_COLORS['text_muted'],
    'fg_disabled':  THEME_COLORS['text_disabled'],
    'accent1':      THEME_COLORS['accent_primary'],
    'accent2':      THEME_COLORS['accent_secondary'],
    'accent3':      THEME_COLORS['accent_tertiary'],
    'accent4':      THEME_COLORS['accent_cyan'],
    'warning':      THEME_COLORS['accent_warning'],
    'critical':     THEME_COLORS['accent_error'],
    'border':       THEME_COLORS['border_default'],
    'border_focus': THEME_COLORS['border_focus'],
    'border_subtle':THEME_COLORS['border_subtle'],
    'select':       THEME_COLORS['selected'],
    'hover':        THEME_COLORS['hover'],
    'active':       THEME_COLORS['active'],
}

# M8.7: Consistent spacing scale for UI polish
SPACING = {
    'xs': 4,       # Between tightly related elements (label ↔ input)
    'sm': 8,       # Compact internal padding
    'md': 12,      # Standard internal padding
    'lg': 16,      # Between groups / sections
    'xl': 24,      # Between major sections
    'xxl': 32,     # Header / footer padding
}

# M8.7: Border configuration for subtle visual separation
BORDERS = {
    'subtle': {'color': '#222222', 'width': 1},
    'default': {'color': '#333333', 'width': 1},
    'accent': {'color': THEME_COLORS['accent_primary'], 'width': 2},
}

# M8.6: Unicode icons for styled tabs
TAB_ICONS = {
    'Setup':                '◆',
    'Track Identification': '◉',
    'Full Mix':             '♪',
    'Analysis':             '▶',
}


# M8.4: Modern typography with font fallback chain
# Font candidate lists ordered by preference (futuristic → system fallbacks)

DISPLAY_FONTS = [
    'Orbitron', 'Rajdhani', 'Exo 2', 'Audiowide', 'Michroma',
    'Aldrich', 'Electrolize', 'Share Tech', 'Titillium Web',
    'Segoe UI', 'Helvetica Neue', 'Arial Black', 'Arial',
]

BODY_FONTS = [
    'Rajdhani', 'Exo 2', 'Titillium Web',
    'Segoe UI', 'Helvetica Neue', 'Calibri', 'Arial',
]

MONO_FONTS = [
    'JetBrains Mono', 'Fira Code', 'Source Code Pro', 'Cascadia Code',
    'Consolas', 'Monaco', 'Courier New',
]

# Global typography state (populated by setup_typography() after Tk root exists)
TYPOGRAPHY = None
_RESOLVED_FONTS = {'display': 'Arial', 'body': 'Arial', 'mono': 'Courier New'}


def _get_available_tk_fonts():
    """Return the set of font families available on the current system."""
    try:
        import tkinter.font as tkfont
        return set(tkfont.families())
    except Exception:
        return set()


def find_best_font(candidates, fallback='Arial'):
    """Find the first available font from a prioritized candidates list."""
    available = _get_available_tk_fonts()
    for name in candidates:
        if name in available:
            return name
    return fallback


def setup_typography():
    """Build the TYPOGRAPHY config with the best available fonts.

    Must be called after tk.Tk() is created so font detection works.
    Returns (TYPOGRAPHY dict, resolved_fonts dict).
    """
    global TYPOGRAPHY, _RESOLVED_FONTS

    display = find_best_font(DISPLAY_FONTS, 'Arial')
    body = find_best_font(BODY_FONTS, 'Arial')
    mono = find_best_font(MONO_FONTS, 'Courier New')

    _RESOLVED_FONTS = {'display': display, 'body': body, 'mono': mono}

    TYPOGRAPHY = {
        # Display / Headers
        'h1':              {'family': display, 'size': 20, 'weight': 'bold'},
        'h2':              {'family': display, 'size': 15, 'weight': 'bold'},
        'h3':              {'family': display, 'size': 13, 'weight': 'bold'},
        'header':          {'family': display, 'size': 14, 'weight': 'bold'},
        'heading_dialog':  {'family': display, 'size': 18, 'weight': 'bold'},
        'heading_help':    {'family': display, 'size': 16, 'weight': 'bold'},
        'subheading':      {'family': display, 'size': 12, 'weight': 'bold'},

        # Body text
        'body':            {'family': body, 'size': 11, 'weight': 'normal'},
        'body_bold':       {'family': body, 'size': 11, 'weight': 'bold'},
        'body_small':      {'family': body, 'size': 10, 'weight': 'normal'},
        'body_small_bold': {'family': body, 'size': 10, 'weight': 'bold'},

        # Captions
        'caption':         {'family': body, 'size': 9,  'weight': 'normal'},
        'caption_tiny':    {'family': body, 'size': 8,  'weight': 'normal'},

        # Buttons
        'button':          {'family': body, 'size': 11, 'weight': 'normal'},
        'button_bold':     {'family': body, 'size': 11, 'weight': 'bold'},
        'button_large':    {'family': body, 'size': 13, 'weight': 'bold'},
        'button_accent':   {'family': body, 'size': 12, 'weight': 'bold'},
        'button_small':    {'family': body, 'size': 10, 'weight': 'normal'},

        # Tabs
        'tab':             {'family': body, 'size': 11, 'weight': 'normal'},
        'tab_selected':    {'family': body, 'size': 14, 'weight': 'bold'},

        # Monospace (log, code, data values)
        'mono':            {'family': mono, 'size': 9,  'weight': 'normal'},
        'mono_large':      {'family': mono, 'size': 14, 'weight': 'bold'},
    }

    # Update BUTTON_PRESETS with resolved fonts (if already defined)
    try:
        _update_button_preset_fonts()
    except NameError:
        pass  # BUTTON_PRESETS not yet defined at import time

    return TYPOGRAPHY, _RESOLVED_FONTS


def get_font(style_name):
    """Return a Tkinter font tuple (family, size, weight) for the given style.

    Falls back to a safe default if typography hasn't been set up yet.
    """
    if TYPOGRAPHY is None:
        return ('Arial', 11, 'normal')
    style = TYPOGRAPHY.get(style_name, TYPOGRAPHY['body'])
    return (style['family'], style['size'], style['weight'])


def _update_button_preset_fonts():
    """Update BUTTON_PRESETS with the resolved typography fonts."""
    BUTTON_PRESETS['primary']['font'] = get_font('button_bold')
    BUTTON_PRESETS['primary_large']['font'] = get_font('button_large')
    BUTTON_PRESETS['secondary']['font'] = get_font('button')
    BUTTON_PRESETS['ghost']['font'] = get_font('button')
    BUTTON_PRESETS['danger']['font'] = get_font('button_bold')
    BUTTON_PRESETS['small']['font'] = get_font('button_small')


def log_typography_info():
    """Print resolved font information for debugging."""
    print(f"[Mix Analyzer] Typography:")
    print(f"  Display font: {_RESOLVED_FONTS['display']}")
    print(f"  Body font:    {_RESOLVED_FONTS['body']}")
    print(f"  Mono font:    {_RESOLVED_FONTS['mono']}")


# M8.1: Neon logo configuration
LOGO_CONFIG = {
    'text': 'MIX ANALYZER',
    'subtitle': 'v2.0 — Visual Mix Diagnostic',
    'font_size': 42,
    'subtitle_font_size': 14,
    'glow_color': (0, 255, 159),      # Turquoise neon (#00FF9F)
    'text_color': (220, 255, 245),     # Slightly tinted white
    'subtitle_color': (184, 184, 184), # Secondary (#B8B8B8) — matches text_secondary
    'bg_color': (13, 13, 13),          # Dark bg (#0D0D0D) — matches THEME_COLORS
    'glow_passes': [
        (20, 0.25),   # (radius, opacity) - wide diffuse outer glow
        (12, 0.4),
        (6, 0.6),
        (3, 0.85),    # tight bright inner glow
    ],
    'padding': 30,
    'line_spacing': 8,
}


def _get_neon_font(size):
    """Try to load a suitable font with fallback chain."""
    from PIL import ImageFont
    candidates = [
        'Orbitron-Bold.ttf', 'Orbitron-Regular.ttf',
        'Rajdhani-Bold.ttf', 'Exo2-Bold.ttf',
        'Audiowide-Regular.ttf',
        'arialbd.ttf', 'Arial Bold.ttf', 'Arial_Bold.ttf',
        'LiberationSans-Bold.ttf', 'DejaVuSans-Bold.ttf',
        'FreeSansBold.ttf',
    ]
    for name in candidates:
        try:
            return ImageFont.truetype(name, size)
        except (OSError, IOError):
            continue
    # Fallback: try system default
    try:
        return ImageFont.truetype('/usr/share/fonts/truetype/dejavu/DejaVuSans-Bold.ttf', size)
    except (OSError, IOError):
        pass
    return ImageFont.load_default()


def create_neon_logo():
    """Generate a PIL image with multi-layer neon glow effect.
    Returns a PIL.Image in RGBA mode."""
    from PIL import Image, ImageDraw, ImageFilter

    cfg = LOGO_CONFIG
    font = _get_neon_font(cfg['font_size'])
    sub_font = _get_neon_font(cfg['subtitle_font_size'])

    # Measure text dimensions
    dummy = Image.new('RGBA', (1, 1))
    draw = ImageDraw.Draw(dummy)
    title_bbox = draw.textbbox((0, 0), cfg['text'], font=font)
    title_w = title_bbox[2] - title_bbox[0]
    title_h = title_bbox[3] - title_bbox[1]
    sub_bbox = draw.textbbox((0, 0), cfg['subtitle'], font=sub_font)
    sub_w = sub_bbox[2] - sub_bbox[0]
    sub_h = sub_bbox[3] - sub_bbox[1]

    pad = cfg['padding']
    content_w = max(title_w, sub_w)
    content_h = title_h + cfg['line_spacing'] + sub_h
    img_w = content_w + pad * 2
    img_h = content_h + pad * 2

    # Title position (centered)
    tx = (img_w - title_w) // 2
    ty = pad
    # Subtitle position (centered, below title)
    sx = (img_w - sub_w) // 2
    sy = pad + title_h + cfg['line_spacing']

    # Start with dark background
    final = Image.new('RGBA', (img_w, img_h), cfg['bg_color'] + (255,))

    # Multi-layer glow for title text
    for radius, opacity in cfg['glow_passes']:
        layer = Image.new('RGBA', (img_w, img_h), (0, 0, 0, 0))
        d = ImageDraw.Draw(layer)
        alpha = int(255 * opacity)
        d.text((tx, ty), cfg['text'], font=font, fill=cfg['glow_color'] + (alpha,))
        layer = layer.filter(ImageFilter.GaussianBlur(radius=radius))
        final = Image.alpha_composite(final, layer)

    # Sharp title text on top
    text_layer = Image.new('RGBA', (img_w, img_h), (0, 0, 0, 0))
    d = ImageDraw.Draw(text_layer)
    d.text((tx, ty), cfg['text'], font=font, fill=cfg['text_color'] + (255,))
    # Subtitle (no glow, just dim text)
    d.text((sx, sy), cfg['subtitle'], font=sub_font, fill=cfg['subtitle_color'] + (255,))
    final = Image.alpha_composite(final, text_layer)

    return final


def _create_logo_widget(parent):
    """Create the neon logo widget with fallback to styled text."""
    try:
        from PIL import ImageTk
        pil_img = create_neon_logo()
        photo = ImageTk.PhotoImage(pil_img)
        label = tk.Label(parent, image=photo, bg=UI_THEME['bg'], bd=0)
        label._logo_photo = photo  # prevent garbage collection
        return label
    except Exception:
        # Fallback: styled text labels
        frame = tk.Frame(parent, bg=UI_THEME['bg'])
        tk.Label(frame, text='MIX ANALYZER',
                 font=get_font('h1'),
                 fg=UI_THEME['accent4'], bg=UI_THEME['bg']).pack()
        tk.Label(frame, text='v2.0 — Visual Mix Diagnostic',
                 font=get_font('body_small'),
                 fg=UI_THEME['fg_dim'], bg=UI_THEME['bg']).pack()
        return frame


# M8.5: Neon progress bar with glow effect
class NeonProgressBar(tk.Canvas):
    """Custom Canvas-based progress bar with neon glow effect.

    Replaces ttk.Progressbar with a cyberpunk-themed bar featuring:
    - Rounded rectangle trough and bar
    - Multi-layer glow effect around the filled portion
    - Highlight shine on top of the bar
    - Percentage text overlay
    - Subtle pulse animation when active
    """

    _GLOW_LAYERS = [
        ('#003322', 3),  # outermost — wide, dark
        ('#005533', 2),  # mid
        ('#007744', 1),  # innermost — tight, brighter
    ]

    _PULSE_COLORS = [
        '#00FF9F',  # base
        '#22FFAE',
        '#44FFBD',
        '#22FFAE',
    ]

    def __init__(self, parent, width=400, height=24, maximum=100,
                 bar_color=None, trough_color=None, glow=True,
                 show_percent=True):
        self._bar_color = bar_color or THEME_COLORS['accent_primary']
        self._trough_color = trough_color or THEME_COLORS['bg_tertiary']
        bg = THEME_COLORS['bg_secondary']

        super().__init__(parent, width=width, height=height,
                         bg=bg, highlightthickness=0, bd=0)

        self._width = width
        self._height = height
        self._maximum = maximum
        self._value = 0
        self._glow = glow
        self._show_percent = show_percent

        # Pulse animation state
        self._animating = False
        self._pulse_idx = 0
        self._after_id = None

        self._draw()

        # Redraw on resize
        self.bind('<Configure>', self._on_resize)

    def _on_resize(self, event):
        """Handle resize events for fill='x' packing."""
        if event.width != self._width:
            self._width = event.width
            self._draw()

    # --- Drawing ----------------------------------------------------------

    def _draw(self):
        """Redraw the entire progress bar."""
        self.delete('all')
        w, h = self._width, self._height
        r = h // 2  # corner radius = half height for pill shape

        # Trough (background)
        self._rounded_rect(1, 1, w - 1, h - 1, r, fill=self._trough_color,
                           outline=THEME_COLORS['border_subtle'])

        frac = self._value / self._maximum if self._maximum else 0
        frac = max(0.0, min(1.0, frac))
        bar_w = int((w - 2) * frac)

        if bar_w > 4:
            # Glow layers (drawn behind the bar)
            if self._glow:
                for color, pad in self._GLOW_LAYERS:
                    self.create_rectangle(
                        1 - pad, 1 - pad,
                        bar_w + 1 + pad, h - 1 + pad,
                        fill='', outline=color, width=1,
                    )

            # Main bar
            self._rounded_rect(1, 1, bar_w + 1, h - 1, r,
                               fill=self._bar_color, outline='')

            # Highlight shine (top quarter of the bar)
            shine_h = max(3, h // 4)
            shine_color = self._lighten(self._bar_color, 0.45)
            self._rounded_rect(3, 3, bar_w - 1, 3 + shine_h, max(2, r // 2),
                               fill=shine_color, outline='')

        # Percentage text overlay
        if self._show_percent:
            pct = int(frac * 100)
            text_color = THEME_COLORS['bg_primary'] if frac > 0.15 else THEME_COLORS['text_primary']
            self.create_text(
                w // 2, h // 2,
                text=f'{pct}%',
                fill=text_color,
                font=get_font('body_small_bold'),
                anchor='center',
            )

    def _rounded_rect(self, x1, y1, x2, y2, radius, **kwargs):
        """Draw a rounded rectangle using a smooth polygon."""
        r = min(radius, (x2 - x1) // 2, (y2 - y1) // 2)
        if r < 1:
            return self.create_rectangle(x1, y1, x2, y2, **kwargs)
        points = [
            x1 + r, y1,
            x2 - r, y1,
            x2, y1,
            x2, y1 + r,
            x2, y2 - r,
            x2, y2,
            x2 - r, y2,
            x1 + r, y2,
            x1, y2,
            x1, y2 - r,
            x1, y1 + r,
            x1, y1,
        ]
        return self.create_polygon(points, smooth=True, **kwargs)

    @staticmethod
    def _lighten(hex_color, amount):
        """Lighten a hex color by blending toward white."""
        r = int(hex_color[1:3], 16)
        g = int(hex_color[3:5], 16)
        b = int(hex_color[5:7], 16)
        r = int(r + (255 - r) * amount)
        g = int(g + (255 - g) * amount)
        b = int(b + (255 - b) * amount)
        return f'#{r:02x}{g:02x}{b:02x}'

    # --- Public API (compatible with ttk.Progressbar) ---------------------

    def __setitem__(self, key, value):
        """Support progress_bar['value'] = pct syntax."""
        if key == 'value':
            self.set_value(value)
        elif key == 'maximum':
            self._maximum = value
            self._draw()
        else:
            super().__setitem__(key, value)

    def __getitem__(self, key):
        """Support progress_bar['value'] syntax."""
        if key == 'value':
            return self._value
        if key == 'maximum':
            return self._maximum
        return super().__getitem__(key)

    def set_value(self, value):
        """Set current progress value (0 to maximum)."""
        value = max(0, min(self._maximum, float(value)))
        if value != self._value:
            self._value = value
            self._draw()
            # Auto-manage pulse animation
            if 0 < value < self._maximum and not self._animating:
                self.start_pulse()
            elif value >= self._maximum or value <= 0:
                self.stop_pulse()

    # --- Pulse animation --------------------------------------------------

    def start_pulse(self):
        """Start subtle pulse animation."""
        if self._animating:
            return
        self._animating = True
        self._pulse_tick()

    def stop_pulse(self):
        """Stop pulse animation and restore base color."""
        self._animating = False
        if self._after_id is not None:
            self.after_cancel(self._after_id)
            self._after_id = None
        self._bar_color = THEME_COLORS.get('accent_primary', '#00FF9F')
        # Only redraw if the widget still exists
        try:
            self._draw()
        except tk.TclError:
            pass

    def _pulse_tick(self):
        """Single animation frame."""
        if not self._animating:
            return
        self._pulse_idx = (self._pulse_idx + 1) % len(self._PULSE_COLORS)
        self._bar_color = self._PULSE_COLORS[self._pulse_idx]
        try:
            self._draw()
            self._after_id = self.after(180, self._pulse_tick)
        except tk.TclError:
            self._animating = False


# Help texts for the info buttons
HELP_TEXTS = {
    'category': """CATEGORY - What kind of sound this track contains.

Categories are organized into families:

DRUMS: Kick, Snare/Clap, Hi-Hat/Cymbal, Tom, Percussion, Drum Loop/Bus
  Use Drum Loop/Bus for pre-mixed drum loops or drum BUS sums.

BASS: Sub Bass, Bass (standard), Acid Bass, 808 / Pitched Bass
  Sub Bass = pure sub frequencies (usually sine-like, 30-80 Hz).
  Bass (standard) = general bass synth or DI bass (40-250 Hz).
  Acid Bass = 303-style resonant saw bass with filter modulation.
  808 = pitched sub bass, hip-hop style, with variable pitch.

SYNTH: Lead, Pluck/Stab, Pad/Drone, Arpeggio/Sequence, Texture/Atmosphere
  Lead = melodic, foreground synth.
  Pluck/Stab = short, percussive synth.
  Pad = sustained, often slow-evolving.
  Arp = repeated melodic pattern.
  Texture = atmospheric, background sound.

GUITAR: Clean, Distorted, Acoustic

VOCAL: Lead, Backing/Harmony, FX/Chop

FX & OTHER: Risers, Noise, Samples, Other

Choose the closest match. If unsure, pick Other.""",

    'type': """TRACK TYPE - How this file relates to your project.

INDIVIDUAL: A single track containing one instrument or one sound.
  Most of your exported tracks should be Individual.

BUS: A summed bus track (e.g. Drums Bus = kick + snare + hats + perc combined).
  Ableton exports both individual tracks AND the bus if you have a group.
  Mark the bus itself as BUS to avoid double-counting in the masking analysis.

FULL MIX: The complete bounce of the entire song (all tracks summed + master bus).
  Only ONE file should be marked as Full Mix.
  This is your reference for the overall sound of the project.
  The global report will give this file its own dedicated section.""",

    'parent_bus': """PARENT BUS - If this individual track is part of a group bus.

When you use groups in Ableton (e.g. a Drums group containing kick, snare, hats),
Ableton exports each individual track AND the summed bus.

Set this field to indicate which BUS this track belongs to, so the global report
can show BUS vs. components consistency.

Leave as None if this track doesn't belong to any bus.""",

    'style': """STYLE - The musical style context for the analysis.

This setting adjusts the interpretation thresholds used in the analysis:
- Target loudness expectations
- Typical crest factor ranges
- Acceptable density levels

Choose the style closest to your project. If unsure, pick Generic.
The style is grouped internally into 7 analytical families (acoustic, rock,
electronic soft, electronic dance, electronic aggressive, urban, pop) that
determine the analytical profile.

Note: The style does NOT change the raw measurements, only their interpretation.""",

    'mix_state': """MIX COMPLETION STATE - Where this bounce is in the production cycle.

Rough mix: First draft, levels only, no real processing
Mix in progress: Active mixing, still iterating
Pre-final mix: Mix nearly done, final adjustments
Final mix: Ready for mastering, no more mix changes
Pre-master: Mastering in progress
Final master: Completed, ready for release

This context helps interpret measurements correctly. For example:
- A limiter on the master bus of a 'Rough mix' is unusual
- A very hot LUFS on a 'Final master' is expected
- Low crest factor on a 'Mix in progress' may be intentional or not""",

    'master_plugins': """MASTER BUS PLUGINS - Effects active on the master bus during export.

Check all that apply. This is important because:
- Compression and limiting reduce dynamic range measurements
- EQ changes the spectral balance
- Stereo imagers affect stereo width and correlation

When the AI analyzes your reports, it will take these into account to
distinguish 'the mix has issues' from 'the master processing is affecting
the measurements'.

If you export with an empty master bus (recommended for mix diagnosis),
leave everything unchecked.""",

    'loudness_target': """LOUDNESS TARGET - What you're aiming for on the final master.

-14 LUFS: Spotify, Apple Music, YouTube (streaming standard)
-12 LUFS: Conservative hot master
-10 LUFS: Common for rock, electronic (pre-streaming levels)
-8 LUFS: Very hot, competitive loudness (not recommended for streaming)
Custom: Your own target or undefined

This helps the AI understand if your current LUFS is on target or not.""",
}


class HelpButton(tk.Button):
    """A small info button that opens a help dialog."""
    def __init__(self, master, help_key, **kwargs):
        super().__init__(master, text='\u24d8', width=2,  # circled i
                          bg=UI_THEME['panel'], fg=UI_THEME['accent1'],
                          activebackground=UI_THEME['panel_light'],
                          activeforeground=UI_THEME['accent1'],
                          relief='flat', bd=0, cursor='hand2',
                          font=get_font('body_bold'),
                          command=lambda: self._show_help(help_key),
                          **kwargs)
        self.help_key = help_key

    def _show_help(self, key):
        text = HELP_TEXTS.get(key, 'No help available for this item.')
        dialog = tk.Toplevel(self.winfo_toplevel())
        dialog.title('Help')
        dialog.configure(bg=UI_THEME['bg'])
        dialog.geometry('560x480')
        dialog.transient(self.winfo_toplevel())

        frame = tk.Frame(dialog, bg=UI_THEME['bg'],
                         padx=SPACING['xl'], pady=SPACING['xl'])
        frame.pack(fill='both', expand=True)

        title_lbl = tk.Label(frame, text='ⓘ  Help',
                              bg=UI_THEME['bg'], fg=UI_THEME['accent1'],
                              font=get_font('heading_help'))
        title_lbl.pack(anchor='w', pady=(0, SPACING['md']))

        text_widget = tk.Text(frame, wrap='word', bg=UI_THEME['panel'],
                               fg=UI_THEME['fg'], font=get_font('body'),
                               bd=0, padx=SPACING['md'], pady=SPACING['md'],
                               relief='flat',
                               highlightbackground=BORDERS['subtle']['color'],
                               highlightthickness=BORDERS['subtle']['width'])
        text_widget.pack(fill='both', expand=True)
        text_widget.insert('1.0', text)
        text_widget.config(state='disabled')

        close_btn = create_neon_button(frame, 'Close', dialog.destroy, preset='primary')
        close_btn.pack(pady=(SPACING['md'], 0))


def setup_ttk_styles():
    """Configure ttk styles for the cyberpunk dark theme (M8.2 + M8.4 typography)."""
    style = ttk.Style()
    style.theme_use('clam')

    # === NOTEBOOK (Tabs) — M8.6: styled tabs with hover & accent ===
    style.configure('TNotebook',
                    background=UI_THEME['bg'],
                    borderwidth=0,
                    tabmargins=[4, 4, 0, 0])

    # Remove default notebook border — borderless modern look
    style.layout('TNotebook', [
        ('Notebook.client', {'sticky': 'nswe'})
    ])

    # Modern tab layout (no focus rectangle)
    style.layout('TNotebook.Tab', [
        ('Notebook.tab', {
            'sticky': 'nswe',
            'children': [
                ('Notebook.padding', {
                    'side': 'top',
                    'sticky': 'nswe',
                    'children': [
                        ('Notebook.label', {'side': 'left', 'sticky': ''}),
                    ]
                })
            ]
        })
    ])

    style.configure('TNotebook.Tab',
                    background=UI_THEME['panel'],
                    foreground=UI_THEME['fg_dim'],
                    padding=[20, 10],
                    font=get_font('tab'),
                    borderwidth=0)
    style.map('TNotebook.Tab',
              background=[
                  ('selected', UI_THEME['bg']),
                  ('active', UI_THEME['hover']),
                  ('!selected', UI_THEME['panel']),
              ],
              foreground=[
                  ('selected', UI_THEME['accent1']),
                  ('active', UI_THEME['fg']),
                  ('!selected', UI_THEME['fg_dim']),
              ],
              font=[('selected', get_font('tab_selected')),
                    ('!selected', get_font('tab'))],
              padding=[('selected', [22, 12]),
                       ('!selected', [20, 10])],
              expand=[('selected', [0, 0, 0, 2])])

    # === FRAMES ===
    style.configure('TFrame', background=UI_THEME['bg'])
    style.configure('Panel.TFrame', background=UI_THEME['panel'])
    style.configure('Card.TFrame', background=UI_THEME['panel'])

    # === LABELS ===
    style.configure('TLabel',
                    background=UI_THEME['bg'],
                    foreground=UI_THEME['fg'],
                    font=get_font('body'))
    style.configure('Title.TLabel',
                    background=UI_THEME['bg'],
                    foreground=UI_THEME['accent1'],
                    font=get_font('h1'))
    style.configure('SubTitle.TLabel',
                    background=UI_THEME['bg'],
                    foreground=UI_THEME['accent2'],
                    font=get_font('h2'))
    style.configure('Section.TLabel',
                    background=UI_THEME['bg'],
                    foreground=UI_THEME['fg'],
                    font=get_font('h3'))
    style.configure('Dim.TLabel',
                    background=UI_THEME['bg'],
                    foreground=UI_THEME['fg_dim'],
                    font=get_font('body_small'))
    style.configure('Panel.TLabel',
                    background=UI_THEME['panel'],
                    foreground=UI_THEME['fg'],
                    font=get_font('body'))
    style.configure('PanelDim.TLabel',
                    background=UI_THEME['panel'],
                    foreground=UI_THEME['fg_dim'],
                    font=get_font('body_small'))
    style.configure('Accent.TLabel',
                    background=UI_THEME['bg'],
                    foreground=UI_THEME['accent1'])
    style.configure('Header.TLabel',
                    background=UI_THEME['bg'],
                    foreground=UI_THEME['fg'],
                    font=get_font('header'))

    # M8.4: Hierarchical heading styles
    style.configure('H1.TLabel',
                    background=UI_THEME['bg'],
                    foreground=THEME_COLORS['text_primary'],
                    font=get_font('h1'))
    style.configure('H2.TLabel',
                    background=UI_THEME['bg'],
                    foreground=THEME_COLORS['text_primary'],
                    font=get_font('h2'))
    style.configure('H3.TLabel',
                    background=UI_THEME['bg'],
                    foreground=THEME_COLORS['text_primary'],
                    font=get_font('h3'))
    style.configure('Body.TLabel',
                    background=UI_THEME['bg'],
                    foreground=THEME_COLORS['text_primary'],
                    font=get_font('body'))
    style.configure('Caption.TLabel',
                    background=UI_THEME['bg'],
                    foreground=THEME_COLORS['text_secondary'],
                    font=get_font('caption'))
    style.configure('Mono.TLabel',
                    background=UI_THEME['bg'],
                    foreground=THEME_COLORS['accent_primary'],
                    font=get_font('mono'))

    # M8.4: Value display styles with accent colors
    style.configure('Value.TLabel',
                    background=UI_THEME['bg'],
                    foreground=THEME_COLORS['accent_primary'],
                    font=get_font('mono_large'))
    style.configure('ValueWarning.TLabel',
                    background=UI_THEME['bg'],
                    foreground=THEME_COLORS['accent_warning'],
                    font=get_font('mono_large'))
    style.configure('ValueDanger.TLabel',
                    background=UI_THEME['bg'],
                    foreground=THEME_COLORS['accent_error'],
                    font=get_font('mono_large'))

    # === ENTRIES ===
    style.configure('TEntry',
                    fieldbackground=UI_THEME['bg_input'],
                    foreground=UI_THEME['fg'],
                    bordercolor=UI_THEME['border'],
                    lightcolor=UI_THEME['border'],
                    darkcolor=UI_THEME['border'],
                    insertcolor=UI_THEME['accent1'])
    style.map('TEntry',
              fieldbackground=[
                  ('focus', UI_THEME['panel_light']),
                  ('disabled', UI_THEME['panel'])
              ])

    # === COMBOBOX ===
    style.configure('TCombobox',
                    fieldbackground=UI_THEME['bg_input'],
                    background=UI_THEME['panel'],
                    foreground=UI_THEME['fg'],
                    bordercolor=UI_THEME['border'],
                    arrowcolor=UI_THEME['accent1'],
                    selectbackground=UI_THEME['select'],
                    selectforeground=UI_THEME['fg'])
    style.map('TCombobox',
              fieldbackground=[
                  ('readonly', UI_THEME['bg_input']),
                  ('disabled', UI_THEME['panel'])
              ],
              foreground=[
                  ('readonly', UI_THEME['fg']),
                  ('disabled', UI_THEME['fg_disabled'])
              ])

    # === BUTTONS ===
    style.configure('TButton',
                    background=UI_THEME['panel_light'],
                    foreground=UI_THEME['fg'],
                    bordercolor=UI_THEME['accent1'],
                    focuscolor=UI_THEME['accent1'],
                    font=get_font('button'),
                    padding=8)
    style.map('TButton',
              background=[
                  ('active', UI_THEME['accent1']),
                  ('pressed', UI_THEME['accent2']),
                  ('disabled', UI_THEME['panel_light'])
              ],
              foreground=[
                  ('active', UI_THEME['bg']),
                  ('disabled', UI_THEME['fg_disabled'])
              ])

    style.configure('Accent.TButton',
                    background=UI_THEME['accent1'],
                    foreground=UI_THEME['bg'],
                    font=get_font('button_accent'),
                    padding=12)
    style.map('Accent.TButton',
              background=[('active', UI_THEME['accent2'])],
              foreground=[('active', UI_THEME['fg'])])

    # M8.3: Additional button style variants
    style.configure('Secondary.TButton',
                    background=UI_THEME['panel_light'],
                    foreground=UI_THEME['fg'],
                    bordercolor=UI_THEME['border'],
                    font=get_font('button'),
                    padding=8)
    style.map('Secondary.TButton',
              background=[
                  ('active', UI_THEME['hover']),
                  ('pressed', UI_THEME['active']),
                  ('disabled', UI_THEME['panel_light'])
              ],
              foreground=[
                  ('disabled', UI_THEME['fg_disabled'])
              ])

    style.configure('Small.TButton',
                    background=UI_THEME['panel_light'],
                    foreground=UI_THEME['fg'],
                    bordercolor=UI_THEME['accent1'],
                    focuscolor=UI_THEME['accent1'],
                    font=get_font('button_small'),
                    padding=[8, 4])
    style.map('Small.TButton',
              background=[
                  ('active', UI_THEME['accent1']),
                  ('pressed', UI_THEME['accent2'])
              ],
              foreground=[
                  ('active', UI_THEME['bg'])
              ])

    style.configure('Danger.TButton',
                    background=UI_THEME['critical'],
                    foreground=UI_THEME['fg'],
                    font=get_font('button_bold'),
                    padding=8)
    style.map('Danger.TButton',
              background=[('active', '#DD4444')],
              foreground=[('active', UI_THEME['fg'])])

    # === CHECKBUTTON ===
    style.configure('TCheckbutton',
                    background=UI_THEME['bg'],
                    foreground=UI_THEME['fg'],
                    font=get_font('body'),
                    focuscolor=UI_THEME['accent1'])
    style.map('TCheckbutton',
              background=[('active', UI_THEME['bg'])])
    style.configure('Panel.TCheckbutton',
                    background=UI_THEME['panel'],
                    foreground=UI_THEME['fg'],
                    font=get_font('body'))
    style.map('Panel.TCheckbutton',
              background=[('active', UI_THEME['panel'])])

    # === RADIOBUTTON ===
    style.configure('TRadiobutton',
                    background=UI_THEME['bg'],
                    foreground=UI_THEME['fg'],
                    font=get_font('body'),
                    focuscolor=UI_THEME['accent1'])
    style.map('TRadiobutton',
              background=[('active', UI_THEME['bg'])])

    # === PROGRESSBAR (ttk fallback — main bar uses NeonProgressBar Canvas) ===
    style.configure('TProgressbar',
                    background=UI_THEME['accent1'],
                    troughcolor=UI_THEME['panel'],
                    borderwidth=0,
                    thickness=20)
    style.configure('Horizontal.TProgressbar',
                    background=UI_THEME['accent1'],
                    troughcolor=UI_THEME['panel'],
                    borderwidth=0)

    # === SCROLLBAR ===
    style.configure('Vertical.TScrollbar',
                    background=UI_THEME['panel'],
                    troughcolor=UI_THEME['bg'],
                    bordercolor=UI_THEME['border'],
                    arrowcolor=UI_THEME['accent1'])
    style.map('Vertical.TScrollbar',
              background=[
                  ('active', UI_THEME['hover']),
                  ('pressed', UI_THEME['active'])
              ])
    style.configure('Horizontal.TScrollbar',
                    background=UI_THEME['panel'],
                    troughcolor=UI_THEME['bg'],
                    bordercolor=UI_THEME['border'],
                    arrowcolor=UI_THEME['accent1'])

    # === LABELFRAME ===
    style.configure('TLabelframe',
                    background=UI_THEME['bg'],
                    bordercolor=UI_THEME['border'])
    style.configure('TLabelframe.Label',
                    background=UI_THEME['bg'],
                    foreground=UI_THEME['fg'])

    # === SEPARATOR ===
    style.configure('TSeparator',
                    background=UI_THEME['border'])

    # === TREEVIEW ===
    style.configure('Treeview',
                    background=UI_THEME['panel'],
                    foreground=UI_THEME['fg'],
                    fieldbackground=UI_THEME['panel'],
                    borderwidth=0)
    style.configure('Treeview.Heading',
                    background=UI_THEME['panel_light'],
                    foreground=UI_THEME['fg'])
    style.map('Treeview',
              background=[('selected', UI_THEME['select'])],
              foreground=[('selected', UI_THEME['accent1'])])

    # === SCALE ===
    style.configure('TScale',
                    background=UI_THEME['bg'],
                    troughcolor=UI_THEME['panel'],
                    borderwidth=0)

    return style


def apply_dark_theme_to_tk_widget(widget, widget_type='default'):
    """Apply the cyberpunk dark theme to native tk widgets (non-ttk)."""
    common = {
        'bg': UI_THEME['bg_input'],
        'fg': UI_THEME['fg'],
        'relief': 'flat',
        'highlightbackground': UI_THEME['border'],
        'highlightcolor': UI_THEME['border_focus'],
        'highlightthickness': 1,
    }

    if widget_type == 'text':
        common.update({
            'insertbackground': UI_THEME['accent1'],
            'selectbackground': UI_THEME['select'],
            'selectforeground': UI_THEME['fg'],
        })
    elif widget_type == 'listbox':
        common.update({
            'selectbackground': UI_THEME['select'],
            'selectforeground': UI_THEME['accent1'],
        })
    elif widget_type == 'canvas':
        common = {
            'bg': UI_THEME['bg'],
            'highlightthickness': 0,
        }

    widget.configure(**common)


# M8.3: Custom neon button presets and factory

BUTTON_PRESETS = {
    'primary': {
        'bg': UI_THEME['accent1'],
        'fg': UI_THEME['bg'],
        'activebackground': UI_THEME['accent2'],
        'activeforeground': UI_THEME['fg'],
        'font': ('Calibri', 11, 'bold'),
        'padx': 20,
        'pady': 6,
        '_hover_bg': UI_THEME['accent3'],
    },
    'primary_large': {
        'bg': UI_THEME['accent1'],
        'fg': UI_THEME['bg'],
        'activebackground': UI_THEME['accent2'],
        'activeforeground': UI_THEME['fg'],
        'font': ('Calibri', 13, 'bold'),
        'padx': 28,
        'pady': 10,
        '_hover_bg': UI_THEME['accent3'],
    },
    'secondary': {
        'bg': UI_THEME['panel_light'],
        'fg': UI_THEME['fg'],
        'activebackground': UI_THEME['hover'],
        'activeforeground': UI_THEME['fg'],
        'font': ('Calibri', 11),
        'padx': 20,
        'pady': 6,
        '_hover_bg': UI_THEME['active'],
    },
    'ghost': {
        'bg': UI_THEME['bg'],
        'fg': UI_THEME['accent1'],
        'activebackground': UI_THEME['panel'],
        'activeforeground': UI_THEME['accent1'],
        'font': ('Calibri', 11),
        'padx': 16,
        'pady': 6,
        '_hover_bg': UI_THEME['panel'],
    },
    'danger': {
        'bg': UI_THEME['critical'],
        'fg': UI_THEME['fg'],
        'activebackground': '#CC3333',
        'activeforeground': UI_THEME['fg'],
        'font': ('Calibri', 11, 'bold'),
        'padx': 20,
        'pady': 6,
        '_hover_bg': '#DD4444',
    },
    'small': {
        'bg': UI_THEME['panel_light'],
        'fg': UI_THEME['fg'],
        'activebackground': UI_THEME['hover'],
        'activeforeground': UI_THEME['fg'],
        'font': ('Calibri', 10),
        'padx': 12,
        'pady': 3,
        '_hover_bg': UI_THEME['active'],
    },
}


def create_neon_button(parent, text, command, preset='primary', **overrides):
    """Create a styled button with hover effects.

    Args:
        parent: Parent widget
        text: Button text
        command: Click callback
        preset: Key from BUTTON_PRESETS ('primary', 'secondary', 'ghost', etc.)
        **overrides: Any tk.Button config to override the preset
    Returns:
        tk.Button with hover bindings attached
    """
    cfg = dict(BUTTON_PRESETS.get(preset, BUTTON_PRESETS['primary']))
    hover_bg = cfg.pop('_hover_bg')
    cfg.update(overrides)
    # Remove any remaining private keys from overrides
    cfg = {k: v for k, v in cfg.items() if not k.startswith('_')}

    btn = tk.Button(parent, text=text, command=command,
                    relief='flat', bd=0, cursor='hand2', **cfg)

    normal_bg = cfg['bg']

    def on_enter(e):
        btn.configure(bg=hover_bg)

    def on_leave(e):
        btn.configure(bg=normal_bg)

    btn.bind('<Enter>', on_enter)
    btn.bind('<Leave>', on_leave)
    return btn


# M8.7: UI polish helper functions — cards, separators, section headers

def create_card(parent, title=None, padding='md', bg=None):
    """Create a card with subtle border and distinct background.

    Args:
        parent: Parent widget
        title: Optional card title
        padding: Key from SPACING for internal padding
        bg: Override background color

    Returns:
        tuple: (outer_frame, content_frame)
    """
    pad = SPACING.get(padding, SPACING['md'])
    card_bg = bg or UI_THEME['panel']

    outer = tk.Frame(
        parent, bg=card_bg,
        highlightbackground=BORDERS['subtle']['color'],
        highlightthickness=BORDERS['subtle']['width'],
    )

    if title:
        title_frame = tk.Frame(outer, bg=card_bg)
        title_frame.pack(fill='x', padx=pad, pady=(pad, SPACING['xs']))
        tk.Label(
            title_frame, text=title, bg=card_bg,
            fg=UI_THEME['accent1'], font=get_font('h3'),
        ).pack(anchor='w')
        sep = tk.Frame(outer, bg=BORDERS['default']['color'], height=1)
        sep.pack(fill='x', padx=pad)

    content = tk.Frame(outer, bg=card_bg)
    content.pack(fill='both', expand=True, padx=pad, pady=pad)

    return outer, content


def create_separator(parent, orientation='horizontal', color=None):
    """Create a thin visual separator line.

    Args:
        parent: Parent widget
        orientation: 'horizontal' or 'vertical'
        color: Line color (default: border_default)

    Returns:
        tk.Frame: The separator widget (caller must pack/grid it)
    """
    color = color or BORDERS['default']['color']
    if orientation == 'horizontal':
        sep = tk.Frame(parent, bg=color, height=1)
    else:
        sep = tk.Frame(parent, bg=color, width=1)
    return sep


def create_section_header(parent, title, subtitle=None, bg=None):
    """Create a section header with an accent bar on the left.

    Args:
        parent: Parent widget
        title: Section title text
        subtitle: Optional subtitle text
        bg: Background color override

    Returns:
        tk.Frame: The header widget (caller must pack/grid it)
    """
    bg = bg or UI_THEME['bg']
    header = tk.Frame(parent, bg=bg)

    accent_bar = tk.Frame(header, bg=UI_THEME['accent1'], width=4)
    accent_bar.pack(side='left', fill='y', padx=(0, SPACING['md']))

    text_frame = tk.Frame(header, bg=bg)
    text_frame.pack(side='left', fill='both', expand=True)

    tk.Label(text_frame, text=title, bg=bg,
             fg=UI_THEME['accent2'], font=get_font('h2')).pack(anchor='w')

    if subtitle:
        tk.Label(text_frame, text=subtitle, bg=bg,
                 fg=UI_THEME['fg_dim'], font=get_font('caption')).pack(anchor='w')

    return header


class MixAnalyzerApp:
    def __init__(self, root):
        self.root = root
        self.root.title('Mix Analyzer v2.0')
        self.root.geometry('1280x820')
        self.root.configure(bg=UI_THEME['bg'])
        self.root.minsize(1100, 700)

        # Set default colors for all native tk widgets
        self.root.option_add('*Background', UI_THEME['bg'])
        self.root.option_add('*Foreground', UI_THEME['fg'])

        # M8.4: Initialize typography (must happen before styles and UI)
        setup_typography()
        log_typography_info()

        setup_ttk_styles()

        # State
        self.input_folder = tk.StringVar()
        self.output_folder = tk.StringVar()
        self.style = tk.StringVar(value='Industrial')

        # Track configs: dict filename -> config dict
        self.track_configs = {}
        # Order of files as discovered
        self.track_order = []
        # Currently selected track in detail panel
        self.selected_track = None

        # Full mix state
        self.mix_state = tk.StringVar(value=MIX_STATES[0])
        self.mix_loudness_target = tk.StringVar(value=LOUDNESS_TARGETS[0])
        self.mix_note = tk.StringVar()
        self.mix_plugins = {p: tk.BooleanVar(value=False) for p in MASTER_PLUGINS}

        # Report options
        # Excel export mode: 'full' (all sheets + individual tracks),
        #   'globals' (all global sheets, no individual tracks),
        #   'ai_optimized' (AI Context + complementary globals only)
        self.excel_export_mode = tk.StringVar(value='full')
        self.image_quality = tk.StringVar(value='standard')

        # Analysis results
        self.analysis_results = None
        self.output_dir_after_run = None
        self.cancel_requested = False
        self._analysis_start_time = None

        self._build_ui()

    @property
    def generate_individual_sheets(self):
        """Backward compatibility: True when export mode is 'full'."""
        return self.excel_export_mode.get() == 'full'

    def _build_ui(self):
        # Top title with neon logo (M8.1) — M8.7: consistent header spacing
        title_frame = ttk.Frame(self.root,
                                padding=(SPACING['xl'], SPACING['md'],
                                         SPACING['xl'], SPACING['sm']))
        title_frame.pack(fill='x')

        self._logo_widget = _create_logo_widget(title_frame)
        self._logo_widget.pack(side='left')

        help_btn = create_neon_button(title_frame, 'HELP',
                                      self._show_main_help, preset='primary',
                                      padx=SPACING['lg'], pady=SPACING['xs'])
        help_btn.pack(side='right', padx=(SPACING['md'], 0))

        # M8.7: Separator between header and tabs
        header_sep = create_separator(self.root, color=BORDERS['subtle']['color'])
        header_sep.pack(fill='x', padx=SPACING['xl'])

        # Notebook — M8.6: styled tabs with icons
        self.notebook = ttk.Notebook(self.root, style='TNotebook')
        self.notebook.pack(fill='both', expand=True,
                           padx=SPACING['lg'], pady=(SPACING['sm'], SPACING['md']))

        self.tab_setup = ttk.Frame(self.notebook, style='TFrame')
        self.tab_tracks = ttk.Frame(self.notebook, style='TFrame')
        self.tab_fullmix = ttk.Frame(self.notebook, style='TFrame')
        self.tab_analysis = ttk.Frame(self.notebook, style='TFrame')

        self.notebook.add(self.tab_setup,
                          text=f"{TAB_ICONS.get('Setup', '')}  1. Setup")
        self.notebook.add(self.tab_tracks,
                          text=f"{TAB_ICONS.get('Track Identification', '')}  2. Track Identification")
        self.notebook.add(self.tab_fullmix,
                          text=f"{TAB_ICONS.get('Full Mix', '')}  3. Full Mix")
        self.notebook.add(self.tab_analysis,
                          text=f"{TAB_ICONS.get('Analysis', '')}  4. Analysis")

        self._build_setup_tab()
        self._build_tracks_tab()
        self._build_fullmix_tab()
        self._build_analysis_tab()

    # ------------------------------------------------------------------
    # MAIN HELP WINDOW
    # ------------------------------------------------------------------
    def _show_main_help(self):
        dialog = tk.Toplevel(self.root)
        dialog.title('Mix Analyzer — Help')
        dialog.configure(bg=UI_THEME['bg'])
        dialog.geometry('780x680')
        dialog.transient(self.root)

        # M8.7: consistent dialog padding
        frame = tk.Frame(dialog, bg=UI_THEME['bg'],
                         padx=SPACING['xl'], pady=SPACING['xl'])
        frame.pack(fill='both', expand=True)

        tk.Label(frame, text='Mix Analyzer — Help',
                  bg=UI_THEME['bg'], fg=UI_THEME['accent1'],
                  font=get_font('heading_dialog')).pack(
            anchor='w', pady=(0, SPACING['md']))

        # M8.7: subtle border around text area
        text_widget = tk.Text(frame, wrap='word', bg=UI_THEME['panel'],
                               fg=UI_THEME['fg'], font=get_font('body'),
                               bd=0, padx=SPACING['lg'], pady=SPACING['lg'],
                               relief='flat', spacing1=2, spacing3=2,
                               highlightbackground=BORDERS['subtle']['color'],
                               highlightthickness=BORDERS['subtle']['width'])
        text_widget.pack(fill='both', expand=True)

        # Configure heading tag
        text_widget.tag_configure('heading', font=get_font('header'),
                                   foreground=UI_THEME['accent2'],
                                   spacing1=14, spacing3=6)
        text_widget.tag_configure('subheading', font=get_font('subheading'),
                                   foreground=UI_THEME['accent1'],
                                   spacing1=10, spacing3=4)

        sections = [
            ('heading', '1. INSTALLATION'),
            ('body', (
                'Python 3.13 or higher — download from https://www.python.org/downloads/\n\n'
                'Install dependencies (open a terminal / command prompt):\n'
                '  py -m pip install numpy scipy librosa matplotlib soundfile openpyxl pyloudnorm\n\n'
                'Troubleshooting (Windows):\n'
                '  - If "py" is not recognized, make sure Python was added to PATH during '
                'installation. Re-run the installer and check "Add Python to PATH".\n'
                '  - If pip fails, try:  python -m pip install --upgrade pip\n'
            )),
            ('heading', '2. EXPORTING AUDIO FROM ABLETON LIVE 12'),
            ('body', (
                'File > Export Audio/Video (Ctrl+Shift+R)\n\n'
                'Recommended settings:\n'
                '  - File Type: WAV\n'
                '  - Bit Depth: 24-bit\n'
                '  - Sample Rate: 44100 Hz (or your project rate)\n'
                '  - Rendered Track: All Individual Tracks\n'
                '  - Convert to Mono: OFF\n'
                '  - Normalize: OFF (critical — normalization distorts measurements)\n'
                '  - Render as Loop: OFF\n\n'
                'Export all tracks to a single folder. Make sure your Ableton tracks are '
                'clearly named before exporting (e.g. "Kick Main", "Bass 303", "Lead Synth") '
                'so the auto-detect feature can identify categories.\n'
            )),
            ('heading', '3. USING MIX ANALYZER'),
            ('body', (
                'Step 1 — Setup tab:\n'
                '  Load the folder containing your exported audio files. Choose your '
                'musical style.\n\n'
                'Step 2 — Track Identification tab:\n'
                '  Mark your full song bounce as "Full Mix" (only one allowed). Run '
                '"Auto-detect all" to automatically categorize tracks by name. Review '
                'and correct any categories that were not detected.\n\n'
                'Step 3 — Full Mix tab:\n'
                '  Set the mix completion state, active master bus plugins, loudness '
                'target, and any notes about the mix.\n\n'
                'Step 4 — Analysis tab:\n'
                '  Click "RUN ANALYSIS". '
                'Wait for the Excel report to be generated. Open the output folder or generate '
                'an AI prompt when done.\n'
            )),
            ('heading', '4. SHARING THE REPORT WITH CLAUDE'),
            ('body', (
                '1. Click "Generate AI Analysis Prompt" after analysis completes.\n'
                '2. Click "Copy to Clipboard".\n'
                '3. Open a new Claude conversation (claude.ai).\n'
                '4. Drag the XLSX report file into the conversation.\n'
                '5. Paste the prompt and send.\n'
                '6. Claude will analyze your reports and provide specific, data-driven '
                'mixing recommendations tailored to your style.\n'
            )),
        ]

        for tag, content in sections:
            if tag == 'heading':
                text_widget.insert('end', content + '\n', 'heading')
            else:
                text_widget.insert('end', content + '\n')

        text_widget.config(state='disabled')

        create_neon_button(frame, 'Close', dialog.destroy,
                          preset='primary').pack(pady=(SPACING['md'], 0))

    # ------------------------------------------------------------------
    # TAB 1 - SETUP
    # ------------------------------------------------------------------
    def _build_setup_tab(self):
        frame = ttk.Frame(self.tab_setup, padding=SPACING['xl'])
        frame.pack(fill='both', expand=True)

        # M8.7: Section header with accent bar
        header = create_section_header(frame, 'Project Setup',
                                       subtitle='Configure input/output folders and style')
        header.grid(row=0, column=0, columnspan=4, sticky='we',
                    pady=(0, SPACING['lg']))

        # Input folder
        ttk.Label(frame, text='Bounces folder (WAV/AIFF):').grid(
            row=1, column=0, sticky='w', pady=SPACING['sm'])
        ttk.Entry(frame, textvariable=self.input_folder, width=70).grid(
            row=1, column=1, sticky='we', pady=SPACING['sm'],
            padx=(SPACING['md'], SPACING['xs']))
        ttk.Button(frame, text='Browse...', command=self._browse_input).grid(
            row=1, column=2, pady=SPACING['sm'])

        # Output folder
        ttk.Label(frame, text='Output folder for reports:').grid(
            row=2, column=0, sticky='w', pady=SPACING['sm'])
        ttk.Entry(frame, textvariable=self.output_folder, width=70).grid(
            row=2, column=1, sticky='we', pady=SPACING['sm'],
            padx=(SPACING['md'], SPACING['xs']))
        ttk.Button(frame, text='Browse...', command=self._browse_output).grid(
            row=2, column=2, pady=SPACING['sm'])

        # Style
        style_row = ttk.Frame(frame)
        style_row.grid(row=3, column=0, columnspan=4, sticky='we',
                       pady=SPACING['sm'])
        ttk.Label(style_row, text='Musical style:').pack(side='left')
        ttk.Combobox(style_row, textvariable=self.style,
                     values=STYLES, state='readonly', width=35, height=50).pack(
            side='left', padx=(SPACING['md'], SPACING['xs']))
        HelpButton(style_row, 'style').pack(side='left', padx=(0, SPACING['xs']))

        # M8.7: Separator before warning
        sep1 = create_separator(frame)
        sep1.grid(row=4, column=0, columnspan=4, sticky='we',
                  pady=(SPACING['lg'], SPACING['sm']))

        # Warning box for filename naming — M8.7: refined padding
        warning_frame = tk.Frame(frame, bg=UI_THEME['panel'], bd=0,
                                   highlightbackground=UI_THEME['warning'],
                                   highlightthickness=1)
        warning_frame.grid(row=5, column=0, columnspan=4, sticky='we',
                          pady=(SPACING['sm'], SPACING['md']))

        tk.Label(warning_frame, text='[!] TRACK NAMING IMPORTANT',
                  bg=UI_THEME['panel'], fg=UI_THEME['warning'],
                  font=get_font('body_bold')).pack(
            anchor='w', padx=SPACING['lg'], pady=(SPACING['md'], SPACING['xs']))

        warning_text = (
            "Before exporting from Ableton, make sure your tracks are clearly named. "
            "Auto-detection of categories uses the track name.\n\n"
            "Good naming examples:\n"
            "  * Kick Main.wav, Kick Sub.wav, Snare.wav, Hi-Hat Closed.wav\n"
            "  * Sub Bass.wav, Acid Bass 303.wav, Bass Reese.wav\n"
            "  * Lead Synth.wav, Pad Evolving.wav, Arp Sequence.wav\n\n"
            "Unnamed tracks (ElevenLabs_2026..., Splice_xyz..., Track 23...) will "
            "require manual categorization in the next tab.\n\n"
            "Ableton export settings:\n"
            "  * Rendered Track: 'All Individual Tracks'\n"
            "  * File Type: WAV, 24-bit, 44.1 or 48 kHz\n"
            "  * Normalize: OFF (critical!)\n"
            "  * Recommended: disable master bus plugins before bouncing"
        )
        tk.Label(warning_frame, text=warning_text,
                  bg=UI_THEME['panel'], fg=UI_THEME['fg'],
                  font=get_font('body_small'), justify='left',
                  wraplength=900).pack(
            anchor='w', padx=SPACING['lg'], pady=(0, SPACING['md']))

        # Status
        self.setup_status = ttk.Label(frame, text='No folder selected yet.',
                                        style='Dim.TLabel')
        self.setup_status.grid(row=6, column=0, columnspan=4, sticky='w',
                              pady=(SPACING['lg'], SPACING['xs']))

        # Load button
        ttk.Button(frame, text='Load tracks from folder',
                   style='Accent.TButton',
                   command=self._load_tracks).grid(
            row=7, column=0, columnspan=4, sticky='we',
            pady=(SPACING['sm'], 0))

        frame.columnconfigure(1, weight=1)

    def _browse_input(self):
        folder = filedialog.askdirectory(title='Select bounces folder')
        if folder:
            self.input_folder.set(folder)
            if not self.output_folder.get():
                self.output_folder.set(os.path.join(folder, 'reports'))

    def _browse_output(self):
        folder = filedialog.askdirectory(title='Select output folder')
        if folder:
            self.output_folder.set(folder)

    def _load_tracks(self):
        folder = self.input_folder.get()
        if not folder or not os.path.isdir(folder):
            messagebox.showerror('Error', 'Please select a valid input folder.')
            return

        exts = {'.wav', '.aiff', '.aif', '.flac', '.ogg'}
        files = sorted([f for f in os.listdir(folder)
                         if os.path.splitext(f)[1].lower() in exts])

        if not files:
            messagebox.showerror('Error', 'No audio files found in the selected folder.')
            return

        self.track_order = files

        # Try to load existing config
        config_path = os.path.join(folder, 'mix_analyzer_config.json')
        existing_config = {}
        if os.path.exists(config_path):
            try:
                with open(config_path, 'r', encoding='utf-8') as f:
                    existing_config = json.load(f)
            except Exception as e:
                existing_config = {}
                print(f"[Mix Analyzer] Warning: could not read config: {e}")

        # Build track configs (merge existing + auto-detect for new)
        self.track_configs = {}
        for fname in files:
            if fname in existing_config.get('tracks', {}):
                self.track_configs[fname] = existing_config['tracks'][fname]
            else:
                self.track_configs[fname] = {
                    'include': True,
                    'type': 'Individual',
                    'category': auto_detect_category(fname),
                    'parent_bus': 'None',
                }

        # Load existing full mix config
        if 'full_mix' in existing_config:
            fm = existing_config['full_mix']
            self.mix_state.set(fm.get('state', MIX_STATES[0]))
            self.mix_loudness_target.set(fm.get('loudness_target', LOUDNESS_TARGETS[0]))
            self.mix_note.set(fm.get('note', ''))
            for p, var in self.mix_plugins.items():
                var.set(p in fm.get('plugins', []))

        if 'style' in existing_config:
            self.style.set(existing_config['style'])

        if 'excel_export_mode' in existing_config:
            self.excel_export_mode.set(existing_config['excel_export_mode'])
        elif 'generate_individual_sheets' in existing_config:
            # Migrate old boolean config to new 3-way mode
            self.excel_export_mode.set(
                'full' if existing_config['generate_individual_sheets'] else 'globals')

        if 'image_quality' in existing_config:
            self.image_quality.set(existing_config['image_quality'])

        self.setup_status.config(
            text=f"Loaded {len(files)} audio files. "
                 f"Auto-detected categories for "
                 f"{sum(1 for c in self.track_configs.values() if c['category'] != '(not set)')} tracks."
        )

        self._refresh_tracks_list()
        self._refresh_fullmix_tab()
        self.notebook.select(self.tab_tracks)

        # Persist config immediately so auto-detected defaults are saved
        self._save_config()

    # ------------------------------------------------------------------
    # TAB 2 - TRACK IDENTIFICATION
    # ------------------------------------------------------------------
    def _build_tracks_tab(self):
        # Split: left = list, right = details — M8.7: uniform spacing
        paned = tk.PanedWindow(self.tab_tracks, orient='horizontal',
                                bg=UI_THEME['bg'], sashwidth=6,
                                sashrelief='flat', bd=0)
        paned.pack(fill='both', expand=True,
                   padx=SPACING['lg'], pady=SPACING['lg'])

        # LEFT: List of tracks
        left = tk.Frame(paned, bg=UI_THEME['bg'])
        paned.add(left, minsize=350)

        ttk.Label(left, text='Tracks in folder',
                  style='SubTitle.TLabel').pack(
            anchor='w', pady=(0, SPACING['sm']))

        # Action buttons
        action_row = tk.Frame(left, bg=UI_THEME['bg'])
        action_row.pack(fill='x', pady=(0, SPACING['sm']))
        ttk.Button(action_row, text='Auto-detect all', style='Small.TButton',
                   command=self._auto_detect_all).pack(
            side='left', padx=(0, SPACING['xs']))
        ttk.Button(action_row, text='Include all', style='Small.TButton',
                   command=lambda: self._set_all_include(True)).pack(
            side='left', padx=SPACING['xs'])
        ttk.Button(action_row, text='Exclude all', style='Small.TButton',
                   command=lambda: self._set_all_include(False)).pack(
            side='left', padx=SPACING['xs'])

        # Listbox with scrollbar
        list_frame = tk.Frame(left, bg=UI_THEME['bg'])
        list_frame.pack(fill='both', expand=True)

        self.tracks_listbox = tk.Listbox(list_frame,
                                           bg=UI_THEME['panel'],
                                           fg=UI_THEME['fg'],
                                           selectbackground=UI_THEME['select'],
                                           selectforeground=UI_THEME['accent1'],
                                           font=get_font('mono'),
                                           bd=0, highlightthickness=1,
                                           highlightbackground=UI_THEME['border'],
                                           highlightcolor=UI_THEME['accent1'],
                                           activestyle='none')
        self.tracks_listbox.pack(side='left', fill='both', expand=True)

        self.tracks_scrollbar = ttk.Scrollbar(list_frame, orient='vertical',
                                    command=self.tracks_listbox.yview)
        self.tracks_listbox.config(yscrollcommand=self.tracks_scrollbar.set)
        self.tracks_listbox.bind('<<ListboxSelect>>', self._on_track_selected)

        # Show/hide scrollbar based on content overflow
        def _update_scrollbar(event=None):
            self.root.after(50, self._check_scrollbar_visibility)
        list_frame.bind('<Configure>', _update_scrollbar)

        # Legend — M8.7: consistent spacing
        legend_frame = tk.Frame(left, bg=UI_THEME['bg'])
        legend_frame.pack(fill='x', pady=(SPACING['sm'], 0))
        tk.Label(legend_frame, text='Legend:',
                  bg=UI_THEME['bg'], fg=UI_THEME['fg_dim'],
                  font=get_font('caption')).pack(side='left')
        for symbol, meaning, color in [
            ('[X]', 'Included', UI_THEME['fg']),
            ('[-]', 'Excluded', UI_THEME['fg_dim']),
            ('[B]', 'BUS', UI_THEME['accent3']),
            ('[M]', 'Full Mix', UI_THEME['accent2']),
        ]:
            tk.Label(legend_frame, text=f' {symbol}={meaning}',
                      bg=UI_THEME['bg'], fg=color,
                      font=get_font('caption_tiny')).pack(side='left', padx=2)

        # RIGHT: Detail panel — M8.7: card-style with border + uniform padding
        right = tk.Frame(paned, bg=UI_THEME['panel'],
                         padx=SPACING['xl'], pady=SPACING['lg'],
                         highlightbackground=BORDERS['subtle']['color'],
                         highlightthickness=BORDERS['subtle']['width'])
        paned.add(right, minsize=500)

        self.details_title = tk.Label(right, text='No track selected',
                                        bg=UI_THEME['panel'], fg=UI_THEME['accent1'],
                                        font=get_font('header'),
                                        wraplength=500, justify='left')
        self.details_title.pack(anchor='w', pady=(0, SPACING['sm']))

        # M8.7: Separator under title
        title_sep = create_separator(right, color=BORDERS['default']['color'])
        title_sep.pack(fill='x', pady=(0, SPACING['lg']))

        self.details_frame = tk.Frame(right, bg=UI_THEME['panel'])
        self.details_frame.pack(fill='both', expand=True)

        # Build the detail widgets (will be shown/hidden based on selection)
        self._build_detail_widgets()

    def _build_detail_widgets(self):
        """Build the per-track detail widgets (hidden until a track is selected)."""
        f = self.details_frame
        _field_pady = SPACING['sm']  # M8.7: uniform vertical spacing for fields
        _field_padx = (SPACING['md'], SPACING['xs'])

        # Include checkbox
        self.detail_include = tk.BooleanVar()
        tk.Checkbutton(f, text='Include this track in the analysis',
                        variable=self.detail_include,
                        bg=UI_THEME['panel'], fg=UI_THEME['fg'],
                        selectcolor=UI_THEME['panel_light'],
                        activebackground=UI_THEME['panel'],
                        activeforeground=UI_THEME['fg'],
                        font=get_font('body'),
                        command=self._on_detail_change).grid(
            row=0, column=0, columnspan=3, sticky='w', pady=_field_pady)

        # Type
        tk.Label(f, text='Track Type:',
                  bg=UI_THEME['panel'], fg=UI_THEME['fg_dim'],
                  font=get_font('body')).grid(
            row=1, column=0, sticky='w', pady=_field_pady)
        self.detail_type = tk.StringVar()
        self.detail_type_combo = ttk.Combobox(f, textvariable=self.detail_type,
                                                 values=TRACK_TYPES, state='readonly',
                                                 width=20, height=50)
        self.detail_type_combo.grid(row=1, column=1, sticky='w',
                                    pady=_field_pady, padx=_field_padx)
        self.detail_type_combo.bind('<<ComboboxSelected>>', self._on_type_change)
        HelpButton(f, 'type').grid(row=1, column=2, sticky='w', pady=_field_pady)

        # Category
        tk.Label(f, text='Category:',
                  bg=UI_THEME['panel'], fg=UI_THEME['fg_dim'],
                  font=get_font('body')).grid(
            row=2, column=0, sticky='w', pady=_field_pady)
        self.detail_category = tk.StringVar()
        # Build hierarchical list
        cat_values = []
        for family, items in CATEGORIES.items():
            for item in items:
                cat_values.append(item)
        cat_values.append('(not set)')
        self.detail_category_combo = ttk.Combobox(f, textvariable=self.detail_category,
                                                     values=cat_values, state='readonly',
                                                     width=35, height=50)
        self.detail_category_combo.grid(row=2, column=1, sticky='w',
                                        pady=_field_pady, padx=_field_padx)
        self.detail_category_combo.bind('<<ComboboxSelected>>', self._on_detail_change)
        HelpButton(f, 'category').grid(row=2, column=2, sticky='w',
                                       pady=_field_pady)

        # Parent BUS
        tk.Label(f, text='Parent BUS:',
                  bg=UI_THEME['panel'], fg=UI_THEME['fg_dim'],
                  font=get_font('body')).grid(
            row=3, column=0, sticky='w', pady=_field_pady)
        self.detail_parent_bus = tk.StringVar()
        self.detail_parent_bus_combo = ttk.Combobox(f, textvariable=self.detail_parent_bus,
                                                       values=['None'], state='readonly',
                                                       width=35, height=50)
        self.detail_parent_bus_combo.grid(row=3, column=1, sticky='w',
                                          pady=_field_pady, padx=_field_padx)
        self.detail_parent_bus_combo.bind('<<ComboboxSelected>>', self._on_detail_change)
        HelpButton(f, 'parent_bus').grid(row=3, column=2, sticky='w',
                                         pady=_field_pady)

        # M8.7: Info box — separator with consistent spacing
        info_sep = tk.Frame(f, bg=UI_THEME['border'], height=1)
        info_sep.grid(row=4, column=0, columnspan=3, sticky='we',
                      pady=(SPACING['xl'], SPACING['lg']))

        tk.Label(f, text='Quick notes:',
                  bg=UI_THEME['panel'], fg=UI_THEME['accent2'],
                  font=get_font('body_bold')).grid(row=5, column=0, sticky='w')

        quick_tips = (
            "* Mark the bounce of your complete song as 'Full Mix' (only one allowed).\n"
            "* Mark bus sum tracks (e.g. Drums bus) as 'BUS' to exclude them from masking.\n"
            "* Individual tracks that belong to a bus should have their Parent BUS set.\n"
            "* Exclude any track you don't want in the analysis (reference tracks, etc.)."
        )
        tk.Label(f, text=quick_tips,
                  bg=UI_THEME['panel'], fg=UI_THEME['fg_dim'],
                  font=get_font('body_small'), justify='left', wraplength=500).grid(
            row=6, column=0, columnspan=3, sticky='w',
            pady=(SPACING['xs'], 0))

        # Hide all detail widgets initially
        for child in f.winfo_children():
            pass  # Will be shown when a track is selected

        self._detail_widgets_enabled(False)

    def _detail_widgets_enabled(self, enabled):
        state = 'normal' if enabled else 'disabled'
        combo_state = 'readonly' if enabled else 'disabled'
        try:
            self.detail_type_combo.configure(state=combo_state)
            self.detail_category_combo.configure(state=combo_state)
            self.detail_parent_bus_combo.configure(state=combo_state)
        except Exception:
            pass

    def _refresh_tracks_list(self):
        self.tracks_listbox.delete(0, 'end')

        # Compute longest common prefix to strip from display (>= 3 chars)
        common_prefix = ''
        if len(self.track_order) > 1:
            ref = self.track_order[0]
            for i in range(len(ref)):
                ch = ref[i]
                if all(len(fn) > i and fn[i] == ch for fn in self.track_order):
                    common_prefix += ch
                else:
                    break
            if len(common_prefix) < 3:
                common_prefix = ''

        for fname in self.track_order:
            cfg = self.track_configs[fname]
            # Build prefix
            if not cfg['include']:
                prefix = '[-]'
            elif cfg['type'] == 'Full Mix':
                prefix = '[M]'
            elif cfg['type'] == 'BUS':
                prefix = '[B]'
            else:
                prefix = '[X]'
            cat = cfg['category']
            if cat == '(not set)':
                cat = '?'
            # Strip common prefix for display, except Full Mix keeps full name
            if common_prefix and cfg['type'] != 'Full Mix':
                short_name = fname[len(common_prefix):]
            else:
                short_name = fname
            display_name = short_name if len(short_name) <= 50 else short_name[:47] + '...'
            line = f"{prefix} {display_name}   ({cat})"
            self.tracks_listbox.insert('end', line)

            # Color the line via item config
            idx = self.tracks_listbox.size() - 1
            if cfg['type'] == 'Full Mix':
                self.tracks_listbox.itemconfig(idx, foreground=UI_THEME['accent2'])
            elif cfg['type'] == 'BUS':
                self.tracks_listbox.itemconfig(idx, foreground=UI_THEME['accent3'])
            elif not cfg['include']:
                self.tracks_listbox.itemconfig(idx, foreground=UI_THEME['fg_dim'])
            else:
                self.tracks_listbox.itemconfig(idx, foreground=UI_THEME['fg'])

        # Rebuild parent BUS dropdown options (only tracks marked as BUS)
        bus_options = ['None'] + [f for f in self.track_order
                                    if self.track_configs[f]['type'] == 'BUS']
        self.detail_parent_bus_combo.configure(values=bus_options)
        self._check_scrollbar_visibility()

    def _check_scrollbar_visibility(self):
        """Show scrollbar only when listbox content overflows."""
        lb = self.tracks_listbox
        sb = self.tracks_scrollbar
        try:
            visible_height = lb.winfo_height()
            line_height = lb.bbox(0)[3] if lb.size() > 0 and lb.bbox(0) else 16
            total_height = lb.size() * line_height
            if total_height > visible_height:
                sb.pack(side='right', fill='y')
            else:
                sb.pack_forget()
        except Exception:
            sb.pack(side='right', fill='y')

    def _on_track_selected(self, event=None):
        selection = self.tracks_listbox.curselection()
        if not selection:
            return
        idx = selection[0]
        fname = self.track_order[idx]
        self.selected_track = fname
        cfg = self.track_configs[fname]

        self.details_title.config(text=fname)
        self.detail_include.set(cfg['include'])
        self.detail_type.set(cfg['type'])
        self.detail_category.set(cfg['category'])
        self.detail_parent_bus.set(cfg.get('parent_bus', 'None'))

        self._detail_widgets_enabled(True)
        # Disable category/parent_bus if Full Mix or BUS
        self._apply_type_constraints()

    def _apply_type_constraints(self):
        t = self.detail_type.get()
        if t == 'Full Mix':
            self.detail_category_combo.configure(state='disabled')
            self.detail_parent_bus_combo.configure(state='disabled')
            self.detail_category.set('(not set)')
            self.detail_parent_bus.set('None')
        elif t == 'BUS':
            self.detail_category_combo.configure(state='readonly')
            self.detail_parent_bus_combo.configure(state='disabled')
            self.detail_parent_bus.set('None')
        else:  # Individual
            self.detail_category_combo.configure(state='readonly')
            self.detail_parent_bus_combo.configure(state='readonly')

    def _on_type_change(self, event=None):
        if not self.selected_track:
            return
        new_type = self.detail_type.get()
        if new_type == 'Full Mix':
            # Ensure only one Full Mix
            for fname, cfg in self.track_configs.items():
                if fname != self.selected_track and cfg['type'] == 'Full Mix':
                    cfg['type'] = 'Individual'
        self._apply_type_constraints()
        self._on_detail_change()

    def _on_detail_change(self, event=None):
        if not self.selected_track:
            return
        cfg = self.track_configs[self.selected_track]
        cfg['include'] = self.detail_include.get()
        cfg['type'] = self.detail_type.get()
        cfg['category'] = self.detail_category.get()
        cfg['parent_bus'] = self.detail_parent_bus.get()
        self._refresh_tracks_list()
        self._refresh_fullmix_tab()
        # Restore selection
        try:
            idx = self.track_order.index(self.selected_track)
            self.tracks_listbox.selection_set(idx)
        except Exception:
            pass
        self._save_config()

    def _auto_detect_all(self):
        if not messagebox.askokcancel(
                'Auto-detect',
                'Auto-detect will overwrite any category labels you have manually set.\n\n'
                'Continue?'):
            return
        project_name = None
        for fname, cfg in self.track_configs.items():
            if cfg.get('type') == 'Full Mix':
                project_name = os.path.splitext(fname)[0]
                break
        if not project_name and self.track_order:
            names = [os.path.splitext(f)[0] for f in self.track_order]
            prefix = os.path.commonprefix(names).strip(' -_')
            if len(prefix) >= 3:
                project_name = prefix
        for fname in self.track_order:
            self.track_configs[fname]['category'] = auto_detect_category(fname, project_name)
        self._refresh_tracks_list()
        if self.selected_track:
            self._on_track_selected_refresh()
        self._save_config()
        messagebox.showinfo('Auto-detect',
                             f"Auto-detection complete.\n"
                             f"{sum(1 for c in self.track_configs.values() if c['category'] != '(not set)')}"
                             f" of {len(self.track_configs)} tracks have a detected category.")

    def _set_all_include(self, include):
        for cfg in self.track_configs.values():
            cfg['include'] = include
        self._refresh_tracks_list()
        self._save_config()

    def _on_track_selected_refresh(self):
        if self.selected_track:
            cfg = self.track_configs[self.selected_track]
            self.detail_include.set(cfg['include'])
            self.detail_type.set(cfg['type'])
            self.detail_category.set(cfg['category'])
            self.detail_parent_bus.set(cfg.get('parent_bus', 'None'))

    # ------------------------------------------------------------------
    # TAB 3 - FULL MIX
    # ------------------------------------------------------------------
    def _build_fullmix_tab(self):
        frame = ttk.Frame(self.tab_fullmix, padding=SPACING['xl'])
        frame.pack(fill='both', expand=True)

        # M8.7: Section header with accent bar
        header = create_section_header(frame, 'Full Mix Configuration',
                                       subtitle='Master bus settings and mix context')
        header.grid(row=0, column=0, columnspan=3, sticky='we',
                    pady=(0, SPACING['lg']))

        self.fullmix_status = ttk.Label(
            frame,
            text='No track marked as Full Mix yet. Go to the Track Identification tab '
                 'and set Type = Full Mix for the bounce of your complete song.',
            style='Dim.TLabel', wraplength=900)
        self.fullmix_status.grid(row=1, column=0, columnspan=3, sticky='w',
                                 pady=(0, SPACING['lg']))

        # Details frame (shown when a full mix is selected)
        self.fullmix_details = tk.Frame(frame, bg=UI_THEME['bg'])
        self.fullmix_details.grid(row=2, column=0, columnspan=3, sticky='nwe',
                                  pady=SPACING['sm'])

        # Mix state
        row = 0
        state_row = tk.Frame(self.fullmix_details, bg=UI_THEME['bg'])
        state_row.grid(row=row, column=0, columnspan=3, sticky='w',
                       pady=SPACING['sm'])
        tk.Label(state_row, text='Mix completion state:',
                  bg=UI_THEME['bg'], fg=UI_THEME['fg'],
                  font=get_font('body')).pack(side='left')
        ttk.Combobox(state_row, textvariable=self.mix_state,
                     values=MIX_STATES, state='readonly', width=25, height=50).pack(
            side='left', padx=(SPACING['md'], SPACING['xs']))
        HelpButton(state_row, 'mix_state').pack(side='left')

        # M8.7: Separator before plugins
        row += 1
        sep_plugins = create_separator(self.fullmix_details)
        sep_plugins.grid(row=row, column=0, columnspan=3, sticky='we',
                         pady=SPACING['md'])

        # Plugins
        row += 1
        tk.Label(self.fullmix_details, text='Active master bus plugins:',
                  bg=UI_THEME['bg'], fg=UI_THEME['fg'],
                  font=get_font('body')).grid(
            row=row, column=0, sticky='nw', pady=(SPACING['sm'], SPACING['xs']))
        HelpButton(self.fullmix_details, 'master_plugins').grid(
            row=row, column=1, sticky='nw',
            pady=(SPACING['sm'], SPACING['xs']), padx=SPACING['xs'])

        row += 1
        plugins_frame = tk.Frame(self.fullmix_details, bg=UI_THEME['bg'])
        plugins_frame.grid(row=row, column=0, columnspan=3, sticky='w',
                          padx=SPACING['xl'])
        for i, plugin in enumerate(MASTER_PLUGINS):
            tk.Checkbutton(plugins_frame, text=plugin,
                            variable=self.mix_plugins[plugin],
                            bg=UI_THEME['bg'], fg=UI_THEME['fg'],
                            selectcolor=UI_THEME['panel_light'],
                            activebackground=UI_THEME['bg'],
                            activeforeground=UI_THEME['fg'],
                            font=get_font('body'),
                            command=self._save_config).grid(
                row=0, column=i, sticky='w', padx=(0, SPACING['xl']))

        # M8.7: Separator before loudness
        row += 1
        sep_loudness = create_separator(self.fullmix_details)
        sep_loudness.grid(row=row, column=0, columnspan=3, sticky='we',
                          pady=SPACING['md'])

        # Loudness target
        row += 1
        target_row = tk.Frame(self.fullmix_details, bg=UI_THEME['bg'])
        target_row.grid(row=row, column=0, columnspan=3, sticky='w',
                        pady=SPACING['sm'])
        tk.Label(target_row, text='Loudness target:',
                  bg=UI_THEME['bg'], fg=UI_THEME['fg'],
                  font=get_font('body')).pack(side='left')
        ttk.Combobox(target_row, textvariable=self.mix_loudness_target,
                     values=LOUDNESS_TARGETS, state='readonly', width=35, height=50).pack(
            side='left', padx=(SPACING['md'], SPACING['xs']))
        HelpButton(target_row, 'loudness_target').pack(side='left')

        # Note
        row += 1
        tk.Label(self.fullmix_details, text='Note (optional):',
                  bg=UI_THEME['bg'], fg=UI_THEME['fg'],
                  font=get_font('body')).grid(
            row=row, column=0, sticky='w', pady=(SPACING['lg'], SPACING['xs']))
        row += 1
        ttk.Entry(self.fullmix_details, textvariable=self.mix_note,
                   width=80).grid(row=row, column=0, columnspan=3, sticky='we')

        # Save changes on mix state/target/note updates
        self.mix_state.trace_add('write', lambda *a: self._save_config())
        self.mix_loudness_target.trace_add('write', lambda *a: self._save_config())
        self.mix_note.trace_add('write', lambda *a: self._save_config())

    def _refresh_fullmix_tab(self):
        full_mix_files = [f for f in self.track_order
                            if self.track_configs.get(f, {}).get('type') == 'Full Mix']
        if full_mix_files:
            self.fullmix_status.config(
                text=f"Full Mix file: {full_mix_files[0]}",
                foreground=UI_THEME['accent2'])
        else:
            self.fullmix_status.config(
                text='No track marked as Full Mix yet. Go to the Track Identification tab '
                     'and set Type = Full Mix for the bounce of your complete song.',
                foreground=UI_THEME['fg_dim'])

    # ------------------------------------------------------------------
    # TAB 4 - ANALYSIS
    # ------------------------------------------------------------------
    def _build_analysis_tab(self):
        frame = ttk.Frame(self.tab_analysis, padding=SPACING['xl'])
        frame.pack(fill='both', expand=True)

        # M8.7: Section header with accent bar
        header = create_section_header(frame, 'Run Analysis',
                                       subtitle='Analyze tracks and generate Excel report')
        header.grid(row=0, column=0, columnspan=3, sticky='we',
                    pady=(0, SPACING['lg']))

        # Summary — M8.7: card-style with border
        summary_card = tk.Frame(frame, bg=UI_THEME['panel'],
                                highlightbackground=BORDERS['subtle']['color'],
                                highlightthickness=BORDERS['subtle']['width'])
        summary_card.grid(row=1, column=0, columnspan=3, sticky='we',
                         pady=(0, SPACING['sm']))

        self.analysis_summary = tk.Text(summary_card, height=5,
                                          bg=UI_THEME['panel'],
                                          fg=UI_THEME['fg'],
                                          font=get_font('body_small'),
                                          bd=0, padx=SPACING['lg'],
                                          pady=SPACING['md'], relief='flat',
                                          state='disabled')
        self.analysis_summary.pack(fill='x')

        # Controls row: Refresh
        controls_row = ttk.Frame(frame)
        controls_row.grid(row=2, column=0, columnspan=3, sticky='we',
                         pady=(SPACING['xs'], SPACING['md']))

        ttk.Button(controls_row, text='Refresh summary',
                    command=self._refresh_analysis_summary).pack(side='left')

        # M8.7: Separator before options
        sep_opts = create_separator(frame)
        sep_opts.grid(row=3, column=0, columnspan=3, sticky='we',
                      pady=(0, SPACING['md']))

        # Report options — M8.7: card-style
        options_card = tk.Frame(frame, bg=UI_THEME['panel'],
                                highlightbackground=BORDERS['subtle']['color'],
                                highlightthickness=BORDERS['subtle']['width'])
        options_card.grid(row=4, column=0, columnspan=3, sticky='we',
                         pady=(0, SPACING['md']))

        options_inner = tk.Frame(options_card, bg=UI_THEME['panel'])
        options_inner.pack(fill='x', padx=SPACING['md'], pady=SPACING['sm'])

        # Excel export mode — 3-way radio group
        export_frame = tk.LabelFrame(options_inner, text='Excel Export Mode',
                                      bg=UI_THEME['panel'],
                                      fg=UI_THEME['fg'],
                                      font=get_font('body_small'))
        export_frame.pack(side='left', fill='x', expand=True)

        export_options = [
            ('full',
             'Full report \u2014 global sheets + one per track',
             'Complete report with all global and individual track sheets.\n'
             'Largest file, best for human deep-dive review.'),
            ('globals',
             'Globals only \u2014 no individual track sheets',
             'All global sheets but skips per-track analyses.\n'
             'Medium file size, good for quick overview.'),
            ('ai_optimized',
             'AI-optimized \u2014 AI Context + essentials only',
             'Smallest file: AI Context + Anomalies, Full Mix Context,\n'
             'Freq Conflicts, Mix Health Score, AI Prompt.\n'
             'Excludes visual/redundant sheets. Best for AI analysis\n'
             'with minimal token cost.'),
        ]
        for value, label, tooltip_text in export_options:
            rb = tk.Radiobutton(export_frame, text=label,
                                 variable=self.excel_export_mode,
                                 value=value,
                                 bg=UI_THEME['panel'], fg=UI_THEME['fg'],
                                 selectcolor=UI_THEME['bg'],
                                 activebackground=UI_THEME['panel'],
                                 activeforeground=UI_THEME['fg'],
                                 font=get_font('body_small'))
            rb.pack(anchor='w', padx=SPACING['sm'], pady=2)

        tk.Label(options_inner, text='  Image quality:',
                 bg=UI_THEME['panel'], fg=UI_THEME['fg'],
                 font=get_font('body_small')).pack(
            side='left', padx=(SPACING['lg'], 2))
        quality_menu = tk.OptionMenu(options_inner, self.image_quality,
                                      'standard', 'high')
        quality_menu.config(bg=UI_THEME['panel'], fg=UI_THEME['fg'],
                            font=get_font('body_small'), highlightthickness=0)
        quality_menu.pack(side='left')

        # Run + Cancel row
        run_row = ttk.Frame(frame)
        run_row.grid(row=5, column=0, columnspan=3, sticky='we',
                     pady=(SPACING['xs'], SPACING['md']))

        self.run_button = ttk.Button(run_row, text='>>> RUN ANALYSIS <<<',
                                       style='Accent.TButton',
                                       command=self._run_analysis)
        self.run_button.pack(side='left', fill='x', expand=True)

        self.cancel_button = ttk.Button(run_row, text='Cancel',
                                          style='Secondary.TButton',
                                          command=self._request_cancel,
                                          state='disabled')
        self.cancel_button.pack(side='left', padx=(SPACING['md'], 0))

        # Multi-level progress display — M8.7: card-style with border
        progress_frame = tk.Frame(frame, bg=UI_THEME['panel'],
                                  padx=SPACING['md'], pady=SPACING['md'],
                                  highlightbackground=BORDERS['subtle']['color'],
                                  highlightthickness=BORDERS['subtle']['width'])
        progress_frame.grid(row=6, column=0, columnspan=3, sticky='we',
                           pady=(0, SPACING['md']))

        self.progress_bar = NeonProgressBar(progress_frame, height=24,
                                              maximum=100, glow=True,
                                              show_percent=True)
        self.progress_bar.pack(fill='x', pady=(0, SPACING['sm']))

        self.progress_step = tk.Label(progress_frame, text='Step: Idle',
                                        bg=UI_THEME['panel'], fg=UI_THEME['accent1'],
                                        font=get_font('body_small_bold'), anchor='w')
        self.progress_step.pack(fill='x')

        self.progress_substep = tk.Label(progress_frame, text='Substep: —',
                                           bg=UI_THEME['panel'], fg=UI_THEME['fg'],
                                           font=get_font('body_small'), anchor='w')
        self.progress_substep.pack(fill='x')

        progress_bottom = tk.Frame(progress_frame, bg=UI_THEME['panel'])
        progress_bottom.pack(fill='x', pady=(SPACING['xs'], 0))

        self.progress_counter = tk.Label(progress_bottom, text='',
                                           bg=UI_THEME['panel'], fg=UI_THEME['fg_dim'],
                                           font=get_font('body_small'), anchor='w')
        self.progress_counter.pack(side='left')

        self.progress_eta = tk.Label(progress_bottom, text='',
                                       bg=UI_THEME['panel'], fg=UI_THEME['fg_dim'],
                                       font=get_font('body_small'), anchor='e')
        self.progress_eta.pack(side='right')

        # Post-run buttons
        buttons_row = ttk.Frame(frame)
        buttons_row.grid(row=7, column=0, columnspan=3, sticky='w',
                        pady=(0, SPACING['md']))

        self.ai_button = ttk.Button(buttons_row, text='Generate AI Analysis Prompt',
                                      style='Secondary.TButton',
                                      command=self._show_ai_prompt,
                                      state='disabled')
        self.ai_button.pack(side='left')

        self.open_folder_button = ttk.Button(buttons_row, text='Open output folder',
                                                style='Secondary.TButton',
                                                command=self._open_output_folder,
                                                state='disabled')
        self.open_folder_button.pack(side='left', padx=(SPACING['md'], 0))

        # M8.7: Separator before log
        sep_log = create_separator(frame)
        sep_log.grid(row=8, column=0, columnspan=3, sticky='we',
                     pady=(0, SPACING['sm']))

        # Log
        ttk.Label(frame, text='Log:', style='Dim.TLabel').grid(
            row=9, column=0, columnspan=3, sticky='w',
            pady=(SPACING['xs'], SPACING['xs']))

        log_frame = tk.Frame(frame, bg=UI_THEME['bg'],
                             highlightbackground=BORDERS['subtle']['color'],
                             highlightthickness=BORDERS['subtle']['width'])
        log_frame.grid(row=10, column=0, columnspan=3, sticky='nsew',
                      pady=(0, SPACING['xs']))

        self.log_text = scrolledtext.ScrolledText(
            log_frame, height=12, bg=UI_THEME['bg_input'], fg=UI_THEME['accent4'],
            font=get_font('mono'), bd=0, padx=SPACING['md'], pady=SPACING['sm'],
            insertbackground=UI_THEME['accent1'],
            selectbackground=UI_THEME['select'],
            selectforeground=UI_THEME['fg'])
        self.log_text.pack(fill='both', expand=True)

        frame.columnconfigure(0, weight=1)
        frame.columnconfigure(1, weight=1)
        frame.rowconfigure(10, weight=1)

    def _refresh_analysis_summary(self):
        if not self.track_configs:
            text = "No tracks loaded. Go to Setup tab and load a folder."
        else:
            included = [f for f, c in self.track_configs.items() if c['include']]
            full_mix = [f for f in included
                         if self.track_configs[f]['type'] == 'Full Mix']
            buses = [f for f in included
                      if self.track_configs[f]['type'] == 'BUS']
            individuals = [f for f in included
                            if self.track_configs[f]['type'] == 'Individual']
            not_set = [f for f in included
                        if self.track_configs[f]['type'] == 'Individual'
                        and self.track_configs[f]['category'] == '(not set)']

            # Family breakdown
            family_counts = {}
            for f in individuals:
                cat = self.track_configs[f]['category']
                family = CATEGORY_FAMILY.get(cat, 'Unknown')
                family_counts[family] = family_counts.get(family, 0) + 1

            text = (
                f"Style: {self.style.get()}\n"
                f"Included tracks: {len(included)} "
                f"({len(individuals)} individual | {len(buses)} BUS | {len(full_mix)} full mix)\n"
                f"Excluded tracks: {len(self.track_configs) - len(included)}\n"
                f"Individual tracks with no category set: {len(not_set)}\n"
                f"Families: {', '.join(f'{k}={v}' for k, v in sorted(family_counts.items()))}"
            )
        self.analysis_summary.config(state='normal')
        self.analysis_summary.delete('1.0', 'end')
        self.analysis_summary.insert('1.0', text)
        self.analysis_summary.config(state='disabled')

    def log(self, msg):
        """Thread-safe log: if called from worker, schedule via root.after."""
        import threading as _th
        def _do_log():
            self.log_text.insert('end', msg + '\n')
            self.log_text.see('end')
        if _th.current_thread() is _th.main_thread():
            _do_log()
        else:
            self.root.after(0, _do_log)

    def _update_progress(self, pct, step='', substep='', counter='', eta=''):
        """Thread-safe progress update."""
        def _do():
            self.progress_bar['value'] = pct
            if step:
                self.progress_step.config(text=f'Step: {step}')
            if substep:
                self.progress_substep.config(text=f'Substep: {substep}')
            self.progress_counter.config(text=counter)
            self.progress_eta.config(text=eta)
        self.root.after(0, _do)

    def _compute_eta(self, completed, total):
        """Return formatted ETA string."""
        import time
        if completed <= 0 or not self._analysis_start_time:
            return ''
        elapsed = time.time() - self._analysis_start_time
        remaining = (elapsed / completed) * (total - completed)
        mins, secs = divmod(int(remaining), 60)
        return f'ETA: {mins:02d}:{secs:02d}'

    def _request_cancel(self):
        self.cancel_requested = True
        self.cancel_button.configure(state='disabled')
        self.log("Cancel requested — stopping after current track...")

    def _reset_progress(self):
        """Reset progress display to idle state."""
        self.progress_bar['value'] = 0
        self.progress_step.config(text='Step: Idle')
        self.progress_substep.config(text='Substep: —')
        self.progress_counter.config(text='')
        self.progress_eta.config(text='')

    def _run_analysis(self):
        if not self.input_folder.get() or not self.output_folder.get():
            messagebox.showerror('Error', 'Please set input and output folders in Setup tab.')
            return
        if not self.track_configs:
            messagebox.showerror('Error', 'No tracks loaded. Load tracks in Setup tab first.')
            return

        self._save_config()
        self.log_text.delete('1.0', 'end')
        self.cancel_requested = False
        self._reset_progress()
        self.run_button.configure(state='disabled')
        self.ai_button.configure(state='disabled')
        self.open_folder_button.configure(state='disabled')
        self.cancel_button.configure(state='normal')

        import time
        self._analysis_start_time = time.time()

        def worker():
            try:
                self._do_analysis()
            except Exception as e:
                if 'CancelledError' not in type(e).__name__:
                    err_msg = f"ERROR: {e}"
                    self.root.after(0, lambda: self.log(err_msg))
                    traceback.print_exc()
            finally:
                def _cleanup():
                    self.run_button.configure(state='normal')
                    self.cancel_button.configure(state='disabled')
                    self.cancel_requested = False
                self.root.after(0, _cleanup)

        threading.Thread(target=worker, daemon=True).start()

    def _do_analysis(self):
        import time

        input_folder = Path(self.input_folder.get())
        output_folder = Path(self.output_folder.get())
        output_folder.mkdir(parents=True, exist_ok=True)
        # Determine project name from Full Mix track (fallback: folder name)
        project_name = None
        for fname, cfg in self.track_configs.items():
            if cfg.get('type') == 'Full Mix':
                project_name = os.path.splitext(fname)[0]
                break
        if not project_name:
            project_name = input_folder.name
        safe_project = "".join(c for c in project_name if c.isalnum() or c in ' _-').strip() or 'Project'
        date_str = datetime.now().strftime('%Y-%m-%d')
        time_str = datetime.now().strftime('%H-%M')
        report_prefix = f"{safe_project}_MixAnalyzer_{date_str}_{time_str}"

        included_files = [f for f in self.track_order
                           if self.track_configs[f]['include']]

        # Build full mix info
        active_plugins = [p for p, v in self.mix_plugins.items() if v.get()]
        full_mix_info = {
            'state': self.mix_state.get(),
            'plugins': active_plugins,
            'loudness_target': self.mix_loudness_target.get(),
            'note': self.mix_note.get(),
        }

        # Compute total steps for progress: analyze + Excel generation
        n_tracks = len(included_files)
        total_steps = n_tracks + 1  # analysis + Excel
        completed_steps = 0

        self.log(f"Starting analysis: {n_tracks} tracks — format: Excel")
        self.log(f"Output folder: {output_folder}")
        self.log("-" * 60)

        self._update_progress(0, 'Analyzing tracks', '', f'[0/{n_tracks}]', '')

        analyses_with_info = []
        generated_files = []

        for i, fname in enumerate(included_files, 1):
            if self.cancel_requested:
                self.log("CANCELLED by user.")
                self._update_progress(0, 'Cancelled', '', '', '')
                # Clean up partial files
                for gf in generated_files:
                    try:
                        if gf.exists():
                            gf.unlink()
                    except Exception:
                        pass
                return

            filepath = input_folder / fname
            cfg = self.track_configs[fname]
            eta = self._compute_eta(completed_steps, total_steps)
            self._update_progress(
                int(completed_steps / total_steps * 100),
                'Analyzing tracks',
                f'Analyzing: {fname}',
                f'[{i}/{n_tracks}]',
                eta)
            self.log(f"[{i}/{n_tracks}] Analyzing: {fname}")
            try:
                is_full_mix = (cfg['type'] == 'Full Mix')
                analysis = analyze_track(str(filepath), compute_tempo=is_full_mix)
                ti = {
                    'type': cfg['type'],
                    'category': cfg['category'],
                    'parent_bus': cfg.get('parent_bus', 'None'),
                    'name': fname,
                }
                analyses_with_info.append((analysis, ti))
            except Exception as e:
                self.log(f"    ERROR: {e}")
                traceback.print_exc()
            completed_steps += 1

        # --- Excel generation ---
        if analyses_with_info:
            if self.cancel_requested:
                self.log("CANCELLED by user.")
                self._update_progress(0, 'Cancelled', '', '', '')
                for gf in generated_files:
                    try:
                        if gf.exists():
                            gf.unlink()
                    except Exception:
                        pass
                return

            export_mode = self.excel_export_mode.get()
            mode_labels = {
                'full': 'full mode (all sheets + individual tracks)',
                'globals': 'globals only (no individual track sheets)',
                'ai_optimized': 'AI-optimized (AI Context + essentials)',
            }
            mode_label = mode_labels.get(export_mode, export_mode)
            self._update_progress(
                int(completed_steps / total_steps * 100),
                'Generating Excel report', '', '', '')
            self.log("-" * 60)
            self.log(f"Generating Excel report ({mode_label})...")
            try:
                # Temporarily store results so _build_ai_prompt can use them
                old_results = self.analysis_results
                self.analysis_results = analyses_with_info
                ai_prompt = self._build_ai_prompt()
                self.analysis_results = old_results

                xlsx_path = output_folder / f'{report_prefix}.xlsx'
                if xlsx_path.exists():
                    xlsx_path.unlink()
                generate_excel_report(
                    analyses_with_info, str(xlsx_path), self.style.get(),
                    full_mix_info=full_mix_info, ai_prompt=ai_prompt,
                    log_fn=self.log,
                    export_mode=export_mode,
                    image_quality=self.image_quality.get())
                generated_files.append(xlsx_path)
                self.log(f"Excel report: {xlsx_path.name}")
            except Exception as e:
                self.log(f"ERROR Excel report: {e}")
                traceback.print_exc()
            completed_steps += 1

        # Done
        self._update_progress(100, 'Done', '', '', '')
        self.log("=" * 60)
        self.log(f"DONE: {len(analyses_with_info)} tracks processed — Excel report generated")
        self.log(f"Location: {output_folder}")

        self.analysis_results = analyses_with_info
        self.output_dir_after_run = output_folder

        def _enable_buttons():
            self.ai_button.configure(state='normal')
            self.open_folder_button.configure(state='normal')
        self.root.after(0, _enable_buttons)

    def _open_output_folder(self):
        if self.output_dir_after_run and os.path.isdir(self.output_dir_after_run):
            try:
                if sys.platform == 'win32':
                    os.startfile(str(self.output_dir_after_run))
                elif sys.platform == 'darwin':
                    subprocess.call(['open', str(self.output_dir_after_run)])
                else:
                    subprocess.call(['xdg-open', str(self.output_dir_after_run)])
            except Exception as e:
                messagebox.showerror('Error', f"Could not open folder: {e}")

    # ------------------------------------------------------------------
    # CONFIG SAVE/LOAD
    # ------------------------------------------------------------------
    def _save_config(self):
        folder = self.input_folder.get()
        if not folder or not os.path.isdir(folder):
            return
        config_path = os.path.join(folder, 'mix_analyzer_config.json')
        try:
            active_plugins = [p for p, v in self.mix_plugins.items() if v.get()]
            config = {
                'style': self.style.get(),
                'tracks': self.track_configs,
                'excel_export_mode': self.excel_export_mode.get(),
                'image_quality': self.image_quality.get(),
                'full_mix': {
                    'state': self.mix_state.get(),
                    'plugins': active_plugins,
                    'loudness_target': self.mix_loudness_target.get(),
                    'note': self.mix_note.get(),
                }
            }
            with open(config_path, 'w', encoding='utf-8') as f:
                json.dump(config, f, indent=2)
        except Exception as e:
            print(f"[Mix Analyzer] Warning: could not save config: {e}")

    # ------------------------------------------------------------------
    # AI PROMPT
    # ------------------------------------------------------------------
    def _show_ai_prompt(self):
        if not self.analysis_results:
            return

        prompt = self._build_ai_prompt()

        dialog = tk.Toplevel(self.root)
        dialog.title('AI Analysis Prompt')
        dialog.configure(bg=UI_THEME['bg'])
        dialog.geometry('900x700')
        dialog.transient(self.root)

        # M8.7: consistent dialog padding
        frame = tk.Frame(dialog, bg=UI_THEME['bg'],
                         padx=SPACING['xl'], pady=SPACING['xl'])
        frame.pack(fill='both', expand=True)

        tk.Label(frame, text='AI Analysis Prompt',
                  bg=UI_THEME['bg'], fg=UI_THEME['accent1'],
                  font=get_font('heading_help')).pack(
            anchor='w', pady=(0, SPACING['xs']))

        tk.Label(frame,
                  text='1. Click "Copy to Clipboard" below\n'
                       '2. Open a new conversation on Claude.ai (or similar AI)\n'
                       '3. Attach the XLSX report file from your output folder\n'
                       '4. Paste this prompt in the message and send',
                  bg=UI_THEME['bg'], fg=UI_THEME['fg_dim'],
                  font=get_font('body_small'), justify='left').pack(
            anchor='w', pady=(0, SPACING['md']))

        # M8.7: subtle border around text area
        text_widget = scrolledtext.ScrolledText(
            frame, wrap='word', bg=UI_THEME['panel'], fg=UI_THEME['fg'],
            font=get_font('mono'), bd=0, padx=SPACING['md'],
            pady=SPACING['md'], relief='flat',
            insertbackground=UI_THEME['fg'],
            highlightbackground=BORDERS['subtle']['color'],
            highlightthickness=BORDERS['subtle']['width'])
        text_widget.pack(fill='both', expand=True)
        text_widget.insert('1.0', prompt)

        def copy_to_clipboard():
            dialog.clipboard_clear()
            dialog.clipboard_append(prompt)
            copy_btn.config(text='Copied!')
            dialog.after(1500, lambda: copy_btn.config(text='Copy to Clipboard'))

        btn_row = tk.Frame(frame, bg=UI_THEME['bg'])
        btn_row.pack(fill='x', pady=(SPACING['md'], 0))

        copy_btn = create_neon_button(btn_row, 'Copy to Clipboard',
                                      copy_to_clipboard, preset='primary',
                                      pady=SPACING['sm'])
        copy_btn.pack(side='left')

        create_neon_button(btn_row, 'Close', dialog.destroy,
                          preset='secondary', pady=SPACING['sm']).pack(
            side='left', padx=(SPACING['md'], 0))

    def _build_ai_prompt(self):
        """Generate the contextualized AI prompt."""
        if not self.analysis_results:
            return ""

        # Count types
        individuals = [a for a, ti in self.analysis_results if ti['type'] == 'Individual']
        buses = [a for a, ti in self.analysis_results if ti['type'] == 'BUS']
        full_mixes = [a for a, ti in self.analysis_results if ti['type'] == 'Full Mix']

        # Track inventory by category
        cat_inventory = {}
        for a, ti in self.analysis_results:
            if ti['type'] == 'Individual':
                cat = ti.get('category', '(not set)')
                cat_inventory.setdefault(cat, []).append(a['filename'])

        cat_inventory_str = ""
        for cat in sorted(cat_inventory.keys()):
            files = cat_inventory[cat]
            cat_inventory_str += f"\n- {cat}: {len(files)} track(s)"
            for f in files[:3]:
                cat_inventory_str += f"\n    * {f}"
            if len(files) > 3:
                cat_inventory_str += f"\n    * ... and {len(files) - 3} more"

        # Full mix context
        full_mix_context = ""
        if full_mixes:
            active_plugins = [p for p, v in self.mix_plugins.items() if v.get()]
            plugins_str = ', '.join(active_plugins) if active_plugins else 'None'
            full_mix_context = (
                f"\n\nFULL MIX CONTEXT:\n"
                f"- File: {full_mixes[0]['filename']}\n"
                f"- Completion state: {self.mix_state.get()}\n"
                f"- Active master bus plugins: {plugins_str}\n"
                f"- Loudness target: {self.mix_loudness_target.get()}\n"
                f"- Note: {self.mix_note.get() if self.mix_note.get() else 'None'}"
            )

        prompt = f"""You are a senior mixing and mastering engineer specialized in {self.style.get()} music. I am providing you with an Excel analysis report generated by the Mix Analyzer tool for one of my music projects. I need your help diagnosing the mix and proposing concrete, actionable improvements.

PROJECT CONTEXT:
- Style: {self.style.get()}
- Total tracks analyzed: {len(self.analysis_results)}
- Individual tracks: {len(individuals)}
- BUS tracks: {len(buses)}
- Full Mix bounce: {len(full_mixes)}{full_mix_context}

TRACK INVENTORY (Individual tracks by category):{cat_inventory_str}

ATTACHED EXCEL REPORT:
- One comprehensive XLSX report with multiple sheets: per-track analysis, Summary, Dashboard, Global Comparison, Full Mix Analysis, Freq Conflicts, Track Comparison, Mix Health Score, Version Tracking, and AI Prompt

YOUR TASK:
Please analyze the reports and provide:

1. **Overall diagnosis**: What is the current state of the mix? What are the main issues you can see from the objective measurements across all tracks?

2. **Per-category analysis**: For each track category (Drums, Bass, Synths, etc.), highlight specific tracks that stand out positively or negatively.

3. **Masking and frequency conflicts**: Using the masking matrix in the global report, identify which tracks are competing for the same frequency ranges and suggest specific carving strategies.

4. **Dynamic range and compression**: Look at crest factors across the project. Identify tracks that are over-compressed relative to their role, and tracks that could benefit from more control.

5. **Stereo field coherence**: Review the stereo panorama of individual tracks. Identify stereo field issues, unexpected mono elements, or wide elements that may cause phase problems.

6. **Full Mix coherence**: Compare the Full Mix measurements to what the individual tracks suggest should be achievable. Is the Full Mix a faithful sum, or does it suggest master bus processing is altering the balance significantly?

7. **Concrete action items**: Provide a prioritized list of specific, non-generic actions I should take. Avoid boilerplate advice. Each recommendation should reference specific tracks or measurements from the reports.

CRITICAL INSTRUCTIONS:
- Base ALL your recommendations strictly on what is visible in the attached Excel report. Do not make assumptions about content you cannot see.
- Avoid generic mixing advice. Do not say things like "add a highpass at 80 Hz on bass tracks" unless you can point to a specific measurement in a specific report that justifies it.
- When a recommendation is subjective or style-dependent, explicitly acknowledge it.
- Consider the {self.style.get()} aesthetic. What would be a "problem" in another style may be intentional here.
- Cross-reference observations across multiple tracks. The power of this analysis is the multi-track view, not individual assessments.
- If certain measurements are affected by the master bus plugins listed above, factor that into your interpretation.

Start your analysis with a concise executive summary (3-4 sentences) before diving into details."""

        return prompt
# ============================================================================
# MAIN ENTRY POINT
# ============================================================================


def main():
    root = tk.Tk()
    app = MixAnalyzerApp(root)
    root.mainloop()


if __name__ == '__main__':
    main()
