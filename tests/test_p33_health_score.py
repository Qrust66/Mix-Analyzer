"""
Validation tests for P3.3 — Mix Health Score sheet.
Run with: python3 tests/test_p33_health_score.py
"""
import sys
import os
import tempfile
import numpy as np

sys.path.insert(0, os.path.dirname(os.path.dirname(os.path.abspath(__file__))))

# --- Dependencies from mix_analyzer.py ---

FREQ_BANDS = [
    ('sub', 20, 60), ('bass', 60, 250), ('low_mid', 250, 500),
    ('mid', 500, 2000), ('high_mid', 2000, 4000), ('presence', 4000, 8000),
    ('air', 8000, 20000),
]

BAND_LABELS = {
    'sub': 'Sub (20-60 Hz)', 'bass': 'Bass (60-250 Hz)',
    'low_mid': 'Low-Mid (250-500 Hz)', 'mid': 'Mid (500-2000 Hz)',
    'high_mid': 'High-Mid (2-4 kHz)', 'presence': 'Presence (4-8 kHz)',
    'air': 'Air (8-20 kHz)',
}


def _xl_write_header(ws, title, subtitle=''):
    from openpyxl.styles import Font, PatternFill, Alignment
    ws.merge_cells('A1:J1')
    ws['A1'] = title
    ws['A1'].font = Font(name='Calibri', size=16, bold=True, color='00D9FF')
    ws['A1'].fill = PatternFill('solid', fgColor='0A0A12')
    ws['A1'].alignment = Alignment(horizontal='left')
    if subtitle:
        ws.merge_cells('A2:J2')
        ws['A2'] = subtitle
        ws['A2'].font = Font(name='Calibri', size=11, color='8888A0')
        ws['A2'].fill = PatternFill('solid', fgColor='0A0A12')
    return 4


def detect_anomalies(analysis):
    """Minimal replica of detect_anomalies for testing."""
    anomalies = []
    L = analysis['loudness']
    stereo = analysis['stereo']
    if L['peak_db'] > -0.3:
        anomalies.append(('critical', 'Clipping risk'))
    elif L['peak_db'] > -1.0:
        anomalies.append(('warning', 'Low headroom'))
    if L.get('true_peak_db', L['peak_db']) > 0.0:
        anomalies.append(('critical', 'Inter-sample clipping'))
    if stereo['is_stereo'] and stereo.get('correlation', 1.0) < 0.0:
        anomalies.append(('warning', 'Phase issue'))
    if L.get('crest_factor', 12) < 5 and L.get('rms_db', -20) > -30:
        anomalies.append(('warning', 'Heavy compression'))
    return anomalies


# --- Load functions under test ---

with open(os.path.join(os.path.dirname(os.path.dirname(os.path.abspath(__file__))),
                       'mix_analyzer.py')) as f:
    source = f.read()

# Extract _apply_clean_layout helper (M5.1)
acl_start = source.find('def _apply_clean_layout(')
acl_end = source.find('\n\ndef ', acl_start + 1)
exec(source[acl_start:acl_end], globals())

# Extract the 5 helper functions + generate_health_score_sheet
start = source.find('def _calc_loudness_score(')
end = source.find('\ndef generate_excel_report(')
exec(source[start:end], globals())


# --- Mock data helpers ---

def _make_band_energies(sub=5, bass=20, low_mid=15, mid=25, high_mid=15, presence=12, air=8):
    return {'sub': sub, 'bass': bass, 'low_mid': low_mid, 'mid': mid,
            'high_mid': high_mid, 'presence': presence, 'air': air}


def _make_analysis(name, lufs=-14.0, peak=-3.0, true_peak=-2.5, crest=12.0,
                   plr=13.0, psr=10.0, rms_db=-20.0, width=0.5,
                   correlation=0.9, is_stereo=True, dominant='mid',
                   band_energies=None, peaks=None):
    if band_energies is None:
        band_energies = _make_band_energies()
    if peaks is None:
        peaks = []
    return {
        'filename': name,
        'loudness': {
            'lufs_integrated': lufs, 'peak_db': peak, 'true_peak_db': true_peak,
            'crest_factor': crest, 'plr': plr, 'psr': psr, 'rms_db': rms_db,
        },
        'spectrum': {
            'dominant_band': dominant, 'band_energies': band_energies, 'peaks': peaks,
        },
        'stereo': {
            'is_stereo': is_stereo, 'width_overall': width, 'correlation': correlation,
            'width_per_band': {'sub': 0.1, 'bass': 0.2, 'low_mid': 0.3, 'mid': 0.4,
                               'high_mid': 0.5, 'presence': 0.6, 'air': 0.7},
        },
    }


def make_good_mix():
    """A healthy mix: balanced, good dynamics, safe levels."""
    tracks = []
    for i, name in enumerate(['Kick', 'Snare', 'Bass', 'Vocal', 'Guitar']):
        tracks.append((_make_analysis(name, lufs=-14.0 - i * 0.5, peak=-3.0, crest=12.0),
                       {'type': 'Individual', 'category': 'Instruments'}))
    tracks.append((_make_analysis('Full Mix', lufs=-14.0, peak=-1.5, true_peak=-1.2,
                                   crest=11.0, plr=12.0, width=0.55),
                   {'type': 'Full Mix', 'category': ''}))
    return tracks


def make_bad_mix():
    """A problematic mix: crushed dynamics, clipping, unbalanced."""
    tracks = []
    for i, name in enumerate(['Kick', 'Snare', 'Bass']):
        tracks.append((_make_analysis(name, lufs=-8.0, peak=-0.1, crest=4.5,
                                       rms_db=-10.0,
                                       band_energies=_make_band_energies(sub=60, bass=25, mid=10,
                                                                          low_mid=3, high_mid=1,
                                                                          presence=0.5, air=0.5)),
                       {'type': 'Individual', 'category': 'Instruments'}))
    tracks.append((_make_analysis('Full Mix', lufs=-6.0, peak=0.5, true_peak=1.0,
                                   crest=3.5, plr=6.0, width=0.9, correlation=-0.1,
                                   band_energies=_make_band_energies(sub=55, bass=30, mid=10,
                                                                      low_mid=3, high_mid=1,
                                                                      presence=0.5, air=0.5)),
                   {'type': 'Full Mix', 'category': ''}))
    return tracks


# --- Tests ---

def test_sheet_creation():
    """Sheet 'Mix Health Score' is created in the workbook."""
    from openpyxl import Workbook
    wb = Workbook()
    generate_health_score_sheet(wb, make_good_mix(), log_fn=lambda m: None)
    assert 'Mix Health Score' in wb.sheetnames


def test_tab_color():
    """Tab color is green (#3DFFAA)."""
    from openpyxl import Workbook
    wb = Workbook()
    generate_health_score_sheet(wb, make_good_mix(), log_fn=lambda m: None)
    ws = wb['Mix Health Score']
    assert '3DFFAA' in ws.sheet_properties.tabColor.rgb


def test_global_score_in_a3():
    """Global score is written in A3 as a number 0-100."""
    from openpyxl import Workbook
    wb = Workbook()
    generate_health_score_sheet(wb, make_good_mix(), log_fn=lambda m: None)
    ws = wb['Mix Health Score']
    score = ws['A3'].value
    assert isinstance(score, (int, float)), f"Score should be numeric, got {type(score)}"
    assert 0 <= score <= 100, f"Score out of range: {score}"


def test_good_mix_high_score():
    """A healthy mix should score >= 60."""
    from openpyxl import Workbook
    wb = Workbook()
    generate_health_score_sheet(wb, make_good_mix(), log_fn=lambda m: None)
    ws = wb['Mix Health Score']
    assert ws['A3'].value >= 60, f"Good mix scored only {ws['A3'].value}"


def test_bad_mix_low_score():
    """A problematic mix should score < 50."""
    from openpyxl import Workbook
    wb = Workbook()
    generate_health_score_sheet(wb, make_bad_mix(), log_fn=lambda m: None)
    ws = wb['Mix Health Score']
    assert ws['A3'].value < 50, f"Bad mix scored {ws['A3'].value}, expected < 50"


def test_title():
    """Title 'MIX HEALTH SCORE' in A1."""
    from openpyxl import Workbook
    wb = Workbook()
    generate_health_score_sheet(wb, make_good_mix(), log_fn=lambda m: None)
    ws = wb['Mix Health Score']
    assert ws['A1'].value == 'MIX HEALTH SCORE'


def test_date_and_track_count():
    """Metadata in A5 (date) and A6 (track count) are present."""
    from openpyxl import Workbook
    wb = Workbook()
    generate_health_score_sheet(wb, make_good_mix(), log_fn=lambda m: None)
    ws = wb['Mix Health Score']
    assert 'Calculated on:' in ws['A5'].value
    assert '5 Individual' in ws['A6'].value
    assert 'Full Mix' in ws['A6'].value


def test_disclaimer():
    """Disclaimer note in A7."""
    from openpyxl import Workbook
    wb = Workbook()
    generate_health_score_sheet(wb, make_good_mix(), log_fn=lambda m: None)
    ws = wb['Mix Health Score']
    assert 'heuristics' in ws['A7'].value.lower()
    assert 'artistic' in ws['A7'].value.lower()


def test_five_categories():
    """5 category rows in the sub-scores table (rows 11-15)."""
    from openpyxl import Workbook
    wb = Workbook()
    generate_health_score_sheet(wb, make_good_mix(), log_fn=lambda m: None)
    ws = wb['Mix Health Score']
    expected = ['Loudness', 'Dynamics', 'Spectral Balance', 'Stereo Image', 'Anomalies']
    for i, name in enumerate(expected):
        row = 11 + i
        assert ws.cell(row=row, column=1).value == name, \
            f"Row {row}: expected '{name}', got '{ws.cell(row=row, column=1).value}'"


def test_sub_scores_are_numeric():
    """Each sub-score in column B is a number 0-100."""
    from openpyxl import Workbook
    wb = Workbook()
    generate_health_score_sheet(wb, make_good_mix(), log_fn=lambda m: None)
    ws = wb['Mix Health Score']
    for row in range(11, 16):
        val = ws.cell(row=row, column=2).value
        assert isinstance(val, (int, float)), f"Row {row} score not numeric: {val}"
        assert 0 <= val <= 100, f"Row {row} score out of range: {val}"


def test_weights_sum_to_100():
    """Category weights sum to 100%."""
    from openpyxl import Workbook
    wb = Workbook()
    generate_health_score_sheet(wb, make_good_mix(), log_fn=lambda m: None)
    ws = wb['Mix Health Score']
    weights = []
    for row in range(11, 16):
        w_str = ws.cell(row=row, column=3).value  # e.g. "20%"
        w_val = int(w_str.replace('%', ''))
        weights.append(w_val)
    assert sum(weights) == 100, f"Weights sum to {sum(weights)}, expected 100"


def test_weighted_total_row():
    """WEIGHTED TOTAL row is present with the global score."""
    from openpyxl import Workbook
    wb = Workbook()
    generate_health_score_sheet(wb, make_good_mix(), log_fn=lambda m: None)
    ws = wb['Mix Health Score']
    assert ws.cell(row=17, column=1).value == 'WEIGHTED TOTAL'
    assert ws.cell(row=17, column=2).value == ws['A3'].value


def test_category_headers():
    """Category table headers in row 10."""
    from openpyxl import Workbook
    wb = Workbook()
    generate_health_score_sheet(wb, make_good_mix(), log_fn=lambda m: None)
    ws = wb['Mix Health Score']
    assert ws.cell(row=10, column=1).value == 'Category'
    assert ws.cell(row=10, column=2).value == 'Score'
    assert ws.cell(row=10, column=3).value == 'Weight'
    assert ws.cell(row=10, column=4).value == 'Notes'


def test_notes_column():
    """Notes column (D) has text for each category."""
    from openpyxl import Workbook
    wb = Workbook()
    generate_health_score_sheet(wb, make_good_mix(), log_fn=lambda m: None)
    ws = wb['Mix Health Score']
    for row in range(11, 16):
        note = ws.cell(row=row, column=4).value
        assert note is not None and len(str(note)) > 0, f"Row {row}: empty note"


def test_detail_sections_present():
    """Detail sections exist after row 20 for all 5 categories."""
    from openpyxl import Workbook
    wb = Workbook()
    generate_health_score_sheet(wb, make_good_mix(), log_fn=lambda m: None)
    ws = wb['Mix Health Score']
    detail_headers = []
    for row in range(20, 80):
        val = ws.cell(row=row, column=1).value
        if val and 'Details' in str(val):
            detail_headers.append(val)
    assert 'Loudness Details' in detail_headers
    assert 'Dynamics Details' in detail_headers
    assert 'Spectral Balance Details' in detail_headers
    assert 'Stereo Image Details' in detail_headers
    assert 'Anomalies Details' in detail_headers


def test_detail_columns():
    """Detail sub-tables have Metric/Value/Ideal Range/Contribution headers."""
    from openpyxl import Workbook
    wb = Workbook()
    generate_health_score_sheet(wb, make_good_mix(), log_fn=lambda m: None)
    ws = wb['Mix Health Score']
    # Find the first detail header row
    for row in range(20, 80):
        if ws.cell(row=row, column=1).value == 'Metric':
            assert ws.cell(row=row, column=2).value == 'Value'
            assert ws.cell(row=row, column=3).value == 'Ideal Range'
            assert ws.cell(row=row, column=4).value == 'Contribution'
            return
    assert False, "No detail column headers found"


def test_freeze_panes():
    """Freeze panes at A11."""
    from openpyxl import Workbook
    wb = Workbook()
    generate_health_score_sheet(wb, make_good_mix(), log_fn=lambda m: None)
    ws = wb['Mix Health Score']
    assert ws.freeze_panes == 'A11'


def test_conditional_formatting():
    """Color scale on sub-score column."""
    from openpyxl import Workbook
    wb = Workbook()
    generate_health_score_sheet(wb, make_good_mix(), log_fn=lambda m: None)
    ws = wb['Mix Health Score']
    assert len(ws.conditional_formatting._cf_rules) >= 1


def test_column_widths():
    """Column widths are set."""
    from openpyxl import Workbook
    wb = Workbook()
    generate_health_score_sheet(wb, make_good_mix(), log_fn=lambda m: None)
    ws = wb['Mix Health Score']
    assert ws.column_dimensions['A'].width == 25
    assert ws.column_dimensions['D'].width == 30


def test_bus_excluded():
    """BUS tracks are excluded from scoring."""
    from openpyxl import Workbook
    wb = Workbook()
    tracks = make_good_mix()
    tracks.append((_make_analysis('Drums Bus', lufs=-8.0, peak=-0.1, crest=3.0),
                   {'type': 'BUS', 'category': 'Drums'}))
    generate_health_score_sheet(wb, tracks, log_fn=lambda m: None)
    ws = wb['Mix Health Score']
    assert '5 Individual' in ws['A6'].value  # still 5, not 6


def test_no_tracks():
    """Graceful handling when no tracks available."""
    from openpyxl import Workbook
    wb = Workbook()
    generate_health_score_sheet(wb, [], log_fn=lambda m: None)
    ws = wb['Mix Health Score']
    assert ws is not None


def test_individuals_only_no_fullmix():
    """Works with only Individual tracks, no Full Mix."""
    from openpyxl import Workbook
    wb = Workbook()
    tracks = [(_make_analysis(f'Track_{i}', lufs=-14.0 - i),
               {'type': 'Individual', 'category': 'X'}) for i in range(3)]
    generate_health_score_sheet(wb, tracks, log_fn=lambda m: None)
    ws = wb['Mix Health Score']
    score = ws['A3'].value
    assert isinstance(score, (int, float)) and 0 <= score <= 100
    assert 'Full Mix' not in ws['A6'].value


def test_loudness_score_helper():
    """_calc_loudness_score returns valid score and details."""
    individuals = [(_make_analysis(f'T{i}', lufs=-14.0 - i * 0.3), {}) for i in range(4)]
    full_mix = _make_analysis('FM', lufs=-14.0, peak=-2.0, true_peak=-1.5)
    score, details, note = _calc_loudness_score(individuals, full_mix)
    assert 0 <= score <= 100
    assert len(details) == 3  # coherence, target, peak
    assert isinstance(note, str)


def test_dynamics_score_helper():
    """_calc_dynamics_score returns valid score and details."""
    individuals = [(_make_analysis(f'T{i}', crest=12.0), {}) for i in range(3)]
    full_mix = _make_analysis('FM', plr=12.0)
    score, details, note = _calc_dynamics_score(individuals, full_mix)
    assert 0 <= score <= 100
    assert len(details) == 3  # crest mean, PLR, crest std
    assert isinstance(note, str)


def test_spectral_balance_score_helper():
    """_calc_spectral_balance_score returns valid score with balanced energies."""
    individuals = [(_make_analysis('T1'), {})]
    score, details, note = _calc_spectral_balance_score(individuals, None)
    assert 0 <= score <= 100
    assert len(details) >= 3  # entropy, sub, air, (maybe dominant)
    assert isinstance(note, str)


def test_spectral_penalty_no_air():
    """Spectral score penalizes when Air Energy is 0."""
    individuals = [(_make_analysis('T1', band_energies=_make_band_energies(air=0)), {})]
    score_no_air, _, _ = _calc_spectral_balance_score(individuals, None)
    individuals2 = [(_make_analysis('T1', band_energies=_make_band_energies(air=10)), {})]
    score_with_air, _, _ = _calc_spectral_balance_score(individuals2, None)
    assert score_no_air < score_with_air


def test_spectral_penalty_dominant_band():
    """Spectral score penalizes when one band > 35%."""
    individuals = [(_make_analysis('T1', band_energies=_make_band_energies(sub=50, bass=20, mid=15,
                                                                            low_mid=5, high_mid=5,
                                                                            presence=3, air=2)), {})]
    score, details, _ = _calc_spectral_balance_score(individuals, None)
    # Should have penalty for dominant band
    has_penalty = any('penalty' in str(d) for d in details)
    assert has_penalty, f"Expected dominant band penalty in details: {details}"


def test_stereo_image_score_helper():
    """_calc_stereo_image_score returns valid score and details."""
    full_mix = _make_analysis('FM', width=0.55)
    score, details, note = _calc_stereo_image_score([], full_mix)
    assert 0 <= score <= 100
    assert len(details) >= 2  # width, sub mono check
    assert isinstance(note, str)


def test_stereo_ideal_width_perfect_score():
    """Width 0.55 (in ideal range 0.4-0.7) should give width_score = 100."""
    full_mix = _make_analysis('FM', width=0.55)
    score, details, _ = _calc_stereo_image_score([], full_mix)
    # Width contributes 60% with score 100, sub mono ~40% with ~100 => ~100
    assert score >= 90, f"Ideal width scored {score}, expected >= 90"


def test_anomalies_score_clean():
    """Clean tracks with no anomalies score 100."""
    tracks = [(_make_analysis(f'T{i}', peak=-3.0, true_peak=-2.5, crest=12.0),
               {'type': 'Individual', 'category': 'X'}) for i in range(3)]
    score, details, note = _calc_anomalies_score(tracks)
    assert score == 100
    assert note == 'Clean'


def test_anomalies_score_with_issues():
    """Tracks with anomalies have reduced score."""
    tracks = [(_make_analysis('T1', peak=0.5, true_peak=1.0, crest=4.0, rms_db=-10.0),
               {'type': 'Individual', 'category': 'X'})]
    score, details, note = _calc_anomalies_score(tracks)
    assert score < 100
    assert 'anomalies' in note


def test_anomalies_excludes_bus():
    """BUS tracks are skipped in anomaly counting."""
    tracks = [
        (_make_analysis('Good', peak=-3.0, true_peak=-2.0), {'type': 'Individual', 'category': 'X'}),
        (_make_analysis('BadBus', peak=0.5, true_peak=1.0, crest=4.0, rms_db=-10.0),
         {'type': 'BUS', 'category': 'X'}),
    ]
    score, _, _ = _calc_anomalies_score(tracks)
    assert score == 100  # BUS anomalies ignored


def test_excel_save_roundtrip():
    """Workbook saves and re-opens without error."""
    from openpyxl import Workbook, load_workbook
    wb = Workbook()
    generate_health_score_sheet(wb, make_good_mix(), log_fn=lambda m: None)
    tmp = tempfile.mktemp(suffix='.xlsx')
    try:
        wb.save(tmp)
        assert os.path.getsize(tmp) > 0
        wb2 = load_workbook(tmp)
        assert 'Mix Health Score' in wb2.sheetnames
        ws2 = wb2['Mix Health Score']
        assert ws2['A1'].value == 'MIX HEALTH SCORE'
        assert ws2.freeze_panes == 'A11'
    finally:
        os.unlink(tmp)


# --- Runner ---

if __name__ == '__main__':
    tests = [
        test_sheet_creation,
        test_tab_color,
        test_global_score_in_a3,
        test_good_mix_high_score,
        test_bad_mix_low_score,
        test_title,
        test_date_and_track_count,
        test_disclaimer,
        test_five_categories,
        test_sub_scores_are_numeric,
        test_weights_sum_to_100,
        test_weighted_total_row,
        test_category_headers,
        test_notes_column,
        test_detail_sections_present,
        test_detail_columns,
        test_freeze_panes,
        test_conditional_formatting,
        test_column_widths,
        test_bus_excluded,
        test_no_tracks,
        test_individuals_only_no_fullmix,
        test_loudness_score_helper,
        test_dynamics_score_helper,
        test_spectral_balance_score_helper,
        test_spectral_penalty_no_air,
        test_spectral_penalty_dominant_band,
        test_stereo_image_score_helper,
        test_stereo_ideal_width_perfect_score,
        test_anomalies_score_clean,
        test_anomalies_score_with_issues,
        test_anomalies_excludes_bus,
        test_excel_save_roundtrip,
    ]

    passed = 0
    failed = 0
    for test in tests:
        try:
            test()
            print(f'  [OK] {test.__doc__}')
            passed += 1
        except Exception as e:
            print(f'  [FAIL] {test.__doc__}')
            print(f'         {e}')
            failed += 1

    print(f'\n{passed}/{passed + failed} tests passed.')
    if failed:
        sys.exit(1)
