"""
Validation tests for P3.1 — Frequency Conflict Detector sheet.
Run with: python3 tests/test_p31_freq_conflicts.py
"""
import sys
import os
import tempfile
import numpy as np
import librosa

sys.path.insert(0, os.path.dirname(os.path.dirname(os.path.abspath(__file__))))

# --- Extract dependencies from mix_analyzer.py ---

FREQ_BANDS_HIRES = [
    ('20-32 Hz', 20, 32), ('32-50 Hz', 32, 50), ('50-80 Hz', 50, 80),
    ('80-125 Hz', 80, 125), ('125-160 Hz', 125, 160), ('160-200 Hz', 160, 200),
    ('200-250 Hz', 200, 250), ('250-315 Hz', 250, 315), ('315-400 Hz', 315, 400),
    ('400-500 Hz', 400, 500), ('500-630 Hz', 500, 630), ('630-800 Hz', 630, 800),
    ('800-1k Hz', 800, 1000), ('1-1.25 kHz', 1000, 1250), ('1.25-1.6 kHz', 1250, 1600),
    ('1.6-2 kHz', 1600, 2000), ('2-2.5 kHz', 2000, 2500), ('2.5-3.15 kHz', 2500, 3150),
    ('3.15-4 kHz', 3150, 4000), ('4-5 kHz', 4000, 5000), ('5-8 kHz', 5000, 8000),
    ('8-20 kHz', 8000, 20000),
]


def compute_hires_band_energies(mono, sr):
    n_fft = 8192
    S = np.abs(librosa.stft(mono, n_fft=n_fft, hop_length=n_fft // 4))
    spectrum_mean = np.mean(S, axis=1)
    freqs = librosa.fft_frequencies(sr=sr, n_fft=n_fft)
    total_energy = np.sum(spectrum_mean ** 2) + 1e-12
    energies = {}
    for idx_b, (label, flow, fhigh) in enumerate(FREQ_BANDS_HIRES):
        if idx_b == len(FREQ_BANDS_HIRES) - 1:
            mask = (freqs >= flow) & (freqs <= fhigh)
        else:
            mask = (freqs >= flow) & (freqs < fhigh)
        if np.any(mask):
            energies[label] = 100 * float(np.sum(spectrum_mean[mask] ** 2)) / total_energy
        else:
            energies[label] = 0.0
    return energies


def _xl_write_header(ws, title, subtitle=''):
    from openpyxl.styles import Font, PatternFill, Alignment
    header_fill = PatternFill('solid', fgColor='0A0A12')
    accent_font = Font(name='Calibri', size=16, bold=True, color='00D9FF')
    sub_font = Font(name='Calibri', size=11, color='8888A0')
    ws.merge_cells('A1:J1')
    ws['A1'] = title
    ws['A1'].font = accent_font
    ws['A1'].fill = header_fill
    ws['A1'].alignment = Alignment(horizontal='left')
    if subtitle:
        ws.merge_cells('A2:J2')
        ws['A2'] = subtitle
        ws['A2'].font = sub_font
        ws['A2'].fill = header_fill
    return 4


# --- Load the function under test ---

with open(os.path.join(os.path.dirname(os.path.dirname(os.path.abspath(__file__))),
                       'mix_analyzer.py')) as f:
    source = f.read()

acl_start = source.find('def _apply_clean_layout(')
acl_end = source.find('\n\ndef ', acl_start + 1)
exec(source[acl_start:acl_end], globals())

# Extract _xl_add_sheet_nav helper (M7.5)
xsn_start = source.find('def _xl_add_sheet_nav(')
xsn_end = source.find('\n\ndef ', xsn_start + 1)
exec(source[xsn_start:xsn_end], globals())

start = source.find('def generate_freq_conflicts_sheet(')
end = source.find('\ndef generate_excel_report(')
exec(source[start:end], globals())


# --- Test helpers ---

def make_mock_tracks(n=5, sr=22050, duration=2.0):
    """Create mock Individual tracks with distinct frequency content."""
    n_samples = int(sr * duration)
    tracks = []
    for i in range(n):
        freq = 80 * (i + 1)
        t = np.linspace(0, duration, n_samples)
        mono = (np.sin(2 * np.pi * freq * t) * 0.8).astype(np.float32)
        mono += 0.05 * np.random.randn(n_samples).astype(np.float32)
        tracks.append((
            {'filename': f'Track_{i+1}', '_mono': mono, 'sample_rate': sr},
            {'type': 'Individual', 'category': 'Instruments'}
        ))
    return tracks


def make_full_test_set():
    """Create Individual + BUS + Full Mix tracks."""
    sr = 22050
    n_samples = int(sr * 2.0)
    tracks = make_mock_tracks(5, sr)
    # BUS (should be excluded)
    tracks.append((
        {'filename': 'Drums Bus', '_mono': np.random.randn(n_samples).astype(np.float32) * 0.3,
         'sample_rate': sr},
        {'type': 'BUS', 'category': 'Drums'}
    ))
    # Full Mix (should be excluded)
    tracks.append((
        {'filename': 'Full Mix Bounce', '_mono': np.random.randn(n_samples).astype(np.float32) * 0.5,
         'sample_rate': sr},
        {'type': 'Full Mix', 'category': ''}
    ))
    return tracks


# --- Tests ---

def test_sheet_creation():
    """Sheet 'Freq Conflicts' is created in the workbook."""
    from openpyxl import Workbook
    wb = Workbook()
    generate_freq_conflicts_sheet(wb, make_mock_tracks(), log_fn=lambda m: None)
    assert 'Freq Conflicts' in wb.sheetnames, f"Sheet not found: {wb.sheetnames}"


def test_parameters():
    """Editable parameters in B2 (threshold) and B3 (min tracks)."""
    from openpyxl import Workbook
    wb = Workbook()
    generate_freq_conflicts_sheet(wb, make_mock_tracks(), log_fn=lambda m: None)
    ws = wb['Freq Conflicts']
    assert ws['A2'].value == 'Conflict threshold (% of max band energy)'
    assert ws['B2'].value == 15.0
    assert ws['A3'].value == 'Min tracks for conflict'
    assert ws['B3'].value == 2


def test_custom_parameters():
    """Custom threshold and min_tracks values are respected."""
    from openpyxl import Workbook
    wb = Workbook()
    generate_freq_conflicts_sheet(wb, make_mock_tracks(), default_threshold=25.0,
                                   default_min_tracks=3, log_fn=lambda m: None)
    ws = wb['Freq Conflicts']
    assert ws['B2'].value == 25.0
    assert ws['B3'].value == 3


def test_track_filtering():
    """Only Individual tracks are included; BUS and Full Mix are excluded."""
    from openpyxl import Workbook
    wb = Workbook()
    tracks = make_full_test_set()  # 5 Individual + 1 BUS + 1 Full Mix
    generate_freq_conflicts_sheet(wb, tracks, log_fn=lambda m: None)
    ws = wb['Freq Conflicts']
    # 5 track columns + Conflict count + Status
    assert ws.cell(row=5, column=7).value == 'Conflict count'
    assert ws.cell(row=5, column=8).value == 'Status'
    # Track names in columns 2-6
    for col in range(2, 7):
        name = ws.cell(row=5, column=col).value
        assert name is not None
        assert 'Bus' not in name and 'Full Mix' not in name


def test_22_bands():
    """All 22 FREQ_BANDS_HIRES bands are present."""
    from openpyxl import Workbook
    wb = Workbook()
    generate_freq_conflicts_sheet(wb, make_mock_tracks(), log_fn=lambda m: None)
    ws = wb['Freq Conflicts']
    assert ws.cell(row=6, column=1).value == '20-32 Hz'
    assert ws.cell(row=27, column=1).value == '8-20 kHz'
    # Count data rows
    band_count = 0
    for r in range(6, 28):
        if ws.cell(row=r, column=1).value:
            band_count += 1
    assert band_count == 22


def test_data_normalized():
    """Data values are normalized to 0-100 range."""
    from openpyxl import Workbook
    wb = Workbook()
    generate_freq_conflicts_sheet(wb, make_mock_tracks(), log_fn=lambda m: None)
    ws = wb['Freq Conflicts']
    for row in range(6, 28):
        for col in range(2, 7):
            val = ws.cell(row=row, column=col).value
            assert isinstance(val, float), f"Cell ({row},{col}) not float: {type(val)}"
            assert 0 <= val <= 100, f"Cell ({row},{col}) out of range: {val}"


def test_countif_formula():
    """COUNTIF formula references $B$2 for dynamic threshold."""
    from openpyxl import Workbook
    wb = Workbook()
    generate_freq_conflicts_sheet(wb, make_mock_tracks(), log_fn=lambda m: None)
    ws = wb['Freq Conflicts']
    formula = ws.cell(row=6, column=7).value
    assert 'COUNTIF' in formula
    assert '$B$2' in formula
    assert '">="' in formula


def test_status_formula():
    """IF formula references $B$3 for dynamic min tracks."""
    from openpyxl import Workbook
    wb = Workbook()
    generate_freq_conflicts_sheet(wb, make_mock_tracks(), log_fn=lambda m: None)
    ws = wb['Freq Conflicts']
    formula = ws.cell(row=6, column=8).value
    assert 'IF' in formula
    assert '$B$3' in formula
    assert 'CONFLICT' in formula
    assert 'OK' in formula


def test_freeze_panes():
    """Freeze panes set at B6."""
    from openpyxl import Workbook
    wb = Workbook()
    generate_freq_conflicts_sheet(wb, make_mock_tracks(), log_fn=lambda m: None)
    ws = wb['Freq Conflicts']
    assert ws.freeze_panes == 'B6'


def test_autofilter():
    """Autofilter is set on the data range."""
    from openpyxl import Workbook
    wb = Workbook()
    generate_freq_conflicts_sheet(wb, make_mock_tracks(), log_fn=lambda m: None)
    ws = wb['Freq Conflicts']
    assert ws.auto_filter.ref is not None
    assert ws.auto_filter.ref != ''
    assert 'A5' in ws.auto_filter.ref


def test_conditional_formatting():
    """Conditional formatting rules are applied (color scale, formula, data bar)."""
    from openpyxl import Workbook
    wb = Workbook()
    generate_freq_conflicts_sheet(wb, make_mock_tracks(), log_fn=lambda m: None)
    ws = wb['Freq Conflicts']
    rules = ws.conditional_formatting._cf_rules
    assert len(rules) >= 3


def test_excel_save_roundtrip():
    """Generated workbook saves and re-opens without error."""
    from openpyxl import Workbook, load_workbook
    wb = Workbook()
    generate_freq_conflicts_sheet(wb, make_full_test_set(), log_fn=lambda m: None)

    tmp = tempfile.mktemp(suffix='.xlsx')
    try:
        wb.save(tmp)
        assert os.path.getsize(tmp) > 0

        wb2 = load_workbook(tmp)
        ws2 = wb2['Freq Conflicts']
        assert ws2['B2'].value == 15
        assert ws2.cell(row=6, column=7).value is not None
        assert ws2.freeze_panes == 'B6'
    finally:
        os.unlink(tmp)


def test_empty_individuals():
    """Graceful handling when no Individual tracks exist."""
    from openpyxl import Workbook
    wb = Workbook()
    tracks = [
        ({'filename': 'Bus', '_mono': np.zeros(1000, dtype=np.float32), 'sample_rate': 22050},
         {'type': 'BUS', 'category': 'X'}),
    ]
    generate_freq_conflicts_sheet(wb, tracks, log_fn=lambda m: None)
    ws = wb['Freq Conflicts']
    assert ws is not None  # Sheet still created with a message


# --- Runner ---

if __name__ == '__main__':
    tests = [
        test_sheet_creation,
        test_parameters,
        test_custom_parameters,
        test_track_filtering,
        test_22_bands,
        test_data_normalized,
        test_countif_formula,
        test_status_formula,
        test_freeze_panes,
        test_autofilter,
        test_conditional_formatting,
        test_excel_save_roundtrip,
        test_empty_individuals,
    ]

    passed = 0
    failed = 0
    for test in tests:
        try:
            test()
            print(f'  [OK] {test.__doc__}')
            passed += 1
        except Exception as e:
            print(f'  [FAIL] {test.__doc__}')
            print(f'         {e}')
            failed += 1

    print(f'\n{passed}/{passed + failed} tests passed.')
    if failed:
        sys.exit(1)
