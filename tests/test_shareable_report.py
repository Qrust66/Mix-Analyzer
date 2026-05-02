#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
Tests for Phase F10f — peak_trajectories filter + SHAREABLE report generator
(Mix Analyzer v2.8.0+).

Two test classes :

1. ``TestFilterPeakTrajectoriesByThreshold`` — pure unit tests on the
   filter function in feature_storage.py. Trivial assertions, ms each.

2. ``TestGenerateShareableReport`` — combination of mock-based unit
   test for the retry algorithm + e2e integration test using the full
   pipeline (analyze_track → generate_excel_report → generate_shareable_report).

The mock unit test catches the retry logic without requiring real
.xlsx generation. The e2e test catches the threading + sheet recreation
in a real pipeline. Both are needed.
"""

import os
import sys
from unittest.mock import patch

import pytest
from openpyxl import load_workbook

sys.path.insert(0, os.path.dirname(os.path.dirname(os.path.abspath(__file__))))

from feature_storage import filter_peak_trajectories_by_threshold
from spectral_evolution import PeakTrajectory


FIXTURES_DIR = os.path.dirname(os.path.abspath(__file__))


# ===========================================================================
# Unit tests — filter_peak_trajectories_by_threshold (4 cases)
# ===========================================================================


def _make_traj(amps_db):
    """Build a PeakTrajectory with the given list of amplitudes (dB).
    All points share dummy frame_idx + freq for simplicity."""
    points = [(i, 1000.0 + i, amp) for i, amp in enumerate(amps_db)]
    return PeakTrajectory(points=points)


class TestFilterPeakTrajectoriesByThreshold:
    """Pure unit tests on filter_peak_trajectories_by_threshold."""

    def test_drops_below_threshold_keeps_above(self):
        """3 trajectories at -50, -65, -80 mean amp ; threshold -60
        keeps only the -50 one."""
        t1 = _make_traj([-50.0, -50.0, -50.0])  # mean = -50
        t2 = _make_traj([-60.0, -70.0])           # mean = -65
        t3 = _make_traj([-80.0, -80.0])           # mean = -80
        result = filter_peak_trajectories_by_threshold([t1, t2, t3], -60.0)
        assert len(result) == 1
        assert result[0] is t1

    def test_empty_input_returns_empty(self):
        assert filter_peak_trajectories_by_threshold([], -60.0) == []

    def test_threshold_minus_inf_keeps_all(self):
        """threshold -90 (lower than -80, the deepest in fixture) keeps all 3."""
        t1 = _make_traj([-50.0])
        t2 = _make_traj([-65.0])
        t3 = _make_traj([-80.0])
        result = filter_peak_trajectories_by_threshold([t1, t2, t3], -90.0)
        assert len(result) == 3
        assert result == [t1, t2, t3]  # order preserved

    def test_relative_order_preserved(self):
        """Even if input is not sorted, the filter preserves insertion order
        among kept trajectories (stability)."""
        t_a = _make_traj([-50.0])  # kept
        t_b = _make_traj([-90.0])  # dropped
        t_c = _make_traj([-55.0])  # kept
        t_d = _make_traj([-100.0])  # dropped
        t_e = _make_traj([-50.0])  # kept
        result = filter_peak_trajectories_by_threshold(
            [t_a, t_b, t_c, t_d, t_e], -70.0,
        )
        assert result == [t_a, t_c, t_e]


# ===========================================================================
# Mock-based unit test — generate_shareable_report retry algorithm
# ===========================================================================


class TestGenerateShareableReportRetry:
    """Mock-based test for the retry-on-size-too-big loop in
    generate_shareable_report. Avoids the ~15-25 sec cost of real
    audio analysis + workbook generation."""

    def test_retry_loop_picks_first_threshold_that_fits(self, tmp_path):
        """Mock the .xlsx file size : returns > 25 MB at thresholds
        -60, -55 ; ≤ 25 MB at -50. Assert the function returns -50.

        We mock at the Path.stat() level so the algorithm sees fake
        sizes without us having to write real big .xlsx files.
        """
        from mix_analyzer import generate_shareable_report
        from openpyxl import Workbook

        # Build a minimal FULL.xlsx with the 3 sheets the algorithm
        # touches : _track_peak_trajectories, _track_valley_trajectories,
        # _analysis_config. Empty content is fine — the mock controls size.
        full_xlsx = tmp_path / "full.xlsx"
        wb = Workbook()
        wb.create_sheet('_track_peak_trajectories')
        wb.create_sheet('_track_valley_trajectories')
        ws_cfg = wb.create_sheet('_analysis_config')
        # Pre-populate the 2 keys the algo modifies
        ws_cfg.cell(row=1, column=1, value='peak_threshold_db')
        ws_cfg.cell(row=1, column=2, value=-70.0)
        ws_cfg.cell(row=2, column=1, value='is_shareable_version')
        ws_cfg.cell(row=2, column=2, value=False)
        wb.save(str(full_xlsx))

        # Empty analyses_with_info — the algo iterates 0 tracks but the
        # workbook surgery still happens (delete + recreate empty sheets,
        # modify config cells). Sufficient to test the retry loop.
        analyses_with_info = []

        share_xlsx = tmp_path / "share.xlsx"

        # Mock Path.stat() to return controlled sizes :
        # - FULL.xlsx : 50 MB (> 25 target → no early return, forces retry)
        # - SHAREABLE per attempt :
        #   - Attempt 1 (threshold -60) → 30 MB (too big)
        #   - Attempt 2 (threshold -55) → 28 MB (too big)
        #   - Attempt 3 (threshold -50) → 24 MB (fits, return)
        share_sizes_mb = [30.0, 28.0, 24.0]
        size_idx = [0]
        original_stat = type(share_xlsx).stat

        def _mock_stat(self):
            class _FakeStat:
                pass
            real = original_stat(self)
            fs = _FakeStat()
            for attr in dir(real):
                if not attr.startswith('_'):
                    try:
                        setattr(fs, attr, getattr(real, attr))
                    except (AttributeError, TypeError):
                        pass
            # FULL must report as big enough that the early-return doesn't
            # fire — this test specifically exercises the retry path.
            if self.resolve() == full_xlsx.resolve():
                fs.st_size = int(50.0 * 1024 * 1024)
            elif self.resolve() == share_xlsx.resolve():
                fs.st_size = int(share_sizes_mb[
                    min(size_idx[0], len(share_sizes_mb) - 1)
                ] * 1024 * 1024)
                size_idx[0] += 1
            else:
                fs.st_size = real.st_size
            return fs

        with patch.object(type(share_xlsx), 'stat', _mock_stat):
            result_path, final_threshold = generate_shareable_report(
                full_xlsx_path=str(full_xlsx),
                analyses_with_info=analyses_with_info,
                output_path=str(share_xlsx),
                target_size_mb=25.0,
                initial_threshold_db=-60.0,
            )

        assert result_path == share_xlsx
        assert final_threshold == -50.0, (
            f"Expected to settle at -50 (3rd attempt fits 24<=25 MB), got {final_threshold}"
        )

    def test_warning_when_no_threshold_fits(self, tmp_path):
        """Mock all 5 attempts to return > 25 MB. Function must NOT crash
        but return the last attempt (-40) with a warning."""
        from mix_analyzer import generate_shareable_report
        from openpyxl import Workbook

        full_xlsx = tmp_path / "full.xlsx"
        wb = Workbook()
        wb.create_sheet('_track_peak_trajectories')
        wb.create_sheet('_track_valley_trajectories')
        ws_cfg = wb.create_sheet('_analysis_config')
        ws_cfg.cell(row=1, column=1, value='peak_threshold_db')
        ws_cfg.cell(row=1, column=2, value=-70.0)
        ws_cfg.cell(row=2, column=1, value='is_shareable_version')
        ws_cfg.cell(row=2, column=2, value=False)
        wb.save(str(full_xlsx))

        share_xlsx = tmp_path / "share.xlsx"
        original_stat = type(share_xlsx).stat

        # FULL + all 5 SHAREABLE attempts return 50 MB (> 25 MB target).
        # FULL > target prevents early return, forcing the retry path
        # all the way to exhaustion.
        def _mock_always_big(self):
            class _FakeStat:
                pass
            real = original_stat(self)
            fs = _FakeStat()
            for attr in dir(real):
                if not attr.startswith('_'):
                    try:
                        setattr(fs, attr, getattr(real, attr))
                    except (AttributeError, TypeError):
                        pass
            if self.resolve() in (full_xlsx.resolve(), share_xlsx.resolve()):
                fs.st_size = int(50.0 * 1024 * 1024)
            else:
                fs.st_size = real.st_size
            return fs

        warnings_logged = []

        def _capture_log(msg):
            if 'WARNING' in msg or 'unreachable' in msg:
                warnings_logged.append(msg)

        with patch.object(type(share_xlsx), 'stat', _mock_always_big):
            result_path, final_threshold = generate_shareable_report(
                full_xlsx_path=str(full_xlsx),
                analyses_with_info=[],
                output_path=str(share_xlsx),
                target_size_mb=25.0,
                initial_threshold_db=-60.0,
                log_fn=_capture_log,
            )

        # Last attempt should be -40 (most selective in the candidate list)
        assert final_threshold == -40.0
        # Warning must have been logged
        assert len(warnings_logged) >= 1, (
            f"Expected a warning log when target unreachable, got nothing. "
            f"All log entries: {warnings_logged}"
        )


# ===========================================================================
# E2E integration test — full pipeline (analyze + FULL + SHAREABLE)
# ===========================================================================


class TestGenerateShareableReportE2E:
    """Slow integration test (~25-35 sec) running the full pipeline :
    analyze_track → generate_excel_report (FULL) → generate_shareable_report
    (SHAREABLE). Asserts both files exist + SHAREABLE has updated
    _analysis_config + the trajectory sheets are smaller (filtering applied)."""

    def test_full_pipeline_produces_full_and_shareable_with_filtering(
        self, tmp_path,
    ):
        from mix_analyzer import (
            analyze_track,
            generate_excel_report,
            generate_shareable_report,
        )
        from resolution_presets import RESOLUTION_PRESETS

        wav = os.path.join(FIXTURES_DIR, '01_flat_wideband.wav')
        preset = RESOLUTION_PRESETS["standard"]

        # 1. Analyse audio
        analysis = analyze_track(wav, compute_tempo=False, preset=preset)
        track_info = {
            'type': 'Individual', 'category': 'test',
            'name': '01_flat_wideband.wav', 'parent_bus': 'None',
        }

        # 2. Generate FULL report
        full_path = tmp_path / "report_full.xlsx"
        generate_excel_report(
            analyses_with_info=[(analysis, track_info)],
            output_path=str(full_path),
            style_name='industrial',
            preset=preset,
            peak_threshold_db=-70.0,
        )
        assert full_path.exists()

        # 3. Generate SHAREABLE from FULL
        share_path = tmp_path / "report_shareable.xlsx"
        # Use a high target so the early-return path triggers (FULL
        # already fits) — that's enough to verify the metadata flag
        # propagation works in the happy case.
        result_path, final_threshold = generate_shareable_report(
            full_xlsx_path=str(full_path),
            analyses_with_info=[(analysis, track_info)],
            output_path=str(share_path),
            target_size_mb=100.0,  # generous target = early-return fires
            initial_threshold_db=-60.0,
        )

        # 4. Both files exist + SHAREABLE got a valid threshold
        assert result_path == share_path
        assert share_path.exists()
        # Phase F10f audit fix : with target=100 MB and a tiny test fixture,
        # the early-return path runs (FULL <= target) so the threshold
        # returned is the initial value, not a retry result.
        assert final_threshold == -60.0

        # 5. SHAREABLE _analysis_config updated correctly (is_shareable
        # flag must be flipped even on the early-return path)
        wb = load_workbook(str(share_path), read_only=True)
        ws_cfg = wb['_analysis_config']
        config = {
            ws_cfg.cell(row=r, column=1).value: ws_cfg.cell(row=r, column=2).value
            for r in range(1, ws_cfg.max_row + 1)
        }
        assert config['is_shareable_version'] is True

        # 6. SHAREABLE has the same essential sheets as FULL
        for required_sheet in [
            'Index', '_analysis_config',
            '_track_peak_trajectories', '_track_valley_trajectories',
        ]:
            assert required_sheet in wb.sheetnames, (
                f"SHAREABLE missing sheet : {required_sheet}"
            )

    def test_full_pipeline_filter_actually_reduces_xlsx_size(self, tmp_path):
        """⭐ Phase F10f audit fix — explicit assertion that the filter
        REDUCES the xlsx size (catches bugs where filter has no effect).

        Forces all retries by setting target_size_mb to a value smaller
        than the FULL — the algorithm must iterate, dropping more rows
        each attempt, and the final SHAREABLE must be strictly smaller
        than the FULL.
        """
        from mix_analyzer import (
            analyze_track,
            generate_excel_report,
            generate_shareable_report,
        )
        from resolution_presets import RESOLUTION_PRESETS

        wav = os.path.join(FIXTURES_DIR, '01_flat_wideband.wav')
        preset = RESOLUTION_PRESETS["standard"]

        analysis = analyze_track(wav, compute_tempo=False, preset=preset)
        track_info = {
            'type': 'Individual', 'category': 'test',
            'name': '01_flat_wideband.wav', 'parent_bus': 'None',
        }

        full_path = tmp_path / "report_full.xlsx"
        generate_excel_report(
            analyses_with_info=[(analysis, track_info)],
            output_path=str(full_path),
            style_name='industrial',
            preset=preset,
            peak_threshold_db=-70.0,
        )
        full_size_bytes = full_path.stat().st_size

        # Set target to 99 % of FULL size : forces at least 1 retry
        # (the algorithm can't early-return because FULL > target).
        # The filter must actually drop trajectories (the -40 threshold
        # is selective enough) so the final SHAREABLE is strictly smaller.
        target_mb_below_full = (full_size_bytes / (1024 * 1024)) * 0.99

        share_path = tmp_path / "report_shareable.xlsx"
        result_path, final_threshold = generate_shareable_report(
            full_xlsx_path=str(full_path),
            analyses_with_info=[(analysis, track_info)],
            output_path=str(share_path),
            target_size_mb=target_mb_below_full,
            initial_threshold_db=-60.0,
        )

        # The retry path executed (NOT the early return)
        assert final_threshold in [-60, -55, -50, -45, -40], (
            f"Retry path should have set final_threshold to a candidate, "
            f"got {final_threshold}"
        )

        # ⭐ Critical assertion : SHAREABLE must be strictly smaller than
        # FULL. If filter_peak_trajectories_by_threshold ever returns the
        # input list unchanged (silent regression), the SHAREABLE size
        # would equal FULL size and this assertion would fire.
        share_size_bytes = share_path.stat().st_size
        assert share_size_bytes < full_size_bytes, (
            f"SHAREABLE ({share_size_bytes} bytes) must be strictly "
            f"smaller than FULL ({full_size_bytes} bytes) — filter did "
            f"not reduce data, possible filter regression."
        )


if __name__ == '__main__':
    pytest.main([__file__, '-v'])
