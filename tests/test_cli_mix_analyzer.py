#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
Tests for Phase F10g — argparse CLI mode of mix_analyzer.py
(Mix Analyzer v2.8.0+).

Two test classes :

1. ``TestCLIArgparseValidation`` — fast subprocess tests on argparse
   behaviour (--help, --version, invalid inputs). No audio analysis
   triggered ; all complete in < 5 seconds total.

2. ``TestCLIE2E`` — slow integration tests that actually run the
   analysis pipeline end-to-end via subprocess. Each test :
   - Stages a tiny WAV fixture in a tmp input dir
   - Invokes ``python mix_analyzer.py --input-dir <tmp> ...``
   - Asserts the produced .xlsx file(s) on disk
   - ~30-60 sec per test

Pattern reused from tests/test_apply_mix_decisions_cli.py.
"""

import os
import shutil
import subprocess
import sys
from pathlib import Path

import pytest


_REPO_ROOT = Path(__file__).resolve().parent.parent
_MIX_ANALYZER_SCRIPT = _REPO_ROOT / 'mix_analyzer.py'
_FIXTURES_DIR = Path(__file__).resolve().parent
_TEST_WAV = _FIXTURES_DIR / '01_flat_wideband.wav'


def _run_cli(extra_args, timeout=180):
    """Helper : run ``python mix_analyzer.py <extra_args>`` and return
    the CompletedProcess. Sets cwd to repo root so relative imports work."""
    cmd = [sys.executable, str(_MIX_ANALYZER_SCRIPT)] + list(extra_args)
    return subprocess.run(
        cmd, capture_output=True, text=True, timeout=timeout,
        cwd=str(_REPO_ROOT),
    )


# ===========================================================================
# Argparse validation tests (fast — no audio analysis triggered)
# ===========================================================================


class TestCLIArgparseValidation:
    """Subprocess tests on argparse parsing + custom validators.
    All tests complete in < 1 sec each (no analyze_track called)."""

    def test_help_works(self):
        """--help exits 0 and prints usage with all the F10 flags."""
        result = _run_cli(['--help'], timeout=30)
        assert result.returncode == 0
        assert 'Mix Analyzer' in result.stdout
        assert '--resolution' in result.stdout
        assert '--peak-threshold' in result.stdout
        assert '--no-shareable' in result.stdout
        assert '--input-dir' in result.stdout
        # F10 preset choices are listed
        for preset in ['economy', 'standard', 'fine', 'ultra', 'maximum']:
            assert preset in result.stdout, (
                f"Preset {preset!r} missing from --help output"
            )

    def test_version_prints_v2_8_0(self):
        """--version prints the canonical VERSION."""
        result = _run_cli(['--version'], timeout=30)
        assert result.returncode == 0
        # argparse prints version to stdout
        assert 'v2.8.0' in result.stdout

    def test_invalid_resolution_rejected_by_argparse(self):
        """--resolution <invalid> exits != 0 with a clear error."""
        result = _run_cli(
            ['--input-dir', 'nonexistent_dir', '--resolution', 'invalid'],
            timeout=30,
        )
        assert result.returncode != 0
        assert (
            'invalid choice' in result.stderr.lower()
            or 'choose from' in result.stderr.lower()
        )

    def test_invalid_peak_threshold_out_of_range(self):
        """--peak-threshold +50 (positive, out of [-80, -40]) → argparse
        rejects via custom type validator."""
        result = _run_cli(
            ['--input-dir', 'nonexistent_dir', '--peak-threshold', '50'],
            timeout=30,
        )
        assert result.returncode != 0
        assert '[-80, -40]' in result.stderr or 'out of range' in result.stderr

    def test_input_dir_missing_returns_exit_1(self):
        """Valid argparse but non-existent --input-dir → exit 1."""
        result = _run_cli(
            ['--input-dir', 'definitely_nonexistent_dir_xyz'],
            timeout=30,
        )
        assert result.returncode == 1
        assert 'not found' in result.stderr.lower() or 'not a directory' in result.stderr.lower()


# ===========================================================================
# E2E integration tests (slow — real audio analysis + .xlsx generation)
# ===========================================================================


class TestCLIE2E:
    """Run the full pipeline through the CLI and verify the produced .xlsx
    files. ~30-60 sec each due to audio analysis cost."""

    def test_no_shareable_skips_shareable_xlsx(self, tmp_path):
        """--no-shareable → only FULL.xlsx is generated (no SHAREABLE)."""
        input_dir = tmp_path / "input"
        input_dir.mkdir()
        shutil.copy(str(_TEST_WAV), str(input_dir / 'test.wav'))
        output_dir = tmp_path / "output"

        result = _run_cli([
            '--input-dir', str(input_dir),
            '--output-dir', str(output_dir),
            '--no-shareable',
            '--resolution', 'standard',
        ])
        assert result.returncode == 0, (
            f"CLI failed.\nstdout:\n{result.stdout}\nstderr:\n{result.stderr}"
        )
        xlsxes = sorted(output_dir.glob('*.xlsx'))
        assert len(xlsxes) == 1, (
            f"Expected exactly 1 xlsx (FULL only, no SHAREABLE), got "
            f"{len(xlsxes)} : {[p.name for p in xlsxes]}"
        )
        # File name follows the F10g convention <project>_MixAnalyzer_full.xlsx
        assert 'full' in xlsxes[0].name.lower()
        assert 'shareable' not in xlsxes[0].name.lower()

    def test_e2e_full_pipeline_produces_full_and_shareable_with_ultra_preset(
        self, tmp_path,
    ):
        """⭐ End-to-end : CLI with ultra preset produces FULL + SHAREABLE
        .xlsx ; both contain _analysis_config sheet with preset=ultra."""
        from openpyxl import load_workbook

        input_dir = tmp_path / "input_e2e"
        input_dir.mkdir()
        shutil.copy(str(_TEST_WAV), str(input_dir / 'test.wav'))
        output_dir = tmp_path / "output_e2e"

        result = _run_cli([
            '--input-dir', str(input_dir),
            '--output-dir', str(output_dir),
            '--resolution', 'ultra',
            '--peak-threshold', '-65',
            '--shareable-target-mb', '100',  # generous target — first attempt fits
            '--shareable-initial-threshold', '-60',
        ])
        assert result.returncode == 0, (
            f"CLI failed.\nstdout:\n{result.stdout}\nstderr:\n{result.stderr}"
        )

        # 2 .xlsx files generated (FULL + SHAREABLE)
        xlsxes = sorted(output_dir.glob('*.xlsx'))
        assert len(xlsxes) == 2, (
            f"Expected 2 xlsx (FULL + SHAREABLE), got "
            f"{len(xlsxes)} : {[p.name for p in xlsxes]}"
        )
        full_path = next(p for p in xlsxes if 'full' in p.name.lower())
        share_path = next(p for p in xlsxes if 'shareable' in p.name.lower())

        # FULL has _analysis_config with preset=ultra + threshold=-65
        wb = load_workbook(str(full_path), read_only=True)
        assert '_analysis_config' in wb.sheetnames
        ws = wb['_analysis_config']
        config = {
            ws.cell(row=r, column=1).value: ws.cell(row=r, column=2).value
            for r in range(1, ws.max_row + 1)
        }
        assert config['preset_name'] == 'ultra', (
            f"FULL _analysis_config preset_name = {config['preset_name']}, "
            f"expected 'ultra'"
        )
        assert config['cqt_target_fps'] == 12   # ultra preset CQT params
        assert config['stft_n_fft'] == 16384    # ultra preset STFT params
        assert config['peak_threshold_db'] == -65.0
        assert config['is_shareable_version'] is False

        # SHAREABLE has is_shareable_version=True (early-return path since
        # target=100 MB is huge for a 1-WAV report)
        wb_share = load_workbook(str(share_path), read_only=True)
        assert '_analysis_config' in wb_share.sheetnames
        ws_share = wb_share['_analysis_config']
        config_share = {
            ws_share.cell(row=r, column=1).value:
                ws_share.cell(row=r, column=2).value
            for r in range(1, ws_share.max_row + 1)
        }
        assert config_share['is_shareable_version'] is True


if __name__ == '__main__':
    pytest.main([__file__, '-v'])
