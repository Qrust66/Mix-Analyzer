#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""Tests for the Sections Timeline sheet (Feature 3 Phase C).

Sheet was renamed from ``_sections_timeline`` to ``Sections Timeline`` and
positioned right after the Index tab in v2.6.x — see SECTIONS_TIMELINE_SHEET_NAME
in section_detector.py.
"""

from __future__ import annotations

import os
import sys
from dataclasses import dataclass, field
from typing import List, Tuple

import numpy as np
import pytest

sys.path.insert(0, os.path.abspath(os.path.join(os.path.dirname(__file__), "..")))

from openpyxl import Workbook  # noqa: E402

from section_detector import (  # noqa: E402
    SECTION_LISTING_THRESHOLD_DB,
    SECTIONS_TIMELINE_SHEET_NAME,
    Section,
    _is_track_active_in_section,
    _peak_max_per_track,
    build_sections_timeline_sheet,
    detect_accumulations_in_section,
    detect_conflicts_in_section,
    enrich_sections_with_track_stats,
    generate_observations,
    get_zone_order,
)


# ---------------------------------------------------------------------------
# Fixtures
# ---------------------------------------------------------------------------

@dataclass
class _FakePeakTraj:
    points: List[Tuple[int, float, float]] = field(default_factory=list)


def _make_section(index: int, start: int, end: int, total_energy_db: float) -> Section:
    return Section(
        index=index,
        name=f"Section {index}",
        start_bucket=start,
        end_bucket=end,
        start_seconds=float(start),
        end_seconds=float(end),
        start_beats=float(start * 2),
        end_beats=float(end * 2),
        total_energy_db=total_energy_db,
    )


def _zone_arrays_for(level_db: float, n_frames: int, zones_with_energy: List[str]) -> dict:
    """Build a {zone: array(n_frames)} dict where listed zones sit at level_db
    and every other zone is silent (-120 dB)."""
    zones = {
        z: np.full(n_frames, -120.0, dtype=float)
        for z in get_zone_order()
    }
    for z in zones_with_energy:
        zones[z] = np.full(n_frames, float(level_db), dtype=float)
    return zones


# ---------------------------------------------------------------------------
# Section enrichment + tracks_active / track_energy
# ---------------------------------------------------------------------------

def test_enrich_populates_tracks_active_and_track_energy():
    n = 30
    sections = [_make_section(1, 0, n - 1, total_energy_db=-20.0)]
    # Pad at -20 dB is above the default presence threshold (-30 dB); Pad at
    # -35 dB would be below and correctly treated as inactive now (Fix 2).
    all_tracks = {
        "Kick":      _zone_arrays_for(-10.0, n, ["sub", "low"]),
        "Pad":       _zone_arrays_for(-20.0, n, ["mid", "presence"]),
        "Silent":    _zone_arrays_for(-80.0, n, ["air"]),
    }

    enrich_sections_with_track_stats(sections, all_tracks)

    section = sections[0]
    # Kick and Pad are active (present above -30 dB in at least one zone).
    assert set(section.tracks_active) == {"Kick", "Pad"}
    # track_energy has every track and every zone
    assert set(section.track_energy.keys()) == {"Kick", "Pad", "Silent"}
    for zones in section.track_energy.values():
        assert set(zones.keys()) == set(get_zone_order())
    # sanity on values (track_energy stores nanmax now, not mean)
    assert section.track_energy["Kick"]["sub"] == pytest.approx(-10.0, abs=0.01)
    assert section.track_energy["Pad"]["mid"] == pytest.approx(-20.0, abs=0.01)
    # track_presence is populated and consistent with tracks_active
    assert section.track_presence["Kick"]["sub"] == pytest.approx(1.0, abs=0.01)
    assert section.track_presence["Silent"]["air"] == pytest.approx(0.0, abs=0.01)


# ---------------------------------------------------------------------------
# Conflict detection
# ---------------------------------------------------------------------------

def test_detect_conflicts_scores_and_severity_order():
    n = 30
    section = _make_section(1, 0, n - 1, total_energy_db=-10.0)

    # Pair 1: near-identical high energy in the same zone -> critical
    # Pair 2: both above threshold but 20 dB apart -> moderate
    # Pair 3: one track below threshold -> skipped
    all_tracks = {
        "TrackA": _zone_arrays_for(-5.0,  n, ["low"]),
        "TrackB": _zone_arrays_for(-6.0,  n, ["low"]),
        "TrackC": _zone_arrays_for(-15.0, n, ["mid"]),
        "TrackD": _zone_arrays_for(-35.0, n, ["mid"]),
        "TrackE": _zone_arrays_for(-60.0, n, ["presence"]),
    }
    enrich_sections_with_track_stats([section], all_tracks)

    conflicts = detect_conflicts_in_section(section)

    # Must surface the critical A/B conflict first.
    assert conflicts[0]["track_a"] == "TrackA"
    assert conflicts[0]["track_b"] == "TrackB"
    assert conflicts[0]["severity"] == "critical"
    assert conflicts[0]["zone"] == "low"

    # TrackC/TrackD: overlap = 1 - 20/30 = 0.333, level_weight = (-25+40)/40 = 0.375
    #               score = 0.125 -> below 0.4 => skipped
    pair_c_d = [c for c in conflicts if {c["track_a"], c["track_b"]} == {"TrackC", "TrackD"}]
    assert pair_c_d == [], "weak conflict must be skipped"

    # Pairs involving TrackE (-60 dB) must never appear.
    assert all("TrackE" not in (c["track_a"], c["track_b"]) for c in conflicts)

    # severities are in order critical-first
    severity_order = [c["severity"] for c in conflicts]
    assert severity_order == sorted(
        severity_order, key=lambda s: 0 if s == "critical" else 1
    )


# ---------------------------------------------------------------------------
# Accumulation detection
# ---------------------------------------------------------------------------

def _simultaneous_peak_traj(freqs: List[float], buckets: List[int], amp_db: float = -5.0):
    """Build a PeakTrajectory with peaks at each (bucket, freq) pair for one track."""
    points = []
    for bucket in buckets:
        for f in freqs:
            points.append((bucket, f, amp_db))
    return _FakePeakTraj(points=points)


def test_detect_accumulations_clusters_simultaneous_peaks_within_semitone():
    """5 tracks with peaks near 247 Hz, all simultaneous over buckets 10..14
    (5 buckets >= default min_duration 3) => one accumulation."""
    section = _make_section(1, 0, 99, total_energy_db=-15.0)
    freqs = [246.0, 247.5, 248.0, 249.0, 250.0]  # within 1 semitone of each other
    peaks = {}
    for i, f in enumerate(freqs):
        track = f"Track{chr(ord('A') + i)}"
        peaks[track] = [_simultaneous_peak_traj([f], buckets=list(range(10, 15)))]
    # Far-away peak for one of the tracks (unrelated freq) must not interfere.
    peaks["TrackE"].append(_simultaneous_peak_traj([4000.0], buckets=[60]))
    # Out-of-section track must not count.
    peaks["Outsider"] = [_simultaneous_peak_traj([247.0], buckets=[500])]

    accs = detect_accumulations_in_section(
        section, peaks, min_tracks_simultaneous=4, min_duration_buckets=3
    )

    assert len(accs) == 1
    assert accs[0]["n_tracks_simultaneous"] == 5
    assert accs[0]["duration_buckets"] == 5
    assert accs[0]["freq_hz"] == pytest.approx(248.0, abs=2.0)
    assert accs[0]["start_bucket"] == 10
    assert accs[0]["end_bucket"] == 14
    assert set(accs[0]["track_names"]) == {
        "TrackA", "TrackB", "TrackC", "TrackD", "TrackE",
    }


def test_detect_accumulations_default_min_tracks_is_six():
    """Production default requires >=6 simultaneous tracks."""
    section = _make_section(1, 0, 99, total_energy_db=-15.0)
    # 5 simultaneous tracks on buckets 10..14 -> below default, no accumulation
    five_tracks = {
        f"T{i}": [_simultaneous_peak_traj([247.0 + 0.1 * i], buckets=list(range(10, 15)))]
        for i in range(5)
    }
    assert detect_accumulations_in_section(section, five_tracks) == []

    # Adding a 6th simultaneous track pushes us over the threshold.
    six_tracks = dict(five_tracks)
    six_tracks["T5"] = [_simultaneous_peak_traj([247.5], buckets=list(range(10, 15)))]
    accs = detect_accumulations_in_section(section, six_tracks)
    assert len(accs) == 1
    assert accs[0]["n_tracks_simultaneous"] == 6
    assert accs[0]["duration_buckets"] == 5


def test_build_sheet_warns_on_bus_or_full_mix_tracks(caplog):
    n = 20
    sections = [_make_section(1, 0, n - 1, total_energy_db=-10.0)]
    all_tracks = {
        "Kick":             _zone_arrays_for(-5.0,  n, ["sub"]),
        "BUS Drums":        _zone_arrays_for(-5.0,  n, ["low"]),  # suspect
        "Song_HIRES_ALL":   _zone_arrays_for(-5.0,  n, ["mid"]),  # suspect
    }
    warnings: list = []
    wb = Workbook()
    build_sections_timeline_sheet(
        workbook=wb,
        sections=sections,
        all_tracks_zone_energy=all_tracks,
        log_fn=lambda msg: warnings.append(msg),
    )
    joined = "\n".join(warnings)
    assert "WARNING" in joined
    assert "BUS Drums" in joined
    assert "Song_HIRES_ALL" in joined


# ---------------------------------------------------------------------------
# Observations
# ---------------------------------------------------------------------------

def test_observations_cap_at_five_and_cover_priority():
    n = 30
    sections = [
        _make_section(1, 0, 9,  total_energy_db=-40.0),   # low energy
        _make_section(2, 10, 19, total_energy_db=-20.0),  # mid
        _make_section(3, 20, 29, total_energy_db=-5.0),   # top energy (section under test)
    ]
    # lots of active tracks in section 3 to exceed density thresholds
    all_tracks = {
        f"Track{i:02d}": _zone_arrays_for(-5.0, n, ["low"]) for i in range(25)
    }
    # ensure some tracks appear only in section 1 (silent in section 3)
    all_tracks["Absentee"] = (
        {z: np.full(n, -120.0, dtype=float) for z in get_zone_order()}
    )
    all_tracks["Absentee"]["presence"] = np.concatenate(
        [np.full(10, -5.0), np.full(20, -120.0)]
    )

    enrich_sections_with_track_stats(sections, all_tracks)
    conflicts = detect_conflicts_in_section(sections[2])
    observations = generate_observations(
        sections[2],
        conflicts=conflicts,
        accumulations=[],
        all_sections=sections,
        all_tracks_zone_energy=all_tracks,
    )

    assert len(observations) <= 5
    assert len(observations) >= 2
    # Must mention density / energy qualification
    assert any("dense" in o.lower() or "densite" in o.lower() or "tracks actives" in o.lower()
               for o in observations)


# ---------------------------------------------------------------------------
# Sheet rendering
# ---------------------------------------------------------------------------

def test_build_sheet_silent_noop_when_no_sections():
    wb = Workbook()
    build_sections_timeline_sheet(
        workbook=wb,
        sections=[],
        all_tracks_zone_energy={},
    )
    assert SECTIONS_TIMELINE_SHEET_NAME not in wb.sheetnames


def test_build_sheet_contains_master_view_and_per_section_blocks():
    n = 30
    sections = [
        _make_section(1, 0, 14,  total_energy_db=-25.0),
        _make_section(2, 15, 29, total_energy_db=-10.0),
    ]
    all_tracks = {
        "Kick":    _zone_arrays_for(-5.0,  n, ["sub", "low"]),
        "Bass":    _zone_arrays_for(-7.0,  n, ["low"]),
        "Pad":     _zone_arrays_for(-20.0, n, ["mid"]),
    }

    wb = Workbook()
    build_sections_timeline_sheet(
        workbook=wb,
        sections=sections,
        all_tracks_zone_energy=all_tracks,
    )
    assert SECTIONS_TIMELINE_SHEET_NAME in wb.sheetnames

    ws = wb[SECTIONS_TIMELINE_SHEET_NAME]
    text = "\n".join(
        " | ".join(str(c.value) for c in row if c.value is not None)
        for row in ws.iter_rows()
    )

    # master view
    assert "VUE MAITRE" in text
    assert "Conflits crit." in text
    assert "Conflits mod." in text

    # per-section blocks
    assert "SECTION 1" in text
    assert "SECTION 2" in text
    assert "TRACKS ACTIVES PAR ZONE D'ENERGIE DOMINANTE" in text
    assert "CONFLITS DE FREQUENCES" in text
    assert "ACCUMULATIONS" in text
    assert "PEAK MAX PAR TRACK" in text
    assert "OBSERVATIONS" in text

    # no formula anywhere (read-only guarantee)
    for row in ws.iter_rows():
        for cell in row:
            if cell.value is not None and isinstance(cell.value, str):
                assert not cell.value.lstrip().startswith("="), (
                    f"cell {cell.coordinate} contains a formula: {cell.value!r}"
                )


def test_build_sheet_user_visible_name():
    """Sheet name must not start with an underscore (underscore-prefixed
    sheets read as 'internal / technical' to Excel users). v2.6.x."""
    assert not SECTIONS_TIMELINE_SHEET_NAME.startswith("_")
    assert SECTIONS_TIMELINE_SHEET_NAME == "Sections Timeline"


def test_build_sheet_positioned_right_after_index():
    """The sheet must land at index 1 in the tab bar (right after Index at 0)
    so the user finds it in the top 3 visible tabs — v2.6.x user-facing move."""
    wb = Workbook()
    # Simulate a real-world workbook order: Index first, then a few global
    # sheets, then the Sections Timeline is created (late in the pipeline).
    wb.active.title = "Index"
    wb.create_sheet("Summary")
    wb.create_sheet("Dashboard")
    wb.create_sheet("Mix Health Score")

    n = 10
    sections = [_make_section(1, 0, 9, total_energy_db=-20.0)]
    all_tracks = {"Kick": _zone_arrays_for(-5.0, n, ["sub"])}
    build_sections_timeline_sheet(
        workbook=wb, sections=sections, all_tracks_zone_energy=all_tracks,
    )

    assert wb.sheetnames[0] == "Index"
    assert wb.sheetnames[1] == SECTIONS_TIMELINE_SHEET_NAME, (
        f"expected Sections Timeline at position 1, got order: {wb.sheetnames}"
    )


def test_build_sheet_replaces_existing_sheet_at_correct_position():
    """Re-running the build on a workbook that already has the sheet should
    replace it AND keep it at position 1, not append at the end."""
    wb = Workbook()
    wb.active.title = "Index"
    wb.create_sheet("Dashboard")

    n = 10
    sections = [_make_section(1, 0, 9, total_energy_db=-15.0)]
    all_tracks = {"Kick": _zone_arrays_for(-5.0, n, ["sub"])}

    build_sections_timeline_sheet(
        workbook=wb, sections=sections, all_tracks_zone_energy=all_tracks,
    )
    build_sections_timeline_sheet(  # second build — should replace cleanly
        workbook=wb, sections=sections, all_tracks_zone_energy=all_tracks,
    )

    assert wb.sheetnames.count(SECTIONS_TIMELINE_SHEET_NAME) == 1
    assert wb.sheetnames[1] == SECTIONS_TIMELINE_SHEET_NAME


# ---------------------------------------------------------------------------
# Per-section track listing via TRACK PEAK (v2.6.x, fixes empty tracks_active)
# ---------------------------------------------------------------------------

def test_peak_max_per_track_lists_all_tracks_above_audibility_threshold():
    """When TRACK PEAK data is passed in, every track with TRACK PEAK above
    -60 dB AND non-zero active_fraction must appear — even if
    section.tracks_active is empty (the bug this fixes). Below-threshold
    peaks, silent slices, and bleed-only tracks (no zone activity) are
    filtered."""
    n = 20
    section = _make_section(1, 0, n - 1, total_energy_db=-15.0)
    assert section.tracks_active == []  # sanity: intentionally empty

    # Every track listed below has zones well above the -30 dB presence
    # threshold, so active_fraction > 0. The filtering in this test
    # exercises the TRACK PEAK gate, not the active_fraction gate (a
    # separate test covers that).
    all_tracks = {
        "Kick":   _zone_arrays_for(-5.0,  n, ["sub"]),
        "Bass":   _zone_arrays_for(-10.0, n, ["low"]),
        "Pad":    _zone_arrays_for(-20.0, n, ["mid"]),
        "Atmos":  _zone_arrays_for(-70.0, n, ["air"]),   # inactive (below presence)
        "Silent": _zone_arrays_for(-120., n, []),
    }
    peak_by_section = {
        "Kick":   {1: -15.0},  # pass both gates
        "Bass":   {1: -22.0},  # pass both gates
        "Pad":    {1: -55.0},  # passes peak gate, active_fraction > 0
        "Atmos":  {1: -75.0},  # below -60 peak gate -> filtered
        "Silent": {1: None},   # None -> filtered
    }

    rows = _peak_max_per_track(
        section, all_tracks,
        all_tracks_peak_trajectories=None,
        active_threshold_db=-30.0,
        all_tracks_peak_by_section=peak_by_section,
    )

    names = [r["track"] for r in rows]
    assert names == ["Kick", "Bass", "Pad"], (
        "expected Kick/Bass/Pad (Atmos filtered by peak gate, Silent by None). "
        f"Got: {names}"
    )


def test_peak_max_per_track_sorted_by_track_peak_descending():
    """Loudest-first ordering: the mix engineer reads hero -> support -> atmos
    from top to bottom."""
    n = 10
    section = _make_section(1, 0, n - 1, total_energy_db=-15.0)
    # All tracks active (zones above -30 dB presence threshold) so the
    # active_fraction gate lets them all through; the test focuses on order.
    all_tracks = {
        "Quiet": _zone_arrays_for(-25.0, n, ["mid"]),
        "Loud":  _zone_arrays_for(-5.0,  n, ["low"]),
        "Mid":   _zone_arrays_for(-15.0, n, ["body"]),
    }
    peak_by_section = {
        "Quiet": {1: -45.0},
        "Loud":  {1: -10.0},
        "Mid":   {1: -25.0},
    }
    rows = _peak_max_per_track(
        section, all_tracks,
        all_tracks_peak_trajectories=None,
        active_threshold_db=-30.0,
        all_tracks_peak_by_section=peak_by_section,
    )
    assert [r["track"] for r in rows] == ["Loud", "Mid", "Quiet"]


def test_peak_max_per_track_legacy_fallback_when_no_track_peak_data():
    """Backwards compatibility: without all_tracks_peak_by_section we must
    still work (iterate section.tracks_active, sort by AMP ZONE)."""
    n = 10
    section = _make_section(1, 0, n - 1, total_energy_db=-10.0)
    all_tracks = {
        "A": _zone_arrays_for(-5.0, n, ["low"]),
        "B": _zone_arrays_for(-10.0, n, ["mid"]),
    }
    enrich_sections_with_track_stats([section], all_tracks)

    rows = _peak_max_per_track(
        section, all_tracks,
        all_tracks_peak_trajectories=None,
        active_threshold_db=-30.0,
        all_tracks_peak_by_section=None,  # legacy path
    )
    names = {r["track"] for r in rows}
    # Both tracks are active (above -30 dB presence threshold)
    assert names == {"A", "B"}
    # No track_peak_db column when no data was provided
    for r in rows:
        assert r["track_peak_db"] is None


def test_section_listing_threshold_constant_is_minus_sixty_db():
    """Pin the audibility threshold: -60 dB is the standard mix-audibility
    floor. A future tweak should be deliberate — document it as a test."""
    assert SECTION_LISTING_THRESHOLD_DB == -60.0


def test_peak_max_per_track_filters_bleed_with_zero_active_fraction():
    """Second gate: a track with TRACK PEAK above -60 dB but no zone above
    the presence threshold (zero active frames) is filtered — it's bleed,
    reverb tail, or plugin noise floor, not a true contributor to the
    section. Regression guard for Acid Drops Tambourine Hi-Hat which
    showed up in Chorus 1/2 at -20 dB (reverb tail) with 0% activity."""
    n = 20
    section = _make_section(1, 0, n - 1, total_energy_db=-15.0)
    # Kick: genuinely active (loud, above -30 dB presence threshold)
    # Tambourine bleed: TRACK PEAK above -60 but zone_energy never above -30
    # in any zone (simulates a reverb tail audible on the WAV but with no
    # sustained spectral signature).
    all_tracks = {
        "Kick":      _zone_arrays_for(-5.0,  n, ["sub"]),
        "TambourineBleed": _zone_arrays_for(-45.0, n, ["air"]),  # below -30
    }
    peak_by_section = {
        "Kick":              {1: -10.0},
        "TambourineBleed":   {1: -20.0},  # above -60, would pass first gate
    }
    rows = _peak_max_per_track(
        section, all_tracks,
        all_tracks_peak_trajectories=None,
        active_threshold_db=-30.0,
        all_tracks_peak_by_section=peak_by_section,
    )
    names = [r["track"] for r in rows]
    assert names == ["Kick"], (
        f"expected TambourineBleed filtered (0 active frames), got {names}"
    )


def test_peak_max_per_track_keeps_track_with_brief_activity():
    """A track that's active only briefly (e.g. 1% of the section, like
    Acid Drops Spoon Percussion in Build 1) must still be listed — any
    non-zero active_fraction passes the second gate."""
    n = 100
    section = _make_section(1, 0, n - 1, total_energy_db=-15.0)
    # Build zone_arrays where "BriefHit" has 1 loud frame (above -30 dB)
    # and 99 silent frames — active_fraction should be 0.01 (non-zero)
    brief_zones = {z: np.full(n, -100.0) for z in get_zone_order()}
    brief_zones["sub"][50] = -10.0  # 1 frame well above -30 dB threshold
    all_tracks = {
        "Loud":     _zone_arrays_for(-5.0, n, ["low"]),
        "BriefHit": brief_zones,
    }
    peak_by_section = {
        "Loud":     {1: -10.0},
        "BriefHit": {1: -15.0},
    }
    rows = _peak_max_per_track(
        section, all_tracks,
        all_tracks_peak_trajectories=None,
        active_threshold_db=-30.0,
        all_tracks_peak_by_section=peak_by_section,
    )
    names = [r["track"] for r in rows]
    assert "BriefHit" in names, (
        f"a 1-frame-active track must be listed, got {names}"
    )


# ---------------------------------------------------------------------------
# Conflict severity helper — threshold boundaries
# ---------------------------------------------------------------------------

def test_conflict_severity_boundaries_match_spec():
    # Minimum level for a "critical" pair under our formula:
    # overlap=1 (equal energies) + level_weight=(avg+40)/40
    # score=1*level_weight; we need score>=0.7 => avg_db >= -12
    n = 10
    section = _make_section(1, 0, n - 1, total_energy_db=0.0)

    critical_tracks = {
        "A": _zone_arrays_for(-10.0, n, ["low"]),
        "B": _zone_arrays_for(-10.0, n, ["low"]),
    }
    enrich_sections_with_track_stats([section], critical_tracks)
    confs = detect_conflicts_in_section(section)
    assert confs and confs[0]["severity"] == "critical"
    assert confs[0]["score"] >= 0.7 - 1e-6


# ---------------------------------------------------------------------------
# Fix 2 — presence-ratio activity gate (NaN-friendly)
# ---------------------------------------------------------------------------

def test_fix2_masked_track_still_active():
    """Track with 50% of buckets masked (NaN) and 50% at -20 dB must stay active.

    Before Fix 2 the mean would collapse toward -120 dB and the track would
    drop out; the presence-ratio gate only counts non-NaN buckets.
    """
    n_zones = 9
    n_buckets = 20
    track_ze = {z: np.full(n_buckets, np.nan) for z in get_zone_order()}
    # 10/20 = 50% presence in Low at -20 dB (above default -30 dB threshold)
    track_ze["low"] = np.concatenate([np.full(10, -20.0), np.full(10, np.nan)])

    assert _is_track_active_in_section(track_ze) is True


def test_fix2_track_all_masked_inactive():
    """An entirely-NaN track is inactive."""
    track_ze = {z: np.full(20, np.nan) for z in get_zone_order()}
    assert _is_track_active_in_section(track_ze) is False


def test_fix2_track_low_energy_inactive():
    """Track with finite energy below threshold everywhere is inactive."""
    track_ze = {z: np.full(20, -50.0) for z in get_zone_order()}
    assert _is_track_active_in_section(track_ze) is False


def test_fix2_tracks_active_and_accumulations_coherent():
    """After Fix 2, the tracks_active count must not be wildly smaller than the
    number of tracks contributing accumulations. Previously Acid Drops showed
    3 active tracks side by side with 32-track accumulations.
    """
    n = 30
    sections = [_make_section(1, 0, n - 1, total_energy_db=-10.0)]
    # Ten tracks, each playing audibly in one zone or another, each with
    # 30% NaN — enough masking that the OLD mean-based rule would reject
    # most of them. Fix 2 must keep them all active.
    all_tracks = {}
    rng = np.random.default_rng(0)
    zone_targets = get_zone_order()
    for i in range(10):
        zones = {z: np.full(n, np.nan) for z in zone_targets}
        # 21/30 = 70% presence at a level clearly above the -30 dB threshold.
        audible = np.full(21, -10.0) + rng.normal(0, 0.5, size=21)
        zones[zone_targets[i % len(zone_targets)]] = np.concatenate(
            [audible, np.full(9, np.nan)]
        )
        all_tracks[f"T{i:02d}"] = zones

    enrich_sections_with_track_stats(sections, all_tracks)
    assert len(sections[0].tracks_active) == 10, (
        f"expected all 10 half-masked tracks active, got "
        f"{sections[0].tracks_active}"
    )


def test_fix2_presence_params_exposed():
    """presence_threshold_db and min_presence_ratio must be callable overrides."""
    n = 30
    sections = [_make_section(1, 0, n - 1, total_energy_db=-20.0)]
    all_tracks = {
        "Quiet": _zone_arrays_for(-35.0, n, ["mid"]),
    }
    # With defaults (-30 dB, 20% ratio): Quiet at -35 dB is not present.
    enrich_sections_with_track_stats(sections, all_tracks)
    assert sections[0].tracks_active == []
    # Override to a lower threshold: Quiet now counts as present.
    enrich_sections_with_track_stats(
        sections, all_tracks, presence_threshold_db=-40.0
    )
    assert sections[0].tracks_active == ["Quiet"]


# ---------------------------------------------------------------------------
# Fix 3 — accumulations require temporal simultaneity over a minimum duration
# ---------------------------------------------------------------------------

def test_fix3_accumulation_requires_simultaneity():
    """5 tracks with peaks at 370 Hz but at DIFFERENT buckets -> no accumulation.

    Before Fix 3, the old algo counted each track as contributing whenever
    any of its peaks fell within the section, producing impossible counts
    (Acid Drops: 32 tracks at 370 Hz).
    """
    section = _make_section(1, 0, 99, total_energy_db=-15.0)
    # Each track has exactly one peak at 370 Hz, at its own dedicated bucket,
    # with no overlap between tracks.
    peaks = {
        f"Track{chr(ord('A') + i)}": [
            _FakePeakTraj(points=[(10 + 10 * i, 370.0, -5.0)])
        ]
        for i in range(5)
    }
    accs = detect_accumulations_in_section(
        section, peaks, min_tracks_simultaneous=4, min_duration_buckets=3
    )
    assert accs == [], (
        f"no simultaneous bucket has >=4 tracks; expected empty, got {accs}"
    )


def test_fix3_accumulation_with_simultaneity():
    """5 tracks with peaks at 370 Hz, all simultaneous over buckets 20..24,
    yields one accumulation with duration 5."""
    section = _make_section(1, 0, 99, total_energy_db=-5.0)
    peaks = {}
    for i in range(5):
        peaks[f"T{i}"] = [_simultaneous_peak_traj([370.0], buckets=list(range(20, 25)))]

    accs = detect_accumulations_in_section(
        section, peaks, min_tracks_simultaneous=5, min_duration_buckets=3
    )
    assert len(accs) == 1
    acc = accs[0]
    assert acc["n_tracks_simultaneous"] == 5
    assert acc["duration_buckets"] == 5
    assert acc["freq_hz"] == pytest.approx(370.0, abs=0.1)
    assert acc["start_bucket"] == 20
    assert acc["end_bucket"] == 24
    assert set(acc["track_names"]) == {f"T{i}" for i in range(5)}


def test_fix3_accumulation_duration_minimum():
    """Runs shorter than ``min_duration_buckets`` are rejected; exactly-minimum
    runs are kept."""
    section = _make_section(1, 0, 99, total_energy_db=-5.0)

    # Case A: 6 tracks simultaneous for 2 buckets only -> reject (below default 3).
    peaks_short = {
        f"T{i}": [_simultaneous_peak_traj([370.0], buckets=[20, 21])]
        for i in range(6)
    }
    assert detect_accumulations_in_section(section, peaks_short) == []

    # Case B: 6 tracks simultaneous for exactly 3 buckets -> keep.
    peaks_exact = {
        f"T{i}": [_simultaneous_peak_traj([370.0], buckets=[20, 21, 22])]
        for i in range(6)
    }
    accs_exact = detect_accumulations_in_section(section, peaks_exact)
    assert len(accs_exact) == 1
    assert accs_exact[0]["duration_buckets"] == 3


def test_fix3_reports_max_simult_when_count_varies_in_run():
    """When the simultaneous count oscillates inside a run, n_tracks_simultaneous
    reports the MAX count seen in the run."""
    section = _make_section(1, 0, 99, total_energy_db=-5.0)
    # 6 baseline tracks for all 5 buckets, plus 2 extra tracks joining at bucket 22.
    peaks = {
        f"T{i}": [_simultaneous_peak_traj([370.0], buckets=list(range(20, 25)))]
        for i in range(6)
    }
    peaks["T6"] = [_simultaneous_peak_traj([370.0], buckets=[22])]
    peaks["T7"] = [_simultaneous_peak_traj([370.0], buckets=[22])]

    accs = detect_accumulations_in_section(section, peaks)
    assert len(accs) == 1
    assert accs[0]["n_tracks_simultaneous"] == 8  # 6 baseline + 2 at bucket 22
    assert accs[0]["duration_buckets"] == 5
    # The union includes the two bucket-22 tracks even though they only play once.
    assert set(accs[0]["track_names"]) >= {f"T{i}" for i in range(8)}


def test_fix3_amplitude_filter_honoured():
    """Peaks below min_amplitude_db must not contribute to accumulations."""
    section = _make_section(1, 0, 99, total_energy_db=-5.0)
    # 6 tracks but half the peaks are quieter than -25 dB default; only
    # the loud peaks contribute.
    peaks = {}
    for i in range(6):
        amp = -5.0 if i < 3 else -40.0
        peaks[f"T{i}"] = [
            _simultaneous_peak_traj([370.0], buckets=list(range(20, 25)), amp_db=amp)
        ]
    # Only 3 loud tracks -> below min 6 -> no accumulation.
    assert detect_accumulations_in_section(section, peaks) == []
