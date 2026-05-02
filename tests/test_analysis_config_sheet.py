#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
Tests for the new ``_analysis_config`` sheet (Phase F10e, Mix Analyzer v2.8.0).

Two test classes :

1. ``TestAnalysisConfigSheetUnit`` — fast unit tests on the
   ``_build_analysis_config_sheet`` function in isolation. Build a
   workbook in memory, call the writer, assert on cells. Validates
   the schema (14 keys), the values matching the preset, and the
   ``is_shareable`` toggle.

2. ``TestAnalysisConfigSheetE2E`` — integration test that runs the
   full pipeline ``analyze_track + generate_excel_report`` and asserts
   that ``_analysis_config`` is present in the resulting .xlsx with
   the right values. Validates the threading (preset reaches the new
   sheet writer through the orchestrator).

The unit tests run in milliseconds. The e2e test takes ~15-25 sec
(real audio analysis on a fixture WAV). Both are needed : unit tests
catch schema bugs, e2e catches threading bugs.
"""

import os
import sys

import pytest
from openpyxl import Workbook, load_workbook

# Ensure repo root is on sys.path so we can import mix_analyzer
sys.path.insert(0, os.path.dirname(os.path.dirname(os.path.abspath(__file__))))

from mix_analyzer import (
    VERSION,
    _build_analysis_config_sheet,
    analyze_track,
    generate_excel_report,
)
from resolution_presets import RESOLUTION_PRESETS


FIXTURES_DIR = os.path.dirname(os.path.abspath(__file__))


# ===========================================================================
# Unit tests — schema + values, no audio analysis
# ===========================================================================


class TestAnalysisConfigSheetUnit:
    """Fast schema/values tests on _build_analysis_config_sheet in isolation."""

    @staticmethod
    def _build_and_read(preset, peak_threshold_db=-70.0, sample_rate=44100,
                        is_shareable=False):
        """Helper : build the sheet in a fresh in-memory workbook and
        return ``{key: value}`` dict from the 2-column layout."""
        wb = Workbook()
        # Default sheet exists ; the new sheet is added next to it.
        _build_analysis_config_sheet(
            wb, preset=preset, peak_threshold_db=peak_threshold_db,
            sample_rate=sample_rate, is_shareable=is_shareable,
        )
        ws = wb['_analysis_config']
        return {
            ws.cell(row=r, column=1).value: ws.cell(row=r, column=2).value
            for r in range(1, ws.max_row + 1)
        }

    def test_sheet_present_with_14_keys(self):
        """The sheet exists in the workbook + has exactly 14 key/value
        rows + sheet is hidden + correct sheet name."""
        wb = Workbook()
        _build_analysis_config_sheet(
            wb, preset=RESOLUTION_PRESETS["standard"],
            peak_threshold_db=-70.0, sample_rate=44100,
        )
        assert '_analysis_config' in wb.sheetnames
        ws = wb['_analysis_config']
        assert ws.sheet_state == 'hidden', (
            "_analysis_config must be hidden (cohérence avec sheets metadata)"
        )
        assert ws.max_row == 14, (
            f"Expected 14 key/value rows per spec v1.3 §5.6, got {ws.max_row}"
        )

    def test_values_match_standard_preset(self):
        """At preset=standard (v2.7.0 backward compat), the 14 values
        must reflect the standard preset's parameters exactly."""
        config = self._build_and_read(RESOLUTION_PRESETS["standard"])
        assert config['preset_name'] == 'standard'
        assert config['stft_n_fft'] == 8192
        assert config['stft_hop_samples'] == 2048   # = 8192 // 4
        assert config['cqt_target_fps'] == 6
        assert config['cqt_bins_per_octave'] == 24
        assert config['cqt_n_bins'] == 256
        assert config['sample_rate'] == 44100
        assert config['peak_threshold_db'] == -70.0
        assert config['is_shareable_version'] is False
        assert config['mix_analyzer_version'] == f'v{VERSION}'
        # generated_at is an ISO timestamp ; just verify it's non-empty string
        assert isinstance(config['generated_at'], str)
        assert len(config['generated_at']) > 0

    def test_values_match_ultra_preset(self):
        """At preset=ultra, the 14 values must reflect ultra's
        parameters (n_fft=16384, fps=12, bins/oct=36)."""
        config = self._build_and_read(RESOLUTION_PRESETS["ultra"])
        assert config['preset_name'] == 'ultra'
        assert config['stft_n_fft'] == 16384
        assert config['stft_hop_samples'] == 4096   # = 16384 // 4
        assert config['cqt_target_fps'] == 12
        assert config['cqt_bins_per_octave'] == 36
        assert config['cqt_n_bins'] == 384

    def test_is_shareable_toggle(self):
        """is_shareable=True (used by F10f for the SHAREABLE report)
        must flip the corresponding key/value pair."""
        config_full = self._build_and_read(
            RESOLUTION_PRESETS["standard"], is_shareable=False,
        )
        config_share = self._build_and_read(
            RESOLUTION_PRESETS["standard"], is_shareable=True,
        )
        assert config_full['is_shareable_version'] is False
        assert config_share['is_shareable_version'] is True

    def test_default_preset_none_resolves_to_standard(self):
        """preset=None must resolve to RESOLUTION_PRESETS["standard"]
        (v2.7.0 backward compat — same default as analyze_track)."""
        config_none = self._build_and_read(preset=None)
        config_std = self._build_and_read(RESOLUTION_PRESETS["standard"])
        # All keys (except generated_at which is timestamp-dependent) must match
        assert config_none['preset_name'] == config_std['preset_name'] == 'standard'
        assert config_none['stft_n_fft'] == config_std['stft_n_fft']
        assert config_none['cqt_target_fps'] == config_std['cqt_target_fps']


# ===========================================================================
# E2E integration test — full pipeline through generate_excel_report
# ===========================================================================


class TestAnalysisConfigSheetE2E:
    """Slow integration test (~15-25 sec) : runs analyze_track on a real
    WAV fixture, then generate_excel_report, then asserts the
    _analysis_config sheet appears in the .xlsx with correct values.

    This test catches THREADING bugs (preset not propagated from
    analyze_track to generate_excel_report to _build_analysis_config_sheet)
    that the unit tests cannot see.
    """

    def test_full_pipeline_writes_analysis_config_sheet_with_ultra_preset(
        self, tmp_path,
    ):
        wav = os.path.join(FIXTURES_DIR, '01_flat_wideband.wav')
        preset = RESOLUTION_PRESETS["ultra"]

        # 1. Run audio analysis with ultra preset
        analysis = analyze_track(wav, compute_tempo=False, preset=preset)
        track_info = {
            'type': 'Individual', 'category': 'test',
            'name': '01_flat_wideband.wav', 'parent_bus': 'None',
        }

        # 2. Generate the Excel report with ultra preset + custom threshold
        output = tmp_path / "report_ultra.xlsx"
        generate_excel_report(
            analyses_with_info=[(analysis, track_info)],
            output_path=str(output),
            style_name='industrial',
            preset=preset,
            peak_threshold_db=-65.0,
        )

        # 3. Open the .xlsx and assert _analysis_config is present + valid
        assert output.exists()
        wb = load_workbook(str(output), read_only=True)
        assert '_analysis_config' in wb.sheetnames
        ws = wb['_analysis_config']
        config = {
            ws.cell(row=r, column=1).value: ws.cell(row=r, column=2).value
            for r in range(1, ws.max_row + 1)
        }
        # Threading worked : preset reached the sheet writer
        assert config['preset_name'] == 'ultra'
        assert config['cqt_target_fps'] == 12
        assert config['stft_n_fft'] == 16384
        assert config['peak_threshold_db'] == -65.0
        # Sample rate detected from the analysis dict (not hardcoded 44100)
        assert config['sample_rate'] == analysis['sample_rate']
        # Version stamp present
        assert config['mix_analyzer_version'] == f'v{VERSION}'

        # 4. Bonus : Index sheet has the new Build Info block
        ws_index = wb['Index']
        index_rows = [
            (ws_index.cell(row=r, column=1).value,
             ws_index.cell(row=r, column=2).value)
            for r in range(1, min(20, ws_index.max_row + 1))
        ]
        # Find the Build Info "Preset used" row
        preset_row = next(
            (val for key, val in index_rows if key == 'Preset used'), None,
        )
        assert preset_row == 'ultra', (
            f"Index Build Info missing or wrong preset : got {preset_row!r}"
        )


if __name__ == '__main__':
    pytest.main([__file__, '-v'])
