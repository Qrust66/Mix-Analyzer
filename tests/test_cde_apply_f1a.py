#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""Tests for Feature 1 phase F1a — CDE diagnostic → clustering data prep.

F1a covers the **read-only** half of Feature 1: group CDE
``CorrectionDiagnostic`` instances by target track, cluster them by
frequency, match peak trajectories to the clusters, and parse the
``_track_peak_trajectories`` sheet of a Mix Analyzer Excel report.

None of these steps mutate the ``.als`` file or create any EQ8 —
F1b/F1c ship the write half.
"""

from __future__ import annotations

import os
import sys
from datetime import datetime
from pathlib import Path

import pytest

sys.path.insert(0, os.path.abspath(os.path.join(os.path.dirname(__file__), "..")))

from cde_engine import (  # noqa: E402
    CDE_VERSION,
    CorrectionDiagnostic,
    CorrectionRecipe,
    ProblemMeasurement,
    SectionContext,
    TFPContext,
)
from spectral_evolution import PeakTrajectory  # noqa: E402
from tfp_parser import Function, Importance  # noqa: E402

from cde_apply import (  # noqa: E402
    DEFAULT_FREQ_CLUSTER_TOLERANCE_SEMITONES,
    MAX_CDE_BANDS,
    CdeApplicationReport,
    FreqCluster,
    _cluster_diagnostics_by_frequency,
    _group_diagnostics_by_target,
    _match_peak_trajectories_to_cluster,
    load_peak_trajectories_from_excel,
)


# ---------------------------------------------------------------------------
# Fixtures — build synthetic diagnostics without running the detector
# ---------------------------------------------------------------------------

def _diag(
    approach: str,
    target_track: str,
    frequency_hz: float,
    gain_db: float = -3.0,
    q: float = 4.0,
    applies_to_sections=None,
    issue_type: str = "masking_conflict",
    diag_id: str = None,
    severity: str = "moderate",
    primary_none: bool = False,
) -> CorrectionDiagnostic:
    """Build a minimal diagnostic suitable for F1a tests."""
    applies_to_sections = list(applies_to_sections or ["Drop 1"])
    diag_id = diag_id or f"TEST_{target_track}_{int(frequency_hz)}HZ"

    measurement = ProblemMeasurement(
        frequency_hz=frequency_hz,
        peak_db=-5.0,
        duration_in_section_s=1.0,
        duration_ratio_in_section=0.5,
        is_audible_fraction=0.5,
        severity_score=0.5,
        masking_score=0.5,
    )
    tfp = TFPContext(
        track_a_role=(Importance.H, Function.R),
        track_b_role=(Importance.S, Function.H),
        role_compatibility="dominant_support",
    )
    sec_ctx = SectionContext(
        section_name=applies_to_sections[0],
        section_duration_s=10.0,
        tracks_active_count=4,
        conflicts_in_section=1,
        coherence_score=None,
    )
    primary = None if primary_none else CorrectionRecipe(
        target_track=target_track,
        device="EQ8 — Peak Resonance",
        approach=approach,
        parameters={
            "frequency_hz": float(frequency_hz),
            "gain_db": float(gain_db),
            "q": float(q),
            "active_in_sections": applies_to_sections,
        },
        applies_to_sections=applies_to_sections,
        rationale="test",
        confidence="medium",
    )
    return CorrectionDiagnostic(
        diagnostic_id=diag_id,
        timestamp=datetime(2026, 4, 22, 0, 0, 0),
        cde_version=CDE_VERSION,
        track_a=target_track,
        track_b=None,
        section=applies_to_sections[0],
        issue_type=issue_type,
        severity=severity,
        measurement=measurement,
        tfp_context=tfp,
        section_context=sec_ctx,
        diagnosis_text="",
        primary_correction=primary,
        fallback_correction=None,
    )


# ---------------------------------------------------------------------------
# Dataclass defaults
# ---------------------------------------------------------------------------

def test_cde_application_report_starts_empty():
    r = CdeApplicationReport()
    assert r.applied == []
    assert r.skipped == []
    assert r.warnings == []
    assert r.devices_created == {}
    assert r.envelopes_written == 0
    assert r.backup_path is None
    assert r.sidechain_count() == 0


def test_cde_application_report_sidechain_count_recognises_both_forms():
    """``sidechain_count`` must count skip reasons that mention either
    'Kickstart 2' or 'sidechain' (case-insensitive)."""
    r = CdeApplicationReport()
    r.skipped.append(("A", "requires Kickstart 2 — scope Feature 1.5"))
    r.skipped.append(("B", "sidechain approach deferred"))
    r.skipped.append(("C", "primary_correction is None"))
    assert r.sidechain_count() == 2


# ---------------------------------------------------------------------------
# _group_diagnostics_by_target
# ---------------------------------------------------------------------------

def test_group_by_target_collects_cut_approaches_per_track():
    diags = [
        _diag("static_dip",      "Kick",   62,  diag_id="D1"),
        _diag("musical_dip",     "Kick",   150, diag_id="D2"),
        _diag("reciprocal_cuts", "Snare",  500, diag_id="D3"),
        _diag("static_dip",      "Snare",  2000, diag_id="D4"),
    ]
    grouped = _group_diagnostics_by_target(diags)
    assert set(grouped.keys()) == {"Kick", "Snare"}
    assert [d.diagnostic_id for d in grouped["Kick"]] == ["D1", "D2"]
    assert [d.diagnostic_id for d in grouped["Snare"]] == ["D3", "D4"]


def test_group_by_target_filters_sidechain_with_kickstart2_reason():
    diags = [
        _diag("static_dip", "Kick",    62,  diag_id="D1"),
        _diag("sidechain",  "Bass",    60,  diag_id="D_SIDECHAIN"),
    ]
    report = CdeApplicationReport()
    grouped = _group_diagnostics_by_target(diags, report=report)
    assert "Bass" not in grouped
    assert "Kick" in grouped
    # The sidechain skip reason mentions Kickstart 2 explicitly (Feature 1.5).
    assert len(report.skipped) == 1
    diag_id, reason = report.skipped[0]
    assert diag_id == "D_SIDECHAIN"
    assert "Kickstart 2" in reason


def test_group_by_target_filters_primary_none_with_reason():
    """R2-skipped diagnostics (primary=None) are excluded with a
    dedicated skip reason, so the report can count them separately."""
    diags = [
        _diag("static_dip", "Kick",  62,  diag_id="D_OK"),
        _diag("static_dip", "Kick2", 60,  diag_id="D_R2_SKIP", primary_none=True),
    ]
    report = CdeApplicationReport()
    grouped = _group_diagnostics_by_target(diags, report=report)
    assert list(grouped.keys()) == ["Kick"]
    assert report.skipped == [("D_R2_SKIP", "primary_correction is None")]


# ---------------------------------------------------------------------------
# _cluster_diagnostics_by_frequency
# ---------------------------------------------------------------------------

def test_cluster_by_frequency_single_cluster_within_tolerance():
    """3 diagnostics at 244 / 247 / 249 Hz — within 2 semitones → 1 cluster."""
    diags = [
        _diag("static_dip", "T", 244, gain_db=-3.0, q=4.0, diag_id="A"),
        _diag("static_dip", "T", 247, gain_db=-6.0, q=4.0, diag_id="B"),
        _diag("static_dip", "T", 249, gain_db=-4.0, q=3.0, diag_id="C"),
    ]
    clusters = _cluster_diagnostics_by_frequency(diags)
    assert len(clusters) == 1
    c = clusters[0]
    # Median of [244, 247, 249] = 247
    assert c.centroid_hz == pytest.approx(247, abs=0.01)
    assert sorted(c.member_ids) == ["A", "B", "C"]
    # Severest (most negative) gain is -6 from diag B.
    assert c.severest_gain_db == pytest.approx(-6.0)
    # Median Q of [4, 4, 3] = 4
    assert c.median_q == pytest.approx(4.0)


def test_cluster_by_frequency_splits_distant_bands():
    """Three well-separated freqs → three clusters."""
    diags = [
        _diag("static_dip", "T", 62,   diag_id="LOW"),
        _diag("static_dip", "T", 500,  diag_id="MID"),
        _diag("static_dip", "T", 4000, diag_id="HIGH"),
    ]
    clusters = _cluster_diagnostics_by_frequency(diags)
    assert len(clusters) == 3
    centroids = sorted(c.centroid_hz for c in clusters)
    assert centroids == pytest.approx([62, 500, 4000])


def test_cluster_by_frequency_boundary_at_exactly_tolerance():
    """Freqs exactly ``tolerance_semitones`` apart are accepted in the
    same cluster (``<=`` comparison, not ``<``)."""
    # 247 Hz vs 247 * 2^(2/12) = 277.27 Hz → exactly 2 semitones apart.
    f2 = 247.0 * 2.0 ** (2.0 / 12.0)
    diags = [
        _diag("static_dip", "T", 247.0, diag_id="A"),
        _diag("static_dip", "T", f2,    diag_id="B"),
    ]
    clusters = _cluster_diagnostics_by_frequency(
        diags, tolerance_semitones=2.0,
    )
    assert len(clusters) == 1


def test_cluster_by_frequency_union_of_applies_to_sections():
    """The cluster's ``applies_to_sections`` is the sorted union of
    every member's list."""
    diags = [
        _diag("static_dip", "T", 247, applies_to_sections=["Drop 1", "Drop 2"],
              diag_id="A"),
        _diag("static_dip", "T", 247, applies_to_sections=["Build 2"],
              diag_id="B"),
        _diag("static_dip", "T", 248, applies_to_sections=["Drop 2"],
              diag_id="C"),
    ]
    clusters = _cluster_diagnostics_by_frequency(diags)
    assert len(clusters) == 1
    assert clusters[0].applies_to_sections == ["Build 2", "Drop 1", "Drop 2"]


def test_cluster_by_frequency_empty_input_returns_empty_list():
    assert _cluster_diagnostics_by_frequency([]) == []


def test_cluster_by_frequency_drops_diagnostics_with_no_frequency():
    """Diagnostics whose ``measurement.frequency_hz`` is ``None`` or
    non-positive cannot ride an EQ8 band and are silently dropped."""
    d_ok = _diag("static_dip", "T", 247, diag_id="OK")
    # Build a diagnostic with freq=0 (invalid)
    d_bad = _diag("static_dip", "T", 247, diag_id="BAD")
    d_bad.measurement.frequency_hz = 0.0
    clusters = _cluster_diagnostics_by_frequency([d_ok, d_bad])
    assert len(clusters) == 1
    assert clusters[0].member_ids == ["OK"]


# ---------------------------------------------------------------------------
# _match_peak_trajectories_to_cluster
# ---------------------------------------------------------------------------

def test_match_peak_trajectories_keeps_trajectories_within_tolerance():
    cluster = FreqCluster(
        centroid_hz=247.0, diagnostics=[],
        severest_gain_db=-6.0, median_q=4.0, applies_to_sections=[],
    )
    # 2 trajectories around 247 Hz, 1 far away (500 Hz)
    trajs = [
        PeakTrajectory(points=[(0, 246.8, -5.0), (1, 247.0, -4.0)]),
        PeakTrajectory(points=[(0, 249.5, -6.0), (1, 250.0, -5.0)]),
        PeakTrajectory(points=[(0, 500.0, -3.0), (1, 500.5, -2.0)]),
    ]
    matched = _match_peak_trajectories_to_cluster(
        trajs, cluster, tolerance_semitones=2.0,
    )
    assert len(matched) == 2
    # The far trajectory is filtered out.
    means = sorted(t.mean_freq for t in matched)
    assert means[0] == pytest.approx(246.9, abs=0.01)
    assert means[1] == pytest.approx(249.75, abs=0.01)


def test_match_peak_trajectories_returns_empty_when_no_overlap():
    cluster = FreqCluster(
        centroid_hz=62.0, diagnostics=[],
        severest_gain_db=-3.0, median_q=3.0, applies_to_sections=[],
    )
    trajs = [
        PeakTrajectory(points=[(0, 5000.0, -5.0)]),
        PeakTrajectory(points=[(0, 8000.0, -4.0)]),
    ]
    matched = _match_peak_trajectories_to_cluster(trajs, cluster)
    assert matched == []


# ---------------------------------------------------------------------------
# load_peak_trajectories_from_excel — round-trip on a minimal fixture
# ---------------------------------------------------------------------------

def test_load_peak_trajectories_from_excel_parses_fixture(tmp_path: Path):
    """Build a minimal ``.xlsx`` in memory with the expected sheet and
    header, then verify the reader reconstructs the trajectories."""
    from openpyxl import Workbook

    wb = Workbook()
    wb.remove(wb.active)
    ws = wb.create_sheet("_track_peak_trajectories")
    ws.append(("Track", "Traj#", "Frame", "Time (s)", "Freq (Hz)", "Amp (dB)"))
    # Track A — two trajectories
    ws.append(("A.wav", 1, 0,  0.0,  100.0, -5.0))
    ws.append(("A.wav", 1, 1,  0.1,  100.5, -4.5))
    ws.append(("A.wav", 2, 0,  0.0,  247.0, -10.0))
    ws.append(("A.wav", 2, 1,  0.1,  247.5, -9.5))
    # Track B — one trajectory
    ws.append(("B.wav", 1, 10, 1.0,  500.0, -3.0))
    ws.append(("B.wav", 1, 11, 1.1,  500.5, -2.5))

    xlsx = tmp_path / "fixture.xlsx"
    wb.save(str(xlsx))

    result = load_peak_trajectories_from_excel(xlsx)
    assert set(result.keys()) == {"A.wav", "B.wav"}
    assert len(result["A.wav"]) == 2   # two trajectories
    assert len(result["B.wav"]) == 1

    traj_a1 = result["A.wav"][0]
    assert traj_a1.points == [(0, 100.0, -5.0), (1, 100.5, -4.5)]
    assert traj_a1.mean_freq == pytest.approx(100.25)

    traj_b1 = result["B.wav"][0]
    assert traj_b1.points == [(10, 500.0, -3.0), (11, 500.5, -2.5)]


def test_load_peak_trajectories_from_excel_applies_name_filter(tmp_path: Path):
    """The ``track_name_filter`` callable lets the caller normalise
    WAV filenames into Ableton EffectiveName format (e.g. strip the
    ``Acid_Drops `` prefix and the ``.wav`` suffix)."""
    from openpyxl import Workbook

    wb = Workbook()
    wb.remove(wb.active)
    ws = wb.create_sheet("_track_peak_trajectories")
    ws.append(("Track", "Traj#", "Frame", "Time (s)", "Freq (Hz)", "Amp (dB)"))
    ws.append(("Acid_Drops [A_T] Ambience.wav", 1, 0, 0.0, 247.0, -8.0))
    ws.append(("Acid_Drops [A_T] Ambience.wav", 1, 1, 0.1, 247.5, -7.5))

    xlsx = tmp_path / "fixture.xlsx"
    wb.save(str(xlsx))

    def _normalize(name: str) -> str:
        # Strip project prefix + TFP underscore form + .wav → mimic cde_apply's intent.
        s = name.removeprefix("Acid_Drops ")
        if s.lower().endswith(".wav"):
            s = s[:-4]
        s = s.replace("[A_T]", "[A/T]")
        return s

    result = load_peak_trajectories_from_excel(
        xlsx, track_name_filter=_normalize,
    )
    assert list(result.keys()) == ["[A/T] Ambience"]


def test_load_peak_trajectories_from_excel_raises_on_missing_sheet(tmp_path: Path):
    from openpyxl import Workbook

    wb = Workbook()
    # No _track_peak_trajectories sheet — only the default 'Sheet'.
    xlsx = tmp_path / "wrong.xlsx"
    wb.save(str(xlsx))

    with pytest.raises(ValueError, match="_track_peak_trajectories"):
        load_peak_trajectories_from_excel(xlsx)


# ---------------------------------------------------------------------------
# Constants sanity
# ---------------------------------------------------------------------------

def test_max_cde_bands_leaves_0_and_7_for_hpf_lpf_convention():
    """MAX_CDE_BANDS must be 6 — bands 0 and 7 are reserved by
    ``eq8_automation._find_available_band`` for HPF/LPF."""
    assert MAX_CDE_BANDS == 6
    assert DEFAULT_FREQ_CLUSTER_TOLERANCE_SEMITONES == 2.0
