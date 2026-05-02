#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
Tests for spectral_evolution.py — Phase 1 & 2 feature extraction.

Uses the WAV test fixtures already present in tests/.
"""

import sys
import os
import pytest
import numpy as np
import soundfile as sf

sys.path.insert(0, os.path.dirname(os.path.dirname(os.path.abspath(__file__))))

from spectral_evolution import (
    generate_matrix,
    extract_zone_energy,
    extract_spectral_descriptors,
    extract_peak_trajectories,
    extract_valley_trajectories,
    extract_crest_by_zone,
    extract_delta_spectrum,
    extract_transients,
    extract_all_features,
    SpectralMatrix,
    ZoneEnergy,
    SpectralDescriptors,
    TrackFeatures,
    ZONE_RANGES,
    CQT_N_BINS,
    TARGET_FRAMES_PER_SEC,
)

FIXTURES_DIR = os.path.join(os.path.dirname(os.path.abspath(__file__)))


def _load_mono(filename: str):
    """Load a fixture file and return (mono, sr)."""
    path = os.path.join(FIXTURES_DIR, filename)
    data, sr = sf.read(path, always_2d=True)
    mono = np.mean(data, axis=1).astype(np.float32)
    return mono, sr


# ---------------------------------------------------------------------------
# Phase 1: Matrix generation
# ---------------------------------------------------------------------------

class TestGenerateMatrix:
    def test_shape_and_types(self):
        mono, sr = _load_mono('01_flat_wideband.wav')
        m = generate_matrix(mono, sr)
        assert isinstance(m, SpectralMatrix)
        assert m.cqt_db.ndim == 2
        assert m.n_bins <= CQT_N_BINS
        assert m.n_bins > 200  # should still have good resolution
        assert len(m.freqs) == m.n_bins
        assert len(m.times) == m.n_frames

    def test_frame_rate(self):
        mono, sr = _load_mono('01_flat_wideband.wav')
        m = generate_matrix(mono, sr)
        duration = len(mono) / sr
        actual_fps = m.n_frames / duration
        assert 3.0 <= actual_fps <= 12.0, f"Unexpected frame rate: {actual_fps:.1f} fps"

    def test_frequency_range(self):
        mono, sr = _load_mono('01_flat_wideband.wav')
        m = generate_matrix(mono, sr)
        assert m.freqs[0] >= 15.0, "Lowest bin should be near 20 Hz"
        assert m.freqs[-1] <= 25000.0, "Highest bin should be near 20 kHz"

    def test_values_are_db(self):
        mono, sr = _load_mono('01_flat_wideband.wav')
        m = generate_matrix(mono, sr)
        assert m.cqt_db.max() <= 10.0, "Max dB should be near 0 dBFS"
        assert m.cqt_db.min() >= -120.0, "Min dB should be above floor"


# ---------------------------------------------------------------------------
# Phase 1: Zone energy
# ---------------------------------------------------------------------------

class TestZoneEnergy:
    def test_all_zones_present(self):
        mono, sr = _load_mono('01_flat_wideband.wav')
        m = generate_matrix(mono, sr)
        ze = extract_zone_energy(m)
        assert isinstance(ze, ZoneEnergy)
        for zone_name in ZONE_RANGES:
            assert zone_name in ze.zones, f"Missing zone: {zone_name}"
            assert len(ze.zones[zone_name]) == m.n_frames

    def test_energy_values_are_finite(self):
        mono, sr = _load_mono('02_wideband_one_resonance_500hz.wav')
        m = generate_matrix(mono, sr)
        ze = extract_zone_energy(m)
        for zone_name, curve in ze.zones.items():
            assert np.all(np.isfinite(curve)), f"Non-finite values in zone {zone_name}"

    def test_resonance_visible_in_zone(self):
        """500 Hz resonance should make 'mud' zone (200-500 Hz) louder than 'air'."""
        mono, sr = _load_mono('02_wideband_one_resonance_500hz.wav')
        m = generate_matrix(mono, sr)
        ze = extract_zone_energy(m)
        mud_mean = np.mean(ze.zones['mud'])
        air_mean = np.mean(ze.zones['air'])
        assert mud_mean > air_mean, "500 Hz resonance should boost mud zone"

    def test_sub_bass_resonance(self):
        """80 Hz resonance should boost 'sub' zone."""
        mono, sr = _load_mono('06_sub_bass_80hz_resonance.wav')
        m = generate_matrix(mono, sr)
        ze = extract_zone_energy(m)
        sub_mean = np.mean(ze.zones['sub'])
        air_mean = np.mean(ze.zones['air'])
        assert sub_mean > air_mean, "80 Hz resonance should boost sub zone"


# ---------------------------------------------------------------------------
# Phase 1: Spectral descriptors
# ---------------------------------------------------------------------------

class TestSpectralDescriptors:
    def test_shape_and_range(self):
        mono, sr = _load_mono('01_flat_wideband.wav')
        m = generate_matrix(mono, sr)
        sd = extract_spectral_descriptors(m)
        assert isinstance(sd, SpectralDescriptors)
        assert len(sd.centroid) == m.n_frames
        assert len(sd.spread) == m.n_frames
        assert len(sd.flatness) == m.n_frames
        assert len(sd.low_rolloff) == m.n_frames
        assert len(sd.high_rolloff) == m.n_frames

    def test_centroid_reasonable(self):
        mono, sr = _load_mono('01_flat_wideband.wav')
        m = generate_matrix(mono, sr)
        sd = extract_spectral_descriptors(m)
        median_centroid = np.median(sd.centroid)
        assert 200 < median_centroid < 15000, f"Centroid {median_centroid:.0f} Hz seems off"

    def test_flatness_range(self):
        mono, sr = _load_mono('01_flat_wideband.wav')
        m = generate_matrix(mono, sr)
        sd = extract_spectral_descriptors(m)
        assert np.all(sd.flatness >= 0), "Flatness should be >= 0"
        assert np.all(sd.flatness <= 1.1), "Flatness should be ~<= 1"

    def test_rolloff_ordering(self):
        """Low rolloff should generally be below high rolloff."""
        mono, sr = _load_mono('01_flat_wideband.wav')
        m = generate_matrix(mono, sr)
        sd = extract_spectral_descriptors(m)
        mean_low = np.mean(sd.low_rolloff)
        mean_high = np.mean(sd.high_rolloff)
        assert mean_low < mean_high, "Low rolloff should be below high rolloff"


# ---------------------------------------------------------------------------
# Phase 2: Trajectories
# ---------------------------------------------------------------------------

class TestTrajectories:
    def test_peak_trajectories_from_resonance(self):
        """Track with 500 Hz resonance should have a trajectory near 500 Hz."""
        mono, sr = _load_mono('02_wideband_one_resonance_500hz.wav')
        m = generate_matrix(mono, sr)
        peaks = extract_peak_trajectories(m, n_peaks=6, min_duration_frames=3)
        assert len(peaks) > 0, "Should detect at least one trajectory"
        freqs = [p.mean_freq for p in peaks]
        has_500 = any(350 < f < 650 for f in freqs)
        assert has_500, f"Should have a trajectory near 500 Hz, got {freqs}"

    def test_three_resonances(self):
        """Should detect trajectories near 500, 1500, 4000 Hz."""
        mono, sr = _load_mono('03_wideband_three_resonances.wav')
        m = generate_matrix(mono, sr)
        peaks = extract_peak_trajectories(m, n_peaks=6, min_duration_frames=3)
        freqs = sorted([p.mean_freq for p in peaks])
        # Check at least 2 of the 3 expected frequencies are represented
        near_500 = any(350 < f < 700 for f in freqs)
        near_1500 = any(1000 < f < 2000 for f in freqs)
        near_4000 = any(3000 < f < 5500 for f in freqs)
        found = sum([near_500, near_1500, near_4000])
        assert found >= 2, f"Expected >=2 of [500,1500,4000], got freqs={freqs}"

    def test_valley_trajectories_exist(self):
        mono, sr = _load_mono('01_flat_wideband.wav')
        m = generate_matrix(mono, sr)
        valleys = extract_valley_trajectories(m, n_valleys=4, min_duration_frames=3)
        # Flat noise may or may not produce stable valleys; just check no crash
        assert isinstance(valleys, list)


# ---------------------------------------------------------------------------
# Phase 2: Dynamics
# ---------------------------------------------------------------------------

class TestDynamics:
    def test_crest_by_zone_keys(self):
        mono, sr = _load_mono('01_flat_wideband.wav')
        m = generate_matrix(mono, sr)
        crest = extract_crest_by_zone(m)
        for zone_name in ZONE_RANGES:
            assert zone_name in crest
            assert len(crest[zone_name]) == m.n_frames

    def test_delta_spectrum_shape(self):
        mono, sr = _load_mono('01_flat_wideband.wav')
        m = generate_matrix(mono, sr)
        delta = extract_delta_spectrum(m)
        assert len(delta) == m.n_frames
        assert delta[0] == 0.0

    def test_transient_detection(self):
        """Percussion fixture should have transient events."""
        mono, sr = _load_mono('04_spoon_percussion_no_anomaly.wav')
        m = generate_matrix(mono, sr)
        delta = extract_delta_spectrum(m)
        events = extract_transients(m, delta)
        assert isinstance(events, list)
        assert len(events) > 0, "Percussion track should have transient events"


# ---------------------------------------------------------------------------
# Full pipeline
# ---------------------------------------------------------------------------

class TestFullPipeline:
    def test_extract_all_features(self):
        mono, sr = _load_mono('01_flat_wideband.wav')
        feat = extract_all_features(mono, sr)
        assert isinstance(feat, TrackFeatures)
        assert feat.zone_energy is not None
        assert feat.descriptors is not None
        assert feat.peak_trajectories is not None
        assert feat.valley_trajectories is not None
        assert feat.crest_by_zone is not None
        assert feat.delta_spectrum is not None
        assert feat.transient_events is not None

    def test_energy_conservation_sanity(self):
        """Zone energy total should be roughly consistent across flat noise frames."""
        mono, sr = _load_mono('01_flat_wideband.wav')
        feat = extract_all_features(mono, sr)
        # For flat wideband noise, zone energies should be relatively stable
        for zone_name in ['mid', 'presence']:
            curve = feat.zone_energy.zones[zone_name]
            std = np.std(curve)
            assert std < 10.0, f"Zone {zone_name} too variable for flat noise: std={std:.1f} dB"


# ---------------------------------------------------------------------------
# Feature storage
# ---------------------------------------------------------------------------

class TestFeatureStorage:
    def test_build_all_sheets(self):
        """Build all v2.5 sheets into a workbook without errors."""
        from openpyxl import Workbook
        from feature_storage import build_all_v25_sheets

        mono, sr = _load_mono('01_flat_wideband.wav')
        feat = extract_all_features(mono, sr)
        ti = {'name': 'test_track.wav', 'type': 'Individual'}
        features_with_info = [(feat, ti)]

        wb = Workbook()
        build_all_v25_sheets(wb, features_with_info, log_fn=print)

        sheet_names = wb.sheetnames
        assert '_track_zone_energy' in sheet_names
        assert '_track_spectral_descriptors' in sheet_names
        assert '_track_peak_trajectories' in sheet_names
        assert '_track_valley_trajectories' in sheet_names
        assert '_track_crest_by_zone' in sheet_names
        assert '_track_transients' in sheet_names

        # Verify sheets are hidden
        for sn in ['_track_zone_energy', '_track_spectral_descriptors']:
            assert wb[sn].sheet_state == 'hidden'

    def test_zone_energy_sheet_content(self):
        """Verify zone energy sheet has data in cells."""
        from openpyxl import Workbook
        from feature_storage import build_v25_zone_energy_sheet

        mono, sr = _load_mono('02_wideband_one_resonance_500hz.wav')
        m = generate_matrix(mono, sr)
        ze = extract_zone_energy(m)
        feat = TrackFeatures(zone_energy=ze, descriptors=extract_spectral_descriptors(m))
        ti = {'name': 'resonance_500.wav', 'type': 'Individual'}

        wb = Workbook()
        build_v25_zone_energy_sheet(wb, [(feat, ti)], log_fn=print)

        ws = wb['_track_zone_energy']
        # Row 1 = header, rows 2-10 = 9 zones
        assert ws.cell(row=1, column=1).value == 'resonance_500.wav'
        assert ws.cell(row=2, column=1).value is not None  # first zone label
        assert ws.cell(row=2, column=2).value is not None  # first data cell


# ===========================================================================
# Phase F10b — preset-aware CQT pipeline tests (5 cases)
# ===========================================================================
#
# These tests exercise the new ``preset`` parameter of ``generate_matrix``
# and ``extract_all_features``. The most critical is the byte-identity
# non-regression test : it guarantees that any existing script that does
# NOT pass a preset (= the default `None` → standard) gets exactly the
# v2.7.0 output, frame-by-frame, sample-by-sample.

from resolution_presets import RESOLUTION_PRESETS


class TestPhaseF10bPresetSupport:
    """Phase F10b — preset parameter on generate_matrix + extract_all_features."""

    def test_no_preset_byte_identical_to_explicit_standard(self):
        """The most important test of F10b : default behaviour preserved.

        ``generate_matrix(mono, sr)`` without preset must produce a result
        byte-identical to ``generate_matrix(mono, sr, preset=standard)``.
        Both must equal the v2.7.0 output (the 22 pre-existing tests
        verify the latter empirically).

        If this test ever fails after a refactor of spectral_evolution.py,
        backward compatibility is broken — every script that doesn't pass
        ``--resolution`` will start producing different output.
        """
        mono, sr = _load_mono('01_flat_wideband.wav')
        m_default = generate_matrix(mono, sr)
        m_standard = generate_matrix(
            mono, sr, preset=RESOLUTION_PRESETS["standard"],
        )
        # Strict byte-identity (np.array_equal). If librosa ever introduces
        # float non-determinism this would fall back to
        # np.allclose(rtol=0, atol=1e-15) — but for now we keep the strict
        # check because it catches REAL regressions silently.
        assert np.array_equal(m_default.cqt_db, m_standard.cqt_db), (
            "cqt_db diverges between no-preset and explicit-standard "
            "— backward compat broken"
        )
        assert np.array_equal(m_default.freqs, m_standard.freqs)
        assert np.array_equal(m_default.times, m_standard.times)
        assert m_default.hop_length == m_standard.hop_length
        assert m_default.sr == m_standard.sr

    def test_preset_ultra_changes_hop(self):
        """Ultra preset (12 fps CQT) → hop ≈ sr / 12 instead of sr / 6."""
        mono, sr = _load_mono('01_flat_wideband.wav')
        m_standard = generate_matrix(
            mono, sr, preset=RESOLUTION_PRESETS["standard"],
        )
        m_ultra = generate_matrix(
            mono, sr, preset=RESOLUTION_PRESETS["ultra"],
        )
        # Standard = 6 fps target, ultra = 12 fps target. Ultra hop should
        # be approximately half of standard hop (modulo rounding + 512 floor).
        expected_ultra_hop = max(round(sr / 12), 512)
        assert m_ultra.hop_length == expected_ultra_hop
        # Sanity : ultra hop is roughly half of standard
        assert m_ultra.hop_length < m_standard.hop_length
        # Frame count scales inversely : ultra has more frames
        assert m_ultra.n_frames > m_standard.n_frames

    def test_preset_ultra_changes_n_bins_effective(self):
        """Ultra preset (36 bins/oct) → n_bins effective scales with bpo,
        capped by Nyquist.

        At sr=44.1 kHz, max_octaves ≈ 10.007. Standard (24 bpo) caps at
        floor(10.007 × 24) = 240. Ultra (36 bpo) caps at
        floor(10.007 × 36) = 360.
        """
        mono, sr = _load_mono('01_flat_wideband.wav')
        m_standard = generate_matrix(
            mono, sr, preset=RESOLUTION_PRESETS["standard"],
        )
        m_ultra = generate_matrix(
            mono, sr, preset=RESOLUTION_PRESETS["ultra"],
        )
        # Compute expected n_bins from the audit math
        max_octaves = np.log2((sr / 2.0) / 20.0) - 0.1
        expected_standard_bins = min(256, int(np.floor(max_octaves * 24)))
        expected_ultra_bins = min(384, int(np.floor(max_octaves * 36)))
        assert m_standard.n_bins == expected_standard_bins
        assert m_ultra.n_bins == expected_ultra_bins
        # Sanity : ultra has 1.5× more bins than standard (36/24 ratio)
        assert m_ultra.n_bins > m_standard.n_bins

    def test_preset_maximum_pushes_to_highest_resolution(self):
        """Maximum preset (24 fps + 48 bpo) — highest resolution tier.

        At sr=44.1 kHz : hop ≈ 1838, n_bins ≈ 480 (= floor(10.007 × 48)
        capped by 512 requested).
        """
        mono, sr = _load_mono('01_flat_wideband.wav')
        m_max = generate_matrix(
            mono, sr, preset=RESOLUTION_PRESETS["maximum"],
        )
        expected_hop = max(round(sr / 24), 512)
        max_octaves = np.log2((sr / 2.0) / 20.0) - 0.1
        expected_bins = min(512, int(np.floor(max_octaves * 48)))
        assert m_max.hop_length == expected_hop
        assert m_max.n_bins == expected_bins

    def test_extract_all_features_with_ultra_preset_returns_full_features(self):
        """End-to-end : extract_all_features with non-default preset must
        return a complete TrackFeatures (peak_trajectories non-empty,
        all sub-fields populated)."""
        mono, sr = _load_mono('02_wideband_one_resonance_500hz.wav')
        feat = extract_all_features(
            mono, sr, preset=RESOLUTION_PRESETS["ultra"],
        )
        assert isinstance(feat, TrackFeatures)
        # Resonance fixture has a clear peak — must be tracked at any preset
        assert feat.peak_trajectories is not None
        assert len(feat.peak_trajectories) > 0
        # Other fields populated
        assert feat.zone_energy is not None
        assert feat.descriptors is not None
        assert feat.crest_by_zone is not None
        assert feat.delta_spectrum is not None


if __name__ == '__main__':
    pytest.main([__file__, '-v'])
