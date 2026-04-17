#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
Tests for automation_map.py — v2.5.1 automation extraction and masking.

Uses synthetic minimal ALS fixtures with various automation scenarios.
"""

import gzip
import os
import sys
import xml.etree.ElementTree as ET

import numpy as np
import pytest

sys.path.insert(0, os.path.dirname(os.path.dirname(os.path.abspath(__file__))))

from automation_map import (
    AutomationCurve,
    TrackAutomationMap,
    AUDIBILITY_THRESHOLD,
    extract_track_automations,
    extract_all_track_automations,
    resample_effective_gain,
    resample_audibility,
    _interpolate_at,
    _utility_gain_to_linear,
)


# ---------------------------------------------------------------------------
# Fixture helpers
# ---------------------------------------------------------------------------

def _create_als_with_automations(path, track_configs, tempo=120.0):
    """Create a minimal .als with configurable automation per track.

    track_configs: list of dicts, each with:
        - name: str
        - volume_automation: list of (time_beats, value) or None for static
        - volume_static: float (default 0.85)
        - speaker_automation: list of (time_beats, value) or None
        - speaker_static: float (default 1.0)
        - utility_gains: list of dicts with 'automation' and 'static' keys
    """
    root = ET.Element("Ableton")
    root.set("Creator", "Test")
    root.set("SchemaChangeCount", "1")

    live_set = ET.SubElement(root, "LiveSet")
    tempo_container = ET.SubElement(live_set, "Tempo")
    ET.SubElement(tempo_container, "Manual").set("Value", str(tempo))

    tracks = ET.SubElement(live_set, "Tracks")
    next_id = 1000

    for cfg in track_configs:
        track = ET.SubElement(tracks, "AudioTrack")
        track.set("Id", str(next_id))
        next_id += 1

        ET.SubElement(track, "EffectiveName").set("Value", cfg["name"])

        # DeviceChain structure
        dc_outer = ET.SubElement(track, "DeviceChain")

        # MixerDevice with Volume and Speaker
        mixer = ET.SubElement(dc_outer, "Mixer")

        # Volume
        volume = ET.SubElement(mixer, "Volume")
        vol_static = cfg.get("volume_static", 0.85)
        ET.SubElement(volume, "Manual").set("Value", str(vol_static))
        vol_target_id = str(next_id)
        ET.SubElement(volume, "AutomationTarget").set("Id", vol_target_id)
        next_id += 1

        # Speaker
        speaker = ET.SubElement(mixer, "Speaker")
        spk_static = cfg.get("speaker_static", 1.0)
        ET.SubElement(speaker, "Manual").set("Value", str(int(spk_static)))
        spk_target_id = str(next_id)
        ET.SubElement(speaker, "AutomationTarget").set("Id", spk_target_id)
        next_id += 1

        # Devices (Utility / StereoGain)
        dc_inner = ET.SubElement(dc_outer, "DeviceChain")
        devices = ET.SubElement(dc_inner, "Devices")

        utility_target_ids = []
        for util_cfg in cfg.get("utility_gains", []):
            sg = ET.SubElement(devices, "StereoGain")
            sg.set("Id", str(next_id))
            next_id += 1
            gain_param = ET.SubElement(sg, "Gain")
            util_static = util_cfg.get("static", 0.0)  # dB
            ET.SubElement(gain_param, "Manual").set("Value", str(util_static))
            util_target_id = str(next_id)
            ET.SubElement(gain_param, "AutomationTarget").set("Id", util_target_id)
            next_id += 1
            utility_target_ids.append((util_target_id, util_cfg))

        # EQ8 (minimal, for compatibility)
        eq8 = ET.SubElement(devices, "Eq8")
        eq8.set("Id", str(next_id))
        next_id += 1

        # AutomationEnvelopes
        auto_env = ET.SubElement(track, "AutomationEnvelopes")
        envelopes = ET.SubElement(auto_env, "Envelopes")

        # Write volume automation if specified
        if cfg.get("volume_automation"):
            _add_envelope(envelopes, vol_target_id, cfg["volume_automation"], next_id)
            next_id += len(cfg["volume_automation"]) + 1

        # Write speaker automation if specified
        if cfg.get("speaker_automation"):
            _add_envelope(envelopes, spk_target_id, cfg["speaker_automation"], next_id)
            next_id += len(cfg["speaker_automation"]) + 1

        # Write utility automation if specified
        for util_target_id, util_cfg in utility_target_ids:
            if util_cfg.get("automation"):
                _add_envelope(envelopes, util_target_id, util_cfg["automation"], next_id)
                next_id += len(util_cfg["automation"]) + 1

    xml_str = ET.tostring(root, encoding="unicode", xml_declaration=True)
    with gzip.open(str(path), "wb") as f:
        f.write(xml_str.encode("utf-8"))
    return path


def _add_envelope(envelopes_node, target_id, events, start_id):
    """Add an AutomationEnvelope with FloatEvents."""
    envelope = ET.SubElement(envelopes_node, "AutomationEnvelope")
    env_target = ET.SubElement(envelope, "EnvelopeTarget")
    ET.SubElement(env_target, "PointeeId").set("Value", str(target_id))
    automation = ET.SubElement(envelope, "Automation")
    events_node = ET.SubElement(automation, "Events")

    # Pre-song default
    default_ev = ET.SubElement(events_node, "FloatEvent")
    default_ev.set("Id", str(start_id))
    default_ev.set("Time", "-63072000")
    default_ev.set("Value", str(events[0][1]))

    for i, (t, v) in enumerate(events):
        ev = ET.SubElement(events_node, "FloatEvent")
        ev.set("Id", str(start_id + i + 1))
        ev.set("Time", str(t))
        ev.set("Value", str(v))


# ---------------------------------------------------------------------------
# Test 1: Utility gain extraction
# ---------------------------------------------------------------------------

class TestExtractUtilityGain:
    def test_utility_mute_mid_track(self, tmp_path):
        """Utility gain drops to -60 dB at beat 32 → effective gain drops in second half."""
        als_path = tmp_path / "test.als"
        _create_als_with_automations(als_path, [
            {
                "name": "Toms Rack",
                "volume_static": 0.5,
                "utility_gains": [{
                    "static": 0.0,
                    "automation": [
                        (0.0, 0.0),     # 0 dB → linear 1.0
                        (32.0, -60.0),  # -60 dB → linear 0.001 at beat 32
                        (64.0, -60.0),  # stays muted
                    ],
                }],
            }
        ], tempo=120.0)

        result = extract_all_track_automations(als_path)
        assert "Toms Rack" in result
        auto_map = result["Toms Rack"]

        # At beat 0: volume=0.5, utility=1.0 → effective ~0.5
        assert auto_map.effective_gain[0] > 0.3
        # After beat 32 (midpoint): utility ~ 0.001 → effective < 0.01
        three_quarter_idx = int(len(auto_map.effective_gain) * 0.75)
        assert auto_map.effective_gain[three_quarter_idx] < 0.01

    def test_multiple_utilities_multiply(self, tmp_path):
        """Two Utility devices: gains multiply."""
        als_path = tmp_path / "test.als"
        _create_als_with_automations(als_path, [
            {
                "name": "Track1",
                "volume_static": 1.0,
                "utility_gains": [
                    {"static": -6.0},   # ~0.5 linear
                    {"static": -6.0},   # ~0.5 linear
                ],
            }
        ], tempo=120.0)

        result = extract_all_track_automations(als_path)
        auto_map = result["Track1"]

        # 0.5 * 0.5 = 0.25
        expected = _utility_gain_to_linear(-6.0) ** 2
        assert abs(auto_map.effective_gain[0] - expected) < 0.05


# ---------------------------------------------------------------------------
# Test 2: Volume fader extraction
# ---------------------------------------------------------------------------

class TestExtractVolumeFader:
    def test_volume_fader_automation(self, tmp_path):
        """Volume fader automates from 0.85 to 0.0 at beat 32."""
        als_path = tmp_path / "test.als"
        _create_als_with_automations(als_path, [
            {
                "name": "Lead Synth",
                "volume_automation": [
                    (0.0, 0.85),
                    (32.0, 0.0),
                ],
            }
        ], tempo=120.0)

        result = extract_all_track_automations(als_path)
        auto_map = result["Lead Synth"]

        # Start should be ~0.85
        assert auto_map.effective_gain[0] == pytest.approx(0.85, abs=0.05)
        # After beat 32, gain should drop to 0
        late_idx = len(auto_map.effective_gain) - 1
        assert auto_map.effective_gain[late_idx] < 0.01


# ---------------------------------------------------------------------------
# Test 3: Effective gain combines correctly
# ---------------------------------------------------------------------------

class TestEffectiveGainCombines:
    def test_volume_times_utility(self, tmp_path):
        """effective_gain = volume × utility."""
        als_path = tmp_path / "test.als"
        _create_als_with_automations(als_path, [
            {
                "name": "Bass",
                "volume_static": 0.5,
                "utility_gains": [{"static": -6.0}],  # ~0.5 linear
            }
        ], tempo=120.0)

        result = extract_all_track_automations(als_path)
        auto_map = result["Bass"]

        utility_linear = _utility_gain_to_linear(-6.0)
        expected = 0.5 * utility_linear
        assert auto_map.effective_gain[0] == pytest.approx(expected, abs=0.05)


# ---------------------------------------------------------------------------
# Test 4: Audibility mask
# ---------------------------------------------------------------------------

class TestAudibilityMask:
    def test_inaudible_below_threshold(self, tmp_path):
        """Frames with effective_gain < 0.001 are not audible."""
        als_path = tmp_path / "test.als"
        _create_als_with_automations(als_path, [
            {
                "name": "Muted Track",
                "volume_static": 0.0001,  # way below threshold
            }
        ], tempo=120.0)

        result = extract_all_track_automations(als_path)
        auto_map = result["Muted Track"]
        assert not auto_map.is_audible.any()

    def test_audible_above_threshold(self, tmp_path):
        """Normal track should be fully audible."""
        als_path = tmp_path / "test.als"
        _create_als_with_automations(als_path, [
            {
                "name": "Normal Track",
                "volume_static": 0.85,
            }
        ], tempo=120.0)

        result = extract_all_track_automations(als_path)
        auto_map = result["Normal Track"]
        assert auto_map.is_audible.all()


# ---------------------------------------------------------------------------
# Test 5: EQ8 skips muted sections
# ---------------------------------------------------------------------------

class TestEQ8SkipsMuted:
    def test_mask_by_audibility_removes_muted_breakpoints(self):
        """_mask_by_audibility drops breakpoints in inaudible sections."""
        from eq8_automation import _mask_by_audibility

        # Create a mock automation map where second half is muted
        auto_map = TrackAutomationMap(
            track_name="Test",
            effective_gain=np.concatenate([
                np.ones(64),
                np.zeros(64),
            ]),
            effective_gain_times=np.linspace(0, 10, 128),
            is_audible=np.concatenate([
                np.ones(64, dtype=bool),
                np.zeros(64, dtype=bool),
            ]),
        )

        # Breakpoints spanning 0 to 10 seconds (at 120 BPM: 0 to 20 beats)
        breakpoints = [(float(i), 100.0) for i in range(20)]
        filtered = _mask_by_audibility(breakpoints, auto_map, tempo=120.0)

        # First ~10 beats (= 5 seconds) should be kept, rest dropped
        assert len(filtered) < len(breakpoints)
        assert len(filtered) > 0
        # All kept breakpoints should be in the audible region
        for t_beats, _ in filtered:
            t_sec = t_beats * (60.0 / 120.0)
            assert t_sec < 6.0  # generous margin

    def test_mask_by_audibility_no_map(self):
        """Without automation_map, all breakpoints pass through."""
        from eq8_automation import _mask_by_audibility
        breakpoints = [(0.0, 1.0), (10.0, 2.0), (20.0, 3.0)]
        result = _mask_by_audibility(breakpoints, None, tempo=120.0)
        assert result == breakpoints


# ---------------------------------------------------------------------------
# Test 6: Automation map sheet
# ---------------------------------------------------------------------------

class TestAutomationMapSheet:
    def test_sheet_creation(self, tmp_path):
        """build_automation_map_sheet creates the sheet with correct structure."""
        from openpyxl import Workbook
        from feature_storage import build_automation_map_sheet

        wb = Workbook()

        # Create a simple automation map
        auto_maps = {
            "Kick": TrackAutomationMap(
                track_name="Kick",
                curves=[
                    AutomationCurve(
                        target_id=100,
                        param_name="Volume",
                        device_name="MixerDevice",
                        times_beats=np.array([0.0]),
                        values=np.array([0.85]),
                    ),
                ],
                effective_gain=np.full(128, 0.85),
                effective_gain_times=np.linspace(0, 10, 128),
                is_audible=np.ones(128, dtype=bool),
            ),
        }

        build_automation_map_sheet(wb, auto_maps, n_buckets=64, duration=10.0)

        assert '_track_automation_map' in wb.sheetnames
        ws = wb['_track_automation_map']
        assert ws.sheet_state == 'hidden'

        # Header row
        assert ws.cell(1, 1).value == 'Track'
        assert ws.cell(1, 2).value == 'Parameter'
        assert ws.cell(1, 3).value == 'Device'

        # Data rows: Volume curve + effective_gain + is_audible = 3 rows
        assert ws.cell(2, 1).value == 'Kick'
        assert ws.cell(2, 2).value == 'Volume'
        assert ws.cell(3, 2).value == 'effective_gain'
        assert ws.cell(4, 2).value == 'is_audible'


# ---------------------------------------------------------------------------
# Test helpers
# ---------------------------------------------------------------------------

class TestResampleEffectiveGain:
    def test_resample_basic(self):
        """Resampling preserves constant gain."""
        auto_map = TrackAutomationMap(
            track_name="Test",
            effective_gain=np.full(100, 0.75),
            effective_gain_times=np.linspace(0, 10, 100),
            is_audible=np.ones(100, dtype=bool),
        )
        target_times = np.linspace(0, 10, 64)
        resampled = resample_effective_gain(auto_map, target_times)
        assert len(resampled) == 64
        np.testing.assert_allclose(resampled, 0.75, atol=0.01)

    def test_resample_audibility(self):
        """resample_audibility returns bool mask."""
        auto_map = TrackAutomationMap(
            track_name="Test",
            effective_gain=np.concatenate([np.full(50, 0.8), np.full(50, 0.0)]),
            effective_gain_times=np.linspace(0, 10, 100),
            is_audible=np.concatenate([np.ones(50, dtype=bool), np.zeros(50, dtype=bool)]),
        )
        target_times = np.linspace(0, 10, 10)
        audible = resample_audibility(auto_map, target_times)
        assert audible[0] == True
        assert audible[-1] == False


class TestInterpolateAt:
    def test_step_interpolation(self):
        """Step interpolation holds last value."""
        times = np.array([0.0, 4.0, 8.0])
        values = np.array([1.0, 0.5, 0.0])
        query = np.array([2.0, 5.0, 10.0])
        result = _interpolate_at(times, values, query)
        assert result[0] == 1.0   # between 0 and 4 → holds 1.0
        assert result[1] == 0.5   # between 4 and 8 → holds 0.5
        assert result[2] == 0.0   # after 8 → holds 0.0


class TestBucketLevelMasking:
    """Round-trip: is_audible=0 on N buckets → zone_energy has N None per zone."""

    def test_muted_buckets_produce_none(self):
        """Muted bucket centers → all frames in that bucket NaN → None after nanmean."""
        from feature_storage import _downsample_frames

        n_frames = 721
        duration = 120.0
        n_buckets = 64

        # Simulate automation map: buckets 10,11,12 are inaudible
        auto_map = TrackAutomationMap(
            track_name="Test",
            effective_gain=np.ones(200),
            effective_gain_times=np.linspace(0, duration, 200),
            is_audible=np.ones(200, dtype=bool),
        )
        # Mute the time range covering buckets 10-12 (center times ~19.7-23.4s)
        bucket_step = duration / n_buckets
        for i in range(200):
            t = auto_map.effective_gain_times[i]
            if 10 * bucket_step <= t < 13 * bucket_step:
                auto_map.effective_gain[i] = 0.0
                auto_map.is_audible[i] = False

        # Compute bucket-level audibility
        bucket_centers = np.array([(i + 0.5) * bucket_step for i in range(n_buckets)])
        bucket_audible = resample_audibility(auto_map, bucket_centers)
        n_muted_buckets = int((~bucket_audible).sum())
        assert n_muted_buckets == 3

        # Apply bucket-level mask to frames
        edges = np.linspace(0, n_frames, n_buckets + 1, dtype=int)
        zone_data = np.ones(n_frames) * -25.0
        for b in range(n_buckets):
            if not bucket_audible[b]:
                zone_data[edges[b]:edges[b + 1]] = np.nan

        # Downsample and verify: 3 muted buckets → 3 None
        ds = _downsample_frames(zone_data, n_buckets)
        none_count = sum(1 for v in ds if v is None)
        assert none_count == n_muted_buckets, f"Expected {n_muted_buckets} None, got {none_count}"

    def test_fully_audible_no_none(self):
        """Fully audible track has zero None buckets."""
        from feature_storage import _downsample_frames

        zone_data = np.ones(721) * -20.0
        ds = _downsample_frames(zone_data, 64)
        none_count = sum(1 for v in ds if v is None)
        assert none_count == 0


class TestNameMatching:
    """Test the _match_auto_map static method logic."""

    def test_exact_match(self):
        maps = {"Kick 1": TrackAutomationMap(track_name="Kick 1")}
        assert _match("Kick 1", maps) is not None

    def test_strip_prefix(self):
        maps = {"Toms Rack": TrackAutomationMap(track_name="Toms Rack")}
        assert _match("Acid_drops Toms Rack", maps) is not None

    def test_strip_group_number(self):
        maps = {"5-Toms Rack": TrackAutomationMap(track_name="5-Toms Rack")}
        assert _match("Acid_drops Toms Rack", maps) is not None

    def test_substring_match(self):
        maps = {"7-OverHead": TrackAutomationMap(track_name="7-OverHead")}
        assert _match("Acid_drops Toms Overhead", maps) is not None

    def test_no_match(self):
        maps = {"Kick 1": TrackAutomationMap(track_name="Kick 1")}
        assert _match("Acid_drops Completely Different", maps) is None


def _match(stem, auto_maps):
    """Standalone copy of MixAnalyzerApp._match_auto_map for testing."""
    import re
    if stem in auto_maps:
        return auto_maps[stem]
    parts = stem.split(' ', 1)
    wav_clean = parts[1] if len(parts) == 2 else stem
    if wav_clean in auto_maps:
        return auto_maps[wav_clean]
    for als_name, am in auto_maps.items():
        als_clean = re.sub(r'^\d+[\s._-]+', '', als_name)
        if als_clean == wav_clean or als_clean == stem:
            return am
    wav_lower = wav_clean.lower()
    best_match = None
    best_len = 0
    for als_name, am in auto_maps.items():
        als_clean = re.sub(r'^\d+[\s._-]+', '', als_name).lower()
        if len(als_clean) < 3:
            continue
        if als_clean in wav_lower or wav_lower in als_clean:
            if len(als_clean) > best_len:
                best_len = len(als_clean)
                best_match = am
    return best_match


class TestSpeakerAutomation:
    def test_speaker_mute_unmute(self, tmp_path):
        """Speaker automation toggles audibility."""
        als_path = tmp_path / "test.als"
        _create_als_with_automations(als_path, [
            {
                "name": "Vocal",
                "volume_static": 0.85,
                "speaker_automation": [
                    (0.0, 1.0),   # unmuted
                    (32.0, 0.0),  # muted at beat 32
                    (64.0, 1.0),  # unmuted at beat 64
                ],
            }
        ], tempo=120.0)

        result = extract_all_track_automations(als_path)
        auto_map = result["Vocal"]

        # First section: audible
        assert auto_map.effective_gain[0] > 0.5
        # Middle section: muted
        mid = len(auto_map.effective_gain) // 2
        assert auto_map.effective_gain[mid] < AUDIBILITY_THRESHOLD
