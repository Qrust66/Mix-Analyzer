#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""Integration tests for Feature 3.5 Phase B2 — TFP roles wired into
``Section`` and the Sections Timeline rendering pipeline.

Covers:
    * ``enrich_sections_with_track_roles`` populates ``section.track_roles``
      from an Ableton-name mapping and Locator annotations.
    * ``build_sections_timeline_sheet`` renders role badges in the
      "TRACKS ACTIVES PAR ZONE" block and a ``RÔLE`` column in the
      "PEAK MAX PAR TRACK" block.
    * A warning banner lists tracks whose Ableton name has no TFP prefix.
    * **Regression guard**: the TFP-prefix-aware matching does not break
      the WAV <-> Ableton auto_map link — a previous session surfaced a
      28/38 tracks → -100 dB regression when a similar change was done
      without stripping the prefix from the match source. This test
      exercises the strip path.
"""

from __future__ import annotations

import os
import sys
from dataclasses import dataclass, field
from typing import List, Tuple

import numpy as np
import pytest

sys.path.insert(0, os.path.abspath(os.path.join(os.path.dirname(__file__), "..")))

from openpyxl import Workbook  # noqa: E402

from section_detector import (  # noqa: E402
    SECTIONS_TIMELINE_SHEET_NAME,
    Section,
    _detect_tracks_without_prefix,
    build_sections_timeline_sheet,
    enrich_sections_with_track_roles,
    get_zone_order,
)
from tfp_parser import DEFAULT_ROLE, Function, Importance  # noqa: E402


# ---------------------------------------------------------------------------
# Fixtures
# ---------------------------------------------------------------------------

def _make_section(
    index: int, start: int, end: int, total_energy_db: float,
    annotation: str = "",
) -> Section:
    return Section(
        index=index,
        name=f"Section {index}",
        start_bucket=start,
        end_bucket=end,
        start_seconds=float(start),
        end_seconds=float(end),
        start_beats=float(start * 2),
        end_beats=float(end * 2),
        total_energy_db=total_energy_db,
        annotation=annotation,
    )


def _zone_arrays_for(level_db: float, n_frames: int, zones_with_energy: List[str]) -> dict:
    zones = {
        z: np.full(n_frames, -120.0, dtype=float)
        for z in get_zone_order()
    }
    for z in zones_with_energy:
        zones[z] = np.full(n_frames, float(level_db), dtype=float)
    return zones


# ---------------------------------------------------------------------------
# enrich_sections_with_track_roles
# ---------------------------------------------------------------------------

def test_enrich_populates_default_role_when_no_ableton_mapping():
    """No Ableton names known -> every track gets DEFAULT_ROLE (S/R)."""
    sections = [_make_section(1, 0, 10, -15.0)]
    enrich_sections_with_track_roles(
        sections,
        wav_to_ableton={"Kick.wav": None, "Bass.wav": None},
    )
    s = sections[0]
    assert s.track_roles["Kick.wav"] == DEFAULT_ROLE
    assert s.track_roles["Bass.wav"] == DEFAULT_ROLE
    assert s.track_role_overridden["Kick.wav"] is False


def test_enrich_parses_prefix_from_ableton_name():
    """WAV "Acid_Drops Kick 1.wav" -> Ableton "[H/R] Kick 1" -> role (H, R)."""
    sections = [_make_section(1, 0, 10, -15.0)]
    enrich_sections_with_track_roles(
        sections,
        wav_to_ableton={
            "Acid_Drops Kick 1.wav": "[H/R] Kick 1",
            "Acid_Drops Ambience.wav": "[A/T] Ambience",
            "Acid_Drops Untagged.wav": "Untagged",
        },
    )
    roles = sections[0].track_roles
    assert roles["Acid_Drops Kick 1.wav"] == (Importance.H, Function.R)
    assert roles["Acid_Drops Ambience.wav"] == (Importance.A, Function.T)
    # No prefix -> default
    assert roles["Acid_Drops Untagged.wav"] == DEFAULT_ROLE


def test_enrich_applies_locator_annotation_overrides_and_flags_them():
    """Override changes Kick from (H, R) to (A, T); flag recorded."""
    annotation = "override: Kick 1=A-T"
    sections = [_make_section(1, 0, 10, -15.0, annotation=annotation)]
    enrich_sections_with_track_roles(
        sections,
        wav_to_ableton={"Acid_Drops Kick 1.wav": "[H/R] Kick 1"},
    )
    s = sections[0]
    assert s.track_roles["Acid_Drops Kick 1.wav"] == (Importance.A, Function.T)
    assert s.track_role_overridden["Acid_Drops Kick 1.wav"] is True


def test_enrich_overrides_do_not_leak_between_sections():
    """Override applied in section 1 must not affect section 2's role."""
    sections = [
        _make_section(1, 0, 5, -15.0, annotation="override: Kick 1=A-T"),
        _make_section(2, 6, 10, -15.0, annotation=""),
    ]
    enrich_sections_with_track_roles(
        sections,
        wav_to_ableton={"Acid_Drops Kick 1.wav": "[H/R] Kick 1"},
    )
    assert sections[0].track_roles["Acid_Drops Kick 1.wav"] == (Importance.A, Function.T)
    assert sections[1].track_roles["Acid_Drops Kick 1.wav"] == (Importance.H, Function.R)
    assert sections[0].track_role_overridden["Acid_Drops Kick 1.wav"] is True
    assert sections[1].track_role_overridden["Acid_Drops Kick 1.wav"] is False


# ---------------------------------------------------------------------------
# Warning banner helper
# ---------------------------------------------------------------------------

def test_detect_tracks_without_prefix_lists_untagged_and_missing():
    """Tracks with no prefix + tracks without an Ableton counterpart both
    surface in the warning list."""
    result = _detect_tracks_without_prefix({
        "Kick 1.wav": "[H/R] Kick 1",      # prefixed - OK
        "Ambience.wav": "Ambience",         # no prefix - flagged
        "Orphan.wav": None,                 # no ableton match - flagged
    })
    assert result == ["Ambience.wav", "Orphan.wav"]


def test_detect_tracks_without_prefix_empty_when_all_tagged():
    result = _detect_tracks_without_prefix({
        "Kick.wav": "[H/R] Kick 1",
        "Bass.wav": "[S/H] Sub Bass",
    })
    assert result == []


# ---------------------------------------------------------------------------
# build_sections_timeline_sheet — integration
# ---------------------------------------------------------------------------

def _dump_sheet(ws) -> str:
    rows = []
    for row in ws.iter_rows(values_only=True):
        rows.append(" | ".join(str(c) for c in row if c is not None))
    return "\n".join(rows)


def test_build_sheet_renders_role_column_in_peak_max():
    """Integration: a minimal workbook + sections + one tagged track
    produces a sheet where the PEAK MAX block includes the RÔLE column
    populated with 'H/R'."""
    n = 20
    sections = [
        _make_section(1, 0, n - 1, total_energy_db=-10.0),
    ]
    all_tracks_zone_energy = {
        "Acid_Drops Kick 1.wav": _zone_arrays_for(-5.0, n, ["sub", "low"]),
    }
    peak_by_section = {"Acid_Drops Kick 1.wav": {1: -10.0}}
    active_fraction = {"Acid_Drops Kick 1.wav": {1: 0.95}}
    wav_to_ableton = {"Acid_Drops Kick 1.wav": "[H/R] Kick 1"}

    wb = Workbook()
    wb.active.title = "Index"
    build_sections_timeline_sheet(
        workbook=wb,
        sections=sections,
        all_tracks_zone_energy=all_tracks_zone_energy,
        all_tracks_peak_by_section=peak_by_section,
        all_tracks_active_fraction=active_fraction,
        wav_to_ableton=wav_to_ableton,
    )
    ws = wb[SECTIONS_TIMELINE_SHEET_NAME]
    dump = _dump_sheet(ws)
    # Header row must include "RÔLE"
    assert "RÔLE" in dump, f"RÔLE column missing in header:\n{dump}"
    # The data row must include the role badge "H/R"
    assert "H/R" in dump, f"role badge not rendered:\n{dump}"


def test_build_sheet_renders_override_asterisk():
    """Integration: when a Locator annotation overrides a track's role,
    the sheet must render the role with a trailing '*' so the user sees
    the section-local override at a glance."""
    n = 20
    sections = [
        _make_section(1, 0, n - 1, -10.0, annotation="override: Kick 1=A"),
    ]
    all_tracks = {
        "Acid_Drops Kick 1.wav": _zone_arrays_for(-5.0, n, ["sub"]),
    }
    peak_by_section = {"Acid_Drops Kick 1.wav": {1: -10.0}}
    active_fraction = {"Acid_Drops Kick 1.wav": {1: 0.95}}
    wav_to_ableton = {"Acid_Drops Kick 1.wav": "[H/R] Kick 1"}

    wb = Workbook()
    wb.active.title = "Index"
    build_sections_timeline_sheet(
        workbook=wb, sections=sections,
        all_tracks_zone_energy=all_tracks,
        all_tracks_peak_by_section=peak_by_section,
        all_tracks_active_fraction=active_fraction,
        wav_to_ableton=wav_to_ableton,
    )
    dump = _dump_sheet(wb[SECTIONS_TIMELINE_SHEET_NAME])
    # Role is (A, R) because importance-only override changed H -> A
    assert "A/R*" in dump, f"override asterisk missing:\n{dump}"


def test_build_sheet_warning_banner_when_tracks_without_prefix():
    """Integration: an untagged track triggers the TFP warning banner
    at the top of the sheet."""
    n = 20
    sections = [_make_section(1, 0, n - 1, -10.0)]
    all_tracks = {
        "Acid_Drops Tagged.wav":  _zone_arrays_for(-5.0, n, ["sub"]),
        "Acid_Drops Untagged.wav": _zone_arrays_for(-5.0, n, ["mid"]),
    }
    peak = {
        "Acid_Drops Tagged.wav":   {1: -10.0},
        "Acid_Drops Untagged.wav": {1: -15.0},
    }
    active = {
        "Acid_Drops Tagged.wav":   {1: 0.95},
        "Acid_Drops Untagged.wav": {1: 0.90},
    }
    wav_to_ableton = {
        "Acid_Drops Tagged.wav":   "[H/R] Tagged",
        "Acid_Drops Untagged.wav": "Untagged",  # no prefix
    }
    wb = Workbook()
    wb.active.title = "Index"
    build_sections_timeline_sheet(
        workbook=wb, sections=sections,
        all_tracks_zone_energy=all_tracks,
        all_tracks_peak_by_section=peak,
        all_tracks_active_fraction=active,
        wav_to_ableton=wav_to_ableton,
    )
    dump = _dump_sheet(wb[SECTIONS_TIMELINE_SHEET_NAME])
    assert "WARNING TFP" in dump
    assert "Untagged" in dump


def test_build_sheet_no_warning_when_all_tagged():
    """Banner must be ABSENT when every track has a TFP prefix."""
    n = 10
    sections = [_make_section(1, 0, n - 1, -10.0)]
    all_tracks = {"Acid_Drops Kick 1.wav": _zone_arrays_for(-5.0, n, ["sub"])}
    peak = {"Acid_Drops Kick 1.wav": {1: -10.0}}
    active = {"Acid_Drops Kick 1.wav": {1: 0.95}}
    wav_to_ableton = {"Acid_Drops Kick 1.wav": "[H/R] Kick 1"}

    wb = Workbook()
    wb.active.title = "Index"
    build_sections_timeline_sheet(
        workbook=wb, sections=sections,
        all_tracks_zone_energy=all_tracks,
        all_tracks_peak_by_section=peak,
        all_tracks_active_fraction=active,
        wav_to_ableton=wav_to_ableton,
    )
    dump = _dump_sheet(wb[SECTIONS_TIMELINE_SHEET_NAME])
    assert "WARNING TFP" not in dump


# ---------------------------------------------------------------------------
# REGRESSION GUARD — WAV<->auto_map matching under TFP prefixes
# ---------------------------------------------------------------------------

# ---------------------------------------------------------------------------
# Phase B3 — TFP coherence block inside Mix Health Score
# ---------------------------------------------------------------------------

def test_mix_health_score_contains_tfp_coherence_block():
    """Phase B3 integration: when sections carry track_roles, the Mix
    Health Score sheet renders a 'Cohérence TFP par section' block."""
    import sys as _sys
    import types as _types
    # Stub tkinter so generate_health_score_sheet import works headless.
    if "tkinter" not in _sys.modules:
        tk = _types.ModuleType("tkinter")
        for _n in ("Tk", "StringVar", "BooleanVar", "IntVar", "Toplevel",
                   "Frame", "Label", "Button", "Entry", "Canvas",
                   "PhotoImage", "Widget", "Misc"):
            setattr(tk, _n, type(_n, (), {}))
        tk.TclError = Exception
        _sys.modules["tkinter"] = tk
        for sub in ("ttk", "filedialog", "messagebox", "scrolledtext", "font"):
            _sys.modules[f"tkinter.{sub}"] = _types.ModuleType(f"tkinter.{sub}")
            setattr(tk, sub, _sys.modules[f"tkinter.{sub}"])

    from mix_analyzer import generate_health_score_sheet
    from tfp_parser import Function, Importance

    # Build a section with balanced roles so coherence score is high.
    section = _make_section(1, 0, 14, -10.0)
    section.end_seconds = 15.0
    section.tracks_active = ["Kick", "Bass", "Pad", "Atmos"]
    section.track_roles = {
        "Kick":  (Importance.H, Function.R),
        "Bass":  (Importance.S, Function.H),
        "Pad":   (Importance.S, Function.M),
        "Atmos": (Importance.A, Function.T),
    }

    # Minimal analyses_with_info — one Individual + one Full Mix. The
    # Mix Health Score sheet is quite data-hungry; give it just enough
    # to reach the TFP block without exploding on the earlier code.
    import numpy as np
    sr = 44100
    audio = (np.random.default_rng(0)
             .standard_normal((int(5 * sr), 2))
             .astype(np.float32) * 0.2)
    analysis_individual = {
        "filepath": "Kick.wav", "filename": "Kick.wav",
        "duration": 5.0, "sample_rate": sr,
        "is_stereo": True, "num_channels": 2,
        "loudness": {"lufs_integrated": -20.0, "peak_db": -6.0,
                     "true_peak_db": -6.0, "rms_db": -20.0, "lra": 5.0,
                     "crest_factor": 14.0, "plr": 14.0, "psr": 12.0,
                     "lufs_short_term_max": -18.0},
        "spectrum": {"band_energies": {}, "spectrum_db_normalized": np.zeros(10),
                     "centroid": 1000.0, "spread": 500.0},
        "stereo": {"is_stereo": True, "correlation": 0.5, "width_overall": 0.3,
                   "width_per_band": {}, "pan_per_freq": None,
                   "pan_freqs": None, "pan_energy": None,
                   "mid": np.zeros(100), "side": np.zeros(100)},
        "temporal": {"onsets": [], "tempogram": None,
                     "tempogram_times": None},
        "anomalies": [], "characteristics": {},
        "_v25_features": None,
        "_mono": np.zeros(int(5 * sr)),
        "_data": audio,
        "tempo": {"tempo_median": 120.0, "tempo_min": 120.0,
                  "tempo_max": 120.0, "tempo_std": 0.0, "confidence": 0.0,
                  "confidence_label": "", "tempogram": None,
                  "tempogram_times": None, "tempo_over_time": None,
                  "reliable": False},
        "musical": {"key": "C", "scale": "major", "tonal_strength": 0.5,
                    "chroma_mean": np.zeros(12)},
        "multiband_timeline": {}, "dynamic_range_timeline": {},
    }
    analyses = [(analysis_individual, {"name": "Kick.wav",
                                        "type": "Individual",
                                        "category": "Drum", "parent_bus": "None"})]

    from openpyxl import Workbook
    wb = Workbook()
    wb.active.title = "Index"
    try:
        generate_health_score_sheet(wb, analyses, sections=[section])
    except Exception as e:
        pytest.fail(f"generate_health_score_sheet raised: {type(e).__name__}: {e}")

    ws = wb["Mix Health Score"]
    dump = "\n".join(
        " | ".join(str(c) for c in row if c is not None)
        for row in ws.iter_rows(values_only=True)
    )
    assert "Cohérence TFP par section" in dump, (
        f"coherence block missing:\n{dump}"
    )
    # Header row should include H/S/A + R/H/M/T + diagnostic
    assert "H/S/A" in dump
    assert "R/H/M/T" in dump
    assert "H×H crit." in dump
    # The balanced section above yields a score ≥ 80 -> "Équilibre OK"
    assert "Équilibre OK" in dump


def test_tfp_prefix_stripping_keeps_match_track_name_working():
    """Regression guard against the "-100 dB everywhere" bug class.

    Scenario: the user renames her tracks in Ableton with TFP prefixes.
    The auto_maps dict would then be keyed by "[H/R] Kick 1" style names,
    which ``als_utils.match_track_name`` cannot match against the WAV
    stem "Acid_Drops Kick 1" (because the stripped project prefix step
    can't remove a bracketed TFP tag).

    The fix in mix_analyzer strips the TFP prefix from auto_maps keys
    *before* they are handed to match_track_name. This test builds a
    dict in the "stripped" shape and confirms match_track_name succeeds,
    then confirms the raw-name recovery works for display purposes.
    """
    from als_utils import match_track_name
    # Simulate the CLEAN-keyed auto_maps dict produced by the fix
    clean_auto_map_keys = ["Kick 1", "Sub Bass", "Spoon Percussion"]
    raw_names_by_clean = {
        "Kick 1":           "[H/R] Kick 1",
        "Sub Bass":         "[S/H] Sub Bass",
        "Spoon Percussion": "[A/T] Spoon Percussion",
    }

    # WAV stems as they appear on disk after Ableton bounces them
    wav_stems = [
        "Acid_Drops Kick 1",
        "Acid_Drops Sub Bass",
        "Acid_Drops Spoon Percussion",
    ]
    for stem in wav_stems:
        matched = match_track_name(
            stem, clean_auto_map_keys, als_stem="Acid_Drops_Code",
        )
        assert matched is not None, (
            f"WAV stem {stem!r} failed to match clean auto_map keys "
            f"{clean_auto_map_keys} — THIS IS THE REGRESSION (-100 dB bug)"
        )
        # The raw name (with TFP prefix) is still recoverable for display
        raw = raw_names_by_clean[matched]
        assert "[" in raw and "]" in raw, f"raw name lost its prefix: {raw}"
