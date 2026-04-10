"""
Validation tests for P3.4 — Version Tracking sheet.
Run with: python3 tests/test_p34_version_tracking.py
"""
import sys
import os
import tempfile
import shutil
import numpy as np
import datetime

sys.path.insert(0, os.path.dirname(os.path.dirname(os.path.abspath(__file__))))

# --- Dependencies from mix_analyzer.py ---

FREQ_BANDS = [
    ('sub', 20, 60), ('bass', 60, 250), ('low_mid', 250, 500),
    ('mid', 500, 2000), ('high_mid', 2000, 4000), ('presence', 4000, 8000),
    ('air', 8000, 20000),
]

BAND_LABELS = {
    'sub': 'Sub (20-60 Hz)', 'bass': 'Bass (60-250 Hz)',
    'low_mid': 'Low-Mid (250-500 Hz)', 'mid': 'Mid (500-2000 Hz)',
    'high_mid': 'High-Mid (2-4 kHz)', 'presence': 'Presence (4-8 kHz)',
    'air': 'Air (8-20 kHz)',
}


def _xl_write_header(ws, title, subtitle=''):
    from openpyxl.styles import Font, PatternFill, Alignment
    ws.merge_cells('A1:J1')
    ws['A1'] = title
    ws['A1'].font = Font(name='Calibri', size=16, bold=True, color='00D9FF')
    ws['A1'].fill = PatternFill('solid', fgColor='0A0A12')
    ws['A1'].alignment = Alignment(horizontal='left')
    if subtitle:
        ws.merge_cells('A2:J2')
        ws['A2'] = subtitle
        ws['A2'].font = Font(name='Calibri', size=11, color='8888A0')
        ws['A2'].fill = PatternFill('solid', fgColor='0A0A12')
    return 4


def detect_anomalies(analysis):
    """Minimal replica for testing."""
    anomalies = []
    L = analysis['loudness']
    if L['peak_db'] > -0.3:
        anomalies.append(('critical', 'Clipping risk'))
    elif L['peak_db'] > -1.0:
        anomalies.append(('warning', 'Low headroom'))
    if L.get('true_peak_db', L['peak_db']) > 0.0:
        anomalies.append(('critical', 'ISP clipping'))
    if L.get('crest_factor', 12) < 5 and L.get('rms_db', -20) > -30:
        anomalies.append(('warning', 'Heavy compression'))
    return anomalies


# --- Load functions under test ---

with open(os.path.join(os.path.dirname(os.path.dirname(os.path.abspath(__file__))),
                       'mix_analyzer.py')) as f:
    source = f.read()

# Extract from _calc_loudness_score through generate_excel_report
start = source.find('def _calc_loudness_score(')
end = source.find('\ndef generate_excel_report(')
exec(source[start:end], globals())


# --- Mock data helpers ---

def _make_band_energies(sub=5, bass=20, low_mid=15, mid=25, high_mid=15, presence=12, air=8):
    return {'sub': sub, 'bass': bass, 'low_mid': low_mid, 'mid': mid,
            'high_mid': high_mid, 'presence': presence, 'air': air}


def _make_analysis(name, lufs=-14.0, peak=-3.0, true_peak=-2.5, crest=12.0,
                   plr=13.0, psr=10.0, rms_db=-20.0, width=0.5,
                   correlation=0.9, is_stereo=True, dominant='mid',
                   band_energies=None, peaks=None):
    if band_energies is None:
        band_energies = _make_band_energies()
    if peaks is None:
        peaks = []
    return {
        'filename': name,
        'loudness': {
            'lufs_integrated': lufs, 'peak_db': peak, 'true_peak_db': true_peak,
            'crest_factor': crest, 'plr': plr, 'psr': psr, 'rms_db': rms_db,
        },
        'spectrum': {
            'dominant_band': dominant, 'band_energies': band_energies, 'peaks': peaks,
        },
        'stereo': {
            'is_stereo': is_stereo, 'width_overall': width, 'correlation': correlation,
            'width_per_band': {'sub': 0.1, 'bass': 0.2, 'low_mid': 0.3, 'mid': 0.4,
                               'high_mid': 0.5, 'presence': 0.6, 'air': 0.7},
        },
    }


def make_tracks():
    """Standard test set: 3 Individual + 1 Full Mix."""
    tracks = []
    for i, name in enumerate(['Kick', 'Snare', 'Bass']):
        tracks.append((_make_analysis(name, lufs=-14.0 - i, crest=12.0 + i),
                       {'type': 'Individual', 'category': 'Instruments'}))
    tracks.append((_make_analysis('Full Mix', lufs=-14.0, peak=-1.5, true_peak=-1.2,
                                   crest=11.0, plr=12.0, width=0.55),
                   {'type': 'Full Mix', 'category': ''}))
    return tracks


def create_fake_previous_report(output_folder, song_name, date_str,
                                 fm_lufs=-14.5, fm_peak=-2.0, fm_crest=10.0,
                                 fm_width=0.5, ind_crests=None, track_count=3):
    """Create a minimal .xlsx that mimics a previous Mix Analyzer report."""
    from openpyxl import Workbook
    from openpyxl.styles import Font

    wb = Workbook()
    ws = wb.active
    ws.title = 'Index'

    # Summary sheet
    ws_sum = wb.create_sheet('Summary')
    # Header row 4
    headers = ['Track', 'Type', 'Category', 'LUFS', 'Peak (dB)', 'Crest (dB)',
               'Stereo Width', 'Dom. Band', 'Centroid (Hz)', 'Duration (s)']
    for col, h in enumerate(headers, 1):
        ws_sum.cell(row=4, column=col, value=h)

    row = 5
    # Individual tracks
    if ind_crests is None:
        ind_crests = [11.0, 12.0, 13.0]
    for i, crest in enumerate(ind_crests):
        ws_sum.cell(row=row, column=1, value=f'Track_{i+1}')
        ws_sum.cell(row=row, column=2, value='Individual')
        ws_sum.cell(row=row, column=3, value='Instruments')
        ws_sum.cell(row=row, column=4, value=-14.0 - i)
        ws_sum.cell(row=row, column=5, value=-3.0)
        ws_sum.cell(row=row, column=6, value=crest)
        ws_sum.cell(row=row, column=7, value=0.4)
        row += 1

    # Full Mix
    ws_sum.cell(row=row, column=1, value='Full Mix')
    ws_sum.cell(row=row, column=2, value='Full Mix')
    ws_sum.cell(row=row, column=4, value=fm_lufs)
    ws_sum.cell(row=row, column=5, value=fm_peak)
    ws_sum.cell(row=row, column=6, value=fm_crest)
    ws_sum.cell(row=row, column=7, value=fm_width)

    # Anomalies sheet (empty = 0 anomalies)
    ws_anom = wb.create_sheet('Anomalies')
    ws_anom.cell(row=4, column=1, value='Track')

    fname = f'{song_name}_MixAnalyzer_{date_str}.xlsx'
    path = os.path.join(output_folder, fname)
    wb.save(path)
    return path


# --- Tests ---

def test_sheet_creation():
    """Sheet 'Version Tracking' is created."""
    from openpyxl import Workbook
    wb = Workbook()
    generate_version_tracking_sheet(wb, make_tracks(), log_fn=lambda m: None)
    assert 'Version Tracking' in wb.sheetnames


def test_tab_color():
    """Tab color is cyan (#3DAAFF)."""
    from openpyxl import Workbook
    wb = Workbook()
    generate_version_tracking_sheet(wb, make_tracks(), log_fn=lambda m: None)
    ws = wb['Version Tracking']
    assert '3DAAFF' in ws.sheet_properties.tabColor.rgb


def test_title():
    """Title 'VERSION TRACKING' in A1."""
    from openpyxl import Workbook
    wb = Workbook()
    generate_version_tracking_sheet(wb, make_tracks(), log_fn=lambda m: None)
    ws = wb['Version Tracking']
    assert ws['A1'].value == 'VERSION TRACKING'


def test_subtitle_with_song_name():
    """Subtitle includes the song name when provided."""
    from openpyxl import Workbook
    wb = Workbook()
    generate_version_tracking_sheet(wb, make_tracks(), song_name='MySong', log_fn=lambda m: None)
    ws = wb['Version Tracking']
    assert 'MySong' in ws['A3'].value


def test_version_count_single():
    """With no previous reports, shows 1 version (current only)."""
    from openpyxl import Workbook
    wb = Workbook()
    generate_version_tracking_sheet(wb, make_tracks(), log_fn=lambda m: None)
    ws = wb['Version Tracking']
    assert '1' in ws['A5'].value


def test_first_report_message():
    """First report shows explanation message in A7."""
    from openpyxl import Workbook
    wb = Workbook()
    generate_version_tracking_sheet(wb, make_tracks(), log_fn=lambda m: None)
    ws = wb['Version Tracking']
    assert ws['A7'] is not None
    assert 'No previous versions' in str(ws['A7'].value)


def test_metric_header_row():
    """Header row 8 has 'Metric' and column headers."""
    from openpyxl import Workbook
    wb = Workbook()
    generate_version_tracking_sheet(wb, make_tracks(), log_fn=lambda m: None)
    ws = wb['Version Tracking']
    assert ws.cell(row=8, column=1).value == 'Metric'
    # With single version: col 2 = version date, col 3 = delta, col 4 = %, col 5 = trend
    header_b = ws.cell(row=8, column=2).value
    assert '(current)' in header_b


def test_nine_tracked_metrics():
    """9 metric rows present (rows 9-17)."""
    from openpyxl import Workbook
    wb = Workbook()
    generate_version_tracking_sheet(wb, make_tracks(), log_fn=lambda m: None)
    ws = wb['Version Tracking']
    expected = [
        'Full Mix LUFS', 'Full Mix True Peak (dBFS)', 'Full Mix Crest (dB)',
        'Full Mix PLR', 'Full Mix Width', 'Avg Individual Crest (dB)',
        'Anomaly count', 'Mix Health Score', 'Track count',
    ]
    for i, name in enumerate(expected):
        row = 9 + i
        assert ws.cell(row=row, column=1).value == name, \
            f"Row {row}: expected '{name}', got '{ws.cell(row=row, column=1).value}'"


def test_current_metrics_values():
    """Current version metrics are populated with correct values."""
    from openpyxl import Workbook
    wb = Workbook()
    generate_version_tracking_sheet(wb, make_tracks(), log_fn=lambda m: None)
    ws = wb['Version Tracking']
    # Full Mix LUFS = -14.0 (row 9, col 2)
    assert ws.cell(row=9, column=2).value == -14.0
    # Full Mix Crest = 11.0 (row 11, col 2)
    assert ws.cell(row=11, column=2).value == 11.0
    # Track count = 3 (row 17, col 2)
    assert ws.cell(row=17, column=2).value == 3


def test_single_version_deltas_are_dashes():
    """With only 1 version, delta and trend columns show '—'."""
    from openpyxl import Workbook
    wb = Workbook()
    generate_version_tracking_sheet(wb, make_tracks(), log_fn=lambda m: None)
    ws = wb['Version Tracking']
    # Delta abs col = 3 (n_versions=1, so col 1+1+1=3)
    assert ws.cell(row=9, column=3).value == '—'
    # Trend col = 5
    assert ws.cell(row=9, column=5).value == '—'


def test_freeze_panes():
    """Freeze panes at A9."""
    from openpyxl import Workbook
    wb = Workbook()
    generate_version_tracking_sheet(wb, make_tracks(), log_fn=lambda m: None)
    ws = wb['Version Tracking']
    assert ws.freeze_panes == 'A9'


def test_column_widths():
    """Column A width is 30."""
    from openpyxl import Workbook
    wb = Workbook()
    generate_version_tracking_sheet(wb, make_tracks(), log_fn=lambda m: None)
    ws = wb['Version Tracking']
    assert ws.column_dimensions['A'].width == 30


def test_source_files_section():
    """Source files section lists the current report."""
    from openpyxl import Workbook
    wb = Workbook()
    generate_version_tracking_sheet(wb, make_tracks(), log_fn=lambda m: None)
    ws = wb['Version Tracking']
    # Find "Source files" header
    found = False
    for row in range(18, 35):
        val = ws.cell(row=row, column=1).value
        if val and 'Source files' in str(val):
            found = True
            break
    assert found, "Source files section not found"


def test_sparkline_note():
    """Sparkline note is present."""
    from openpyxl import Workbook
    wb = Workbook()
    generate_version_tracking_sheet(wb, make_tracks(), log_fn=lambda m: None)
    ws = wb['Version Tracking']
    found = False
    for row in range(18, 40):
        val = ws.cell(row=row, column=1).value
        if val and 'Sparkline' in str(val):
            found = True
            break
    assert found, "Sparkline note not found"


# --- Helper function tests ---

def test_compute_current_metrics():
    """_compute_current_metrics returns expected keys."""
    metrics = _compute_current_metrics(make_tracks())
    assert 'Full Mix LUFS' in metrics
    assert 'Full Mix True Peak (dBFS)' in metrics
    assert 'Full Mix Crest (dB)' in metrics
    assert 'Full Mix PLR' in metrics
    assert 'Full Mix Width' in metrics
    assert 'Avg Individual Crest (dB)' in metrics
    assert 'Anomaly count' in metrics
    assert 'Mix Health Score' in metrics
    assert 'Track count' in metrics
    assert metrics['Full Mix LUFS'] == -14.0
    assert metrics['Track count'] == 3


def test_compute_current_metrics_no_fullmix():
    """_compute_current_metrics works without Full Mix."""
    tracks = [(_make_analysis(f'T{i}', crest=12.0), {'type': 'Individual', 'category': 'X'})
              for i in range(3)]
    metrics = _compute_current_metrics(tracks)
    assert 'Full Mix LUFS' not in metrics
    assert 'Avg Individual Crest (dB)' in metrics
    assert metrics['Track count'] == 3


def test_compute_trend_higher_is_better():
    """Trend: higher is better metrics."""
    assert _compute_trend(50, 80, 'Mix Health Score') == '↗'
    assert _compute_trend(80, 50, 'Mix Health Score') == '↘'
    assert _compute_trend(80, 81, 'Mix Health Score') == '→'  # < 5% change


def test_compute_trend_lower_is_better():
    """Trend: lower is better metrics."""
    assert _compute_trend(5, 0, 'Anomaly count') == '↗'  # fewer anomalies = good
    assert _compute_trend(0, 5, 'Anomaly count') == '→'  # first_val=0 => stable
    assert _compute_trend(-0.5, -1.5, 'Full Mix True Peak (dBFS)') == '↗'  # lower peak = good


def test_compute_trend_target_metrics():
    """Trend: closer-to-target metrics."""
    # LUFS target = -14. Moving from -18 to -15 = closer = improving
    assert _compute_trend(-18.0, -15.0, 'Full Mix LUFS') == '↗'
    # Moving from -14 to -18 = further = worsening
    assert _compute_trend(-14.0, -18.0, 'Full Mix LUFS') == '↘'
    # Moving from -14.1 to -14.0 = barely moving
    assert _compute_trend(-14.1, -14.0, 'Full Mix LUFS') == '→'


def test_compute_trend_none_values():
    """Trend: None values return '—'."""
    assert _compute_trend(None, 80, 'Mix Health Score') == '—'
    assert _compute_trend(80, None, 'Mix Health Score') == '—'


def test_find_previous_reports():
    """_find_previous_reports finds matching files sorted by date."""
    tmpdir = tempfile.mkdtemp()
    try:
        # Create fake report files
        for date in ['2025-01-01', '2025-03-15', '2025-02-10']:
            path = os.path.join(tmpdir, f'MySong_MixAnalyzer_{date}.xlsx')
            open(path, 'w').close()
        # Non-matching file
        open(os.path.join(tmpdir, 'OtherSong_MixAnalyzer_2025-01-05.xlsx'), 'w').close()

        results = _find_previous_reports(tmpdir, 'MySong')
        assert len(results) == 3
        assert results[0][0] == '2025-01-01'  # sorted chronologically
        assert results[1][0] == '2025-02-10'
        assert results[2][0] == '2025-03-15'
    finally:
        shutil.rmtree(tmpdir)


def test_find_previous_reports_empty():
    """_find_previous_reports returns empty list when no matches."""
    tmpdir = tempfile.mkdtemp()
    try:
        results = _find_previous_reports(tmpdir, 'MySong')
        assert results == []
    finally:
        shutil.rmtree(tmpdir)


def test_extract_metrics_from_report():
    """_extract_metrics_from_report reads metrics from a fake previous report."""
    tmpdir = tempfile.mkdtemp()
    try:
        path = create_fake_previous_report(tmpdir, 'MySong', '2025-01-01',
                                            fm_lufs=-14.5, fm_peak=-2.0, fm_crest=10.0)
        metrics = _extract_metrics_from_report(path)
        assert metrics is not None
        assert metrics['Full Mix LUFS'] == -14.5
        assert metrics['Full Mix Peak (dBFS)'] == -2.0
        assert metrics['Full Mix Crest (dB)'] == 10.0
        assert metrics['Track count'] == 3
        assert metrics['Anomaly count'] == 0
    finally:
        shutil.rmtree(tmpdir)


def test_extract_metrics_corrupt_file():
    """_extract_metrics_from_report returns None for corrupt files."""
    tmpdir = tempfile.mkdtemp()
    try:
        path = os.path.join(tmpdir, 'corrupt.xlsx')
        with open(path, 'w') as f:
            f.write('not a real xlsx')
        result = _extract_metrics_from_report(path, log_fn=lambda m: None)
        assert result is None
    finally:
        shutil.rmtree(tmpdir)


def test_multiple_versions_with_deltas():
    """With previous reports, deltas and trends are computed."""
    from openpyxl import Workbook
    tmpdir = tempfile.mkdtemp()
    try:
        create_fake_previous_report(tmpdir, 'MySong', '2025-01-01',
                                     fm_lufs=-16.0, fm_crest=9.0)
        wb = Workbook()
        generate_version_tracking_sheet(wb, make_tracks(),
                                         output_folder=tmpdir, song_name='MySong',
                                         log_fn=lambda m: None)
        ws = wb['Version Tracking']

        # Should have 2 versions (previous + current)
        assert '2' in ws['A5'].value

        # Version 1 date column (col B) = 2025-01-01
        header_b = ws.cell(row=8, column=2).value
        assert '2025-01-01' in header_b

        # Version 2 (current) = col C
        header_c = ws.cell(row=8, column=3).value
        assert '(current)' in header_c

        # Delta column = col 4, trend = col 6
        # Full Mix LUFS: -16.0 -> -14.0 (closer to -14 target) => trend should be '↗'
        trend_lufs = ws.cell(row=9, column=6).value
        assert trend_lufs == '↗', f"Expected ↗ for LUFS improving, got {trend_lufs}"

        # Delta abs for LUFS (col 4) = -14.0 - (-16.0) = 2.0
        delta_lufs = ws.cell(row=9, column=4).value
        assert delta_lufs == 2.0
    finally:
        shutil.rmtree(tmpdir)


def test_no_output_folder():
    """Works gracefully when output_folder is None."""
    from openpyxl import Workbook
    wb = Workbook()
    generate_version_tracking_sheet(wb, make_tracks(), output_folder=None,
                                     song_name=None, log_fn=lambda m: None)
    ws = wb['Version Tracking']
    assert '1' in ws['A5'].value  # only current version


def test_no_tracks():
    """Works with empty analyses."""
    from openpyxl import Workbook
    wb = Workbook()
    generate_version_tracking_sheet(wb, [], log_fn=lambda m: None)
    ws = wb['Version Tracking']
    assert ws['A1'].value == 'VERSION TRACKING'


def test_source_files_list_previous():
    """Source files section lists previous report files."""
    from openpyxl import Workbook
    tmpdir = tempfile.mkdtemp()
    try:
        create_fake_previous_report(tmpdir, 'MySong', '2025-01-01')
        wb = Workbook()
        generate_version_tracking_sheet(wb, make_tracks(),
                                         output_folder=tmpdir, song_name='MySong',
                                         log_fn=lambda m: None)
        ws = wb['Version Tracking']
        # Find source file entries
        found_prev = False
        found_current = False
        for row in range(18, 40):
            val = ws.cell(row=row, column=2).value
            if val and 'MySong_MixAnalyzer_2025-01-01' in str(val):
                found_prev = True
            status = ws.cell(row=row, column=3).value
            if status and 'in-memory' in str(status):
                found_current = True
        assert found_prev, "Previous report not in source files"
        assert found_current, "Current report not in source files"
    finally:
        shutil.rmtree(tmpdir)


def test_excel_save_roundtrip():
    """Workbook saves and re-opens without error."""
    from openpyxl import Workbook, load_workbook
    wb = Workbook()
    generate_version_tracking_sheet(wb, make_tracks(), log_fn=lambda m: None)
    tmp = tempfile.mktemp(suffix='.xlsx')
    try:
        wb.save(tmp)
        assert os.path.getsize(tmp) > 0
        wb2 = load_workbook(tmp)
        assert 'Version Tracking' in wb2.sheetnames
        ws2 = wb2['Version Tracking']
        assert ws2['A1'].value == 'VERSION TRACKING'
        assert ws2.freeze_panes == 'A9'
    finally:
        os.unlink(tmp)


def test_excel_roundtrip_with_previous():
    """Full roundtrip with previous reports saves correctly."""
    from openpyxl import Workbook, load_workbook
    tmpdir = tempfile.mkdtemp()
    try:
        create_fake_previous_report(tmpdir, 'Song', '2025-01-15', fm_lufs=-15.0)
        create_fake_previous_report(tmpdir, 'Song', '2025-02-20', fm_lufs=-14.5)
        wb = Workbook()
        generate_version_tracking_sheet(wb, make_tracks(),
                                         output_folder=tmpdir, song_name='Song',
                                         log_fn=lambda m: None)
        tmp = os.path.join(tmpdir, 'test_output.xlsx')
        wb.save(tmp)
        wb2 = load_workbook(tmp)
        ws2 = wb2['Version Tracking']
        assert '3' in ws2['A5'].value  # 2 previous + 1 current
    finally:
        shutil.rmtree(tmpdir)


# --- Runner ---

if __name__ == '__main__':
    tests = [
        test_sheet_creation,
        test_tab_color,
        test_title,
        test_subtitle_with_song_name,
        test_version_count_single,
        test_first_report_message,
        test_metric_header_row,
        test_nine_tracked_metrics,
        test_current_metrics_values,
        test_single_version_deltas_are_dashes,
        test_freeze_panes,
        test_column_widths,
        test_source_files_section,
        test_sparkline_note,
        test_compute_current_metrics,
        test_compute_current_metrics_no_fullmix,
        test_compute_trend_higher_is_better,
        test_compute_trend_lower_is_better,
        test_compute_trend_target_metrics,
        test_compute_trend_none_values,
        test_find_previous_reports,
        test_find_previous_reports_empty,
        test_extract_metrics_from_report,
        test_extract_metrics_corrupt_file,
        test_multiple_versions_with_deltas,
        test_no_output_folder,
        test_no_tracks,
        test_source_files_list_previous,
        test_excel_save_roundtrip,
        test_excel_roundtrip_with_previous,
    ]

    passed = 0
    failed = 0
    for test in tests:
        try:
            test()
            print(f'  [OK] {test.__doc__}')
            passed += 1
        except Exception as e:
            print(f'  [FAIL] {test.__doc__}')
            print(f'         {e}')
            failed += 1

    print(f'\n{passed}/{passed + failed} tests passed.')
    if failed:
        sys.exit(1)
