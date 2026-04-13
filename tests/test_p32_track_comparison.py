"""
Validation tests for P3.2 — Track Comparison Tool sheet.
Run with: python3 tests/test_p32_track_comparison.py
"""
import sys
import os
import tempfile
import numpy as np

sys.path.insert(0, os.path.dirname(os.path.dirname(os.path.abspath(__file__))))

# --- Dependencies from mix_analyzer.py ---

FREQ_BANDS = [
    ('sub', 20, 60), ('bass', 60, 250), ('low_mid', 250, 500),
    ('mid', 500, 2000), ('high_mid', 2000, 4000), ('presence', 4000, 8000),
    ('air', 8000, 20000),
]

BAND_LABELS = {
    'sub': 'Sub (20-60 Hz)', 'bass': 'Bass (60-250 Hz)',
    'low_mid': 'Low-Mid (250-500 Hz)', 'mid': 'Mid (500-2000 Hz)',
    'high_mid': 'High-Mid (2-4 kHz)', 'presence': 'Presence (4-8 kHz)',
    'air': 'Air (8-20 kHz)',
}


def _xl_write_header(ws, title, subtitle=''):
    from openpyxl.styles import Font, PatternFill, Alignment
    header_fill = PatternFill('solid', fgColor='0A0A12')
    accent_font = Font(name='Calibri', size=16, bold=True, color='00D9FF')
    sub_font = Font(name='Calibri', size=11, color='8888A0')
    ws.merge_cells('A1:J1')
    ws['A1'] = title
    ws['A1'].font = accent_font
    ws['A1'].fill = header_fill
    ws['A1'].alignment = Alignment(horizontal='left')
    if subtitle:
        ws.merge_cells('A2:J2')
        ws['A2'] = subtitle
        ws['A2'].font = sub_font
        ws['A2'].fill = header_fill
    return 4


# --- Load function under test ---

with open(os.path.join(os.path.dirname(os.path.dirname(os.path.abspath(__file__))),
                       'mix_analyzer.py')) as f:
    source = f.read()

# Extract METRIC_GLOSSARY and ANOMALY_COMMENTS dictionaries
mg_start = source.find('METRIC_GLOSSARY = {')
mg_end = source.find('\n\ndef ', mg_start + 1)
exec(source[mg_start:mg_end], globals())

acl_start = source.find('def _apply_clean_layout(')
acl_end = source.find('\n\ndef ', acl_start + 1)
exec(source[acl_start:acl_end], globals())

# Extract _xl_add_sheet_nav helper (M7.5)
xsn_start = source.find('def _xl_add_sheet_nav(')
xsn_end = source.find('\n\ndef ', xsn_start + 1)
exec(source[xsn_start:xsn_end], globals())

# Extract _xl_add_comment helper
xac_start = source.find('def _xl_add_comment(')
xac_end = source.find('\n\ndef ', xac_start + 1)
exec(source[xac_start:xac_end], globals())

start = source.find('def generate_track_comparison_sheet(')
end = source.find('\ndef generate_excel_report(')
exec(source[start:end], globals())


# --- Mock data helpers ---

def _make_analysis(name, lufs=-14.0, peak=-1.0, crest=12.0, width=0.5,
                   plr=13.0, psr=10.0, dominant='mid', band_energies=None,
                   is_stereo=True):
    """Build a minimal analysis dict matching the structure used by P3.2."""
    if band_energies is None:
        band_energies = {n: round(100.0 / 7, 2) for n, _, _ in FREQ_BANDS}
    return {
        'filename': name,
        'loudness': {
            'lufs_integrated': lufs,
            'peak_db': peak,
            'crest_factor': crest,
            'plr': plr,
            'psr': psr,
        },
        'spectrum': {
            'dominant_band': dominant,
            'band_energies': band_energies,
        },
        'stereo': {
            'is_stereo': is_stereo,
            'width_overall': width,
        },
    }


def make_mock_tracks(n=5):
    """Create n Individual tracks with slightly different metrics."""
    tracks = []
    for i in range(n):
        a = _make_analysis(
            f'Track_{i+1}',
            lufs=-14.0 - i,
            peak=-1.0 - i * 0.5,
            crest=12.0 + i,
            width=0.3 + i * 0.1,
            dominant=['sub', 'bass', 'mid', 'high_mid', 'presence'][i % 5],
        )
        tracks.append((a, {'type': 'Individual', 'category': 'Instruments'}))
    return tracks


def make_full_test_set():
    """Individual + BUS + Full Mix tracks."""
    tracks = make_mock_tracks(5)
    tracks.append((_make_analysis('Drums Bus', lufs=-10),
                   {'type': 'BUS', 'category': 'Drums'}))
    tracks.append((_make_analysis('Full Mix Bounce', lufs=-12),
                   {'type': 'Full Mix', 'category': ''}))
    return tracks


# --- Tests ---

def test_sheet_creation():
    """Sheet 'Track Comparison' is created in the workbook."""
    from openpyxl import Workbook
    wb = Workbook()
    generate_track_comparison_sheet(wb, make_mock_tracks(), log_fn=lambda m: None)
    assert 'Track Comparison' in wb.sheetnames


def test_hidden_data_sheet():
    """Hidden '_track_data' sheet is created with correct data."""
    from openpyxl import Workbook
    wb = Workbook()
    generate_track_comparison_sheet(wb, make_mock_tracks(3), log_fn=lambda m: None)
    assert '_track_data' in wb.sheetnames
    ws_data = wb['_track_data']
    assert ws_data.sheet_state == 'hidden'
    # Header row
    assert ws_data.cell(row=1, column=1).value == 'Track Name'
    assert ws_data.cell(row=1, column=2).value == 'LUFS'
    # Data rows: 3 tracks
    assert ws_data.cell(row=2, column=1).value == 'Track_1'
    assert ws_data.cell(row=3, column=1).value == 'Track_2'
    assert ws_data.cell(row=4, column=1).value == 'Track_3'
    assert ws_data.cell(row=5, column=1).value is None  # no 4th track


def test_data_sheet_metrics():
    """Data sheet contains all 14 metric columns (7 core + 7 bands)."""
    from openpyxl import Workbook
    wb = Workbook()
    generate_track_comparison_sheet(wb, make_mock_tracks(2), log_fn=lambda m: None)
    ws_data = wb['_track_data']
    # 1=Track Name, 2=LUFS, 3=Peak, 4=Crest, 5=Width, 6=PLR, 7=PSR,
    # 8=Dominant Band, 9-15=7 band energies => 15 columns
    assert ws_data.cell(row=1, column=8).value == 'Dominant Band'
    assert ws_data.cell(row=1, column=9).value == 'Sub Energy %'
    assert ws_data.cell(row=1, column=15).value == 'Air Energy %'
    assert ws_data.cell(row=1, column=16).value is None  # no more columns


def test_data_sheet_values():
    """Data sheet values match the analysis input."""
    from openpyxl import Workbook
    wb = Workbook()
    tracks = [(_make_analysis('Kick', lufs=-14.5, peak=-0.8, crest=11.0,
                              width=0.25, plr=13.7, psr=9.5, dominant='sub'),
               {'type': 'Individual', 'category': 'Drums'})]
    generate_track_comparison_sheet(wb, tracks, log_fn=lambda m: None)
    ws_data = wb['_track_data']
    assert ws_data.cell(row=2, column=1).value == 'Kick'
    assert ws_data.cell(row=2, column=2).value == -14.5   # LUFS
    assert ws_data.cell(row=2, column=3).value == -0.8    # Peak
    assert ws_data.cell(row=2, column=4).value == 11.0    # Crest
    assert ws_data.cell(row=2, column=5).value == 0.25    # Width
    assert ws_data.cell(row=2, column=6).value == 13.7    # PLR
    assert ws_data.cell(row=2, column=7).value == 9.5     # PSR
    assert ws_data.cell(row=2, column=8).value == 'Sub (20-60 Hz)'  # Dominant Band


def test_track_filtering():
    """Only Individual tracks appear; BUS and Full Mix are excluded."""
    from openpyxl import Workbook
    wb = Workbook()
    tracks = make_full_test_set()  # 5 Individual + 1 BUS + 1 Full Mix
    generate_track_comparison_sheet(wb, tracks, log_fn=lambda m: None)
    ws_data = wb['_track_data']
    # Should have exactly 5 data rows
    names = []
    for r in range(2, 20):
        v = ws_data.cell(row=r, column=1).value
        if v is not None:
            names.append(v)
    assert len(names) == 5
    assert 'Drums Bus' not in names
    assert 'Full Mix Bounce' not in names


def test_selector_defaults():
    """Track selectors B2-B5 have correct default values."""
    from openpyxl import Workbook
    wb = Workbook()
    generate_track_comparison_sheet(wb, make_mock_tracks(4), log_fn=lambda m: None)
    ws = wb['Track Comparison']
    assert ws['B2'].value == 'Track_1'  # Track A = first
    assert ws['B3'].value == 'Track_2'  # Track B = second
    assert ws['B4'].value == ''         # Track C = empty
    assert ws['B5'].value == ''         # Track D = empty


def test_selector_labels():
    """Track selector labels in column A."""
    from openpyxl import Workbook
    wb = Workbook()
    generate_track_comparison_sheet(wb, make_mock_tracks(), log_fn=lambda m: None)
    ws = wb['Track Comparison']
    assert ws['A2'].value == 'Track A (reference)'
    assert ws['A3'].value == 'Track B'
    assert ws['A4'].value == 'Track C (optional)'
    assert ws['A5'].value == 'Track D (optional)'


def test_note_row():
    """Note about reference track is present."""
    from openpyxl import Workbook
    wb = Workbook()
    generate_track_comparison_sheet(wb, make_mock_tracks(), log_fn=lambda m: None)
    ws = wb['Track Comparison']
    assert 'reference' in ws['A6'].value.lower()


def test_data_validations():
    """Data validations are present on B2-B5."""
    from openpyxl import Workbook
    wb = Workbook()
    generate_track_comparison_sheet(wb, make_mock_tracks(3), log_fn=lambda m: None)
    ws = wb['Track Comparison']
    dvs = ws.data_validations.dataValidation
    assert len(dvs) == 2  # required + optional
    # Collect all cells covered by validations
    all_cells = set()
    for dv in dvs:
        for cell_range in dv.sqref.ranges:
            all_cells.add(str(cell_range))
    assert 'B2' in all_cells
    assert 'B3' in all_cells
    assert 'B4' in all_cells
    assert 'B5' in all_cells


def test_comparison_headers():
    """Comparison table headers in row 9."""
    from openpyxl import Workbook
    wb = Workbook()
    generate_track_comparison_sheet(wb, make_mock_tracks(), log_fn=lambda m: None)
    ws = wb['Track Comparison']
    assert ws.cell(row=9, column=1).value == 'Metric'
    assert ws.cell(row=9, column=2).value == 'Track A'
    assert ws.cell(row=9, column=3).value == 'Track B'
    assert ws.cell(row=9, column=4).value == 'Δ B-A'
    assert ws.cell(row=9, column=5).value == '% B-A'
    assert ws.cell(row=9, column=6).value == 'Track C'
    assert ws.cell(row=9, column=9).value == 'Track D'
    assert ws.cell(row=9, column=11).value == '% D-A'


def test_metric_rows():
    """14 metric rows present (7 core + 7 bands)."""
    from openpyxl import Workbook
    wb = Workbook()
    generate_track_comparison_sheet(wb, make_mock_tracks(), log_fn=lambda m: None)
    ws = wb['Track Comparison']
    # Row 10 = LUFS, Row 23 = Air Energy %
    assert ws.cell(row=10, column=1).value == 'LUFS'
    assert ws.cell(row=11, column=1).value == 'Peak (dBFS)'
    assert ws.cell(row=12, column=1).value == 'Crest Factor (dB)'
    assert ws.cell(row=13, column=1).value == 'Stereo Width'
    assert ws.cell(row=14, column=1).value == 'PLR (dB)'
    assert ws.cell(row=15, column=1).value == 'PSR (dB)'
    assert ws.cell(row=16, column=1).value == 'Dominant Band'
    assert ws.cell(row=17, column=1).value == 'Sub Energy %'
    assert ws.cell(row=23, column=1).value == 'Air Energy %'
    # Nothing after last metric
    assert ws.cell(row=24, column=1).value is None


def test_index_match_formulas():
    """Track value cells use INDEX/MATCH formulas referencing _track_data."""
    from openpyxl import Workbook
    wb = Workbook()
    generate_track_comparison_sheet(wb, make_mock_tracks(), log_fn=lambda m: None)
    ws = wb['Track Comparison']
    # Track A LUFS (B10)
    formula_a = ws.cell(row=10, column=2).value
    assert 'INDEX' in formula_a and 'MATCH' in formula_a
    assert '_track_data' in formula_a
    assert '$B$2' in formula_a
    # Track B LUFS (C10)
    formula_b = ws.cell(row=10, column=3).value
    assert 'INDEX' in formula_b and 'MATCH' in formula_b
    assert '$B$3' in formula_b


def test_delta_formulas():
    """Delta columns contain subtraction formulas."""
    from openpyxl import Workbook
    wb = Workbook()
    generate_track_comparison_sheet(wb, make_mock_tracks(), log_fn=lambda m: None)
    ws = wb['Track Comparison']
    # Δ B-A for LUFS (D10)
    delta = ws.cell(row=10, column=4).value
    assert 'C10' in delta and 'B10' in delta
    # % B-A (E10)
    pct = ws.cell(row=10, column=5).value
    assert 'ABS' in pct or 'D10' in pct


def test_optional_track_conditional_formulas():
    """Track C/D formulas are wrapped in IF($B$4="",...) / IF($B$5="",...) ."""
    from openpyxl import Workbook
    wb = Workbook()
    generate_track_comparison_sheet(wb, make_mock_tracks(), log_fn=lambda m: None)
    ws = wb['Track Comparison']
    # Track C value (F10)
    formula_c = ws.cell(row=10, column=6).value
    assert '$B$4' in formula_c and 'IF' in formula_c
    # Track D value (I10)
    formula_d = ws.cell(row=10, column=9).value
    assert '$B$5' in formula_d and 'IF' in formula_d


def test_dominant_band_text_delta():
    """Dominant Band delta shows 'match'/'differ' instead of numeric delta."""
    from openpyxl import Workbook
    wb = Workbook()
    generate_track_comparison_sheet(wb, make_mock_tracks(), log_fn=lambda m: None)
    ws = wb['Track Comparison']
    # Row 16 = Dominant Band
    delta_dom = ws.cell(row=16, column=4).value
    assert 'match' in delta_dom and 'differ' in delta_dom


def test_freeze_panes():
    """Freeze panes at A10."""
    from openpyxl import Workbook
    wb = Workbook()
    generate_track_comparison_sheet(wb, make_mock_tracks(), log_fn=lambda m: None)
    ws = wb['Track Comparison']
    assert ws.freeze_panes == 'A10'


def test_tab_color():
    """Tab color is orange (#FF8B3D)."""
    from openpyxl import Workbook
    wb = Workbook()
    generate_track_comparison_sheet(wb, make_mock_tracks(), log_fn=lambda m: None)
    ws = wb['Track Comparison']
    assert ws.sheet_properties.tabColor.rgb == '00FF8B3D' or ws.sheet_properties.tabColor.index == 'FF8B3D'


def test_conditional_formatting():
    """Conditional formatting rules are applied (delta color scales + band energy scales)."""
    from openpyxl import Workbook
    wb = Workbook()
    generate_track_comparison_sheet(wb, make_mock_tracks(), log_fn=lambda m: None)
    ws = wb['Track Comparison']
    n_rules = len(ws.conditional_formatting._cf_rules)
    # 3 delta columns + 4 band energy columns = 7 rules minimum
    assert n_rules >= 7, f"Expected >= 7 CF rules, got {n_rules}"


def test_column_widths():
    """Column widths are set."""
    from openpyxl import Workbook
    wb = Workbook()
    generate_track_comparison_sheet(wb, make_mock_tracks(), log_fn=lambda m: None)
    ws = wb['Track Comparison']
    assert ws.column_dimensions['A'].width == 20
    assert ws.column_dimensions['B'].width == 12


def test_empty_individuals():
    """Graceful handling when no Individual tracks exist."""
    from openpyxl import Workbook
    wb = Workbook()
    tracks = [(_make_analysis('Bus'), {'type': 'BUS', 'category': 'X'})]
    generate_track_comparison_sheet(wb, tracks, log_fn=lambda m: None)
    ws = wb['Track Comparison']
    assert ws is not None
    # Should still have _track_data not created
    assert '_track_data' not in wb.sheetnames


def test_single_track():
    """Works with a single Individual track (Track B defaults to Track A)."""
    from openpyxl import Workbook
    wb = Workbook()
    tracks = [(_make_analysis('OnlyTrack'), {'type': 'Individual', 'category': 'X'})]
    generate_track_comparison_sheet(wb, tracks, log_fn=lambda m: None)
    ws = wb['Track Comparison']
    assert ws['B2'].value == 'OnlyTrack'
    assert ws['B3'].value == 'OnlyTrack'  # fallback when only 1 track


def test_excel_save_roundtrip():
    """Workbook saves and re-opens without error."""
    from openpyxl import Workbook, load_workbook
    wb = Workbook()
    generate_track_comparison_sheet(wb, make_full_test_set(), log_fn=lambda m: None)

    tmp = tempfile.mktemp(suffix='.xlsx')
    try:
        wb.save(tmp)
        assert os.path.getsize(tmp) > 0
        wb2 = load_workbook(tmp)
        assert 'Track Comparison' in wb2.sheetnames
        assert '_track_data' in wb2.sheetnames
        ws2 = wb2['Track Comparison']
        assert ws2['B2'].value == 'Track_1'
        assert ws2.freeze_panes == 'A10'
    finally:
        os.unlink(tmp)


# --- Runner ---

if __name__ == '__main__':
    tests = [
        test_sheet_creation,
        test_hidden_data_sheet,
        test_data_sheet_metrics,
        test_data_sheet_values,
        test_track_filtering,
        test_selector_defaults,
        test_selector_labels,
        test_note_row,
        test_data_validations,
        test_comparison_headers,
        test_metric_rows,
        test_index_match_formulas,
        test_delta_formulas,
        test_optional_track_conditional_formulas,
        test_dominant_band_text_delta,
        test_freeze_panes,
        test_tab_color,
        test_conditional_formatting,
        test_column_widths,
        test_empty_individuals,
        test_single_track,
        test_excel_save_roundtrip,
    ]

    passed = 0
    failed = 0
    for test in tests:
        try:
            test()
            print(f'  [OK] {test.__doc__}')
            passed += 1
        except Exception as e:
            print(f'  [FAIL] {test.__doc__}')
            print(f'         {e}')
            failed += 1

    print(f'\n{passed}/{passed + failed} tests passed.')
    if failed:
        sys.exit(1)
