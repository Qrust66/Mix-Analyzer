#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""Tests for Feature 3.6 B1a — Section extension + shared role-filter helper.

B1a covers the foundation pieces the CDE (cde_engine.py, shipped in B1b)
will consume:

    * ``Section.conflicts`` and ``Section.accumulations`` persisted on the
      instance by ``build_sections_timeline_sheet``.
    * ``active_tracks_with_roles(section)`` — the single shared gate used
      by both :mod:`tfp_coherence` and :mod:`cde_engine` to decide which
      tracks "count" in a section.

The shared-helper invariance test guards against Risque 6 of the Feature
3.6 reconnaissance: divergence between tfp_coherence's active-role
filter and the CDE's. Before the refactor both would have been separate
inline expressions; now both paths must go through the same helper.
"""

from __future__ import annotations

import os
import sys
from dataclasses import dataclass, field
from typing import List, Tuple

import numpy as np
import pytest

sys.path.insert(0, os.path.abspath(os.path.join(os.path.dirname(__file__), "..")))

from openpyxl import Workbook  # noqa: E402

from section_detector import (  # noqa: E402
    Section,
    active_tracks_with_roles,
    build_sections_timeline_sheet,
    detect_accumulations_in_section,
    detect_conflicts_in_section,
    enrich_sections_with_track_roles,
    enrich_sections_with_track_stats,
    get_zone_order,
)
from tfp_parser import Function, Importance  # noqa: E402


# ---------------------------------------------------------------------------
# Fixtures
# ---------------------------------------------------------------------------

@dataclass
class _FakePeakTraj:
    points: List[Tuple[int, float, float]] = field(default_factory=list)


def _make_section(index: int, start: int, end: int, total_energy_db: float) -> Section:
    return Section(
        index=index,
        name=f"Section {index}",
        start_bucket=start,
        end_bucket=end,
        start_seconds=float(start),
        end_seconds=float(end),
        start_beats=float(start * 2),
        end_beats=float(end * 2),
        total_energy_db=total_energy_db,
    )


def _zone_arrays_for(level_db: float, n_frames: int, zones_with_energy: List[str]) -> dict:
    zones = {
        z: np.full(n_frames, -120.0, dtype=float)
        for z in get_zone_order()
    }
    for z in zones_with_energy:
        zones[z] = np.full(n_frames, float(level_db), dtype=float)
    return zones


def _simultaneous_peak_traj(freqs: List[float], buckets: List[int], amp_db: float = -5.0):
    points = []
    for bucket in buckets:
        for f in freqs:
            points.append((bucket, f, amp_db))
    return _FakePeakTraj(points=points)


# ---------------------------------------------------------------------------
# Section dataclass defaults — conflicts / accumulations must start empty
# ---------------------------------------------------------------------------

def test_section_defaults_conflicts_and_accumulations_to_empty_lists():
    """A freshly-built Section must expose empty lists (not None) so
    downstream consumers can iterate unconditionally."""
    s = _make_section(1, 0, 10, total_energy_db=-20.0)
    assert s.conflicts == []
    assert s.accumulations == []
    assert isinstance(s.conflicts, list)
    assert isinstance(s.accumulations, list)


# ---------------------------------------------------------------------------
# Persistence — build_sections_timeline_sheet must fill section.conflicts
# and section.accumulations with the detector output for each section.
# ---------------------------------------------------------------------------

def test_build_sheet_persists_conflicts_on_each_section():
    """After the sheet is built, each section carries its conflicts list
    equal to ``detect_conflicts_in_section(section)`` — same semantics,
    so CDE can consume them without re-running the detector."""
    n = 30
    sections = [_make_section(1, 0, n - 1, total_energy_db=-10.0)]

    # Two tracks with near-identical energy in the same zone -> critical.
    all_tracks = {
        "TrackA": _zone_arrays_for(-5.0, n, ["low"]),
        "TrackB": _zone_arrays_for(-6.0, n, ["low"]),
        "TrackC": _zone_arrays_for(-60.0, n, ["presence"]),
    }

    wb = Workbook()
    build_sections_timeline_sheet(
        workbook=wb,
        sections=sections,
        all_tracks_zone_energy=all_tracks,
    )

    s = sections[0]
    assert s.conflicts, "persistence must not leave conflicts empty"
    # The persisted list must be exactly what the detector produces — we
    # re-run it on a fresh equivalent section and expect the same shape.
    fresh = _make_section(1, 0, n - 1, total_energy_db=-10.0)
    enrich_sections_with_track_stats([fresh], all_tracks)
    expected = detect_conflicts_in_section(fresh)
    assert s.conflicts == expected


def test_build_sheet_persists_accumulations_on_each_section():
    """Accumulations detected by ``detect_accumulations_in_section`` must
    land on ``section.accumulations`` untouched."""
    n = 40
    sections = [_make_section(1, 0, n - 1, total_energy_db=-10.0)]
    all_tracks = {
        f"T{i}": _zone_arrays_for(-5.0, n, ["low"]) for i in range(6)
    }
    peaks = {
        f"T{i}": [_simultaneous_peak_traj([247.0 + 0.1 * i], buckets=list(range(10, 15)))]
        for i in range(6)
    }

    wb = Workbook()
    build_sections_timeline_sheet(
        workbook=wb,
        sections=sections,
        all_tracks_zone_energy=all_tracks,
        all_tracks_peak_trajectories=peaks,
    )

    s = sections[0]
    assert s.accumulations, "persistence must not leave accumulations empty"
    # Re-run the detector on an equivalent section to confirm identity.
    fresh = _make_section(1, 0, n - 1, total_energy_db=-10.0)
    enrich_sections_with_track_stats([fresh], all_tracks)
    expected = detect_accumulations_in_section(fresh, peaks)
    assert s.accumulations == expected


def test_build_sheet_persistence_is_idempotent_across_runs():
    """Re-building the sheet must leave section.conflicts / accumulations
    in the same state — guards against leaking state between runs that
    would silently inflate diagnostic counts."""
    n = 30
    sections = [_make_section(1, 0, n - 1, total_energy_db=-10.0)]
    all_tracks = {
        "TrackA": _zone_arrays_for(-5.0, n, ["low"]),
        "TrackB": _zone_arrays_for(-6.0, n, ["low"]),
    }

    wb1 = Workbook()
    build_sections_timeline_sheet(
        workbook=wb1, sections=sections, all_tracks_zone_energy=all_tracks,
    )
    first_conflicts = list(sections[0].conflicts)
    first_accs = list(sections[0].accumulations)

    # Second build on the same Section instances — must not accumulate.
    wb2 = Workbook()
    build_sections_timeline_sheet(
        workbook=wb2, sections=sections, all_tracks_zone_energy=all_tracks,
    )
    assert sections[0].conflicts == first_conflicts
    assert sections[0].accumulations == first_accs


# ---------------------------------------------------------------------------
# Shared helper — active_tracks_with_roles
# ---------------------------------------------------------------------------

def test_active_tracks_with_roles_filters_to_meter_based_active_set():
    """The helper must return only tracks present in ``tracks_active``
    (meter-based gate), dropping everything else even if the role map
    contains more entries (MIDI without WAV, etc.)."""
    s = _make_section(1, 0, 10, total_energy_db=-15.0)
    s.tracks_active = ["Kick", "Bass"]
    s.track_roles = {
        "Kick":     (Importance.H, Function.R),
        "Bass":     (Importance.H, Function.H),
        "UnusedMIDI": (Importance.S, Function.T),  # not in tracks_active
    }
    out = active_tracks_with_roles(s)
    assert set(out.keys()) == {"Kick", "Bass"}
    assert out["Kick"] == (Importance.H, Function.R)
    assert out["Bass"] == (Importance.H, Function.H)


def test_active_tracks_with_roles_drops_tracks_without_role_entry():
    """Tracks listed in tracks_active but missing from track_roles must be
    skipped silently — the CDE will log a warning at its call site."""
    s = _make_section(1, 0, 10, total_energy_db=-15.0)
    s.tracks_active = ["Kick", "Ghost"]
    s.track_roles = {"Kick": (Importance.S, Function.R)}
    out = active_tracks_with_roles(s)
    assert set(out.keys()) == {"Kick"}


def test_active_tracks_with_roles_handles_empty_section():
    """Empty tracks_active / track_roles must yield an empty dict, not
    raise and not return None."""
    s = _make_section(1, 0, 10, total_energy_db=-120.0)
    assert active_tracks_with_roles(s) == {}


# ---------------------------------------------------------------------------
# Shared-helper invariance — Risque 6 guard
# ---------------------------------------------------------------------------
#
# Before the refactor, tfp_coherence.compute_section_coherence_score
# expanded the active-role list inline. B1a replaces that expansion with
# active_tracks_with_roles(). The invariance test pins the helper output
# to the exact shape the old inline code produced so any future
# refactoring on either side fails loudly.

def _legacy_active_roles(section: Section):
    """The pre-refactor inline filter, kept here as the oracle.

    Any change in :func:`active_tracks_with_roles` must continue to
    produce the same role list for the same Section — otherwise the
    coherence score (Feature 3.5) silently shifts.
    """
    tracks_active = list(getattr(section, "tracks_active", []) or [])
    track_roles = getattr(section, "track_roles", {}) or {}
    return [track_roles[t] for t in tracks_active if t in track_roles]


def test_helper_invariance_against_legacy_inline_filter():
    """The helper and the legacy inline expression must produce the same
    ordered sequence of role tuples on a realistic multi-track section."""
    s = _make_section(1, 0, 40, total_energy_db=-12.0)
    s.tracks_active = ["Kick 1", "Sub Bass", "Acid Bass", "Pad"]
    s.track_roles = {
        "Kick 1":     (Importance.H, Function.R),
        "Sub Bass":   (Importance.S, Function.H),
        "Acid Bass":  (Importance.H, Function.M),
        "Pad":        (Importance.A, Function.T),
        # Extra entry that must NOT leak through the filter.
        "Bleed":      (Importance.S, Function.R),
    }

    legacy = _legacy_active_roles(s)
    helper = list(active_tracks_with_roles(s).values())

    assert helper == legacy
    assert (Importance.S, Function.R) not in helper  # "Bleed" excluded
