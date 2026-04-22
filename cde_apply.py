#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
cde_apply.py — v2.7.0 — Feature 1 (CDE → ALS bridge).

Consumes ``CorrectionDiagnostic`` instances and applies them to an
Ableton ``.als`` project as dynamic EQ8 corrections. Sits between
:mod:`cde_engine` (data model) and :mod:`eq8_automation` (low-level
XML writers).

F1a — data prep (read-only):
    - Dataclasses ``FreqCluster`` and ``CdeApplicationReport``
    - Grouping by target track, frequency clustering, peak-trajectory
      matching, Excel reader for ``_track_peak_trajectories``

F1b — section-locked writer:
    - ``write_dynamic_eq8_from_cde_diagnostics(als_path, diagnostics,
      *, peak_follow=False, …)`` — public entry point
    - Atomic reciprocal_cuts validation (both tracks must exist)
    - Reciprocal expansion into primary + secondary pseudo-diagnostics
    - Dense per-frame gain envelope that reads ``severest_gain_db``
      in every section listed by ``cluster.applies_to_sections``,
      0 dB everywhere else
    - Cluster cap at :data:`MAX_CDE_BANDS` with "least severe first"
      skip policy
    - Idempotent: purges any prior ``CDE Correction …`` EQ8 on the
      target track before insertion
    - Preview + confirmation prompt (``_skip_confirmation=True`` for
      tests, ``dry_run=True`` for CLI previews)

F1c (not yet shipped): ``peak_follow=True`` mode — Freq + Gain + Q
automation frame-by-frame driven by the target track's own peak
trajectories. F1c is gated on the F1b.5 field validation.
"""

from __future__ import annotations

import logging
from dataclasses import dataclass, field
from pathlib import Path
from typing import Callable, Dict, List, Optional, Tuple

import numpy as np

from cde_engine import CorrectionDiagnostic
from spectral_evolution import PeakTrajectory

logger = logging.getLogger(__name__)


# ---------------------------------------------------------------------------
# Vocabulary — approach strings the writer can handle vs skip
# ---------------------------------------------------------------------------

# F1 scope: EQ8-cut-style corrections. Every other approach rides a
# different device and is therefore out of scope for this module.
CUT_APPROACHES: Tuple[str, ...] = (
    "static_dip", "musical_dip", "reciprocal_cuts",
)
SIDECHAIN_APPROACH: str = "sidechain"

# Clustering defaults. 2 semitones = ×2^(2/12) ≈ ×1.122 — narrow enough
# to keep distinct tonal centres apart (the user's call after Phase A).
DEFAULT_FREQ_CLUSTER_TOLERANCE_SEMITONES = 2.0

# Peak-trajectory matching — by default shares the same tolerance as
# the cluster itself. Caller can override.
DEFAULT_PEAK_MATCH_TOLERANCE_SEMITONES = 2.0

# EQ8 conventions. Bands 0 and 7 are reserved by
# ``eq8_automation._find_available_band`` for HPF/LPF.
MAX_CDE_BANDS = 6


# ---------------------------------------------------------------------------
# Dataclasses
# ---------------------------------------------------------------------------

@dataclass
class FreqCluster:
    """A bundle of diagnostics that should ride a single EQ8 band.

    Produced by :func:`_cluster_diagnostics_by_frequency` — each cluster's
    member diagnostics share a central frequency within the clustering
    tolerance and will share one band in F1b/F1c.

    Attributes:
        centroid_hz:  median of member frequencies (in Hz).
        diagnostics:  the source ``CorrectionDiagnostic`` list.
        severest_gain_db:  the most-negative ``gain_db`` across members —
            used as the cut depth for the produced band (Rule 3 already
            capped it at CDE-matrix time).
        median_q:  the median ``q`` across members — used as the band Q.
        applies_to_sections: union of every member's
            ``applies_to_sections`` list, deduplicated and sorted.
    """
    centroid_hz: float
    diagnostics: List[CorrectionDiagnostic]
    severest_gain_db: float
    median_q: float
    applies_to_sections: List[str]

    @property
    def member_ids(self) -> List[str]:
        return [d.diagnostic_id for d in self.diagnostics]


@dataclass
class CdeApplicationReport:
    """Summary of one CDE batch application run.

    The writer (F1b/c) populates the fields as it goes. ``skipped`` is
    a list of ``(diagnostic_id, reason)`` pairs so the caller can show
    the user WHY something was not applied without reverse-engineering.
    """
    applied: List[str] = field(default_factory=list)
    skipped: List[Tuple[str, str]] = field(default_factory=list)
    warnings: List[str] = field(default_factory=list)
    # {track_name: device_user_name} — shows which devices were inserted.
    devices_created: Dict[str, str] = field(default_factory=dict)
    envelopes_written: int = 0
    backup_path: Optional[Path] = None

    def sidechain_count(self) -> int:
        """How many diagnostics were skipped because they needed
        Kickstart 2 — reported separately so the user sees what
        Feature 1.5 will need to cover."""
        return sum(
            1 for (_, reason) in self.skipped
            if "Kickstart 2" in reason or "sidechain" in reason.lower()
        )


# ---------------------------------------------------------------------------
# Grouping
# ---------------------------------------------------------------------------

def _group_diagnostics_by_target(
    diagnostics: List[CorrectionDiagnostic],
    report: Optional[CdeApplicationReport] = None,
) -> Dict[str, List[CorrectionDiagnostic]]:
    """Filter cut-approach diagnostics and group them by
    ``primary_correction.target_track``.

    Skipped diagnostics (with reasons appended to ``report.skipped`` when
    ``report`` is provided):

        - ``primary_correction is None``
          (R2 sub-integrity skip, all-Hero accumulation awareness)
        - ``approach == "sidechain"``
          (Kickstart 2 device; Feature 1.5 scope)
        - Any approach outside :data:`CUT_APPROACHES`
          (future detectors; explicit opt-in required)

    Args:
        diagnostics: Raw CDE diagnostics list (e.g. deserialised from
            ``<project>_diagnostics.json``).
        report: Optional report to collect skip reasons into. When
            ``None``, skipped diagnostics are silently dropped.

    Returns:
        ``{target_track: [diagnostic, diagnostic, ...]}`` — track names
        are taken verbatim from ``primary_correction.target_track`` so
        they match the raw WAV filename the CDE stored.
    """
    grouped: Dict[str, List[CorrectionDiagnostic]] = {}
    for d in diagnostics:
        pc = d.primary_correction
        if pc is None:
            if report is not None:
                report.skipped.append((
                    d.diagnostic_id, "primary_correction is None",
                ))
            continue
        if pc.approach == SIDECHAIN_APPROACH:
            if report is not None:
                report.skipped.append((
                    d.diagnostic_id,
                    "requires Kickstart 2 — scope Feature 1.5",
                ))
            continue
        if pc.approach not in CUT_APPROACHES:
            if report is not None:
                report.skipped.append((
                    d.diagnostic_id,
                    f"unsupported approach {pc.approach!r}",
                ))
            continue
        grouped.setdefault(pc.target_track, []).append(d)
    return grouped


# ---------------------------------------------------------------------------
# Frequency clustering
# ---------------------------------------------------------------------------

def _semitone_ratio(tolerance_semitones: float) -> float:
    """Convert a tolerance in semitones to a frequency ratio factor."""
    return 2.0 ** (float(tolerance_semitones) / 12.0)


def _freq_distance_semitones(f1: float, f2: float) -> float:
    """Return ``|log2(f2/f1)| * 12`` — the semitone distance between
    two positive frequencies."""
    if f1 <= 0 or f2 <= 0:
        return float("inf")
    return abs(float(np.log2(f2 / f1))) * 12.0


# Floating-point slack for inclusive boundary comparisons — ``247 Hz``
# and ``247 * 2^(2/12) Hz`` are musically exactly 2 semitones apart but
# the round-trip through ``np.log2`` drifts by ~4e-16.
_CLUSTER_EPSILON_SEMITONES = 1e-9


def _cluster_diagnostics_by_frequency(
    diagnostics: List[CorrectionDiagnostic],
    tolerance_semitones: float = DEFAULT_FREQ_CLUSTER_TOLERANCE_SEMITONES,
) -> List[FreqCluster]:
    """Centroid-link cluster the diagnostics' frequencies.

    Algorithm:

        1. Sort members by their ``measurement.frequency_hz``.
        2. Walk the sorted list, appending each freq to the current
           cluster when its distance to the cluster's **median** is
           within ``tolerance_semitones``; otherwise start a new cluster.
        3. Build one :class:`FreqCluster` per resulting group with:
            - ``centroid_hz``   = median of member frequencies
            - ``severest_gain_db`` = most negative ``gain_db`` among members
            - ``median_q``      = median of member Q values
            - ``applies_to_sections`` = sorted union of each member's
              ``applies_to_sections``

    Members with missing / non-positive frequencies are dropped
    silently — a frequency is mandatory for a cut recipe, so a
    diagnostic without one cannot ride an EQ8 band.

    Args:
        diagnostics: a group of diagnostics on the **same target track**
            (call :func:`_group_diagnostics_by_target` first).
        tolerance_semitones: clustering tolerance (default 2.0).

    Returns:
        Clusters sorted by ``centroid_hz`` ascending.
    """
    if not diagnostics:
        return []

    # Only keep diagnostics with a usable frequency.
    freq_items: List[Tuple[float, CorrectionDiagnostic]] = []
    for d in diagnostics:
        meas = d.measurement
        f = float(meas.frequency_hz) if (meas and meas.frequency_hz) else None
        if f is None or f <= 0:
            continue
        freq_items.append((f, d))
    freq_items.sort(key=lambda x: x[0])

    if not freq_items:
        return []

    # Single-link walk using the running median as the comparison anchor.
    clusters_raw: List[List[Tuple[float, CorrectionDiagnostic]]] = []
    for f, d in freq_items:
        if not clusters_raw:
            clusters_raw.append([(f, d)])
            continue
        current_freqs = [p[0] for p in clusters_raw[-1]]
        centroid = float(np.median(current_freqs))
        if (_freq_distance_semitones(centroid, f)
                <= tolerance_semitones + _CLUSTER_EPSILON_SEMITONES):
            clusters_raw[-1].append((f, d))
        else:
            clusters_raw.append([(f, d)])

    # Materialise the cluster dataclasses.
    out: List[FreqCluster] = []
    for group in clusters_raw:
        freqs = [p[0] for p in group]
        diags = [p[1] for p in group]
        gains = [
            float(d.primary_correction.parameters.get("gain_db", 0.0))
            for d in diags
            if d.primary_correction is not None
        ]
        qs = [
            float(d.primary_correction.parameters.get("q", 1.0))
            for d in diags
            if d.primary_correction is not None
        ]
        sections_set: set = set()
        for d in diags:
            if d.primary_correction is None:
                continue
            sections_set.update(d.primary_correction.applies_to_sections)

        out.append(FreqCluster(
            centroid_hz=float(np.median(freqs)),
            diagnostics=diags,
            severest_gain_db=float(min(gains)) if gains else 0.0,
            median_q=float(np.median(qs)) if qs else 1.0,
            applies_to_sections=sorted(sections_set),
        ))
    return out


# ---------------------------------------------------------------------------
# Peak-trajectory matching
# ---------------------------------------------------------------------------

def _match_peak_trajectories_to_cluster(
    trajectories: List[PeakTrajectory],
    cluster: FreqCluster,
    tolerance_semitones: float = DEFAULT_PEAK_MATCH_TOLERANCE_SEMITONES,
) -> List[PeakTrajectory]:
    """Return the subset of trajectories whose ``mean_freq`` falls
    within ``tolerance_semitones`` of the cluster's centroid.

    Used in F1c (peak-following mode) to decide which trajectories to
    drive the Freq + Gain envelopes with. When no trajectory matches,
    the caller falls back to the section-locked mode (F1b).
    """
    matched: List[PeakTrajectory] = []
    for traj in trajectories or []:
        mean_freq = float(traj.mean_freq)
        if mean_freq <= 0:
            continue
        if (_freq_distance_semitones(cluster.centroid_hz, mean_freq)
                <= tolerance_semitones + _CLUSTER_EPSILON_SEMITONES):
            matched.append(traj)
    return matched


# ---------------------------------------------------------------------------
# Excel reader — _track_peak_trajectories sheet
# ---------------------------------------------------------------------------

_PEAK_SHEET_NAME = "_track_peak_trajectories"
_PEAK_SHEET_HEADER = ("Track", "Traj#", "Frame", "Time (s)",
                      "Freq (Hz)", "Amp (dB)")


def load_peak_trajectories_from_excel(
    xlsx_path,
    track_name_filter: Optional[Callable[[str], str]] = None,
) -> Dict[str, List[PeakTrajectory]]:
    """Parse the ``_track_peak_trajectories`` sheet of a Mix Analyzer
    Excel report.

    The sheet shape (verified on the Acid Drops production report):

        Track | Traj# | Frame | Time (s) | Freq (Hz) | Amp (dB)

    Rows are grouped by ``(Track, Traj#)`` and each group materialises
    as one :class:`spectral_evolution.PeakTrajectory` with its
    ``points = [(frame_idx, freq_hz, amp_db), ...]`` list. Points are
    kept in the sheet's row order (which is already frame-ascending in
    the production reports).

    Args:
        xlsx_path: Path to the Mix Analyzer ``.xlsx`` report.
        track_name_filter: Optional callable applied to each row's
            Track name — useful to normalise ``"Acid_Drops [H_R]
            Kick 1.wav"`` into ``"[H/R] Kick 1"`` so names match the
            Ableton EffectiveName used by ``find_track_by_name``. When
            ``None``, names are kept verbatim.

    Returns:
        ``{track_name: [PeakTrajectory, PeakTrajectory, ...]}``.
        Trajectories preserve the sheet's Traj# ordering implicitly
        (Python dict insertion order).

    Raises:
        FileNotFoundError: if ``xlsx_path`` does not exist.
        ValueError: if the expected sheet is absent from the workbook.
    """
    from openpyxl import load_workbook

    path = Path(xlsx_path)
    if not path.exists():
        raise FileNotFoundError(f"Report not found: {path}")

    wb = load_workbook(str(path), read_only=True, data_only=True)
    if _PEAK_SHEET_NAME not in wb.sheetnames:
        raise ValueError(
            f"Sheet {_PEAK_SHEET_NAME!r} not found in {path.name}. "
            f"Available: {wb.sheetnames}"
        )
    ws = wb[_PEAK_SHEET_NAME]

    grouped: Dict[Tuple[str, int], List[Tuple[int, float, float]]] = {}
    track_order: List[str] = []
    traj_order_seen: set = set()

    rows = ws.iter_rows(values_only=True)
    header = next(rows, None)
    if header is None or tuple(header[: len(_PEAK_SHEET_HEADER)]) != _PEAK_SHEET_HEADER:
        logger.warning(
            "Unexpected header in %s: %r (expected %r). Parsing anyway.",
            _PEAK_SHEET_NAME, header, _PEAK_SHEET_HEADER,
        )

    for row in rows:
        if not row or row[0] is None or row[1] is None:
            continue
        track_raw = str(row[0])
        traj_num_raw = row[1]
        frame_raw = row[2]
        freq_raw = row[4]
        amp_raw = row[5]
        if freq_raw is None or amp_raw is None:
            continue

        track = (track_name_filter(track_raw) if track_name_filter
                 else track_raw)
        try:
            traj_num = int(traj_num_raw)
            frame_idx = int(frame_raw) if frame_raw is not None else 0
            freq_hz = float(freq_raw)
            amp_db = float(amp_raw)
        except (TypeError, ValueError):
            continue

        key = (track, traj_num)
        if key not in grouped:
            grouped[key] = []
            if track not in traj_order_seen:
                track_order.append(track)
                traj_order_seen.add(track)
        grouped[key].append((frame_idx, freq_hz, amp_db))

    # Build PeakTrajectory per (track, traj_num), respecting first-seen
    # track order + ascending traj_num within each track.
    result: Dict[str, List[PeakTrajectory]] = {t: [] for t in track_order}
    for (track, traj_num) in sorted(grouped.keys(), key=lambda k: (track_order.index(k[0]), k[1])):
        result[track].append(PeakTrajectory(points=grouped[(track, traj_num)]))

    return result


# ===========================================================================
# F1b — section-locked writer
# ===========================================================================

# Imports needed only for the writer. Kept here rather than at the top so the
# F1a slice (clustering only) does not pull in the heavy als/EQ8 stack just
# to read an Excel sheet.
from cde_engine import CorrectionRecipe, ProblemMeasurement  # noqa: E402
from als_utils import (  # noqa: E402
    backup_als,
    configure_eq8_band,
    find_track_by_name,
    get_eq8_band,
    get_next_id,
    parse_als,
    read_locators,
    save_als_from_tree,
    _clone_eq8_with_unique_ids,
)
from eq8_automation import (  # noqa: E402
    _extract_tempo,
    _feature_to_breakpoints,
    _remove_existing_envelope,
    _write_validated_env,
)


# Default dense-sampling spacing for the gain envelope. Matches the v3
# script that drove the 247 Hz cut on Ambience and the production
# Peak Resonance density (~0.5 s per event).
DEFAULT_GRID_SPACING_SEC = 0.5


# ---------------------------------------------------------------------------
# Reciprocal validation + expansion
# ---------------------------------------------------------------------------

def _validate_reciprocal_pairs(
    tree,
    diagnostics: List[CorrectionDiagnostic],
    report: CdeApplicationReport,
) -> List[CorrectionDiagnostic]:
    """Filter out reciprocal_cuts whose secondary_cut targets a track
    that doesn't exist in the project. Both halves of the pair are
    skipped together with reason ``reciprocal_cut_missing_track`` so
    the user can spot the broken link.
    """
    valid: List[CorrectionDiagnostic] = []
    for d in diagnostics:
        pc = d.primary_correction
        if pc is None or pc.approach != "reciprocal_cuts":
            valid.append(d)
            continue
        sec = pc.parameters.get("secondary_cut") if isinstance(pc.parameters, dict) else None
        if not isinstance(sec, dict):
            valid.append(d)
            continue
        track_a = pc.target_track
        track_b = sec.get("track", "")
        try:
            find_track_by_name(tree, track_a)
            if track_b:
                find_track_by_name(tree, track_b)
            else:
                raise ValueError("missing secondary track name")
        except ValueError:
            report.skipped.append((
                d.diagnostic_id,
                f"reciprocal_cut_missing_track "
                f"(track_a={track_a!r}, track_b={track_b!r})",
            ))
            continue
        valid.append(d)
    return valid


def _expand_reciprocal_cuts(
    diagnostics: List[CorrectionDiagnostic],
) -> List[CorrectionDiagnostic]:
    """For every reciprocal_cuts diagnostic, append a pseudo-diagnostic
    targeting the secondary track so it ends up in the same grouping
    pipeline. The pseudo carries diagnostic_id ``<orig>_SEC`` so the
    audit trail still groups the atomic pair logically.
    """
    expanded: List[CorrectionDiagnostic] = []
    for d in diagnostics:
        expanded.append(d)
        pc = d.primary_correction
        if pc is None or pc.approach != "reciprocal_cuts":
            continue
        sec = pc.parameters.get("secondary_cut") if isinstance(pc.parameters, dict) else None
        if not isinstance(sec, dict):
            continue
        track_b = sec.get("track", "")
        if not track_b:
            continue

        sec_freq = sec.get("frequency_hz", pc.parameters.get("frequency_hz"))
        sec_gain = sec.get("gain_db", pc.parameters.get("gain_db"))
        sec_q = sec.get("q", pc.parameters.get("q"))

        meas = d.measurement
        pseudo_meas = ProblemMeasurement(
            frequency_hz=float(sec_freq) if sec_freq is not None else None,
            peak_db=meas.peak_db if meas else None,
            duration_in_section_s=meas.duration_in_section_s if meas else 0.0,
            duration_ratio_in_section=meas.duration_ratio_in_section if meas else 0.0,
            is_audible_fraction=meas.is_audible_fraction if meas else 0.0,
            severity_score=meas.severity_score if meas else 0.0,
            masking_score=meas.masking_score if meas else None,
        )
        pseudo_recipe = CorrectionRecipe(
            target_track=track_b,
            device=pc.device,
            approach="reciprocal_cuts",
            parameters={
                "frequency_hz": float(sec_freq) if sec_freq is not None else 0.0,
                "gain_db": float(sec_gain) if sec_gain is not None else 0.0,
                "q": float(sec_q) if sec_q is not None else 1.0,
                "active_in_sections": list(pc.parameters.get("active_in_sections", [])),
            },
            applies_to_sections=list(pc.applies_to_sections),
            rationale=pc.rationale,
            confidence=pc.confidence,
        )
        pseudo = CorrectionDiagnostic(
            diagnostic_id=f"{d.diagnostic_id}_SEC",
            timestamp=d.timestamp,
            cde_version=d.cde_version,
            track_a=track_b,
            track_b=None,
            section=d.section,
            issue_type=d.issue_type,
            severity=d.severity,
            measurement=pseudo_meas,
            tfp_context=d.tfp_context,
            section_context=d.section_context,
            diagnosis_text=d.diagnosis_text,
            primary_correction=pseudo_recipe,
            fallback_correction=None,
        )
        expanded.append(pseudo)
    return expanded


# ---------------------------------------------------------------------------
# Section ranges + dense gain curves
# ---------------------------------------------------------------------------

def _build_section_ranges_beats(locators) -> Dict[str, Tuple[float, float]]:
    """Return ``{section_name: (start_beats, end_beats)}`` from a
    sorted locator list. Each section ends where the next locator
    starts; the final section is given a 32-beat tail (~1 bar at 4/4)
    so a cut that runs to the end of the song doesn't get cropped.
    """
    if not locators:
        return {}
    sorted_locs = sorted(locators, key=lambda L: L["time_beats"])
    ranges: Dict[str, Tuple[float, float]] = {}
    for i, loc in enumerate(sorted_locs):
        start = float(loc["time_beats"])
        if i + 1 < len(sorted_locs):
            end = float(sorted_locs[i + 1]["time_beats"])
        else:
            end = start + 32.0
        ranges[loc["name"]] = (start, end)
    return ranges


def _estimate_song_end_beats(locators) -> float:
    """Last locator + 32 beats fallback. 32 beats default keeps a
    16-second tail at 128 BPM so the envelope covers the very end of
    the song without truncation."""
    if not locators:
        return 32.0
    return max(float(L["time_beats"]) for L in locators) + 32.0


def _build_section_locked_gain_curve(
    cluster: FreqCluster,
    section_ranges: Dict[str, Tuple[float, float]],
    times_sec: np.ndarray,
    tempo_bpm: float,
) -> np.ndarray:
    """Build a per-frame gain curve: ``cluster.severest_gain_db`` for
    frames whose time falls inside any section listed in
    ``cluster.applies_to_sections``, ``0.0`` everywhere else.

    Vectorised numpy mask — cheap even on 5000-frame grids.
    """
    beats_per_sec = float(tempo_bpm) / 60.0
    gain_curve = np.zeros(len(times_sec), dtype=float)
    for section_name in cluster.applies_to_sections:
        rng = section_ranges.get(section_name)
        if rng is None:
            continue
        start_b, end_b = rng
        start_s = start_b / beats_per_sec
        end_s = end_b / beats_per_sec
        mask = (times_sec >= start_s) & (times_sec < end_s)
        gain_curve[mask] = cluster.severest_gain_db
    return gain_curve


# ---------------------------------------------------------------------------
# EQ8 cloning + chain helpers (mirrored from the apply_cde_correction_247hz
# scripts so callers don't depend on them)
# ---------------------------------------------------------------------------

_NEUTRAL_BAND_FREQS_HZ = (60.0, 150.0, 400.0, 1000.0,
                          2500.0, 6000.0, 12000.0, 18000.0)


def _clone_in_project_eq8(tree, user_name: str):
    """Sandbox-friendly clone — used when ``Pluggin Mapping.als`` is
    not deployed locally. Deep-copies an existing EQ8 from anywhere in
    the project, renumbers IDs and resets every band to neutral."""
    import copy
    source = tree.getroot().find(".//Eq8")
    if source is None:
        raise RuntimeError("No Eq8 in project to clone from.")
    eq8 = copy.deepcopy(source)
    next_id = get_next_id(tree)
    for elem in eq8.iter():
        raw = elem.get("Id")
        if raw is not None and raw != "0":
            elem.set("Id", str(next_id))
            next_id += 1
    un = eq8.find("UserName")
    if un is not None:
        un.set("Value", user_name)
    for i in range(8):
        band = eq8.find(f"Bands.{i}/ParameterA")
        if band is None:
            continue
        for tag, value in (("Mode/Manual", "3"),
                           ("Freq/Manual", str(_NEUTRAL_BAND_FREQS_HZ[i])),
                           ("Gain/Manual", "0.0"),
                           ("Q/Manual", "0.7071067095"),
                           ("IsOn/Manual", "false")):
            el = band.find(tag)
            if el is not None:
                el.set("Value", value)
    return eq8


def _build_fresh_eq8(tree, user_name: str):
    """Try the canonical Pluggin Mapping template, fall back to an
    in-project clone when it's not deployed."""
    try:
        return _clone_eq8_with_unique_ids(tree, user_name=user_name)
    except RuntimeError as e:
        if "template file not found" not in str(e):
            raise
        return _clone_in_project_eq8(tree, user_name)


def _devices_container(track):
    devices = track.find(".//DeviceChain/DeviceChain/Devices")
    if devices is None:
        devices = track.find(".//DeviceChain/Devices")
    if devices is None:
        raise RuntimeError("No Devices container on the target track.")
    return devices


def _find_device_by_username(track, username: str):
    if not username:
        return None
    for eq8 in track.findall(".//Eq8"):
        un = eq8.find("UserName")
        if un is not None and un.get("Value") == username:
            return eq8
    return None


def _insert_after(devices, anchor, new_device) -> int:
    children = list(devices)
    if anchor is None:
        devices.append(new_device)
        return len(children)
    anchor_idx = children.index(anchor)
    for c in children:
        devices.remove(c)
    for i, c in enumerate(children):
        devices.append(c)
        if i == anchor_idx:
            devices.append(new_device)
    return anchor_idx + 1


def _purge_existing_cde_eq8(
    track,
    device_user_name_fmt: str,
    report: CdeApplicationReport,
) -> int:
    """Remove every prior ``"CDE Correction (…)"`` EQ8 + its envelopes.

    The match uses the pre-``{n}`` prefix of ``device_user_name_fmt``
    so re-runs don't accumulate stale devices.
    """
    prefix = device_user_name_fmt.split("{")[0]
    devices = _devices_container(track)
    removed = 0
    for child in list(devices):
        if child.tag != "Eq8":
            continue
        un = child.find("UserName")
        if un is None or not un.get("Value", "").startswith(prefix):
            continue
        for at in child.iter("AutomationTarget"):
            tid = at.get("Id")
            if tid is not None:
                _remove_existing_envelope(track, tid)
        devices.remove(child)
        removed += 1
        report.warnings.append(
            f"Purged previous {un.get('Value')!r} on chain"
        )
    return removed


# ---------------------------------------------------------------------------
# Cluster cap + preview
# ---------------------------------------------------------------------------

def _cap_clusters(
    clusters: List[FreqCluster],
    max_bands: int,
    track_name: str,
    report: CdeApplicationReport,
) -> List[FreqCluster]:
    """Keep the ``max_bands`` most severe clusters (most negative
    ``severest_gain_db`` first); record the rest in
    ``report.skipped`` with reason ``"cluster_cap_exceeded"``."""
    if len(clusters) <= max_bands:
        return clusters
    # Sort ascending = most negative first = most severe first.
    by_severity = sorted(clusters, key=lambda c: c.severest_gain_db)
    kept = by_severity[:max_bands]
    dropped = by_severity[max_bands:]
    dropped_freqs = [round(c.centroid_hz) for c in dropped]
    report.warnings.append(
        f"{track_name}: {len(clusters)} clusters exceed cap "
        f"({max_bands}). Dropped {len(dropped)} least severe "
        f"@ {dropped_freqs} Hz."
    )
    for c in dropped:
        for d in c.diagnostics:
            report.skipped.append((
                d.diagnostic_id, "cluster_cap_exceeded",
            ))
    return kept


def _build_preview_text(
    als_path: Path,
    clusters_by_track: Dict[str, List[FreqCluster]],
    report: CdeApplicationReport,
    backup_path_hint: Path,
) -> str:
    sep = "═" * 64
    total_diags = sum(
        len(c.diagnostics)
        for cs in clusters_by_track.values() for c in cs
    )
    crit = mod = minor = 0
    for cs in clusters_by_track.values():
        for c in cs:
            for d in c.diagnostics:
                if d.severity == "critical":
                    crit += 1
                elif d.severity == "moderate":
                    mod += 1
                else:
                    minor += 1

    sidechain_n = report.sidechain_count()
    primary_none_n = sum(
        1 for (_, r) in report.skipped if "primary_correction is None" in r
    )
    other_skip = len(report.skipped) - sidechain_n - primary_none_n

    lines = [
        sep,
        "CDE APPLICATION PREVIEW",
        sep,
        f"Diagnostics à appliquer : {total_diags} "
        f"(critical: {crit}, moderate: {mod}, minor: {minor})",
        f"Diagnostics skippés     : {len(report.skipped)}",
        f"  - sidechain (Feature 1.5) : {sidechain_n}",
        f"  - primary=None (R2/R3)    : {primary_none_n}",
        f"  - autre                   : {other_skip}",
        "",
        f"Tracks affectées : {len(clusters_by_track)}",
    ]
    for track_name, clusters in sorted(clusters_by_track.items()):
        freqs = sorted(int(round(c.centroid_hz)) for c in clusters)
        freq_str = ", ".join(f"{f} Hz" for f in freqs)
        lines.append(
            f"  - {track_name} : {len(clusters)} clusters ({freq_str})"
        )
    lines.append("")
    lines.append(f"Backup sera créé : {backup_path_hint}")
    if report.warnings:
        lines.append("")
        lines.append("Warnings :")
        for w in report.warnings:
            lines.append(f"  - {w}")
    lines.append(sep)
    return "\n".join(lines)


def _prompt_confirmation() -> bool:
    """Ask the user via stdin. Accepts y/yes/o/oui (case-insensitive)
    as confirmation. Anything else is treated as a refusal."""
    try:
        resp = input("Procéder ? (y/N) ").strip().lower()
    except EOFError:
        return False
    return resp in ("y", "yes", "o", "oui")


# ---------------------------------------------------------------------------
# F1c1a — peak-following helpers (pure functions)
# ---------------------------------------------------------------------------
#
# The writer's peak_follow=True mode (wired in F1c1b, smoke-tested in
# F1c1c) drives the band's Freq + Gain + Q envelopes frame-by-frame
# from the target track's own PeakTrajectory list. These four helpers
# handle the data-shaping side of that mode — they are deliberately
# pure (no IO, no XML) so they can be tested in isolation in F1c2.

def _scale_gain_by_amplitude(
    target_gain_db: float,
    peak_amp_db: float,
    threshold_db: float,
) -> float:
    """Scale a cut gain proportionally to how loud the peak is.

    Under ``threshold_db`` → return ``0.0`` (band neutral for this frame).
    At the threshold → return ``0.0``. As ``peak_amp_db`` rises toward
    ``0 dB``, the scale climbs linearly from 0 to 1 and the return
    value approaches the full ``target_gain_db``.

    The formula ``(peak_amp_db - threshold_db) / |threshold_db|`` treats
    the distance above the threshold as a fraction of the threshold's
    magnitude, then clips to ``[0, 1]``. For the default
    ``threshold_db = -30``:

        peak_amp_db = 0 dB   → scale = 1.0 → target_gain_db (full cut)
        peak_amp_db = -15 dB → scale = 0.5 → target_gain_db / 2
        peak_amp_db = -35 dB → below threshold → 0 dB (no cut)

    Args:
        target_gain_db: The cluster's cap (e.g. -6 dB). Usually negative
            — the magnitude is what the hottest peak will receive.
        peak_amp_db: Peak amplitude at this frame, in dB.
        threshold_db: Amplitude floor (negative) below which no cut is
            applied.

    Returns:
        Scaled gain in dB. Always has the same sign as ``target_gain_db``
        (or zero when under threshold).
    """
    if peak_amp_db <= threshold_db:
        return 0.0
    denom = abs(float(threshold_db)) if threshold_db != 0 else 1.0
    scale = (float(peak_amp_db) - float(threshold_db)) / denom
    scale = max(0.0, min(1.0, scale))
    return float(target_gain_db) * scale


def _scale_q_by_peak_width(
    peak_width_hz,
    base_q: float,
    centroid_hz: float,
) -> float:
    """Adaptive Q placeholder — currently returns ``base_q`` unchanged.

    When ``PeakTrajectory`` eventually exposes a peak_width field (see
    methodology §2.4 "Notch dynamique" in
    ``ableton_devices_mapping_v2_3.json``), this stub will be replaced
    by the spec's adaptive formula (Q=14 for surgical-sharp peaks,
    Q=1 for broad ones). The signature stays stable so that F1c1b can
    already call the function; upgrading the logic later will not break
    the writer.

    Args:
        peak_width_hz: Reserved for the future adaptive calculation.
            Currently ignored.
        base_q: The cluster's median Q — returned as-is for now.
        centroid_hz: Reserved (width-to-Q mapping depends on the
            absolute frequency). Currently ignored.

    Returns:
        ``float(base_q)`` unchanged.
    """
    return float(base_q)


def _collect_active_peak_frames(
    trajectories,
    section_ranges_beats,
    times,
    tempo,
    threshold_db: float,
):
    """Collect ``(frame_idx, freq_hz, amp_db)`` points where at least
    one trajectory's peak is active inside one of the allowed section
    ranges and rises above ``threshold_db``.

    When several trajectories emit a peak on the same frame, the one
    with the highest ``amp_db`` wins — the band cannot follow two
    drifting peaks at once.

    Args:
        trajectories: iterable of :class:`spectral_evolution.PeakTrajectory`.
            Each exposes ``points = [(frame_idx, freq_hz, amp_db), …]``.
        section_ranges_beats: iterable of ``(start_beats, end_beats)``
            tuples. A frame's time is kept only if it falls inside at
            least one range (inclusive bounds).
        times: 1-D array of per-frame timestamps in seconds.
            ``times[frame_idx]`` gives the absolute time of the frame.
        tempo: Project tempo in BPM for seconds↔beats conversion.
        threshold_db: Amplitude floor — points with ``amp_db`` ≤ this
            are dropped (the band stays neutral on quiet frames).

    Returns:
        Sorted list of ``(frame_idx, freq_hz, amp_db)`` tuples in
        ascending frame order.
    """
    if not trajectories:
        return []
    beats_per_sec = float(tempo) / 60.0
    ranges = list(section_ranges_beats or [])
    n_frames = len(times) if times is not None else 0

    candidates: Dict[int, Tuple[float, float]] = {}
    for traj in trajectories:
        for point in getattr(traj, "points", None) or []:
            try:
                frame_idx, freq_hz, amp_db = point
            except (TypeError, ValueError):
                continue
            if amp_db <= threshold_db:
                continue
            frame_idx = int(frame_idx)
            if n_frames and (frame_idx < 0 or frame_idx >= n_frames):
                continue
            time_sec = float(times[frame_idx]) if n_frames else 0.0
            time_beats = time_sec * beats_per_sec
            in_range = any(
                start_b <= time_beats <= end_b
                for (start_b, end_b) in ranges
            )
            if not in_range:
                continue
            existing = candidates.get(frame_idx)
            if existing is None or float(amp_db) > existing[1]:
                candidates[frame_idx] = (float(freq_hz), float(amp_db))

    return sorted(
        ((f, fr, amp) for f, (fr, amp) in candidates.items()),
        key=lambda item: item[0],
    )


def _build_peak_following_curves(
    active_frames,
    n_frames: int,
    fallback_freq_hz: float,
    fallback_gain_db: float,
    fallback_q: float,
    target_gain_db: float,
    threshold_db: float,
) -> Tuple[np.ndarray, np.ndarray, np.ndarray]:
    """Build dense per-frame Freq / Gain / Q envelopes for a band in
    peak-following mode.

    Semantics:
        - ``freq_curve`` starts at ``fallback_freq_hz`` everywhere. On
          active frames it takes the peak's actual frequency. Between
          two active frames it is FORWARD-FILLED with the last active
          freq so Ableton never sees a jump back to the fallback while
          the band is cutting.
        - ``gain_curve`` starts at ``fallback_gain_db`` (typically 0,
          i.e. band neutral). On active frames it takes
          :func:`_scale_gain_by_amplitude` of the peak's amp. Between
          active frames it STAYS at fallback — the band releases
          between peaks.
        - ``q_curve`` is fully populated via :func:`_scale_q_by_peak_width`
          on active frames and stays at ``fallback_q`` elsewhere. The
          Q stub currently returns ``fallback_q`` everywhere, so in
          practice the array is constant until the adaptive-Q logic
          lands.

    The forward-fill of ``freq_curve`` only kicks in AFTER the first
    active frame. Frames before the very first peak keep the
    ``fallback_freq_hz`` so Ableton's pre-song default sits at a sane
    value.

    Args:
        active_frames: Output of :func:`_collect_active_peak_frames`.
        n_frames: Length of the output arrays (typically
            ``len(times_sec)``).
        fallback_freq_hz: Freq value outside active peaks (usually the
            cluster centroid).
        fallback_gain_db: Gain value outside active peaks (usually 0).
        fallback_q: Q value outside active peaks (usually the cluster
            median Q).
        target_gain_db: Cluster-level max cut depth passed through to
            :func:`_scale_gain_by_amplitude`.
        threshold_db: Amplitude floor — same semantics as
            :func:`_scale_gain_by_amplitude`.

    Returns:
        ``(freq_curve, gain_curve, q_curve)`` numpy arrays, each of
        length ``n_frames``.
    """
    freq_curve = np.full(n_frames, float(fallback_freq_hz), dtype=float)
    gain_curve = np.full(n_frames, float(fallback_gain_db), dtype=float)
    q_curve = np.full(n_frames, float(fallback_q), dtype=float)

    active_indices: set = set()
    for frame_idx, freq_hz, amp_db in active_frames or []:
        if not (0 <= frame_idx < n_frames):
            continue
        active_indices.add(frame_idx)
        freq_curve[frame_idx] = float(freq_hz)
        gain_curve[frame_idx] = _scale_gain_by_amplitude(
            target_gain_db, amp_db, threshold_db,
        )
        q_curve[frame_idx] = _scale_q_by_peak_width(
            None, fallback_q, freq_hz,
        )

    # Forward-fill the Freq curve only — Gain and Q stay at their
    # fallback values outside active frames so the band is musically
    # neutral when no peak is present.
    if active_indices:
        last_active_freq: Optional[float] = None
        for i in range(n_frames):
            if i in active_indices:
                last_active_freq = float(freq_curve[i])
            elif last_active_freq is not None:
                freq_curve[i] = last_active_freq

    return freq_curve, gain_curve, q_curve


# ---------------------------------------------------------------------------
# Public API — section-locked writer
# ---------------------------------------------------------------------------

def write_dynamic_eq8_from_cde_diagnostics(
    als_path,
    diagnostics: List[CorrectionDiagnostic],
    *,
    peak_trajectories_by_track: Optional[Dict[str, List[PeakTrajectory]]] = None,
    zone_energy_by_track: Optional[Dict[str, Dict[str, np.ndarray]]] = None,
    times: Optional[np.ndarray] = None,
    device_user_name_fmt: str = "CDE Correction ({n} bands)",
    insertion_anchor_username: Optional[str] = "Peak Resonance",
    freq_match_tolerance_semitones: float = DEFAULT_FREQ_CLUSTER_TOLERANCE_SEMITONES,
    peak_follow: bool = False,
    threshold_db: float = -30.0,
    use_fallback_when_primary_none: bool = False,
    dry_run: bool = False,
    automation_map=None,
    grid_spacing_sec: float = DEFAULT_GRID_SPACING_SEC,
    _skip_confirmation: bool = False,
    _preview_writer: Callable[[str], None] = print,
) -> CdeApplicationReport:
    """Apply a batch of CDE diagnostics to ``.als`` as dynamic EQ8 corrections.

    F1b — section-locked mode only (``peak_follow=False``). The
    ``peak_follow=True`` path will ship in F1c after F1b.5 field
    validation; the kwarg already exists for forward compatibility but
    raises ``NotImplementedError`` when set to True in this slice.

    Per target track, diagnostics are clustered by frequency
    (tolerance ``freq_match_tolerance_semitones``); each cluster gets
    its own EQ8 band in a single ``CDE Correction (n bands)`` device
    inserted right after ``insertion_anchor_username`` (defaults to
    the existing ``Peak Resonance`` corrective EQ).

    Each band is configured Bell, ``cluster.centroid_hz``, ``cluster.median_q``,
    ``IsOn=true``, with a manual gain of 0 dB. The Gain envelope is
    a dense per-frame curve (``grid_spacing_sec`` apart) that holds
    ``cluster.severest_gain_db`` whenever the time falls inside any
    section listed by ``cluster.applies_to_sections`` and 0 dB
    everywhere else — same pattern as the production
    ``write_section_aware_eq``.

    Args:
        als_path: Path to the ``.als`` to mutate.
        diagnostics: Source diagnostics (typically
            ``json.load(<project>_diagnostics.json)`` parsed back to
            ``CorrectionDiagnostic`` instances).
        peak_trajectories_by_track: Reserved for F1c (``peak_follow=True``).
        zone_energy_by_track: Reserved for F1c.
        times: Reserved for F1c. F1b builds its own dense time grid.
        device_user_name_fmt: Format string with ``{n}`` placeholder
            for the band count. Used as the device's ``UserName``.
        insertion_anchor_username: Insert the new EQ8 right after the
            track's device with this UserName. ``None`` → append to
            the end of the chain.
        freq_match_tolerance_semitones: Clustering tolerance.
        peak_follow: Set ``True`` to enable peak-following automation
            (F1c). Raises ``NotImplementedError`` in this slice.
        threshold_db: F1c parameter, ignored in F1b.
        use_fallback_when_primary_none: When ``True``, also process
            diagnostics whose ``primary_correction`` is ``None`` by
            falling back to ``fallback_correction`` if available.
            ``False`` (default) skips them.
        dry_run: Print the preview, return the report, do not write.
        automation_map: Optional ``TrackAutomationMap`` for audibility
            masking — passed through to ``_feature_to_breakpoints``
            indirectly when implemented.
        grid_spacing_sec: Dense-sampling grid step. Default 0.5 s
            matches production Peak Resonance density.
        _skip_confirmation: Internal — bypass the interactive prompt
            for tests and CLI ``--yes``.
        _preview_writer: Callable used to print the preview. Tests
            override with a list collector.

    Returns:
        :class:`CdeApplicationReport` with ``applied`` / ``skipped`` /
        ``warnings`` / ``devices_created`` / ``envelopes_written`` /
        ``backup_path`` populated.
    """
    if peak_follow:
        raise NotImplementedError(
            "peak_follow=True ships in F1c after F1b.5 field validation."
        )
    if use_fallback_when_primary_none:
        raise NotImplementedError(
            "use_fallback_when_primary_none not yet supported in F1b."
        )

    als_path = Path(als_path)
    report = CdeApplicationReport()

    if not als_path.exists():
        report.warnings.append(f"als_path not found: {als_path}")
        return report

    # Parse once. Locator + tempo extracted before validation so the
    # preview can show what will happen without writing anything.
    tree = parse_als(str(als_path))
    tempo = _extract_tempo(tree)
    locators = read_locators(str(als_path))
    section_ranges = _build_section_ranges_beats(locators)
    if not section_ranges:
        report.warnings.append(
            "No locators found — section-locked mode requires section "
            "boundaries. Aborting without write."
        )
        return report

    # Reciprocal validation + expansion before grouping.
    diagnostics = _validate_reciprocal_pairs(tree, list(diagnostics), report)
    diagnostics = _expand_reciprocal_cuts(diagnostics)

    # Group + cluster.
    grouped = _group_diagnostics_by_target(diagnostics, report=report)
    clusters_by_track: Dict[str, List[FreqCluster]] = {}
    for track_name, diags in grouped.items():
        clusters = _cluster_diagnostics_by_frequency(
            diags, tolerance_semitones=freq_match_tolerance_semitones,
        )
        clusters = _cap_clusters(clusters, MAX_CDE_BANDS, track_name, report)
        if clusters:
            clusters_by_track[track_name] = clusters

    # Preview — always shown, even in dry_run / _skip_confirmation.
    backup_hint = als_path.with_suffix(".als.v24.bak")
    preview = _build_preview_text(
        als_path, clusters_by_track, report, backup_hint,
    )
    _preview_writer(preview)

    if dry_run:
        report.warnings.append("dry_run=True — no write performed")
        return report

    if not _skip_confirmation:
        if not _prompt_confirmation():
            report.warnings.append("user cancelled at preview prompt")
            return report

    # Backup + write.
    report.backup_path = backup_als(str(als_path))

    song_end_beats = _estimate_song_end_beats(locators)
    beats_per_sec = tempo / 60.0
    song_end_sec = song_end_beats / beats_per_sec
    n_frames = int(np.ceil(song_end_sec / grid_spacing_sec)) + 1
    times_sec = np.arange(n_frames) * grid_spacing_sec

    for track_name, clusters in clusters_by_track.items():
        if not clusters:
            continue
        try:
            track = find_track_by_name(tree, track_name)
        except ValueError:
            for c in clusters:
                for d in c.diagnostics:
                    report.skipped.append((
                        d.diagnostic_id,
                        f"track not found in project: {track_name!r}",
                    ))
            continue

        # Self-heal: purge any prior CDE Correction EQ8 + envelopes.
        _purge_existing_cde_eq8(track, device_user_name_fmt, report)

        # Build + place the new EQ8.
        device_name = device_user_name_fmt.format(n=len(clusters))
        eq8 = _build_fresh_eq8(tree, user_name=device_name)
        devices = _devices_container(track)
        anchor = _find_device_by_username(track, insertion_anchor_username)
        _insert_after(devices, anchor, eq8)
        report.devices_created[track_name] = device_name

        # Configure each cluster on a dedicated band (1, 2, 3, …, 6).
        next_id_counter = [get_next_id(tree)]
        for i, cluster in enumerate(clusters):
            band_index = i + 1  # band 0 reserved for HPF convention
            band_param = get_eq8_band(eq8, band_index)
            configure_eq8_band(
                band_param,
                mode=3,
                freq=cluster.centroid_hz,
                gain=0.0,
                q=cluster.median_q,
            )
            ison = band_param.find("IsOn/Manual")
            if ison is not None:
                ison.set("Value", "true")

            gain_curve = _build_section_locked_gain_curve(
                cluster, section_ranges, times_sec, tempo,
            )
            gain_bps = _feature_to_breakpoints(gain_curve, times_sec, tempo)
            _write_validated_env(
                track, band_param, "Gain", gain_bps, next_id_counter,
            )
            report.envelopes_written += 1

            for d in cluster.diagnostics:
                if d.diagnostic_id not in report.applied:
                    report.applied.append(d.diagnostic_id)

    save_als_from_tree(tree, str(als_path))
    return report
