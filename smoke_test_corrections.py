#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""Smoke test — Premières corrections EQ8 sur Acid Drops.

Applique les étapes 1-5 du smoke_test_prompt.md sur une COPIE du .als :
  1. Safety HPF 30 Hz sur les tracks sans contenu sub
  2. HPF adaptatif sur Toms Rack
  3. Notch dynamique 248 Hz sur Toms Rack
  4. Notch dynamique 248 Hz sur ARP Glitter Box
  5. Détection de masking Kick <-> Sub Bass (pas d'écriture)

Usage : python smoke_test_corrections.py
"""

from __future__ import annotations

import shutil
import sys
from pathlib import Path

import numpy as np
import openpyxl

from als_utils import parse_als
import eq8_automation
from eq8_automation import (
    EQ8SlotFullError,
    TrackNotFoundError,
    detect_masking,
    write_adaptive_hpf,
    write_dynamic_notch,
    write_safety_hpf,
)
from spectral_evolution import PeakTrajectory, ZONE_RANGES

# Bump à chaque essai pour identifier l'.als généré.
# Apparaît dans le nom du fichier ET dans le UserName des EQ8 nouvellement clonés.
SMOKE_TEST_BUILD = "v4-NextPointeeId-bump"

REPO_ROOT = Path(__file__).resolve().parent
PROJECTS = REPO_ROOT / "ableton" / "projects"
ALS_SRC = PROJECTS / "Acid_drops_Code_P14.als"
ALS_DST = PROJECTS / f"Acid_drops_SMOKE_TEST_{SMOKE_TEST_BUILD}.als"
EXCEL = PROJECTS / "Acid_drops_MixAnalyzer_2026-04-16_19-56.xlsx"

SAFETY_TARGETS = [
    "China",
    "Tambourine Hi-Hat",
    "Clap",
    "Guitar PM A",
    "Guitar PM B",
    "Xylo Percussion",
    "Xylo Texture",
    "Harmony Vocal Female",
    "Lead Vocal Hey",
    "Lead Vocal Shhh",
    "Voice FX Dark Wihispers",
]

ZONE_ORDER = list(ZONE_RANGES.keys())
DESC_ORDER = ["centroid", "spread", "flatness", "low_rolloff", "high_rolloff"]
FRAMES_PER_SEC = 6


def parse_bucket(label: str) -> tuple[float, float]:
    label = label.rstrip("s")
    a, b = label.split("-")
    return float(a), float(b)


def load_zone_energy(ws):
    data = {}
    rows = list(ws.iter_rows(values_only=True))
    block = 1 + len(ZONE_ORDER)
    for i in range(0, len(rows), block):
        header = rows[i]
        if not header or not header[0]:
            continue
        track = header[0]
        bucket_labels = [c for c in header[1:] if c]
        times = np.array(
            [(a + b) / 2 for a, b in (parse_bucket(x) for x in bucket_labels)]
        )
        zones = {}
        for j, zone in enumerate(ZONE_ORDER):
            if i + 1 + j >= len(rows):
                break
            zr = rows[i + 1 + j]
            vals = np.array(
                [v if v is not None else -120.0 for v in zr[1 : 1 + len(times)]],
                dtype=float,
            )
            zones[zone] = vals
        data[track] = {"zones": zones, "times": times}
    return data


def load_descriptors(ws):
    data = {}
    rows = list(ws.iter_rows(values_only=True))
    block = 1 + len(DESC_ORDER)
    for i in range(0, len(rows), block):
        header = rows[i]
        if not header or not header[0]:
            continue
        track = header[0]
        bucket_labels = [c for c in header[1:] if c]
        times = np.array(
            [(a + b) / 2 for a, b in (parse_bucket(x) for x in bucket_labels)]
        )
        d = {"times": times}
        for j, name in enumerate(DESC_ORDER):
            if i + 1 + j >= len(rows):
                break
            dr = rows[i + 1 + j]
            vals = np.array(
                [v if v is not None else 0.0 for v in dr[1 : 1 + len(times)]],
                dtype=float,
            )
            d[name] = vals
        data[track] = d
    return data


def load_peak_trajectories(ws):
    data: dict[str, list[PeakTrajectory]] = {}
    cursor: dict[tuple[str, int], PeakTrajectory] = {}
    rows = list(ws.iter_rows(values_only=True))
    for r in rows[1:]:
        if not r or not r[0]:
            continue
        track, traj_idx, frame, _t, freq, amp = r[:6]
        key = (track, int(traj_idx))
        if key not in cursor:
            cursor[key] = PeakTrajectory(points=[])
        cursor[key].points.append((int(frame), float(freq), float(amp)))
    for (track, _), pt in cursor.items():
        data.setdefault(track, []).append(pt)
    return data


def find_excel_track(lookup: dict, als_name: str) -> str | None:
    direct = f"Acid_drops {als_name}.wav"
    if direct in lookup:
        return direct
    for label in lookup:
        if als_name.lower() in label.lower():
            return label
    return None


def pick_peak_near(
    trajectories: list[PeakTrajectory], target_hz: float, tol: float = 0.15
) -> PeakTrajectory | None:
    best, best_dist = None, float("inf")
    for tr in trajectories:
        d = abs(tr.mean_freq - target_hz)
        if d / target_hz < tol and d < best_dist:
            best, best_dist = tr, d
    return best


def main() -> int:
    if not ALS_SRC.exists():
        print(f"ERR: {ALS_SRC} introuvable", file=sys.stderr)
        return 1
    if not EXCEL.exists():
        print(f"ERR: {EXCEL} introuvable", file=sys.stderr)
        return 1

    print(f"=== Smoke test build: {SMOKE_TEST_BUILD} ===")
    print(f"Copie {ALS_SRC.name} -> {ALS_DST.name}")
    shutil.copy2(ALS_SRC, ALS_DST)

    # Tag les EQ8 nouvellement créés avec la build pour les retrouver dans Ableton.
    eq8_automation.NEW_EQ8_USER_NAME = f"MixAnalyzer SMOKE {SMOKE_TEST_BUILD}"

    print(f"Chargement rapport Excel : {EXCEL.name}")
    wb = openpyxl.load_workbook(EXCEL, read_only=True, data_only=True)
    zone_data = load_zone_energy(wb["_track_zone_energy"])
    desc_data = load_descriptors(wb["_track_spectral_descriptors"])
    peaks_data = load_peak_trajectories(wb["_track_peak_trajectories"])
    print(f"  zone_energy: {len(zone_data)} tracks")
    print(f"  spectral_descriptors: {len(desc_data)} tracks")
    print(f"  peak_trajectories: {len(peaks_data)} tracks")

    band_usage: dict[str, set[int]] = {}
    errors: list[str] = []

    # ------------------------------------------------------------------
    # Étape 1 — Safety HPF 30 Hz
    # ------------------------------------------------------------------
    print("\n=== Étape 1 — Safety HPF 30 Hz ===")
    for als_name in SAFETY_TARGETS:
        label = find_excel_track(zone_data, als_name)
        if label is None:
            print(f"[{als_name}] SKIP : pas trouvé dans le rapport Excel")
            continue
        zd = zone_data[label]
        sub = zd["zones"].get("sub")
        if sub is None or len(sub) == 0:
            print(f"[{als_name}] SKIP : pas de zone_energy sub")
            continue
        mean_sub = float(np.mean(sub))
        try:
            report = write_safety_hpf(
                als_path=str(ALS_DST),
                track_id=als_name,
                sub_energy=sub,
                times=zd["times"],
                threshold_db=-30,
                band_index=0,
            )
            print(
                f"[{als_name}] mean_sub={mean_sub:.1f}dB — "
                f"success={report.success} bps={report.breakpoints_written} "
                f"band={report.eq8_band_index}"
            )
            band_usage.setdefault(als_name, set()).add(report.eq8_band_index)
        except (ValueError, TrackNotFoundError, EQ8SlotFullError) as e:
            msg = f"[{als_name}] ERREUR: {type(e).__name__}: {e}"
            print(msg)
            errors.append(msg)

    # ------------------------------------------------------------------
    # Étape 2 — HPF adaptatif sur Toms Rack
    # ------------------------------------------------------------------
    print("\n=== Étape 2 — HPF adaptatif sur Toms Rack ===")
    toms_als = "Toms Rack"
    toms_label = find_excel_track(desc_data, toms_als)
    if toms_label is None:
        msg = f"[{toms_als}] SKIP : pas de spectral_descriptors"
        print(msg)
        errors.append(msg)
    else:
        dd = desc_data[toms_label]
        try:
            report = write_adaptive_hpf(
                als_path=str(ALS_DST),
                track_id=toms_als,
                low_rolloff_curve=dd["low_rolloff"],
                times=dd["times"],
                valley_trajectories=[],
                safety_hz=10.0,
                band_index=0,
            )
            print(
                f"[{toms_als}] success={report.success} "
                f"bps={report.breakpoints_written} "
                f"band={report.eq8_band_index}"
            )
            band_usage.setdefault(toms_als, set()).add(report.eq8_band_index)
        except (ValueError, TrackNotFoundError, EQ8SlotFullError) as e:
            msg = f"[{toms_als}] ERREUR: {type(e).__name__}: {e}"
            print(msg)
            errors.append(msg)

    # ------------------------------------------------------------------
    # Étape 3 — Notch 248 Hz sur Toms Rack
    # ------------------------------------------------------------------
    print("\n=== Étape 3 — Notch dynamique 248 Hz sur Toms Rack ===")
    toms_peaks_label = find_excel_track(peaks_data, toms_als)
    peak_248_toms = None
    if toms_peaks_label:
        peak_248_toms = pick_peak_near(peaks_data[toms_peaks_label], 248.0)
    if peak_248_toms is None:
        msg = (
            f"[{toms_als}] SKIP : aucun peak à ±15% de 248 Hz "
            "dans le rapport (résonance non dominante ici)"
        )
        print(msg)
    else:
        max_frame = max(p[0] for p in peak_248_toms.points)
        times_frames = np.arange(max_frame + 10) / FRAMES_PER_SEC
        try:
            report = write_dynamic_notch(
                als_path=str(ALS_DST),
                track_id=toms_als,
                peak_trajectory=peak_248_toms,
                times=times_frames,
                reduction_db=-3.0,
                proportional=True,
                threshold_db=-40,
                q=8.0,
                band_index=1,
            )
            print(
                f"[{toms_als}] peak mean_freq={peak_248_toms.mean_freq:.1f}Hz "
                f"success={report.success} bps={report.breakpoints_written} "
                f"band={report.eq8_band_index}"
            )
            band_usage.setdefault(toms_als, set()).add(report.eq8_band_index)
        except (ValueError, TrackNotFoundError, EQ8SlotFullError) as e:
            msg = f"[{toms_als}] ERREUR: {type(e).__name__}: {e}"
            print(msg)
            errors.append(msg)

    # ------------------------------------------------------------------
    # Étape 4 — Notch 248 Hz sur ARP Glitter Box
    # ------------------------------------------------------------------
    print("\n=== Étape 4 — Notch dynamique 248 Hz sur ARP Glitter Box ===")
    glitter_als = "ARP Glitter Box"
    glitter_label = find_excel_track(peaks_data, glitter_als)
    peak_248_glitter = None
    if glitter_label:
        peak_248_glitter = pick_peak_near(peaks_data[glitter_label], 248.0)
    if peak_248_glitter is None:
        msg = f"[{glitter_als}] SKIP : aucun peak à ±15% de 248 Hz"
        print(msg)
    else:
        max_frame = max(p[0] for p in peak_248_glitter.points)
        times_frames = np.arange(max_frame + 10) / FRAMES_PER_SEC
        try:
            report = write_dynamic_notch(
                als_path=str(ALS_DST),
                track_id=glitter_als,
                peak_trajectory=peak_248_glitter,
                times=times_frames,
                reduction_db=-2.0,
                proportional=True,
                threshold_db=-40,
                q=8.0,
                band_index=1,
            )
            print(
                f"[{glitter_als}] peak mean_freq={peak_248_glitter.mean_freq:.1f}Hz "
                f"success={report.success} bps={report.breakpoints_written} "
                f"band={report.eq8_band_index}"
            )
            band_usage.setdefault(glitter_als, set()).add(report.eq8_band_index)
        except (ValueError, TrackNotFoundError, EQ8SlotFullError) as e:
            msg = f"[{glitter_als}] ERREUR: {type(e).__name__}: {e}"
            print(msg)
            errors.append(msg)

    # ------------------------------------------------------------------
    # Étape 5 — Masking Kick <-> Sub Bass
    # ------------------------------------------------------------------
    print("\n=== Étape 5 — Masking Kick 1 ↔ Sub Bass ===")
    kick_label = find_excel_track(zone_data, "Kick 1")
    sub_label = find_excel_track(zone_data, "Sub Bass")
    if kick_label and sub_label:
        kd = zone_data[kick_label]
        sd = zone_data[sub_label]
        masking = detect_masking(
            track_a_zone_energy=kd["zones"],
            track_b_zone_energy=sd["zones"],
            times=kd["times"],
            zones=["low", "mud"],
        )
        print(f"severity = {masking.severity:.3f}")
        print(f"zones affectées : {masking.zones}")
        for z, s in masking.scores.items():
            print(
                f"  {z}: mean={float(np.mean(s)):.3f} "
                f"max={float(np.max(s)):.3f}"
            )
    else:
        print(f"SKIP : kick_label={kick_label} sub_label={sub_label}")

    # ------------------------------------------------------------------
    # Validation finale
    # ------------------------------------------------------------------
    print("\n=== Récapitulatif ===")
    print(f"{'Track':40s} bandes EQ8 utilisées")
    for track, bands in sorted(band_usage.items()):
        b = ", ".join(str(x) for x in sorted(bands))
        print(f"  {track:38s} [{b}]")

    bak = ALS_DST.parent / f"{ALS_DST.name}.v24.bak"
    print(f"\nBackup présent : {bak.exists()} ({bak.name})")

    tree = parse_als(str(ALS_DST))
    root = tree.getroot()
    eq8_count = len(root.findall(".//Eq8"))
    env_count = len(root.findall(".//AutomationEnvelope"))
    bp_count = len(root.findall(".//FloatEvent"))
    print(f".als modifié — EQ8={eq8_count} envelopes={env_count} breakpoints={bp_count}")

    if errors:
        print(f"\n⚠ {len(errors)} erreurs survenues :")
        for e in errors:
            print(f"  {e}")
    else:
        print("\n✓ Aucune erreur.")

    return 0


if __name__ == "__main__":
    sys.exit(main())
